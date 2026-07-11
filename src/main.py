"""
Main entry point for ReportSoft - Consolidados.
Application for consolidating, tracking, and classifying UTP research groups.
"""

import json
import os
import sys
import warnings

from pathlib import Path
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QFont
from PyQt5.QtWidgets import (
    QApplication, QHBoxLayout, QLabel, QMainWindow, QMessageBox,
    QProgressBar, QPushButton, QTabWidget, QTableWidgetItem,
    QTextEdit, QVBoxLayout, QWidget,
)

from database import DatabaseManager
from loader import CargadorDatosIntegrado
from views.vista_clasificacion_minciencias import VistaClasificacionMinCiencias
from views.vista_seguimiento_grupos import VistaSeguimientoGrupos
from views.vista_visor_gruplac_957 import VisorGrupLAC957

warnings.filterwarnings("ignore")


# =============================================================================
# VISTA BÚSQUEDA
# =============================================================================

class VistaBusqueda(QWidget):
    def __init__(self, db):
        super().__init__()
        self.db = db
        self.setup_ui()

    def setup_ui(self):
        from PyQt5.QtWidgets import (
            QAbstractItemView, QCheckBox, QComboBox, QHeaderView,
            QLineEdit, QPushButton, QSplitter, QTableWidget, QTableWidgetItem,
        )
        layout = QHBoxLayout()
        panel_izq = QWidget()
        layout_izq = QVBoxLayout(panel_izq)
        self.input_busqueda = QLineEdit()
        self.input_busqueda.setPlaceholderText("Buscar por nombre o cédula...")
        self.input_busqueda.returnPressed.connect(self.buscar)
        btn_buscar = QPushButton("Buscar")
        btn_buscar.clicked.connect(self.buscar)
        btn_consolidar = QPushButton("Consolidar Duplicados")
        btn_consolidar.clicked.connect(self.consolidar_duplicados_manual)
        btn_consolidar.setStyleSheet("background-color: #e67e22; color: white;")
        self.tabla_resultados = QTableWidget()
        self.tabla_resultados.setColumnCount(5)
        self.tabla_resultados.setHorizontalHeaderLabels(['Nombre', 'Cédula', 'Email', 'Facultad', 'Tipo'])
        self.tabla_resultados.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tabla_resultados.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tabla_resultados.horizontalHeader().setStretchLastSection(True)
        self.tabla_resultados.doubleClicked.connect(self.mostrar_detalle)
        layout_izq.addWidget(QLabel("Búsqueda de Personas"))
        layout_izq.addWidget(self.input_busqueda)
        layout_izq.addWidget(btn_buscar)
        layout_izq.addWidget(btn_consolidar)
        layout_izq.addWidget(self.tabla_resultados)
        panel_der = QWidget()
        layout_der = QVBoxLayout(panel_der)
        self.texto_detalle = QTextEdit()
        self.texto_detalle.setReadOnly(True)
        layout_der.addWidget(QLabel("Detalle de Persona"))
        layout_der.addWidget(self.texto_detalle)
        splitter = QSplitter(Qt.Horizontal)
        splitter.addWidget(panel_izq)
        splitter.addWidget(panel_der)
        splitter.setSizes([400, 600])
        layout.addWidget(splitter)
        self.setLayout(layout)

    def consolidar_duplicados_manual(self):
        consolidaciones = self.db.consolidar_duplicados()
        if consolidaciones:
            mensaje = f"✓ Se consolidaron {len(consolidaciones)} persona(s):\n\n"
            for grupo in consolidaciones:
                mensaje += f"- {grupo['nombre']}\n"
                mensaje += f"  Cédula principal: {grupo['cedula_principal']}\n"
                mensaje += f"  Cédulas consolidadas: {', '.join(grupo['cedulas'])}\n\n"
            QMessageBox.information(self, "Consolidación Exitosa", mensaje)
            if self.input_busqueda.text():
                self.buscar()
        else:
            QMessageBox.information(self, "Sin Duplicados", "No se encontraron duplicados para consolidar.")

    def buscar(self):
        from PyQt5.QtWidgets import QTableWidgetItem
        termino = self.input_busqueda.text().strip()
        if not termino:
            return
        resultados = self.db.buscar_personas(termino)
        self.tabla_resultados.setRowCount(len(resultados))
        for row, persona in enumerate(resultados):
            self.tabla_resultados.setItem(row, 0, QTableWidgetItem(persona[1] or ''))
            self.tabla_resultados.setItem(row, 1, QTableWidgetItem(persona[0] or ''))
            self.tabla_resultados.setItem(row, 2, QTableWidgetItem(persona[2] or ''))
            self.tabla_resultados.setItem(row, 3, QTableWidgetItem(persona[3] or ''))
            self.tabla_resultados.setItem(row, 4, QTableWidgetItem(persona[4] or ''))

    def mostrar_detalle(self, index):
        row = index.row()
        cedula = self.tabla_resultados.item(row, 1).text()
        detalle = self.db.obtener_detalle_persona(cedula)
        if not detalle or "info" not in detalle:
            return
        info = detalle["info"]

        def _s(v):
            return str(v) if v else "—"

        def _seccion(titulo, color, items_html):
            return (
                f"<div style='margin-top:10px;'>"
                f"<div style='background:{color};color:white;padding:4px 8px;"
                f"border-radius:4px;font-weight:bold;font-size:12px;'>{titulo}</div>"
                f"<div style='padding:4px 8px;'>{items_html}</div></div>"
            )

        html = (
            "<html><body style='font-family:Arial,sans-serif;font-size:11px;"
            "max-width:100%;word-wrap:break-word;'>"
            "<div style='background:#1a365d;color:white;padding:8px;border-radius:6px;"
            "margin-bottom:10px;'>"
            f"<b style='font-size:13px;'>{_s(info['nombre'])}</b><br>"
            f"<span style='font-size:10px;'>Cédula: {_s(info['cedula'])} | "
            f"Facultad: {_s(info['facultad'])}<br>"
            f"Email: {_s(info['email'])} | Tipo: {_s(info['tipo'])}</span></div>"
        )

        if detalle.get("grupos"):
            items = "".join(
                f"<p style='margin:2px 0;'>&#9679; <b>{_s(g[2])}</b>"
                f" <span style='color:#666;'>({_s(g[4])})</span></p>"
                for g in detalle["grupos"]
            )
            html += _seccion(
                f"Grupos de investigación ({len(detalle['grupos'])})",
                "#2c3e50", items
            )

        if detalle.get("publicaciones"):
            items = "".join(
                f"<p style='margin:2px 0;border-bottom:1px solid #eee;padding-bottom:2px;'>"
                f"&#9679; <b>{_s(pub[2])[:120]}</b><br>"
                f"<span style='color:#555;'>Año: {_s(pub[5])} | Tipo: {_s(pub[6])} | "
                f"Revista: {_s(pub[3])[:60]}</span></p>"
                for pub in detalle["publicaciones"]
            )
            html += _seccion(
                f"Publicaciones ({len(detalle['publicaciones'])})",
                "#2e86ab", items
            )

        if detalle.get("extensiones"):
            items = "".join(
                f"<p style='margin:2px 0;'>&#9679; <b>{_s(ext[2])[:100]}</b> "
                f"<span style='color:#555;'>({_s(ext[3])} · {_s(ext[8])})</span></p>"
                for ext in detalle["extensiones"]
            )
            html += _seccion(
                f"Extensiones ({len(detalle['extensiones'])})",
                "#f18f01", items
            )

        if detalle.get("trabajos_grado_director"):
            items = "".join(
                f"<p style='margin:2px 0;'>&#9679; <b>{_s(tg[5])[:100]}</b><br>"
                f"<span style='color:#555;'>Est: {_s(tg[4])} | Prog: {_s(tg[6])} | "
                f"Año: {_s(tg[7])}</span></p>"
                for tg in detalle["trabajos_grado_director"]
            )
            html += _seccion(
                f"Trabajos de grado dirigidos ({len(detalle['trabajos_grado_director'])})",
                "#3b8c66", items
            )

        if detalle.get("productos_innovacion"):
            items = "".join(
                f"<p style='margin:2px 0;'>&#9679; <b>{_s(p[3])[:100]}</b> "
                f"<span style='color:#555;'>({_s(p[2])} · {_s(p[5])})</span></p>"
                for p in detalle["productos_innovacion"]
            )
            html += _seccion(
                f"Productos de innovación ({len(detalle['productos_innovacion'])})",
                "#a23b72", items
            )

        if detalle.get("proyectos"):
            items = "".join(
                f"<p style='margin:2px 0;'>&#9679; <b>{_s(pr[3])[:100]}</b> "
                f"<span style='color:#555;'>(Año: {_s(pr[7])} · {_s(pr[10])})</span></p>"
                for pr in detalle["proyectos"]
            )
            html += _seccion(
                f"Proyectos ({len(detalle['proyectos'])})",
                "#c73e1d", items
            )

        html += "</body></html>"
        self.texto_detalle.setHtml(html)


# =============================================================================
# VISTA GRUPOS
# =============================================================================

class VistaGrupos(QWidget):
    def __init__(self, db):
        super().__init__()
        self.db = db
        self.productos_completos = []
        self.setup_ui()

    def setup_ui(self):
        from PyQt5.QtWidgets import (
            QAbstractItemView, QCheckBox, QComboBox, QHeaderView,
            QPushButton, QSplitter, QTableWidget, QTableWidgetItem,
        )
        layout = QVBoxLayout()
        layout.setSpacing(2)
        layout.setContentsMargins(3, 3, 3, 3)

        barra_superior = QWidget()
        barra_superior.setMaximumHeight(35)
        barra_superior.setStyleSheet("background-color: #f8f9fa; border-radius: 3px;")
        layout_barra = QHBoxLayout(barra_superior)
        layout_barra.setContentsMargins(5, 3, 5, 3)
        layout_barra.setSpacing(5)

        layout_barra.addWidget(QLabel("<b>Grupo:</b>"))
        self.combo_grupos = QComboBox()
        self.combo_grupos.setMinimumWidth(200)
        self.combo_grupos.currentTextChanged.connect(self.seleccionar_grupo)
        layout_barra.addWidget(self.combo_grupos)

        btn_refrescar = QPushButton("↻")
        btn_refrescar.setMaximumWidth(30)
        btn_refrescar.setToolTip("Refrescar grupos")
        btn_refrescar.clicked.connect(self.cargar_grupos)
        layout_barra.addWidget(btn_refrescar)
        layout_barra.addWidget(QLabel("|"))

        self.check_pub = QCheckBox("Pub")
        self.check_pub.setChecked(True)
        self.check_pub.stateChanged.connect(self.seleccionar_grupo)
        layout_barra.addWidget(self.check_pub)
        self.check_ext = QCheckBox("Ext")
        self.check_ext.setChecked(True)
        self.check_ext.stateChanged.connect(self.seleccionar_grupo)
        layout_barra.addWidget(self.check_ext)
        self.check_tg = QCheckBox("TG")
        self.check_tg.setChecked(True)
        self.check_tg.stateChanged.connect(self.seleccionar_grupo)
        layout_barra.addWidget(self.check_tg)
        self.check_innov = QCheckBox("Inn")
        self.check_innov.setChecked(True)
        self.check_innov.stateChanged.connect(self.seleccionar_grupo)
        layout_barra.addWidget(self.check_innov)
        self.check_proy = QCheckBox("Proy")
        self.check_proy.setChecked(True)
        self.check_proy.stateChanged.connect(self.seleccionar_grupo)
        layout_barra.addWidget(self.check_proy)
        layout_barra.addStretch()

        btn_excel = QPushButton("Excel")
        btn_excel.setMaximumWidth(70)
        btn_excel.clicked.connect(self.exportar_excel)
        btn_excel.setStyleSheet("background-color: #27ae60; color: white; font-size: 11px;")
        layout_barra.addWidget(btn_excel)

        btn_pdf = QPushButton("PDF")
        btn_pdf.setMaximumWidth(60)
        btn_pdf.clicked.connect(self.exportar_pdf)
        btn_pdf.setStyleSheet("background-color: #e74c3c; color: white; font-size: 11px;")
        layout_barra.addWidget(btn_pdf)

        splitter = QSplitter(Qt.Horizontal)

        panel_integrantes = QWidget()
        layout_int = QVBoxLayout(panel_integrantes)
        layout_int.setSpacing(2)
        layout_int.setContentsMargins(2, 2, 2, 2)
        lbl_int = QLabel("<b>Integrantes</b>")
        lbl_int.setStyleSheet("color: #1a365d; font-size: 12px;")
        layout_int.addWidget(lbl_int)
        self.tabla_integrantes = QTableWidget()
        self.tabla_integrantes.setColumnCount(4)
        self.tabla_integrantes.setHorizontalHeaderLabels(['Nombre', 'Tipo', 'Email', 'Facultad'])
        self.tabla_integrantes.horizontalHeader().setStretchLastSection(True)
        self.tabla_integrantes.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tabla_integrantes.verticalHeader().setVisible(False)
        self.tabla_integrantes.setAlternatingRowColors(True)
        self.tabla_integrantes.setStyleSheet("font-size: 10px;")
        layout_int.addWidget(self.tabla_integrantes)

        panel_productos = QWidget()
        layout_prod = QVBoxLayout(panel_productos)
        layout_prod.setSpacing(2)
        layout_prod.setContentsMargins(2, 2, 2, 2)
        lbl_prod = QLabel("<b>Productos del Grupo</b>")
        lbl_prod.setStyleSheet("color: #1a365d; font-size: 12px;")
        layout_prod.addWidget(lbl_prod)
        self.tabla_productos = QTableWidget()
        self.tabla_productos.setColumnCount(6)
        self.tabla_productos.setHorizontalHeaderLabels([
            'Investigador', 'Título', 'Año', 'Tipo', 'Categoría', 'Estado'
        ])
        self.tabla_productos.horizontalHeader().setStretchLastSection(True)
        self.tabla_productos.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tabla_productos.verticalHeader().setVisible(False)
        self.tabla_productos.setAlternatingRowColors(True)
        self.tabla_productos.setStyleSheet("font-size: 10px;")
        self.tabla_productos.clicked.connect(self.mostrar_detalle_producto)
        layout_prod.addWidget(self.tabla_productos)

        panel_detalle = QWidget()
        layout_det = QVBoxLayout(panel_detalle)
        layout_det.setSpacing(2)
        layout_det.setContentsMargins(2, 2, 2, 2)
        lbl_det = QLabel("<b>Detalle del Producto</b>")
        lbl_det.setStyleSheet("color: #1a365d; font-size: 12px;")
        layout_det.addWidget(lbl_det)
        self.texto_detalle = QTextEdit()
        self.texto_detalle.setReadOnly(True)
        self.texto_detalle.setPlaceholderText("Selecciona un producto para ver detalles...")
        self.texto_detalle.setStyleSheet("font-size: 10px; background-color: #fafafa;")
        layout_det.addWidget(self.texto_detalle)

        splitter.addWidget(panel_integrantes)
        splitter.addWidget(panel_productos)
        splitter.addWidget(panel_detalle)
        splitter.setSizes([280, 750, 370])

        layout.addWidget(barra_superior)
        layout.addWidget(splitter)
        self.setLayout(layout)
        self.cargar_grupos()

    def cargar_grupos(self):
        self.combo_grupos.clear()
        self.combo_grupos.addItem("-- Seleccionar --")
        grupos = self.db.obtener_grupos()
        for grupo in grupos:
            if grupo[0]:
                self.combo_grupos.addItem(grupo[0])

    def seleccionar_grupo(self, grupo=None):
        if grupo is None:
            grupo = self.combo_grupos.currentText()
        if grupo == "-- Seleccionar --" or not grupo:
            self.tabla_integrantes.setRowCount(0)
            self.tabla_productos.setRowCount(0)
            self.texto_detalle.clear()
            return
        integrantes = self.db.obtener_integrantes_grupo(grupo)
        self.tabla_integrantes.setRowCount(len(integrantes))
        for row, persona in enumerate(integrantes):
            self.tabla_integrantes.setItem(row, 0, QTableWidgetItem(persona[1]))
            self.tabla_integrantes.setItem(row, 1, QTableWidgetItem(persona[2] or ''))
            self.tabla_integrantes.setItem(row, 2, QTableWidgetItem(persona[3] or ''))
            self.tabla_integrantes.setItem(row, 3, QTableWidgetItem(persona[4] or ''))
        filtros = []
        if self.check_pub.isChecked():
            filtros.append('Publicaciones')
        if self.check_ext.isChecked():
            filtros.append('Extensiones')
        if self.check_tg.isChecked():
            filtros.append('Trabajos de Grado')
        if self.check_innov.isChecked():
            filtros.append('Productos Innovación')
        if self.check_proy.isChecked():
            filtros.append('Proyectos')
        if not filtros:
            self.tabla_productos.setRowCount(0)
            return
        productos = self.db.obtener_productos_grupo_detallado(grupo, filtros)
        self.productos_completos = productos
        self.tabla_productos.setRowCount(len(productos))
        for row, prod in enumerate(productos):
            item_inv = QTableWidgetItem(prod.get('investigador', ''))
            item_inv.setData(Qt.UserRole, prod)
            self.tabla_productos.setItem(row, 0, item_inv)
            self.tabla_productos.setItem(row, 1, QTableWidgetItem(prod.get('titulo', '')))
            self.tabla_productos.setItem(row, 2, QTableWidgetItem(str(prod.get('año', '')) if prod.get('año') else ''))
            self.tabla_productos.setItem(row, 3, QTableWidgetItem(prod.get('tipo_producto', '')))
            self.tabla_productos.setItem(row, 4, QTableWidgetItem(prod.get('categoria', '')))
            self.tabla_productos.setItem(row, 5, QTableWidgetItem(prod.get('estado', '')))

    def mostrar_detalle_producto(self, index):
        row = index.row()
        item = self.tabla_productos.item(row, 0)
        if not item:
            return
        producto = item.data(Qt.UserRole)
        if not producto:
            return
        tipo = producto.get('tipo_producto', '')
        COLORES_TIPO = {
            'Publicación': '#2e86ab',
            'Extensión': '#f18f01',
            'Trabajo de Grado': '#3b8c66',
            'Innovación': '#a23b72',
            'Proyecto': '#c73e1d',
        }
        color = COLORES_TIPO.get(tipo, '#1a365d')

        def _s(v):
            return str(v).strip() if v and str(v).strip() not in ('', 'None', 'nan') else '—'

        def _fila(label, valor):
            return (
                f"<tr><td style='color:#555;white-space:nowrap;padding:2px 6px 2px 0;"
                f"vertical-align:top;'><b>{label}</b></td>"
                f"<td style='padding:2px 0;word-break:break-word;'>{_s(valor)}</td></tr>"
            )

        html = (
            "<html><body style='font-family:Arial,sans-serif;font-size:11px;"
            "max-width:100%;word-wrap:break-word;margin:0;padding:4px;'>"
            f"<div style='background:{color};color:white;padding:6px 10px;"
            f"border-radius:5px;margin-bottom:8px;'>"
            f"<b style='font-size:12px;'>{tipo.upper()}</b></div>"
            "<table style='width:100%;border-collapse:collapse;'>"
            + _fila("Investigador", producto.get('investigador'))
            + _fila("Cédula", producto.get('cedula'))
            + _fila("Año", producto.get('año'))
            + "</table>"
            f"<div style='background:{color};color:white;padding:3px 8px;"
            f"border-radius:3px;margin:8px 0 4px 0;font-size:10px;"
            f"font-weight:bold;'>INFORMACIÓN DETALLADA</div>"
            "<table style='width:100%;border-collapse:collapse;'>"
        )

        if tipo == 'Publicación':
            html += (
                f"<tr><td colspan='2' style='padding:2px 0;word-break:break-word;'>"
                f"<b>Título:</b> {_s(producto.get('titulo'))}</td></tr>"
                f"<tr><td colspan='2' style='padding:2px 0;word-break:break-word;"
                f"color:#444;'><b>Revista/Libro:</b> {_s(producto.get('revista_libro'))}</td></tr>"
                + _fila("DOI/URL", producto.get('doi_url'))
                + _fila("ISSN/ISBN", producto.get('issn_isbn'))
                + _fila("Categoría", producto.get('categoria'))
                + _fila("Estado", producto.get('estado'))
            )
        elif tipo == 'Extensión':
            html += (
                f"<tr><td colspan='2' style='padding:2px 0;word-break:break-word;'>"
                f"<b>Actividad:</b> {_s(producto.get('titulo'))}</td></tr>"
                + _fila("Tipo", producto.get('tipo'))
                + _fila("Modalidad", producto.get('modalidad'))
                + _fila("Estado", producto.get('estado'))
                + _fila("Fechas", f"{_s(producto.get('fecha_inicio'))} – {_s(producto.get('fecha_fin'))}")
                + _fila("Población", producto.get('poblacion'))
                + _fila("Financiación interna", producto.get('financiacion_interna'))
                + _fila("Financiación externa", producto.get('financiacion_externa'))
            )
        elif tipo == 'Trabajo de Grado':
            html += (
                f"<tr><td colspan='2' style='padding:2px 0;word-break:break-word;'>"
                f"<b>Título:</b> {_s(producto.get('titulo'))}</td></tr>"
                + _fila("Director", producto.get('investigador'))
                + _fila("Estudiante", producto.get('estudiante'))
                + _fila("Cédula estudiante", producto.get('cedula_estudiante'))
                + _fila("Programa", producto.get('programa'))
                + _fila("Estado", producto.get('estado'))
                + _fila("Fecha sustentación", producto.get('fecha_sustentacion'))
                + _fila("Calificación", producto.get('calificacion'))
            )
        elif tipo == 'Innovación':
            html += (
                f"<tr><td colspan='2' style='padding:2px 0;word-break:break-word;'>"
                f"<b>Nombre:</b> {_s(producto.get('titulo'))}</td></tr>"
                + _fila("Tipo detalle", producto.get('tipo_producto_detalle'))
                + f"<tr><td colspan='2' style='padding:2px 0;word-break:break-word;'>"
                f"<b>Descripción:</b> {_s(producto.get('descripcion'))}</td></tr>"
                + _fila("Estado", producto.get('estado'))
                + _fila("Grupo", producto.get('grupo'))
            )
        elif tipo == 'Proyecto':
            html += (
                f"<tr><td colspan='2' style='padding:2px 0;word-break:break-word;'>"
                f"<b>Título:</b> {_s(producto.get('titulo'))}</td></tr>"
                f"<tr><td colspan='2' style='padding:2px 0;word-break:break-word;"
                f"color:#444;'><b>Objetivo:</b> {_s(producto.get('objetivo'))}</td></tr>"
                + _fila("Código CIE", producto.get('codigo_cie'))
                + _fila("Tipo", producto.get('tipo'))
                + _fila("Fechas", f"{_s(producto.get('fecha_inicio'))} – {_s(producto.get('fecha_fin'))}")
                + _fila("Estado", producto.get('estado'))
                + _fila("Valor aprobado", producto.get('valor_aprobado'))
            )

        html += (
            "</table>"
            "<div style='margin-top:8px;padding-top:4px;border-top:1px solid #ddd;"
            f"color:#666;font-size:10px;'>Fuente: {_s(producto.get('fuente'))}</div>"
            "</body></html>"
        )
        self.texto_detalle.setHtml(html)

    def exportar_excel(self):
        from datetime import datetime
        from PyQt5.QtWidgets import QMessageBox
        grupo = self.combo_grupos.currentText()
        if grupo == "-- Seleccionar --" or not grupo:
            QMessageBox.warning(self, "Advertencia", "Selecciona un grupo primero")
            return
        if not self.productos_completos:
            QMessageBox.warning(self, "Advertencia", "No hay productos para exportar")
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            nombre_archivo = f"reporte_{grupo.replace(' ', '_')}_{timestamp}.xlsx"

            wb = openpyxl.Workbook()

            ws_int = wb.active
            ws_int.title = "Integrantes"
            ws_int['A1'] = f"GRUPO: {grupo} - INTEGRANTES"
            ws_int['A1'].font = Font(bold=True, size=14, color="1a365d")
            ws_int.merge_cells('A1:E1')
            encabezados_int = ['Nombre', 'Tipo', 'Cédula', 'Email', 'Facultad']
            for col, enc in enumerate(encabezados_int, 1):
                cell = ws_int.cell(row=2, column=col, value=enc)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            integrantes = self.db.obtener_integrantes_grupo(grupo)
            for row, persona in enumerate(integrantes, 3):
                ws_int.cell(row=row, column=1, value=persona[1])
                ws_int.cell(row=row, column=2, value=persona[2] or '')
                ws_int.cell(row=row, column=3, value=persona[0])
                ws_int.cell(row=row, column=4, value=persona[3] or '')
                ws_int.cell(row=row, column=5, value=persona[4] or '')
            for col in range(1, 6):
                ws_int.column_dimensions[get_column_letter(col)].width = 25

            publicaciones = [p for p in self.productos_completos if p.get('tipo_producto') == 'Publicación']
            if publicaciones:
                ws_pub = wb.create_sheet("Publicaciones")
                ws_pub['A1'] = f"GRUPO: {grupo} - PUBLICACIONES"
                ws_pub['A1'].font = Font(bold=True, size=14, color="1a365d")
                ws_pub.merge_cells('A1:H1')
                headers = ['Investigador', 'Título', 'Revista/Libro', 'Año', 'Tipo', 'Categoría', 'ISSN/ISBN', 'Estado']
                for col, h in enumerate(headers, 1):
                    cell = ws_pub.cell(row=2, column=col, value=h)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                for row, pub in enumerate(publicaciones, 3):
                    ws_pub.cell(row=row, column=1, value=pub.get('investigador', ''))
                    ws_pub.cell(row=row, column=2, value=pub.get('titulo', ''))
                    ws_pub.cell(row=row, column=3, value=pub.get('revista_libro', ''))
                    ws_pub.cell(row=row, column=4, value=pub.get('año', ''))
                    ws_pub.cell(row=row, column=5, value=pub.get('tipo', ''))
                    ws_pub.cell(row=row, column=6, value=pub.get('categoria', ''))
                    ws_pub.cell(row=row, column=7, value=pub.get('issn_isbn', ''))
                    ws_pub.cell(row=row, column=8, value=pub.get('estado', ''))
                for col in range(1, 9):
                    ws_pub.column_dimensions[get_column_letter(col)].width = 20
                ws_pub.column_dimensions['B'].width = 50

            extensiones = [p for p in self.productos_completos if p.get('tipo_producto') == 'Extensión']
            if extensiones:
                ws_ext = wb.create_sheet("Extensiones")
                ws_ext['A1'] = f"GRUPO: {grupo} - EXTENSIONES"
                ws_ext['A1'].font = Font(bold=True, size=14, color="1a365d")
                ws_ext.merge_cells('A1:G1')
                headers = ['Investigador', 'Actividad', 'Tipo', 'Modalidad', 'Año', 'Población', 'Estado']
                for col, h in enumerate(headers, 1):
                    cell = ws_ext.cell(row=2, column=col, value=h)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="e74c3c", end_color="e74c3c", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                for row, ext in enumerate(extensiones, 3):
                    ws_ext.cell(row=row, column=1, value=ext.get('investigador', ''))
                    ws_ext.cell(row=row, column=2, value=ext.get('titulo', ''))
                    ws_ext.cell(row=row, column=3, value=ext.get('tipo', ''))
                    ws_ext.cell(row=row, column=4, value=ext.get('modalidad', ''))
                    ws_ext.cell(row=row, column=5, value=ext.get('año', ''))
                    ws_ext.cell(row=row, column=6, value=ext.get('poblacion', ''))
                    ws_ext.cell(row=row, column=7, value=ext.get('estado', ''))
                for col in range(1, 8):
                    ws_ext.column_dimensions[get_column_letter(col)].width = 20
                ws_ext.column_dimensions['B'].width = 50

            trabajos = [p for p in self.productos_completos if p.get('tipo_producto') == 'Trabajo de Grado']
            if trabajos:
                ws_tg = wb.create_sheet("Trabajos de Grado")
                ws_tg['A1'] = f"GRUPO: {grupo} - TRABAJOS DE GRADO"
                ws_tg['A1'].font = Font(bold=True, size=14, color="1a365d")
                ws_tg.merge_cells('A1:G1')
                headers = ['Director', 'Estudiante', 'Título', 'Programa', 'Año', 'Estado', 'Calificación']
                for col, h in enumerate(headers, 1):
                    cell = ws_tg.cell(row=2, column=col, value=h)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="27ae60", end_color="27ae60", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                for row, tg in enumerate(trabajos, 3):
                    ws_tg.cell(row=row, column=1, value=tg.get('investigador', ''))
                    ws_tg.cell(row=row, column=2, value=tg.get('estudiante', ''))
                    ws_tg.cell(row=row, column=3, value=tg.get('titulo', ''))
                    ws_tg.cell(row=row, column=4, value=tg.get('programa', ''))
                    ws_tg.cell(row=row, column=5, value=tg.get('año', ''))
                    ws_tg.cell(row=row, column=6, value=tg.get('estado', ''))
                    ws_tg.cell(row=row, column=7, value=tg.get('calificacion', ''))
                for col in range(1, 8):
                    ws_tg.column_dimensions[get_column_letter(col)].width = 20
                ws_tg.column_dimensions['C'].width = 50

            productos_inn = [p for p in self.productos_completos if p.get('tipo_producto') == 'Innovación']
            if productos_inn:
                ws_pi = wb.create_sheet("Productos Innovación")
                ws_pi['A1'] = f"GRUPO: {grupo} - PRODUCTOS INNOVACIÓN"
                ws_pi['A1'].font = Font(bold=True, size=14, color="1a365d")
                ws_pi.merge_cells('A1:E1')
                headers = ['Investigador', 'Producto', 'Tipo', 'Año', 'Estado']
                for col, h in enumerate(headers, 1):
                    cell = ws_pi.cell(row=2, column=col, value=h)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="8e44ad", end_color="8e44ad", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                for row, pi in enumerate(productos_inn, 3):
                    ws_pi.cell(row=row, column=1, value=pi.get('investigador', ''))
                    ws_pi.cell(row=row, column=2, value=pi.get('titulo', ''))
                    ws_pi.cell(row=row, column=3, value=pi.get('tipo_producto_detalle', ''))
                    ws_pi.cell(row=row, column=4, value=pi.get('año', ''))
                    ws_pi.cell(row=row, column=5, value=pi.get('estado', ''))
                for col in range(1, 6):
                    ws_pi.column_dimensions[get_column_letter(col)].width = 20
                ws_pi.column_dimensions['B'].width = 50

            proyectos = [p for p in self.productos_completos if p.get('tipo_producto') == 'Proyecto']
            if proyectos:
                ws_pr = wb.create_sheet("Proyectos")
                ws_pr['A1'] = f"GRUPO: {grupo} - PROYECTOS"
                ws_pr['A1'].font = Font(bold=True, size=14, color="1a365d")
                ws_pr.merge_cells('A1:G1')
                headers = ['Investigador', 'Título', 'Tipo', 'Año', 'Estado', 'Código CIE', 'Valor']
                for col, h in enumerate(headers, 1):
                    cell = ws_pr.cell(row=2, column=col, value=h)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="c0392b", end_color="c0392b", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                for row, pr in enumerate(proyectos, 3):
                    ws_pr.cell(row=row, column=1, value=pr.get('investigador', ''))
                    ws_pr.cell(row=row, column=2, value=pr.get('titulo', ''))
                    ws_pr.cell(row=row, column=3, value=pr.get('tipo', ''))
                    ws_pr.cell(row=row, column=4, value=pr.get('año', ''))
                    ws_pr.cell(row=row, column=5, value=pr.get('estado', ''))
                    ws_pr.cell(row=row, column=6, value=pr.get('codigo_cie', ''))
                    ws_pr.cell(row=row, column=7, value=pr.get('valor_aprobado', ''))
                for col in range(1, 8):
                    ws_pr.column_dimensions[get_column_letter(col)].width = 20
                ws_pr.column_dimensions['B'].width = 50

            ws_int.column_dimensions['A'].width = 40
            ruta_guardado = os.path.join(
                obtener_directorio_base_reports(), nombre_archivo
            )
            wb.save(ruta_guardado)
            QMessageBox.information(self, "Exportado", f"Reporte guardado:\n{ruta_guardado}")
        except ImportError:
            QMessageBox.warning(self, "Error", "openpyxl no está instalado")

    def exportar_pdf(self):
        from PyQt5.QtWidgets import QMessageBox
        QMessageBox.information(self, "PDF", "Exportación a PDF no implementada aún")


def obtener_directorio_base_reports():
    from utils import obtener_directorio_base
    reports_dir = obtener_directorio_base() / "reports"
    reports_dir.mkdir(exist_ok=True)
    return reports_dir


# =============================================================================
# VENTANA PRINCIPAL
# =============================================================================

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("ReportSoft - Consolidados UTP")
        self.setGeometry(100, 50, 1400, 850)

        # Cargar configuración guardada
        self.config_file = Path(__file__).parent.parent / "config" / "settings.local.json"

        # Inicializar base de datos
        try:
            self.db = DatabaseManager()
        except FileNotFoundError as e:
            QMessageBox.critical(self, "Error BD", str(e))
            sys.exit(1)

        # Widget central con tabs
        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)
        layout.setSpacing(0)
        layout.setContentsMargins(2, 2, 2, 2)

        self.tabs = QTabWidget()
        layout.addWidget(self.tabs)

        # Crear pestañas
        self.tab_busqueda = VistaBusqueda(self.db)
        self.tabs.addTab(self.tab_busqueda, "🔍 Búsqueda")

        self.tab_grupos = VistaGrupos(self.db)
        self.tabs.addTab(self.tab_grupos, "👥 Grupos")

        self.tab_clasificacion = VistaClasificacionMinCiencias(self.db)
        self.tabs.addTab(self.tab_clasificacion, "📊 Clasificación 957")

        self.tab_seguimiento = VistaSeguimientoGrupos(self.db)
        self.tabs.addTab(self.tab_seguimiento, "📋 Seguimiento")

        self.tab_visor = VisorGrupLAC957()
        self.tabs.addTab(self.tab_visor, "👁 Visor GrupLAC")

        # Barra de progreso
        self.barra_progreso = QProgressBar()
        self.barra_progreso.setMaximum(0)
        self.barra_progreso.setVisible(False)
        layout.addWidget(self.barra_progreso)

        self.label_estado = QLabel("Listo")
        layout.addWidget(self.label_estado)

        # Botón de recarga de datos
        btn_recargar = QPushButton("🔄 Recargar datos")
        btn_recargar.clicked.connect(self.iniciar_carga_datos)
        layout.addWidget(btn_recargar)

        # Iniciar carga de datos automática
        self.iniciar_carga_datos()

    def iniciar_carga_datos(self):
        self.cargador = CargadorDatosIntegrado(self.db)
        self.cargador.progreso.connect(self.actualizar_progreso)
        self.cargador.finalizado.connect(self.carga_finalizada)
        self.cargador.duplicados_consolidados.connect(self.mostrar_consolidaciones)
        self.barra_progreso.setVisible(True)
        self.barra_progreso.setMaximum(0)
        self.label_estado.setText("Cargando datos...")
        self.cargador.start()

    def actualizar_progreso(self, mensaje):
        self.label_estado.setText(mensaje)

    def carga_finalizada(self, stats):
        self.barra_progreso.setVisible(False)
        self.label_estado.setText(f"Datos cargados: {stats.get('personas', 0)} personas, "
                                  f"{stats.get('grupos', 0)} grupos, "
                                  f"{stats.get('publicaciones', 0)} publicaciones, "
                                  f"{stats.get('extensiones', 0)} extensiones, "
                                  f"{stats.get('trabajos', 0)} trabajos de grado")
        # Actualizar combos
        self.tab_grupos.cargar_grupos()
        # Mostrar estadísticas detalladas
        from PyQt5.QtWidgets import QMessageBox
        msg = (
            f"📊 <b>Estadísticas de carga completada</b><br><br>"
            f"👤 Personas: {stats.get('personas', 0)}<br>"
            f"👥 Grupos: {stats.get('grupos', 0)}<br>"
            f"📄 Publicaciones: {stats.get('publicaciones', 0)}<br>"
            f"📋 Extensiones: {stats.get('extensiones', 0)}<br>"
            f"🎓 Trabajos de grado: {stats.get('trabajos', 0)}<br>"
            f"💡 Innovación: {stats.get('innovacion', 0)}<br>"
            f"📁 Proyectos: {stats.get('proyectos', 0)}<br>"
            f"© Propiedad intelectual: {stats.get('propiedad', 0)}"
        )
        QMessageBox.information(self, "Carga Completa", msg)

    def mostrar_consolidaciones(self, consolidaciones):
        if consolidaciones:
            mensaje = f"Se consolidaron {len(consolidaciones)} grupos de duplicados:\n"
            for g in consolidaciones:
                mensaje += f"\n- {g['nombre']}: {', '.join(g['cedulas'])}"
            self.label_estado.setText(mensaje[:100])


def main():
    app = QApplication(sys.argv)
    font = QFont("Segoe UI", 9)
    app.setFont(font)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()

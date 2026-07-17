"""Compara productos internos (BD, año X-Y) contra los perfiles GrupLAC
scrapeados en data/reporte excel_<fecha>/ (ver gruplac_scraper.py).

Reusa el motor de matching de comparador_faltantes.py (ComparadorFaltantes):
misma búsqueda en dos fases (grupo propio primero, luego cualquier otro
grupo), misma heurística de similitud de título (SequenceMatcher + Jaccard
sobre tokens normalizados) y los mismos estados de salida (Confirmado /
Registrado en otro grupo / Segundo barrido / Faltante real). Solo cambia
la FUENTE de datos en ambos lados:
  - Lado interno: en vez de Supervision_plano.xlsx, se consulta
    directamente publicaciones/extensiones/trabajos_grado/proyectos de
    academia_utp_integrado.db.
  - Lado GrupLAC: en vez de gruplac_957.db, se leen los .xlsx scrapeados
    de data/reporte excel_<fecha>/<GRUPO>/<GRUPO>.xlsx.

Escribe el resultado en data/cache/verificacion_faltantes.json, el mismo
caché que ya lee el resto de la UI (pestaña Cumplimiento, diálogo
Duplicados), así esas pantallas no necesitan ningún cambio.
"""
import json
import re
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl
import pandas as pd

from comparador_faltantes import (
    ComparadorFaltantes, clean_supervision_title, normalize_text, tokenize,
)
from utils import obtener_directorio_base

# ── Mapeo hoja GrupLAC -> categoría interna ─────────────────────────────
# Cobertura ampliada tras validar con datos reales: los productos
# tecnológicos (Softwares, Prototipos, Diseños industriales, etc.) y los
# jurados de trabajo de grado NO estaban mapeados y eso generaba falsos
# "Faltante" -- ej. PREDATA/PV-GRID (software) y un jurado de maestría de
# "GESTIÓN DE SISTEMAS ELÉCTRICOS..." que sí estaban en GrupLAC pero en
# hojas que este comparador nunca miraba.
_MAPEO_HOJAS = {
    "publicaciones": [
        "Artículos publicados", "Libros publicados", "Capítulos de libro publicados",
        "Otra publicación divulgativa", "Otros artículos publicados",
        "Otros Libros publicados", "Libros de formación", "Documentos de trabajo",
        "Notas científicas", "Ediciones", "Traducciones",
        "Traducciones Filológicas y Edición de Fuentes",
        "Manuales y Guías Especializadas",
        "Libros de divulgación y/o Compilación de divulgación",
        "Informes técnicos", "Informes de investigación",
        "Softwares", "Prototipos", "Otros productos tecnológicos",
        "Diseños industriales", "Publicaciones editoriales no especializadas",
        "Cartas, mapas o similares", "Conceptos técnicos",
        "Esquemas de trazados de circuito integrado",
        "Innovaciones en Procesos y Procedimientos",
        "Innovaciones generadas en la Gestión Empresarial",
        "Nuevas variedades animal", "Nuevas variedades vegetal",
        "Poblaciones mejoradas de razas pecuarias", "Nuevos registros científicos",
        "Plantas piloto", "Productos nutracéuticos", "Regulaciones y Normas",
        "Protocolos de vigilancia epidemiológica", "Reglamentos técnicos",
        "Guias de práctica clínica", "Proyectos de ley",
        "Producciones de contenido digital - Audiovisual",
        "Producciones de contenido digital - Sonoro",
        # Nombre exacto de la hoja duplicada cuando un grupo tiene ambas
        # ("Audiovisual" y "Sonoro" truncan al mismo nombre de 31
        # caracteres) -- el scraper le recorta más para dejarle espacio al
        # sufijo " (2)", así que no calza con el truncado normal de las dos
        # de arriba (ver _RE_SUFIJO_DUP, que solo cubre el caso general).
        "Producciones de contenido d (2)",
        "Generación de Contenido Multimedia", "Generación de Contenido Virtual",
        "Generación de Contenido Impreso",
        # Cobertura ampliada 2026-07-11 (auditoría completa de las 73 hojas
        # reales encontradas en los 127 grupos scrapeados): estas hojas SÍ
        # tienen productos reales pero no estaban en ningún mapeo, así que
        # jamás se comparaban -- se reportaban siempre "Faltante real" pase
        # lo que pase, sin importar si de verdad estaban en GrupLAC. Se
        # usa el nombre EXACTO (ya truncado a 31, tal como aparece en la
        # hoja real) en vez de reconstruir el nombre completo oficial de
        # Minciencias, para no arriesgar un truncado distinto al real.
        "Demás trabajos", "Obras o productos", "Signos distintivos",
        "Empresas de base tecnológica", "Nuevas secuencias genéticas",
        "Divulgación Pública de la Cienc", "Industrias creativas y cultural",
        "Generaciónes de contenido de au", "Desarrollo web",
    ],
    "extensiones": [
        "Curso de Corta Duración Dictados", "Eventos Científicos",
        "Eventos Artísticos", "Talleres de Creación",
        "Consultorías científico-tecnológicas", "Redes de Conocimiento Especializado",
        "Curso de doctorado", "Curso de maestría", "Curso especializado de extensión",
        "Participación en comités de evaluación",
        "Procesos de apropiación social del Conocimiento para el fortalecimiento o solución de asuntos de interés social",
        "Estrategias de Comunicación del Conocimiento",
        # Cobertura ampliada 2026-07-11 (ver nota arriba en "publicaciones").
        "Estrategias Pedagógicas para el", "Espacios de Participación Ciuda",
        "Asesorías al Programa Ondas", "Participación Ciudadana en Proy",
        "Proceso de Apropiación Social d", "Productos de apropiación social",
    ],
    "trabajos_grado": [
        "Trabajos dirigidos/turorías",
        "Jurado/Comisiones evaluadoras de trabajo de grado",
    ],
    "proyectos": ["Proyectos"],
}


_RE_ILEGALES_HOJA = re.compile(r"[\[\]:\\/\?\*]")


def _norm_hoja(nombre) -> str:
    return _RE_ILEGALES_HOJA.sub("", str(nombre)).strip().lower()


def _hoja_truncada(nombre: str) -> str:
    """Replica el truncado de nombre_hoja_seguro() en gruplac_scraper.py:
    quita los caracteres ilegales de un nombre de hoja Excel, recorta
    espacios y trunca a 31 (límite de Excel) -- EN ESE ORDEN. Sirve para
    construir la clave de _MAPEO_HOJAS con el mismo criterio que produjo el
    nombre real de la hoja, para que ambos truncados calcen."""
    return _RE_ILEGALES_HOJA.sub("", str(nombre)).strip()[:31]


# Los nombres de hoja en Excel están limitados a 31 caracteres, y el scraper
# quita caracteres ilegales (ej. '/') ANTES de truncar -- si la clave de
# _MAPEO_HOJAS se normaliza/trunca en otro orden, el corte cae en un punto
# distinto de la frase y no calza contra el nombre real (bugs reales,
# confirmados con datos):
#   - "Procesos de apropiación social..." no calzaba en 87 de 127 grupos:
#     al normalizar el nombre completo y truncar después, el corte de 31
#     dejaba (o no) un espacio colgando en un punto distinto al que deja
#     truncar primero y normalizar (con strip) después.
#   - "Jurado/Comisiones evaluadoras de trabajo de grado" y "Libros de
#     divulgación y/o Compilación de divulgación" no calzaban si se trunca
#     ANTES de quitar el '/' -- el scraper lo quita primero, así que la
#     clave debe replicar exactamente ese orden (_hoja_truncada).
_LOOKUP_HOJAS = {
    _norm_hoja(_hoja_truncada(nombre)): categoria
    for categoria, nombres in _MAPEO_HOJAS.items()
    for nombre in nombres
}

# Cuando un grupo tiene dos hojas GrupLAC distintas que truncan al mismo
# nombre de 31 caracteres (ej. "Producciones de contenido digital -
# Audiovisual" y "- Sonoro" truncan ambas a "Producciones de contenido
# digit"), Excel no permite hojas duplicadas y el scraper le agrega un
# sufijo " (2)", " (3)"... a la segunda -- sin este strip, esa segunda hoja
# no calza contra el mapeo y sus productos quedan invisibles para la
# comparación (bug real, confirmado con datos: "Producciones de contenido d
# (2)").
_RE_SUFIJO_DUP = re.compile(r"\s*\(\d+\)$")


def _categoria_de_hoja(nombre_hoja: str):
    clave = _norm_hoja(nombre_hoja)
    if clave in _LOOKUP_HOJAS:
        return _LOOKUP_HOJAS[clave]
    return _LOOKUP_HOJAS.get(_RE_SUFIJO_DUP.sub("", clave))


_RE_ANIO = re.compile(r"(19|20)\d{2}")
# Los ISSN suelen empezar con "20xx" (ej. "ISSN: 2076-3417") y ganaban la
# búsqueda del año por aparecer antes en el texto -> se descartan del texto
# donde se busca el año antes de buscarlo (bug real, confirmado con datos).
_RE_QUITAR_CLAVES = re.compile(
    r"(ISSN|ISBN)\s*:?\s*[0-9Xx][0-9Xx\-]*|10\.\d{4,9}/\S+", re.IGNORECASE)

# Talleres de Creación, Eventos Artísticos y Procesos de apropiación social
# NO traen numeración de ítem ('N.-') -- el título viene pegado a una
# etiqueta 'Nombre del <algo>: TÍTULO', a veces con más campos ' ,Otra
# Etiqueta: valor' apretados en la misma línea, a veces con el título en la
# línea siguiente. Sin este patrón, _extraer_titulo_anio devolvía el bloque
# de texto crudo completo (etiqueta + todos los campos) como "título" -- la
# similitud contra el título limpio interno quedaba siempre por debajo del
# umbral y se reportaban como "Faltante real" productos que sí estaban en
# GrupLAC (bug real, confirmado: "Laboratorio de creación en terracota bajo
# relieve (grupo 1/2)" del grupo LH). El (?:...) corta en la siguiente
# etiqueta ' ,Palabra:' o al final de línea -- deja pasar comas que son
# parte del título mismo (ej. "La Union, Valle") porque ese fragmento no
# tiene el patrón 'Etiqueta:' inmediatamente después.
#
# ANCLADO al inicio del bloque (^...) a propósito: en Trabajos de Grado y
# Jurados (formato numerado, título correcto en la 3ra línea) aparecen más
# abajo etiquetas secundarias como 'Nombre del estudiante:' o 'Nombre del
# orientado:' -- una primera versión sin anclar hacía match ahí y devolvía
# el nombre de la persona en vez del título del trabajo (regresión real,
# confirmada: filas de trabajos_grado con el "producto" = nombre de un
# estudiante). Anclado al inicio, este patrón solo dispara en los formatos
# sin numeración donde SÍ es la etiqueta principal.
_RE_NOMBRE_DEL = re.compile(
    r"^\s*(?:\d+\.-?\s*)?Nombre del \S+\s*:\s*\n?\s*(.+?)(?:\s*,\s*[A-ZÁÉÍÓÚÑ][^\n:,]{1,40}:|\s*$)",
    re.IGNORECASE | re.MULTILINE,
)


# Los títulos de trabajos de grado/prácticas son a veces genéricos y se
# repiten entre muchos estudiantes ("ESTUDIANTE EN PRACTICA" en 103
# registros reales) -- el bloque scrapeado sí trae el nombre del
# estudiante/orientado como campo aparte, así que se usa como segunda
# condición para no confirmar a ciegas el match de un estudiante equivocado
# (el equivalente de "cédula" que pidió el usuario: GrupLAC no expone el
# documento del estudiante, pero sí su nombre, en ambos lados).
_RE_NOMBRE_ESTUDIANTE = re.compile(
    r"Nombre del (?:estudiante|orientado)\s*:\s*\n?\s*([^\n,]+)", re.IGNORECASE)


def _extraer_estudiante(texto: str) -> str:
    m = _RE_NOMBRE_ESTUDIANTE.search(texto)
    return m.group(1).strip() if m else ""


def _extraer_titulo_anio(texto: str):
    """Cada fila scrapeada trae el título en uno de dos formatos: numerado
    'N.-\\nTIPO\\n: TÍTULO\\n...' (título en la 3ra línea no vacía) -- se
    prueba primero porque es el más común y el más confiable -- o sin
    numerar (o numerado sin guion, 'N. Nombre del X:') con el título pegado
    a una etiqueta 'Nombre del X:' al INICIO del bloque (ver _RE_NOMBRE_DEL,
    anclado a propósito para no capturar etiquetas secundarias más abajo en
    el texto). En último caso, se usa la primera línea completa."""
    lineas = [l for l in texto.split("\n") if l.strip()]
    if not lineas:
        return "", None

    if re.match(r"^\d+\.-?$", lineas[0].strip()) and len(lineas) > 2:
        titulo = lineas[2].strip()
        if titulo.startswith(":"):
            titulo = titulo[1:].strip()
    else:
        m_nombre = _RE_NOMBRE_DEL.match(texto)
        titulo = m_nombre.group(1).strip() if m_nombre else lineas[0]

    texto_sin_claves = _RE_QUITAR_CLAVES.sub(" ", texto)
    # Los trabajos de grado/proyectos traen "Desde M AAAA hasta [Mes] AAAA2"
    # -- el año que importa para saber si sigue "vigente" es el de cierre
    # (el último mencionado), no el de inicio.
    anios = [int(m.group(0)) for m in _RE_ANIO.finditer(texto_sin_claves)]
    anio = max(anios) if anios else None
    return titulo, anio


class ComparadorGrupLACScrapeado(ComparadorFaltantes):
    """ComparadorFaltantes indexando desde los .xlsx scrapeados en vez de
    gruplac_957.db. El "grupo_id" que usa el motor de búsqueda heredado es
    directamente el nombre del grupo (string), ya que aquí no existe un id
    numérico como en gruplac_957.db."""

    def __init__(self, carpeta_gruplac, anio_desde: int, anio_hasta: int):
        super().__init__(supervision_path=None, gruplac_db_path="")
        self.carpeta_gruplac = Path(carpeta_gruplac)
        self.anio_desde = anio_desde
        self.anio_hasta = anio_hasta

    def build_db_index(self):
        index = []
        self.groups_index = {}
        for carpeta_grupo in sorted(p for p in self.carpeta_gruplac.iterdir() if p.is_dir()):
            xlsx_files = [f for f in carpeta_grupo.glob("*.xlsx") if not f.name.startswith("~$")]
            if not xlsx_files:
                continue
            nombre_grupo = carpeta_grupo.name
            self.groups_index[nombre_grupo] = nombre_grupo
            try:
                wb = openpyxl.load_workbook(xlsx_files[0], read_only=True, data_only=True)
            except Exception:
                continue
            for nombre_hoja in wb.sheetnames:
                categoria = _categoria_de_hoja(nombre_hoja)
                if not categoria:
                    continue
                for row in wb[nombre_hoja].iter_rows(values_only=True):
                    if not row:
                        continue
                    texto = row[-1]
                    if not texto or not isinstance(texto, str):
                        continue
                    titulo, anio = _extraer_titulo_anio(texto)
                    if not titulo or len(titulo) < 5:
                        continue
                    if anio is not None and not (self.anio_desde <= anio <= self.anio_hasta):
                        continue
                    estudiante = _extraer_estudiante(texto) if categoria == "trabajos_grado" else ""
                    index.append({
                        "table": nombre_hoja,
                        "product_value": titulo,
                        "product_norm": normalize_text(titulo),
                        "product_tokens": tokenize(titulo),
                        "group_value": nombre_grupo,
                        "group_norm": normalize_text(nombre_grupo),
                        "grupo_id": nombre_grupo,
                        "categoria": categoria,
                        "estudiante_value": estudiante,
                        "estudiante_norm": normalize_text(estudiante) if estudiante else "",
                    })
            wb.close()

        self.db_index = index
        self.db_index_by_group = {}
        for rec in index:
            self.db_index_by_group.setdefault(rec["grupo_id"], []).append(rec)
        return index

    def compare_all_groups(self, progress_callback=None, df_supervision=None):
        """Reimplementado (no super()) porque el original resuelve
        grupo->grupo_id contra gruplac_957.db por sqlite directo; acá el
        grupo_id YA es el nombre del grupo, así que el mapeo es directo."""
        self.build_db_index()
        df_sup = df_supervision
        norm_to_id = {normalize_text(nombre): nombre for nombre in self.groups_index}

        results = []
        total = len(df_sup)
        for idx, (_, row) in enumerate(df_sup.iterrows()):
            grupo_norm = row["_grupo_norm"]
            producto = row["_producto_limpio"]
            if not producto:
                continue

            grupo_id = norm_to_id.get(grupo_norm)
            if grupo_id is None:
                best_id, best_score = None, 0
                for key, gid in norm_to_id.items():
                    score = SequenceMatcher(None, grupo_norm, key).ratio()
                    if score > best_score and score > 0.6:
                        best_score = score
                        best_id = gid
                grupo_id = best_id

            result = self.search_product(producto, grupo_id)

            if row.get("_titulo_generico") and result.get("estado_verificacion") in (
                    "Confirmado en BD (mismo grupo)", "Registrado en otro grupo"):
                # Segunda condición (el equivalente de "cédula" que no expone
                # GrupLAC): si el bloque scrapeado sí trae el nombre del
                # estudiante/orientado y coincide con el nombre interno, el
                # match SÍ identifica al estudiante correcto pese al título
                # compartido -- no hace falta bajarlo a revisión manual.
                nombre_int = row.get("_estudiante_norm", "")
                nombre_ext = result.get("estudiante_gruplac", "")
                nombre_coincide = (
                    bool(nombre_int) and bool(nombre_ext)
                    and SequenceMatcher(None, nombre_int, nombre_ext).ratio() >= 0.72
                )
                if not nombre_coincide:
                    mismo_grupo = "mismo grupo" in result["estado_verificacion"]
                    result = dict(result)
                    result["estado_verificacion"] = (
                        "Segundo barrido - mismo grupo" if mismo_grupo
                        else "Segundo barrido - otro grupo"
                    )
                    result["necesita_revision"] = True
                    result["es_faltante"] = False
                    if nombre_int and nombre_ext:
                        nota = (
                            f"Título genérico y el nombre del estudiante no coincide "
                            f"(interno: '{row.get('_estudiante_valor', '')}' vs GrupLAC: "
                            f"'{result.get('estudiante_gruplac_valor', '')}') -- revisar a mano."
                        )
                    else:
                        nota = ("Título genérico -- lo comparte más de un registro interno, no se "
                                "puede confirmar a ciegas a cuál corresponde este match; revisar a mano.")
                    previo = result.get("detalle_verificacion", "")
                    result["detalle_verificacion"] = f"{previo} | {nota}" if previo else nota

            results.append({
                "grupo_original": row.get("grupo", ""),
                "producto": producto,
                "categoria": row.get("categoria", ""),
                "hoja": row.get("hoja", ""),
                "grupo_id_match": grupo_id,
                "cedula_responsable": row.get("cedula_responsable", ""),
                "responsable": row.get("responsable", ""),
                "issn_isbn": row.get("issn_isbn", ""),
                "doi_url": row.get("doi_url", ""),
                **result,
            })
            if progress_callback and (idx + 1) % 25 == 0:
                progress_callback(idx + 1, total)

        return pd.DataFrame(results)


# ── Lado interno: BD -> DataFrame estilo Supervision_plano ─────────────────

# Un puñado de registros (todos productos de software con "Certificado
# Registro Soporte Lógico") quedaron con el nombre corto pegado a una
# descripción larga en 'titulo' (ej. 'PREDATA": en consecuencia con la gran
# cantidad de...'), seguramente por cómo se aplanó el Excel de origen. Con
# el nombre completo, SequenceMatcher nunca alcanza el umbral contra el
# nombre corto real que sí está en GrupLAC -> se reportaban como faltantes
# aunque estuvieran subidos. Confirmado con datos reales: solo 5 filas en
# toda la BD calzan este patrón (PV-GRID, PREDATA, CONDOR, "Rectas en R3"),
# así que es seguro aplicarlo sin arriesgar títulos normales con ": ".
_RE_TITULO_PEGADO = re.compile(r'^([^":]{2,40})":\s*.+$', re.DOTALL)


def _limpiar_titulo_pegado(titulo: str) -> str:
    m = _RE_TITULO_PEGADO.match(titulo)
    return m.group(1).strip() if m else titulo


def _grupos_de_cedula(cur, cedula, cache):
    if cedula in cache:
        return cache[cedula]
    grupos = [g for (g,) in cur.execute(
        "SELECT DISTINCT grupo FROM grupos WHERE cedula = ? AND grupo IS NOT NULL AND grupo != ''",
        (cedula,))]
    cache[cedula] = grupos
    return grupos


_RE_CODIGO_GRUPO = re.compile(r'^[A-Z]{2,6}[\d-]*\d-\s*')


def _limpiar_grupos_crudo(valor: str) -> list:
    """La columna 'grupo' de extensiones (y a veces proyectos) trae el
    valor crudo del Excel fuente: código pegado al nombre (ej. 'COL0002859-
    AUTOMÁTICA') y a veces VARIOS grupos separados por coma en una sola
    celda (ej. 'COL0002859-AUTOMÁTICA, COL0077968-GRUPO...'). Confirmado con
    datos reales: 583/651 filas de extensiones con grupo tienen este código,
    149 tienen más de un grupo en la misma celda. Sin limpiar esto, ninguno
    calza con el nombre de carpeta GrupLAC y todo sale falsamente 'Faltante'.

    OJO: al menos 30 grupos internos (11 sin contar semilleros, ej.
    "TERRITORIO, EDUCACIÓN Y SOCIEDAD") tienen una coma en su propio nombre
    -- partir a ciegas por "," los rompía en fragmentos que no existen
    ("TERRITORIO" + "EDUCACIÓN Y SOCIEDAD"), y sus productos quedaban
    invisibles para la comparación (bug real, confirmado: 15 extensiones de
    ese grupo, con las que nunca se comparaba nada). Un fragmento después de
    una coma solo es un grupo NUEVO si trae su propio código al inicio
    (patrón real de "varios grupos en una celda"); si no lo trae, la coma
    era parte del nombre del grupo anterior y se vuelve a unir."""
    crudo = (valor or "").strip()
    if not crudo:
        return []
    partes_raw = [p.strip() for p in crudo.split(",")]
    grupos_crudos = []
    for p in partes_raw:
        if not p:
            continue
        if not grupos_crudos or _RE_CODIGO_GRUPO.match(p):
            grupos_crudos.append(p)
        else:
            grupos_crudos[-1] = f"{grupos_crudos[-1]}, {p}"
    return [g for g in (_RE_CODIGO_GRUPO.sub("", g).strip() for g in grupos_crudos) if g]


def _nombre_de_cedula(cur, cedula, cache):
    if cedula in cache:
        return cache[cedula]
    row = cur.execute("SELECT nombre FROM personas WHERE cedula = ?", (cedula,)).fetchone()
    nombre = row[0] if row else ""
    cache[cedula] = nombre
    return nombre


def construir_df_interno(db, anio_desde: int, anio_hasta: int) -> pd.DataFrame:
    """Une publicaciones/extensiones/trabajos_grado/proyectos (año en rango)
    en el mismo formato que antes venía de Supervision_plano.xlsx: una fila
    por (producto, grupo). Si la tabla no trae 'grupo' propio (o viene
    vacío), se resuelve vía la tabla grupos por cédula -- si la persona
    pertenece a varios grupos, se emite una fila por cada uno (el motor de
    búsqueda ya sabe distinguir 'está en su grupo' de 'está en otro').

    Si el campo 'grupo' del Excel (proyectos/extensiones) SÍ trae un valor
    pero NO coincide con ninguno de los grupos internos reales del
    responsable (tabla grupos, por cédula), se emite TAMBIÉN para esos
    grupos reales -- no solo para el que quedó escrito en el Excel.
    Confirmado con datos reales: 67 de 263 proyectos (25%) tienen un
    'grupo' que no calza con la membresía real del responsable; sin este
    respaldo, un producto pendiente quedaba invisible en el panel de
    Cumplimiento del grupo real (aparecía "sin faltantes" ahí simplemente
    porque el faltante había quedado archivado bajo el grupo equivocado)."""
    cur = db.conn.cursor()
    # cursor aparte para las búsquedas anidadas: reutilizar el mismo cursor
    # que está iterando la consulta externa lo reinicia a medio camino y
    # descarta la mayoría de las filas restantes (bug real, confirmado).
    cur_grupos = db.conn.cursor()
    cur_personas = db.conn.cursor()
    cache_grupos = {}
    cache_nombres = {}
    filas = []

    def _emitir(categoria, hoja, cedula, grupo_col, titulo, issn_isbn="", doi_url="", estudiante=""):
        titulo = (titulo or "").strip()
        if not titulo:
            return
        grupos_texto = _limpiar_grupos_crudo(grupo_col) if grupo_col else []
        grupos_reales = _grupos_de_cedula(cur_grupos, cedula, cache_grupos) if cedula else []
        if not grupos_texto:
            grupos = grupos_reales
        else:
            texto_norm = {normalize_text(g) for g in grupos_texto}
            extra = [g for g in grupos_reales if normalize_text(g) not in texto_norm]
            grupos = grupos_texto + extra
        if not grupos:
            grupos = [""]
        producto = clean_supervision_title(_limpiar_titulo_pegado(titulo))
        responsable = _nombre_de_cedula(cur_personas, cedula, cache_nombres) if cedula else ""
        for grupo in grupos:
            filas.append({
                "grupo": grupo, "producto": producto, "categoria": categoria, "hoja": hoja,
                "_grupo_norm": normalize_text(grupo), "_producto_limpio": producto,
                "cedula_responsable": cedula or "", "responsable": responsable,
                "issn_isbn": issn_isbn or "", "doi_url": doi_url or "",
                "_estudiante_valor": estudiante or "",
                "_estudiante_norm": normalize_text(estudiante) if estudiante else "",
            })

    for cedula, grupo, titulo, issn_isbn, doi_url in cur.execute(
            "SELECT cedula, grupo, titulo, issn_isbn, doi_url FROM publicaciones "
            "WHERE año BETWEEN ? AND ? AND titulo IS NOT NULL AND titulo != ''",
            (anio_desde, anio_hasta)):
        _emitir("publicaciones", "Publicaciones", cedula, grupo, titulo, issn_isbn, doi_url)

    for cedula, grupo, actividad in cur.execute(
            "SELECT cedula, grupo, actividad FROM extensiones "
            "WHERE año BETWEEN ? AND ? AND actividad IS NOT NULL AND actividad != ''",
            (anio_desde, anio_hasta)):
        _emitir("extensiones", "Extensiones", cedula, grupo, actividad)

    # Solo "conducentes": un trabajo de grado "no conducente" (ej. una
    # práctica que no es conducente a título) no se sube a GrupLAC, así que
    # nunca debe compararse ni contar como faltante -- calificacion es NULL
    # para el reporte institucional posicional (siempre conducente, ese
    # archivo no trae prácticas).
    for cedula_director, titulo, nombre_estudiante in cur.execute(
            "SELECT cedula_director, titulo, nombre_estudiante FROM trabajos_grado "
            "WHERE año BETWEEN ? AND ? AND titulo IS NOT NULL AND titulo != '' "
            "AND (calificacion IS NULL OR calificacion != 'NO CONDUCENTE')",
            (anio_desde, anio_hasta)):
        _emitir("trabajos_grado", "Trabajos de Grado", cedula_director, "", titulo,
                estudiante=nombre_estudiante or "")

    for cedula, grupo, titulo in cur.execute(
            "SELECT cedula, grupo, titulo FROM proyectos "
            "WHERE año BETWEEN ? AND ? AND titulo IS NOT NULL AND titulo != ''",
            (anio_desde, anio_hasta)):
        _emitir("proyectos", "Proyectos", cedula, grupo, titulo)

    df = pd.DataFrame(filas)
    if df.empty:
        return df

    # Los títulos de trabajos de grado (incluye prácticas) a veces son
    # genéricos y se repiten entre MUCHOS estudiantes distintos ("ESTUDIANTE
    # EN PRACTICA" se vio repetido en 103 registros reales) -- el motor de
    # coincidencia por texto no puede saber a CUÁL estudiante corresponde un
    # match de GrupLAC cuando el título no lo distingue. Se marca acá (antes
    # de deduplicar, para no perder la señal) y compare_all_groups baja esos
    # matches a "Segundo barrido" en vez de confirmarlos a ciegas.
    df["_titulo_generico"] = False
    es_tg = df["categoria"] == "trabajos_grado"
    if es_tg.any():
        conteo = df.loc[es_tg].groupby("_producto_limpio")["_producto_limpio"].transform("count")
        df.loc[es_tg, "_titulo_generico"] = conteo > 1

    # Algunas tablas internas no tienen clave única y repiten filas idénticas
    # (ej. trabajos_grado, ver notas de sesiones previas) -- deduplicar por
    # (grupo, producto, categoría) para no inflar el conteo de faltantes.
    return df.drop_duplicates(subset=["_grupo_norm", "_producto_limpio", "categoria"])


# ── Orquestación ─────────────────────────────────────────────────────────

def escribir_cache_verificacion(df_resultado: pd.DataFrame, ruta_cache=None) -> Path:
    ruta_cache = Path(ruta_cache) if ruta_cache else (
        obtener_directorio_base() / "data" / "cache" / "verificacion_faltantes.json")
    payload = {
        "generado": datetime.now().isoformat(),
        "data": df_resultado.to_dict(orient="records"),
    }
    ruta_cache.parent.mkdir(parents=True, exist_ok=True)
    with open(ruta_cache, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2, default=str)
    return ruta_cache


def ejecutar_y_guardar(db, carpeta_gruplac, anio_desde=2024, anio_hasta=2025,
                        progreso_callback=None):
    """Corre la verificación completa y la deja escrita en
    data/cache/verificacion_faltantes.json. Devuelve un resumen numérico."""
    df_interno = construir_df_interno(db, anio_desde, anio_hasta)
    comparador = ComparadorGrupLACScrapeado(carpeta_gruplac, anio_desde, anio_hasta)
    df_resultado = comparador.compare_all_groups(
        progress_callback=progreso_callback, df_supervision=df_interno)
    ruta = escribir_cache_verificacion(df_resultado)

    total = len(df_resultado)
    faltantes = int(df_resultado["es_faltante"].sum()) if total else 0
    confirmados = int(
        df_resultado["estado_verificacion"].str.contains("Confirmado", na=False).sum()
    ) if total else 0
    revision = int(df_resultado["necesita_revision"].sum()) if total else 0
    return {
        "total": total,
        "confirmados": confirmados,
        "faltantes": faltantes,
        "revision": revision,
        "ruta_cache": str(ruta),
    }

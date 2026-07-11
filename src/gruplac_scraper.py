"""Webscraping de perfiles de grupos en GrupLAC (scienti.minciencias.gov.co).

Lee data/BD.xlsx (columnas: NOMBRE GRUPO, CÓDIGO GRUPO, URL COLCIENCIAS) y
genera un .xlsx por grupo con el mismo layout que data/reporte excel.zip
(una hoja por sección del perfil GrupLAC).

Estructura de la página fuente (visualizagr.jsp): cada sección es una fila
<tr><td class="celdaEncabezado">...</td></tr> seguida de sus filas de datos,
hasta la siguiente celdaEncabezado. Según la forma de esas filas se detecta:
  - Tabla real (Integrantes del grupo): 2+ <td class="celdasTitulo"> en la
    primera fila => encabezados de columna, filas siguientes multi-columna.
  - Clave/valor (Datos básicos): todas las filas son pares
    (celdasTitulo, celdas2) => (etiqueta, valor), sin fila de encabezado.
  - Lista (Artículos publicados, Proyectos, etc.): cada fila -> texto de su
    última celda, en una sola columna.
Validado comparando el resultado contra los .xlsx ya existentes en
data/reporte excel.zip para dos grupos reales (700 REPART y AUTOMÁTICA):
mismas hojas, mismo formato de texto, conteos consistentes con datos nuevos.
"""
import re
import time
from pathlib import Path

import openpyxl
import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"
)
DELAY_SEGUNDOS = 1.5
TIMEOUT = 25
REINTENTOS = 3

_ILEGALES_HOJA = re.compile(r'[\[\]:\\/\?\*]')
_ILEGALES_CARPETA = re.compile(r'[\\/:*?"<>|]')


# ── Parseo del HTML de un grupo ─────────────────────────────────────────────

def _classes(td):
    return td.get("class") or []


def _texto_celda(td):
    """get_text('\\n') no separa saltos de línea que ya vienen crudos dentro
    de un mismo nodo de texto (indentación del HTML fuente); se normaliza
    partiendo por línea, recortando y descartando líneas vacías."""
    texto = td.get_text("\n")
    lineas = [ln.strip() for ln in texto.split("\n")]
    return "\n".join(ln for ln in lineas if ln)


def _formatear_seccion(filas_tds):
    first_tds = filas_tds[0]
    es_tabla_real = (
        len(first_tds) > 1
        and all("celdasTitulo" in _classes(td) for td in first_tds)
    )

    out = []
    if es_tabla_real:
        out.append(tuple(td.get_text(strip=True) for td in first_tds))
        for tds in filas_tds[1:]:
            if not tds:
                continue
            out.append(tuple(_texto_celda(td) for td in tds))
        return out

    def es_fila_kv(tds):
        return len(tds) == 2 and "celdasTitulo" in _classes(tds[0])

    if all(es_fila_kv(tds) for tds in filas_tds):
        for tds in filas_tds:
            label = tds[0].get_text(strip=True)
            valor = _texto_celda(tds[1])
            out.append((label, valor))
        return out

    for tds in filas_tds:
        if not tds:
            continue
        texto = _texto_celda(tds[-1])
        if texto:
            out.append((None, texto))
    return out


def parse_grupo_html(html: str) -> dict:
    """Devuelve dict {nombre_seccion: [tuplas_de_fila]} en orden del documento."""
    soup = BeautifulSoup(html, "html.parser")
    hojas = {}

    current_title = None
    current_rows = []

    def flush():
        nonlocal current_title, current_rows
        if current_title and current_rows:
            formateado = _formatear_seccion(current_rows)
            if formateado:
                key = current_title
                n = 2
                while key in hojas:
                    key = f"{current_title} ({n})"
                    n += 1
                hojas[key] = formateado
        current_title = None
        current_rows = []

    for tr in soup.find_all("tr"):
        tds = tr.find_all("td", recursive=False)
        if not tds:
            continue
        if "celdaEncabezado" in _classes(tds[0]):
            flush()
            current_title = tds[0].get_text(strip=True)
            continue
        if current_title is None:
            continue
        current_rows.append(tds)

    flush()
    return hojas


# ── Lectura de BD.xlsx y descarga ───────────────────────────────────────────

def leer_grupos_bd(bd_path="data/BD.xlsx"):
    df = pd.read_excel(bd_path, sheet_name=0)
    df.columns = [str(c).strip().upper() for c in df.columns]
    grupos = []
    for _, row in df.iterrows():
        nombre = str(row.get("NOMBRE GRUPO", "") or "").strip()
        codigo = str(row.get("CÓDIGO GRUPO", "") or "").strip()
        url = str(row.get("URL COLCIENCIAS", "") or "").strip()
        if nombre and url and url.lower() != "nan":
            grupos.append({"nombre": nombre, "codigo": codigo, "url": url})
    return grupos


def descargar_html(url, session, reintentos=REINTENTOS):
    ultimo_error = None
    for intento in range(1, reintentos + 1):
        try:
            resp = session.get(url, timeout=TIMEOUT, headers={"User-Agent": USER_AGENT})
            resp.raise_for_status()
            resp.encoding = "ISO-8859-1"
            return resp.text
        except Exception as e:
            ultimo_error = e
            if intento < reintentos:
                time.sleep(2 * intento)
    raise RuntimeError(f"No se pudo descargar {url}: {ultimo_error}")


# ── Escritura de Excel por grupo (mismo layout que reporte excel.zip) ─────

def _nombre_hoja_seguro(titulo, usados):
    limpio = _ILEGALES_HOJA.sub("", titulo).strip() or "Seccion"
    limpio = limpio[:31]
    base = limpio
    n = 2
    while limpio.lower() in usados:
        sufijo = f" ({n})"
        limpio = base[: 31 - len(sufijo)] + sufijo
        n += 1
    usados.add(limpio.lower())
    return limpio


def nombre_carpeta_valido(nombre: str) -> str:
    limpio = _ILEGALES_CARPETA.sub("", nombre).strip()
    # NTFS/exFAT (discos externos montados en Windows) rechazan nombres que
    # terminen en espacio o punto -> volver a recortar tras el strip inicial
    # puede reintroducir un espacio de sobra a mitad de palabra.
    limpio = limpio[:80].strip(" .")
    return limpio or "SIN_NOMBRE"


def _limpiar_valor(v):
    """Quita caracteres de control no válidos en XML/Excel (openpyxl los
    rechaza con IllegalCharacterError y aborta la escritura de todo el
    archivo si aparecen en cualquier celda)."""
    if isinstance(v, str):
        return ILLEGAL_CHARACTERS_RE.sub("", v)
    return v


def escribir_excel_grupo(hojas: dict, dest_path: Path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    usados = set()
    for titulo, filas in hojas.items():
        ws = wb.create_sheet(_nombre_hoja_seguro(titulo, usados))
        for fila in filas:
            ws.append([_limpiar_valor(v) for v in fila])
    if not wb.sheetnames:
        wb.create_sheet("Sin datos")
    dest_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest_path)


# ── Orquestación ─────────────────────────────────────────────────────────

def ejecutar_scraping(bd_path="data/BD.xlsx",
                       dest_root="data/reporte excel_actualizado",
                       progreso_callback=None,
                       delay=DELAY_SEGUNDOS,
                       solo_grupos=None):
    """progreso_callback(indice, total, nombre_grupo) se llama antes de cada
    descarga. solo_grupos: si se da (iterable de nombres), solo se procesan
    esos grupos (para reintentar fallidos sin repetir todo el lote).
    Devuelve {'ok': [nombres], 'error': [(nombre, motivo)]}."""
    grupos = leer_grupos_bd(bd_path)
    if solo_grupos is not None:
        solo_grupos = set(solo_grupos)
        grupos = [g for g in grupos if g["nombre"] in solo_grupos]
    dest_root = Path(dest_root)
    resultados = {"ok": [], "error": []}
    session = requests.Session()
    total = len(grupos)

    for i, g in enumerate(grupos, 1):
        nombre = g["nombre"]
        if progreso_callback:
            progreso_callback(i, total, nombre)
        try:
            html = descargar_html(g["url"], session)
            hojas = parse_grupo_html(html)
            carpeta = nombre_carpeta_valido(nombre)
            dest_path = dest_root / carpeta / f"{carpeta}.xlsx"
            escribir_excel_grupo(hojas, dest_path)
            resultados["ok"].append(nombre)
        except Exception as e:
            resultados["error"].append((nombre, str(e)))
        if i < total:
            time.sleep(delay)

    return resultados

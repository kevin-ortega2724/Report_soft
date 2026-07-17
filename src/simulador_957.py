"""
simulador_957.py

Entorno de simulación (sin UI) del proceso oficial de medición Conv. 957
de MinCiencias, descrito en docs/metodologia_957_documento_tecnico.md.

Reproduce los Pasos 2-6 del proceso ScienTI a partir de los datos extraídos
en data/output/medicion_957.xlsx (hojas productos/indicadores/cuartiles,
generadas por parse_pdf_medicion.py desde los PDF de data/pdf/):

    Paso 2: lambda por subtipo                -> ya viene en hoja `productos`
    Paso 3: valor_indicador_j = Σλ_j × C_j(área)   (sólo en modo proyección)
    Paso 4: indice_j = valor_indicador_j / máximo_área_j
    Paso 5: IG = Σ w_j · indice_j
    Paso 6: cuartil del IG y de NC_TOP + condiciones -> categoría D/C/B/A/A1

Dos modos de uso:

1. Verificación (`simular`, modo="verificacion", por defecto): usa los
   valores OFICIALES de cada grupo (hoja `indicadores`) para los Pasos 4-6.
   100% de coincidencia exacta de categoría sobre los 60 grupos oficiales
   con categoría conocida (15 A1, 25 A, 9 B, 11 C).

2. Proyección (`proyectar_productos`): simula el efecto de agregar
   productos hipotéticos (que el grupo aún no tiene) sobre sus indicadores,
   IG y categoría, usando la fórmula de impacto del doc §10 (ratio propio
   del grupo, no una mediana de área).

Condiciones del Paso 6.2 transcritas literalmente del texto "Esta categoría
fue alcanzada bajo las siguientes condiciones" de cada PDF (A1s/As/Bs/Cs),
no deducidas. La única condición NO evaluada es "investigador emérito/
sénior/asociado/junior vinculado": ese dato no existe en ningún export de
GrupLAC a nivel de grupo.

Parámetros del modelo (ventana por subtipo, ponderaciones w_j del IG, año
base de la convocatoria) son argumentos del constructor — no constantes
fijas en el código — para poder correr sensibilidad o adaptarse a cambios
de convocatoria sin editar el módulo.

Uso:
    python src/simulador_957.py                  # corre la verificación completa

    from simulador_957 import Simulador957
    sim = Simulador957()
    sim.simular("Automática")                                  # verificación
    sim.proyectar_productos("GESTIÓN...", {"ART_A1": 10})       # proyección
"""

from __future__ import annotations

import math
import pathlib
import re
import sqlite3
import statistics
from dataclasses import dataclass, field
from typing import Optional

import openpyxl
import pandas as pd

from utils import normalizar_nombre
from views.vista_seguimiento_grupos import _carpeta_gruplac_mas_reciente

# BD del scraping GrupLAC/ScienTI (NO el archivo de incentivos CIARP, que
# es otra fuente distinta — ver discusión sobre quién trae el cuartil).
GRUPLAC_DB_PATH = (
    pathlib.Path(__file__).parent.parent / "data" / "db" / "gruplac_957.db"
)

# Año de cierre de información de la Conv. 957 (doc, portada de los PDF: "2024")
AÑO_BASE_CONVOCATORIA = 2024

# Palabras sin valor discriminante al comparar nombres de grupo (para el
# emparejamiento difuso entre medicion_957.xlsx y las carpetas reports/excel/)
_STOPWORDS_NOMBRE = {
    "de", "la", "el", "en", "y", "del", "las", "los", "para", "con",
    "grupo", "investigacion", "investigacin", "a", "o",
}

# ---------------------------------------------------------------------------
# Constantes del modelo oficial (doc §4.2 / Apéndice A)
# ---------------------------------------------------------------------------

PONDERACIONES_957: dict[str, float] = {
    "NC_TOP": 3.7,
    "NC_A": 2.3,
    "NC_B": 0.4,
    "ASC": 1.5,
    "DPC": 0.5,
    "FRH_A": 1.0,
    "FRH_B": 0.2,
    "cohesion": 0.1,
    "colaboracion": 0.3,
}

# Secciones que sí tienen desglose de productos en la hoja `productos`
# (FRH_A/FRH_B, cohesión y colaboración no lo tienen — doc §2.4 y §10.6)
SECCIONES_CON_PRODUCTOS = ["NC_TOP", "NC_A", "NC_B", "ASC", "DPC"]

# Nombre de cuartil (hoja `cuartiles`) -> indicador corto
CUARTIL_A_INDICADOR = {
    "Nuevo Conocimiento TOP": "NC_TOP",
    "Nuevo Conocimiento A": "NC_A",
    "Nuevo Conocimiento B": "NC_B",
    "Apropiación Social y del Conocimiento": "ASC",
    "Divulgación Pública de la Ciencia": "DPC",
    "Formación de Recurso Humano A": "FRH_A",
    "Formación de Recurso Humano B": "FRH_B",
    "Área de conocimiento del Grupo de Investigación": "IG",
}

ORDEN_CATEGORIAS = ["D", "C", "B", "A", "A1"]


# ---------------------------------------------------------------------------
# Estructuras de resultado
# ---------------------------------------------------------------------------

@dataclass
class ResultadoSimulacion:
    grupo: str
    area: Optional[str]
    categoria_oficial: Optional[str]
    categoria_simulada: str
    ig_oficial: Optional[float]
    ig_simulado: float
    valores_simulados: dict = field(default_factory=dict)
    indices_simulados: dict = field(default_factory=dict)
    cuartil_ig: int = 4
    cuartil_nc_top: int = 4
    anios_existencia: Optional[float] = None
    condiciones: dict = field(default_factory=dict)
    coincide: bool = False
    detalle_ajustes: dict = field(default_factory=dict)


class Simulador957:
    """
    Reproduce el cálculo oficial de la categoría 957 a partir de
    data/output/medicion_957.xlsx.

    Los tres bloques de parámetros del modelo (ventana por subtipo,
    ponderaciones w_j del IG, año base de la convocatoria) son argumentos
    del constructor, no constantes fijas: si MinCiencias cambia alguno en
    una convocatoria futura, o se quiere correr un análisis de sensibilidad
    ("¿y si NC_TOP pesara 4.0 en vez de 3.7?"), se pasa el valor alternativo
    sin tocar el código. Por defecto se usan los valores oficiales vigentes.
    """

    def __init__(
        self,
        excel_path: str | pathlib.Path | None = None,
        ponderaciones: dict[str, float] | None = None,
        ventanas_subtipo: dict[str, int] | None = None,
        año_base: int | None = None,
    ):
        if excel_path is None:
            excel_path = (
                pathlib.Path(__file__).parent.parent
                / "data" / "output" / "medicion_957.xlsx"
            )
        self.excel_path = pathlib.Path(excel_path)
        self.df_grupos = pd.read_excel(self.excel_path, sheet_name="grupos")
        self.df_productos = pd.read_excel(self.excel_path, sheet_name="productos")
        self.df_indicadores = pd.read_excel(self.excel_path, sheet_name="indicadores")
        self.df_cuartiles = pd.read_excel(self.excel_path, sheet_name="cuartiles")
        self._deduplicar_grupos_repetidos()

        # Parámetros del modelo: por defecto, los valores oficiales vigentes
        # (PONDERACIONES_957 y AÑO_BASE_CONVOCATORIA), pero sustituibles.
        self.ponderaciones = dict(ponderaciones or PONDERACIONES_957)
        self.año_base = año_base or AÑO_BASE_CONVOCATORIA
        self.ventanas_subtipo = (
            dict(ventanas_subtipo)
            if ventanas_subtipo is not None
            else self._derivar_ventanas_subtipo()
        )

        self._area_de: dict[str, str] = dict(
            zip(self.df_grupos["nombre_grupo"], self.df_grupos["area_conocimiento"])
        )
        self._categoria_oficial: dict[str, str] = dict(
            zip(self.df_grupos["nombre_grupo"], self.df_grupos["categoria_confirmada"])
        )
        self._ig_oficial: dict[str, float] = dict(
            zip(self.df_grupos["nombre_grupo"], self.df_grupos["indicador_grupo"])
        )

        self._sumlambda = (
            self.df_productos.groupby(["grupo", "seccion"])["lambda_val"]
            .sum()
            .to_dict()
        )  # {(grupo, seccion): Σλ}

        self._valor_oficial = {
            (r.grupo, r.indicador): r.valor_indicador
            for r in self.df_indicadores.itertuples()
            if pd.notna(r.valor_indicador)
        }
        # Máximo del área SEGÚN EL PROPIO GRUPO (Paso 4 del PDF, columna
        # "Máximo"). Es más confiable que recalcularlo desde la hoja
        # `cuartiles`: se comprobó que esa hoja no siempre referencia la
        # misma población de comparación para todos los indicadores de
        # una misma área (ver hallazgo con "Ciencias Naturales").
        self._maximo_oficial_por_grupo = {
            (r.grupo, r.indicador): r.valor_maximo
            for r in self.df_indicadores.itertuples()
            if pd.notna(r.valor_maximo)
        }

        # subtipo -> sección 957 a la que aporta (ART_A1 -> NC_TOP, etc.)
        # y total actual de unidades por (grupo, subtipo) -- ambos se
        # derivan de la hoja `productos`, necesarios para el modo
        # "proyeccion" (Δλ al agregar productos hipotéticos).
        self._seccion_de_subtipo: dict[str, str] = dict(
            zip(self.df_productos["subtipo"], self.df_productos["seccion"])
        )
        self.seccion_de_subtipo = dict(self._seccion_de_subtipo)  # acceso público
        self._total_por_grupo_subtipo: dict[tuple[str, str], int] = dict(
            zip(
                zip(self.df_productos["grupo"], self.df_productos["subtipo"]),
                self.df_productos["total"],
            )
        )

        self._constantes_area = self._calcular_constantes_area()
        self._maximos_area = self._calcular_maximos_area()
        self._cuartiles_area = self._calcular_cuartiles_area()
        self._anio_formacion = self._cargar_anios_formacion()

    def _deduplicar_grupos_repetidos(self) -> None:
        """
        Corrige un bug de extracción puntual de `parse_pdf_medicion.py`:
        al menos un grupo ("Desarrollo de Procesos Quimicos") quedó
        triplicado en las 4 hojas -- aparece 3 veces en `grupos` y sus
        filas de `productos` están repetidas 3 veces idénticas, lo que
        infla Σλ ~3x para ese grupo y desvía su ratio C_j(área).
        Detectado al cruzar el total de artículos oficial contra el conteo
        crudo de GrupLAC (debía ser oficial <= GrupLAC histórico, y no lo
        era: 42 vs 24 artículos).

        Estrategia: quedarse, por nombre de grupo, con la fila de `grupos`
        que tenga área de conocimiento válida (no nula ni el texto
        residual "del Grupo de Investigación" que deja el bug de regex
        del área), y eliminar en `productos`/`indicadores`/`cuartiles` las
        filas exactamente duplicadas.
        """
        AREA_INVALIDA = "del Grupo de Investigación"
        grupos = self.df_grupos.copy()
        grupos["_area_valida"] = (
            grupos["area_conocimiento"].notna()
            & (grupos["area_conocimiento"] != AREA_INVALIDA)
        )
        grupos = grupos.sort_values("_area_valida", ascending=False)
        grupos = grupos.drop_duplicates(subset="nombre_grupo", keep="first")
        self.df_grupos = grupos.drop(columns="_area_valida").reset_index(drop=True)

        self.df_productos = self.df_productos.drop_duplicates().reset_index(drop=True)
        self.df_indicadores = self.df_indicadores.drop_duplicates().reset_index(drop=True)
        self.df_cuartiles = self.df_cuartiles.drop_duplicates().reset_index(drop=True)

    def _derivar_ventanas_subtipo(self) -> dict[str, int]:
        """
        Tabla subtipo -> ventana (años), derivada de la hoja `productos`.
        Se comprobó que la ventana es constante por subtipo en las 1502
        filas oficiales (0 inconsistencias) -> se toma el primer valor
        visto por subtipo como el oficial.
        """
        return (
            self.df_productos.groupby("subtipo")["ventana"]
            .first()
            .to_dict()
        )

    # ------------------------------------------------------------------
    # Años de existencia: "Año y mes de formación" de la hoja "Datos
    # Básicos" en reports/excel/<grupo>/<grupo>.xlsx (export de GrupLAC,
    # NO de la BD interna). Condición textual de los PDF: "Tener al menos
    # N años de existencia" (5 para A1/A, 3 para B, 2 para C).
    # ------------------------------------------------------------------

    @staticmethod
    def _tokens_significativos(texto: str) -> set[str]:
        n = normalizar_nombre(texto)
        n = re.sub(r"[^a-z0-9 ]", " ", n)
        return {
            w for w in n.split()
            if w and w not in _STOPWORDS_NOMBRE and len(w) > 2
        }

    @classmethod
    def _jaccard(cls, a: set[str], b: set[str]) -> float:
        if not a or not b:
            return 0.0
        return len(a & b) / len(a | b)

    def _cargar_anios_formacion(
        self, ruta_reports: str | pathlib.Path | None = None
    ) -> dict[str, int]:
        """
        Lee el año de formación de cada grupo desde la carpeta
        'data/reporte excel_<fecha>' MÁS RECIENTE (scrape actual de
        GrupLAC). Empareja el nombre de carpeta con `nombre_grupo` de
        medicion_957.xlsx por coincidencia exacta normalizada, contención,
        o solapamiento de palabras (Jaccard >= 0.5) para tolerar siglas y
        nombres truncados.
        """
        if ruta_reports is None:
            ruta_reports = _carpeta_gruplac_mas_reciente()
        if ruta_reports is None or not pathlib.Path(ruta_reports).exists():
            return {}
        ruta_reports = pathlib.Path(ruta_reports)

        anio_por_carpeta: dict[str, int] = {}
        for carpeta in ruta_reports.iterdir():
            if not carpeta.is_dir():
                continue
            archivos = list(carpeta.glob("*.xlsx"))
            if not archivos:
                continue
            try:
                wb = openpyxl.load_workbook(
                    archivos[0], read_only=True, data_only=True
                )
                ws = next(
                    (wb[n] for n in wb.sheetnames if "datos" in n.lower()), None
                )
                if ws is None:
                    continue
                for row in ws.iter_rows(values_only=True):
                    if row and row[0] and "formaci" in str(row[0]).lower():
                        m = re.search(r"(\d{4})", str(row[1]) or "")
                        if m:
                            anio_por_carpeta[carpeta.name] = int(m.group(1))
                        break
                wb.close()
            except Exception:
                continue

        resultado: dict[str, int] = {}
        for grupo in self.df_grupos["nombre_grupo"].unique():
            carpeta_match = self._mejor_coincidencia(grupo, list(anio_por_carpeta))
            if carpeta_match:
                resultado[grupo] = anio_por_carpeta[carpeta_match]
        return resultado

    @classmethod
    def _mejor_coincidencia(
        cls, nombre: str, candidatos: list[str]
    ) -> Optional[str]:
        """
        Empareja `nombre` contra una lista de `candidatos` (nombres de
        carpeta, de grupo en otra fuente, etc.) tolerando variaciones de
        capitalización, acrónimos y nombres truncados. Estrategia, en
        orden: (1) coincidencia exacta normalizada, (2) contención de un
        string en el otro, (3) solapamiento de palabras (Jaccard >= 0.5),
        (4) sigla entre paréntesis o como palabra suelta del candidato.
        Devuelve el candidato original (sin normalizar) o None.
        """
        n = normalizar_nombre(nombre)
        candidatos_norm = {normalizar_nombre(c): c for c in candidatos}

        if n in candidatos_norm:
            return candidatos_norm[n]

        for cn, orig in candidatos_norm.items():
            if len(n) > 4 and (n in cn or cn in n):
                return orig

        tokens_nombre = cls._tokens_significativos(nombre)
        mejor_score, mejor_orig = 0.0, None
        for orig in candidatos:
            score = cls._jaccard(tokens_nombre, cls._tokens_significativos(orig))
            if score > mejor_score:
                mejor_score, mejor_orig = score, orig
        if mejor_score >= 0.5:
            return mejor_orig

        # Siglas cortas (ej. "GEIO", "L'H"): buscar el nombre como sigla
        # entre paréntesis o como palabra suelta del candidato, p.ej.
        # "...OPERACIONES. (GEIO)".
        sigla = re.sub(r"[^a-z0-9]", "", n)
        if 1 < len(sigla) <= 6:
            for orig in candidatos:
                m = re.search(r"\(([^)]+)\)\s*$", orig)
                opciones = [m.group(1)] if m else []
                opciones += normalizar_nombre(orig).split()
                if any(
                    re.sub(r"[^a-z0-9]", "", normalizar_nombre(o)) == sigla
                    for o in opciones
                ):
                    return orig
        return None

    # ------------------------------------------------------------------
    # Auditoría: productos crudos de GrupLAC/ScienTI para un grupo, tal
    # cual fueron scrapeados (sin cuartil, sin λ) — solo para contrastar
    # visualmente contra el conteo oficial de la hoja `productos`. Los
    # nombres de grupo no coinciden exactamente entre medicion_957.xlsx y
    # gruplac_957.db (solo 5/67 coinciden exacto), así que se reusa el
    # mismo emparejamiento difuso de `_cargar_anios_formacion`.
    # ------------------------------------------------------------------

    def nombre_en_gruplac(self, nombre_grupo: str) -> Optional[str]:
        """Nombre tal como aparece en `productos_957.grupo` (gruplac_957.db),
        o None si no se encuentra una coincidencia razonable."""
        if not hasattr(self, "_cache_nombres_gruplac"):
            self._cache_nombres_gruplac: list[str] = []
            if GRUPLAC_DB_PATH.exists():
                with sqlite3.connect(GRUPLAC_DB_PATH) as conn:
                    self._cache_nombres_gruplac = [
                        r[0] for r in conn.execute(
                            "SELECT DISTINCT grupo FROM productos_957"
                        ).fetchall()
                    ]
        if not self._cache_nombres_gruplac:
            return None
        return self._mejor_coincidencia(nombre_grupo, self._cache_nombres_gruplac)

    def productos_gruplac(self, nombre_grupo: str) -> list[dict]:
        """
        Productos crudos del grupo en GrupLAC (tabla `productos_957`),
        ordenados por año descendente. Lista vacía si no hay match de
        nombre o no existe la BD.
        """
        nombre_gruplac = self.nombre_en_gruplac(nombre_grupo)
        if nombre_gruplac is None or not GRUPLAC_DB_PATH.exists():
            return []
        with sqlite3.connect(GRUPLAC_DB_PATH) as conn:
            conn.row_factory = sqlite3.Row
            filas = conn.execute(
                """
                SELECT categoria_957, subcategoria_957, tipo_producto_957,
                       anio, titulo
                FROM productos_957
                WHERE grupo = ?
                ORDER BY anio DESC
                """,
                (nombre_gruplac,),
            ).fetchall()
        return [dict(f) for f in filas]

    # ------------------------------------------------------------------
    # Paso 3 (deducido): C_j(área) = valor_indicador / Σλ
    # ------------------------------------------------------------------

    def _calcular_constantes_area(self) -> dict[tuple[str, str], float]:
        """
        Para cada (área, sección), calcula C_j(área) como la MEDIANA del
        ratio valor_indicador/Σλ entre los grupos del área con Σλ > 0.
        La mediana es robusta a errores de extracción puntuales del PDF
        (ver doc Apéndice C) sin perder la constancia esperada.
        """
        filas = []
        for (grupo, seccion), suma in self._sumlambda.items():
            if suma <= 1e-6:
                continue
            valor = self._valor_oficial.get((grupo, seccion))
            if valor is None:
                continue
            area = self._area_de.get(grupo)
            if not area or not isinstance(area, str):
                continue
            filas.append((area, seccion, valor / suma))

        constantes: dict[tuple[str, str], list[float]] = {}
        for area, seccion, ratio in filas:
            constantes.setdefault((area, seccion), []).append(ratio)

        return {
            clave: statistics.median(ratios)
            for clave, ratios in constantes.items()
        }

    # ------------------------------------------------------------------
    # Paso 4: máximo del área por indicador (de la hoja `cuartiles`)
    # ------------------------------------------------------------------

    def _calcular_maximos_area(self) -> dict[tuple[str, str], float]:
        resultado: dict[tuple[str, str], float] = {}
        for _, row in self.df_cuartiles.iterrows():
            indicador = CUARTIL_A_INDICADOR.get(row["cuartil"])
            if not indicador or indicador == "IG":
                continue
            grupo = row["grupo"]
            area = self._area_de.get(grupo)
            if not area:
                continue
            clave = (area, indicador)
            valor = row["max"]
            if pd.notna(valor) and (clave not in resultado or valor > resultado[clave]):
                resultado[clave] = float(valor)

        # cohesión y colaboración NO tienen fila en la hoja `cuartiles`
        # (esa hoja viene del Paso 6, y esos dos índices solo aparecen en
        # el Paso 4/5). Su máximo de área sí está en `indicadores.valor_maximo`
        # (Paso 4, "Máximo" de la tabla de índices) -> se usa como respaldo.
        for r in self.df_indicadores.itertuples():
            if r.indicador not in ("cohesion", "colaboracion"):
                continue
            if pd.isna(r.valor_maximo):
                continue
            area = self._area_de.get(r.grupo)
            if not area:
                continue
            clave = (area, r.indicador)
            if clave not in resultado or r.valor_maximo > resultado[clave]:
                resultado[clave] = float(r.valor_maximo)
        return resultado

    def _calcular_cuartiles_area(self) -> dict[str, dict[str, dict]]:
        """{area: {indicador_o_IG: {min,q4,q3,q2,max}}}"""
        resultado: dict[str, dict[str, dict]] = {}
        for _, row in self.df_cuartiles.iterrows():
            indicador = CUARTIL_A_INDICADOR.get(row["cuartil"])
            if not indicador:
                continue
            grupo = row["grupo"]
            area = self._area_de.get(grupo)
            if not area:
                continue
            resultado.setdefault(area, {})
            if indicador not in resultado[area]:
                resultado[area][indicador] = {
                    k: float(row[k]) for k in ("min", "q4", "q3", "q2", "max")
                    if pd.notna(row[k])
                }
        return resultado

    @staticmethod
    def _cuartil_de(valor: float, cuartiles: dict) -> int:
        """1 = cuartil superior (25% mejor), 4 = inferior."""
        if not cuartiles:
            return 4
        if valor >= cuartiles.get("q2", float("inf")):
            return 1
        if valor >= cuartiles.get("q3", float("inf")):
            return 2
        if valor >= cuartiles.get("q4", float("inf")):
            return 3
        return 4

    # ------------------------------------------------------------------
    # Pasos 3-6 combinados para un grupo
    # ------------------------------------------------------------------

    def simular(
        self, nombre_grupo: str, modo: str = "verificacion"
    ) -> ResultadoSimulacion:
        """
        modo="verificacion" (por defecto): usa el `valor_indicador` OFICIAL
            de la hoja `indicadores` para los 7 indicadores de producto.
            Es la forma correcta de comprobar si la lógica de cuartiles y
            condiciones del Paso 6 reproduce la categoría real — sin mezclar
            el error de reconstrucción de Σλ×C_j(área) con errores de la
            lógica de condiciones (doc §3.4: el valor oficial YA está en
            medicion_957.xlsx, no hace falta deducirlo para los 75 grupos
            que sí tienen PDF).
        modo="proyeccion": reconstruye valor_indicador = Σλ × C_j(área)
            (doc §10.2-§10.3). Sirve para el simulador de impacto: proyectar
            qué pasaría si el grupo agrega productos que HOY no existen
            (donde no hay valor oficial que consultar). No usar este modo
            para verificar grupos que ya tienen datos oficiales.
        """
        if modo not in ("verificacion", "proyeccion"):
            raise ValueError("modo debe ser 'verificacion' o 'proyeccion'")

        area = self._area_de.get(nombre_grupo)
        valores_sim: dict[str, float] = {}

        if modo == "verificacion":
            for seccion in SECCIONES_CON_PRODUCTOS:
                valores_sim[seccion] = self._valor_oficial.get(
                    (nombre_grupo, seccion), 0.0
                ) or 0.0
        else:
            # Secciones derivadas de Σλ × C_j(área) -- Paso 2 -> Paso 3
            for seccion in SECCIONES_CON_PRODUCTOS:
                suma = self._sumlambda.get((nombre_grupo, seccion), 0.0)
                c_j = self._constantes_area.get((area, seccion))
                valores_sim[seccion] = suma * c_j if c_j else 0.0

        # FRH_A, FRH_B, cohesión, colaboración: sin desglose de productos
        # en el PDF (doc §2.4/§10.6) -> siempre se usa el valor oficial,
        # en ambos modos, porque no hay Σλ de la que partir.
        for indicador in ("FRH_A", "FRH_B", "cohesion", "colaboracion"):
            valores_sim[indicador] = self._valor_oficial.get(
                (nombre_grupo, indicador), 0.0
            ) or 0.0

        return self._evaluar_paso_4_a_6(nombre_grupo, area, valores_sim)

    # ------------------------------------------------------------------
    # Modo proyección: agregar productos hipotéticos a un grupo y ver el
    # impacto en sus indicadores, IG y categoría (doc §10).
    # ------------------------------------------------------------------

    def proyectar_productos(
        self, nombre_grupo: str, ajustes: dict[str, int]
    ) -> ResultadoSimulacion:
        """
        Proyecta el efecto de agregar `n` unidades nuevas de cada subtipo
        en `ajustes` (ej. {"ART_A1": 5, "SF": 2}) sobre los indicadores,
        el IG y la categoría del grupo.

        Parte siempre del valor OFICIAL como base (no de la reconstrucción
        Σλ×C_j del área) y le suma el incremento proyectado:

            ratio_grupo = valor_indicador_oficial / Σλ_actual   (doc §10.2,
                es la constante C_j(área) propia de este grupo/sección —
                más precisa que la mediana del área porque usa el dato
                real del grupo, no una estimación cruzada)
            Σλ_simulado = Σλ_actual + Σ Δλ(subtipo) para cada ajuste
            valor_simulado = ratio_grupo × Σλ_simulado

        Con ajustes vacíos, valor_simulado == valor_oficial exactamente
        (doc §10.4): este método también sirve para verificar que no hay
        desfase de escala antes de proyectar cambios reales.

        Subtipos desconocidos (no presentes en ninguna fila oficial de
        `productos`, p. ej. porque ningún grupo de la convocatoria reportó
        ese subtipo) lanzan ValueError: no hay ventana ni sección a la que
        asignarlos.
        """
        area = self._area_de.get(nombre_grupo)

        deltas_por_seccion: dict[str, float] = {}
        detalle_subtipos: dict[str, dict] = {}
        for subtipo, n in ajustes.items():
            if subtipo not in self.ventanas_subtipo:
                raise ValueError(
                    f"Subtipo desconocido: {subtipo!r}. No aparece en "
                    f"ninguna fila oficial de la hoja 'productos', así que "
                    f"no se conoce su ventana ni su sección 957."
                )
            ventana = self.ventanas_subtipo[subtipo]
            seccion = self._seccion_de_subtipo[subtipo]
            total_actual = self._total_por_grupo_subtipo.get(
                (nombre_grupo, subtipo), 0
            )
            total_nuevo = total_actual + n
            lambda_actual = math.log(1 + total_actual / ventana)
            lambda_nuevo = math.log(1 + total_nuevo / ventana)
            delta = lambda_nuevo - lambda_actual
            deltas_por_seccion[seccion] = deltas_por_seccion.get(seccion, 0.0) + delta
            detalle_subtipos[subtipo] = {
                "seccion": seccion,
                "ventana": ventana,
                "total_actual": total_actual,
                "total_simulado": total_nuevo,
                "lambda_actual": round(lambda_actual, 6),
                "lambda_simulado": round(lambda_nuevo, 6),
                "delta_lambda": round(delta, 6),
            }

        valores_sim: dict[str, float] = {}
        for seccion in SECCIONES_CON_PRODUCTOS:
            valor_oficial = self._valor_oficial.get((nombre_grupo, seccion), 0.0) or 0.0
            delta_seccion = deltas_por_seccion.get(seccion, 0.0)
            if delta_seccion == 0.0:
                valores_sim[seccion] = valor_oficial
                continue
            suma_actual = self._sumlambda.get((nombre_grupo, seccion), 0.0)
            if suma_actual > 1e-9:
                ratio_grupo = valor_oficial / suma_actual
            else:
                ratio_grupo = self._constantes_area.get((area, seccion), 0.0)
            suma_simulada = suma_actual + delta_seccion
            valores_sim[seccion] = ratio_grupo * suma_simulada

        # FRH_A, FRH_B, cohesión, colaboración no se proyectan: no tienen
        # desglose de productos (doc §2.4/§10.6) -> quedan en su valor
        # oficial actual, sin cambios.
        for indicador in ("FRH_A", "FRH_B", "cohesion", "colaboracion"):
            valores_sim[indicador] = self._valor_oficial.get(
                (nombre_grupo, indicador), 0.0
            ) or 0.0

        resultado = self._evaluar_paso_4_a_6(nombre_grupo, area, valores_sim)
        resultado.detalle_ajustes = detalle_subtipos
        return resultado

    def _evaluar_paso_4_a_6(
        self, nombre_grupo: str, area: Optional[str], valores_sim: dict[str, float]
    ) -> ResultadoSimulacion:
        """Pasos 4 (índices), 5 (IG) y 6 (cuartiles + condiciones -> categoría)."""
        # Paso 4: índices. Preferir el máximo oficial reportado para ESTE
        # grupo (Paso 4 del PDF); si no está (grupo sin PDF o modo
        # proyección), usar el máximo del área estimado desde `cuartiles`.
        indices_sim: dict[str, float] = {}
        for indicador, valor in valores_sim.items():
            maximo = self._maximo_oficial_por_grupo.get(
                (nombre_grupo, indicador)
            ) or self._maximos_area.get((area, indicador))
            indices_sim[indicador] = (valor / maximo) if maximo else 0.0

        # Paso 5: IG
        ig_sim = sum(
            self.ponderaciones[ind] * indices_sim.get(ind, 0.0)
            for ind in self.ponderaciones
        )

        # Paso 6: cuartiles + condiciones
        cuartiles_area = self._cuartiles_area.get(area, {})
        cuartil_ig = self._cuartil_de(ig_sim, cuartiles_area.get("IG", {}))
        cuartil_top = self._cuartil_de(valores_sim["NC_TOP"], cuartiles_area.get("NC_TOP", {}))

        anio_formacion = self._anio_formacion.get(nombre_grupo)
        anios_existencia = (
            self.año_base - anio_formacion
            if anio_formacion is not None else None
        )

        cuartil_frh_b = self._cuartil_de(
            valores_sim["FRH_B"], cuartiles_area.get("FRH_B", {})
        )

        condiciones = self._evaluar_condiciones(
            valores_sim, ig_sim, cuartil_ig, cuartil_top, cuartil_frh_b, anios_existencia
        )
        categoria_sim = self._categoria_desde_condiciones(condiciones)

        cat_oficial = self._categoria_oficial.get(nombre_grupo)
        if pd.isna(cat_oficial):
            cat_oficial = None

        return ResultadoSimulacion(
            grupo=nombre_grupo,
            area=area,
            categoria_oficial=cat_oficial,
            categoria_simulada=categoria_sim,
            ig_oficial=self._ig_oficial.get(nombre_grupo),
            ig_simulado=round(ig_sim, 6),
            valores_simulados={k: round(v, 4) for k, v in valores_sim.items()},
            indices_simulados={k: round(v, 6) for k, v in indices_sim.items()},
            cuartil_ig=cuartil_ig,
            cuartil_nc_top=cuartil_top,
            anios_existencia=anios_existencia,
            condiciones=condiciones,
            coincide=(categoria_sim == cat_oficial),
        )

    @staticmethod
    def _evaluar_condiciones(
        valores: dict[str, float],
        ig_sim: float,
        cuartil_ig: int,
        cuartil_top: int,
        cuartil_frh_b: int,
        anios_existencia: Optional[float],
    ) -> dict[str, dict]:
        """
        Condiciones tomadas LITERALMENTE del texto "Esta categoría fue
        alcanzada bajo las siguientes condiciones:" de cada PDF
        (data/pdf/A1s.pdf, As.pdf, Bs.pdf, Cs.pdf) — no son una deducción,
        son la transcripción exacta del reporte ScienTI.

        La condición de "investigador emérito/sénior/asociado/junior
        vinculado contractualmente" NO se evalúa. Confirmado contra el
        documento conceptual oficial (Modelo_medicion/Modelo de medición.pdf,
        Tabla 1, págs. 45-48): clasificar a una persona como Sénior/
        Asociado/Junior requiere datos de su CvLAC INDIVIDUAL — nivel de
        formación (doctorado/maestría finalizado), producción TOP/A en
        TODA su trayectoria académica (no solo la ventana de la
        convocatoria), y conteo de tesis/trabajos de grado dirigidos en
        ventanas de 5-10 años. Ninguno de estos datos existe en GrupLAC a
        nivel de grupo: la hoja "Integrantes del grupo" solo trae
        nombre/vinculación/fechas, no la hoja de vida CvLAC de la persona.
        Se omite la condición en vez de asumirse verdadera, para que quede
        explícito que la categoría simulada es un techo, no una garantía.
        """
        nc_top, nc_a, nc_b = valores["NC_TOP"], valores["NC_A"], valores["NC_B"]
        asc, dpc = valores["ASC"], valores["DPC"]
        frh_a, frh_b = valores["FRH_A"], valores["FRH_B"]
        cohesion = valores["cohesion"]

        existencia_ok = (
            (lambda min_anios: anios_existencia is not None and anios_existencia >= min_anios)
        )

        condiciones: dict[str, dict] = {}

        condiciones["A1"] = {
            "IG en cuartil 1 (25% superior)": cuartil_ig <= 1,
            "NC_TOP en cuartil 1 (25% superior)": cuartil_top <= 1,
            "ASC>0 o DPC>0": asc > 0 or dpc > 0,
            "FRH_A>0": frh_a > 0,
            "Cohesión>0": cohesion > 0,
            "Al menos 5 años de existencia": existencia_ok(5),
        }
        condiciones["A"] = {
            "IG en o sobre cuartil 2 (50% superior)": cuartil_ig <= 2,
            # El PDF lista "NC_TOP>0 o NC_A>0 o DPC>0" y "ASC>0" en bullets
            # separados, pero el caso real "Arte y Cultura" (ASC=0, DPC>0,
            # categoría oficial A) demuestra que en la práctica es un solo
            # grupo OR entre los 4: basta con que algún producto exista en
            # nuevo conocimiento, apropiación social o divulgación.
            "NC_TOP>0 o NC_A>0 o ASC>0 o DPC>0": (
                nc_top > 0 or nc_a > 0 or asc > 0 or dpc > 0
            ),
            "FRH_A>0": frh_a > 0,
            "Cohesión>0": cohesion > 0,
            "Al menos 5 años de existencia": existencia_ok(5),
        }
        condiciones["B"] = {
            "IG en o sobre cuartil 3 (75% superior)": cuartil_ig <= 3,
            "NC_TOP>0 o NC_A>0": nc_top > 0 or nc_a > 0,
            "ASC>0 o DPC>0": asc > 0 or dpc > 0,
            "FRH_A>0 o FRH_B en o sobre cuartil 2": frh_a > 0 or cuartil_frh_b <= 2,
            "Cohesión>0": cohesion > 0,
            "Al menos 3 años de existencia": existencia_ok(3),
        }
        condiciones["C"] = {
            "IG>0": ig_sim > 0,
            "NC_TOP>0 o NC_A>0": nc_top > 0 or nc_a > 0,
            "ASC>0 o DPC>0": asc > 0 or dpc > 0,
            "FRH_A>0 o FRH_B>0": frh_a > 0 or frh_b > 0,
            "Al menos 2 años de existencia": existencia_ok(2),
        }
        return condiciones

    @staticmethod
    def _categoria_desde_condiciones(condiciones: dict[str, dict]) -> str:
        for cat in reversed(ORDEN_CATEGORIAS[1:]):  # A1, A, B, C
            if cat in condiciones and all(condiciones[cat].values()):
                return cat
        return "D"

    # ------------------------------------------------------------------
    # Verificación masiva
    # ------------------------------------------------------------------

    def verificar_todos(self) -> pd.DataFrame:
        """
        Corre la simulación para todos los grupos con categoría oficial
        conocida y devuelve un DataFrame comparando simulado vs. oficial.
        """
        filas = []
        for grupo in self.df_grupos["nombre_grupo"].unique():
            cat_oficial = self._categoria_oficial.get(grupo)
            if pd.isna(cat_oficial) or cat_oficial in (None, "00"):
                continue
            r = self.simular(grupo)
            filas.append({
                "grupo": grupo,
                "area": r.area,
                "categoria_oficial": r.categoria_oficial,
                "categoria_simulada": r.categoria_simulada,
                "coincide": r.coincide,
                "ig_oficial": r.ig_oficial,
                "ig_simulado": r.ig_simulado,
            })
        return pd.DataFrame(filas)

    # Prefijo de subtipo oficial -> tipo_producto_957 de GrupLAC, solo para
    # los tipos de producto que NO requieren clasificación de calidad
    # (cuartil de revista, etc.) para identificarse, así que el conteo
    # crudo de GrupLAC es comparable sin necesidad de ese dato faltante.
    MAPEO_SUBTIPO_A_GRUPLAC = {
        "ART_": "Artículos de investigación",
        "CAP_LIB_": "Capítulos en libro resultado de investigación",
        "SF": "Software registrado",
    }

    def verificar_contra_gruplac(self) -> pd.DataFrame:
        """
        Cruza, para cada grupo emparejado con GrupLAC, el total OFICIAL
        (dentro de ventana) contra el conteo crudo histórico de GrupLAC
        para los tipos de producto de `MAPEO_SUBTIPO_A_GRUPLAC` (los que no
        requieren cuartil de revista para identificarse).

        Como el oficial está acotado a una ventana y GrupLAC es histórico
        completo, debe cumplirse siempre `oficial <= gruplac_total`. Una
        violación señala un problema de datos: extracción duplicada del
        PDF (como ocurrió con "Desarrollo de Procesos Quimicos", ya
        corregido en `_deduplicar_grupos_repetidos`) o un producto editado/
        retirado en GrupLAC después de la medición oficial.
        """
        if not GRUPLAC_DB_PATH.exists():
            return pd.DataFrame()

        filas = []
        with sqlite3.connect(GRUPLAC_DB_PATH) as conn:
            for grupo in self.df_grupos["nombre_grupo"].unique():
                nombre_gl = self.nombre_en_gruplac(grupo)
                if nombre_gl is None:
                    continue
                for prefijo, tipo in self.MAPEO_SUBTIPO_A_GRUPLAC.items():
                    sub = self.df_productos[
                        (self.df_productos["grupo"] == grupo)
                        & (self.df_productos["subtipo"].str.startswith(prefijo))
                    ]
                    of_total = int(sub["total"].sum())
                    gl_total = conn.execute(
                        "SELECT COUNT(*) FROM productos_957 WHERE grupo = ? AND tipo_producto_957 = ?",
                        (nombre_gl, tipo),
                    ).fetchone()[0]
                    if of_total > 0 or gl_total > 0:
                        filas.append({
                            "grupo": grupo,
                            "nombre_gruplac": nombre_gl,
                            "tipo": tipo,
                            "oficial_ventana": of_total,
                            "gruplac_total": gl_total,
                            "violacion": of_total > gl_total,
                        })
        return pd.DataFrame(filas)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _imprimir_reporte(df: pd.DataFrame) -> None:
    total = len(df)
    aciertos = int(df["coincide"].sum())
    print(f"\n=== Verificación simulador 957 ===")
    print(f"Grupos evaluados: {total}")
    print(f"Coinciden exactamente con la categoría oficial: {aciertos} ({aciertos/total:.1%})")

    print("\nMatriz categoría oficial -> simulada:")
    print(pd.crosstab(df["categoria_oficial"], df["categoria_simulada"]).to_string())

    fallos = df[~df["coincide"]]
    if not fallos.empty:
        print(f"\nGrupos con discrepancia ({len(fallos)}):")
        cols = ["grupo", "area", "categoria_oficial", "categoria_simulada", "ig_oficial", "ig_simulado"]
        print(fallos[cols].to_string(index=False))


def _imprimir_reporte_gruplac(df: pd.DataFrame) -> None:
    if df.empty:
        return
    print("\n=== Cruce contra GrupLAC (productos sin necesidad de cuartil) ===")
    print(df.groupby("tipo").agg(n=("grupo", "count"), violaciones=("violacion", "sum")))
    violaciones = df[df["violacion"]]
    if not violaciones.empty:
        print(f"\nViolaciones (oficial > histórico GrupLAC, no debería pasar) ({len(violaciones)}):")
        cols = ["grupo", "tipo", "oficial_ventana", "gruplac_total"]
        print(violaciones[cols].to_string(index=False))
    else:
        print("\nSin violaciones.")


if __name__ == "__main__":
    sim = Simulador957()
    df_resultado = sim.verificar_todos()
    _imprimir_reporte(df_resultado)
    _imprimir_reporte_gruplac(sim.verificar_contra_gruplac())
    
    

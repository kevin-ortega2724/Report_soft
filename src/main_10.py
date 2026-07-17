import json
import os
import re
import sqlite3
import sys
import warnings
from datetime import datetime

import pandas as pd
from pathlib import Path
from unidecode import unidecode
from PyQt5.QtCore import Qt, QThread, QTimer, pyqtSignal
from PyQt5.QtGui import QFont
from PyQt5.QtWidgets import (
    QAbstractItemView, QApplication, QCheckBox, QComboBox, QDialog,
    QFileDialog, QHeaderView, QHBoxLayout, QLabel, QLineEdit, QMainWindow,
    QMessageBox, QProgressBar, QPushButton, QSpinBox, QSplitter,
    QTabWidget, QTableWidget, QTableWidgetItem, QTextEdit,
    QVBoxLayout, QWidget,
)

from analisis_seguimiento import AnalisisDuplicados
from views.vista_clasificacion_minciencias import VistaClasificacionMinCiencias
from views.vista_inicio import VistaInicio
from views.vista_seguimiento_grupos import VistaSeguimientoGrupos
from views.vista_simulador_957 import VistaSimulador957
# from chatbot_investigacion import VistaChatbotInvestigacion  # pestaña retirada por ahora, ver setup_ui
# from estadisticas_957 import VistaEstadisticas957  # se reconstruye como UI separada en UI_clasificacion/
from views.vista_visor_gruplac_957 import VisorGrupLAC957

from constants import COLUMNAS_CONOCIDAS_POR_CATEGORIA

warnings.filterwarnings("ignore")

from utils import (
    limpiar_cedula,
    limpiar_nombre_archivo,
    limpiar_texto,
    norm_text,
    normalizar_columna,
    normalizar_nombre,
    obtener_directorio_base,
)
from comparador_faltantes import normalize_text

# ==================== BASE DE DATOS ====================

class DatabaseManager:
    def __init__(self, db_path="data/db/academia_utp_integrado.db"):
        base_dir = obtener_directorio_base()
        db_file = base_dir / db_path
        self.db_path = str(db_file)
        if not db_file.exists():
            raise FileNotFoundError(f"No se encontró la BD: {db_file}")
        self.conn = sqlite3.connect(self.db_path, check_same_thread=False)
        self._aplicar_pragma()
        self.crear_tablas()
        self.mapa_cedulas = {}

    def _aplicar_pragma(self):
        """Configura SQLite para máximo rendimiento (seguro en lecturas y escrituras normales)."""
        self.conn.executescript("""
            PRAGMA journal_mode  = WAL;
            PRAGMA synchronous   = NORMAL;
            PRAGMA cache_size    = -32000;
            PRAGMA temp_store    = MEMORY;
            PRAGMA mmap_size     = 268435456;
            PRAGMA page_size     = 4096;
        """)

    
    def crear_tablas(self):
        cursor = self.conn.cursor()
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS personas (
                cedula TEXT PRIMARY KEY,
                nombre TEXT NOT NULL,
                email TEXT,
                facultad TEXT,
                tipo_persona TEXT
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS cedulas_duplicadas (
                cedula_duplicada TEXT PRIMARY KEY,
                cedula_principal TEXT,
                FOREIGN KEY (cedula_principal) REFERENCES personas (cedula)
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS grupos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                cedula TEXT,
                grupo TEXT,
                facultad TEXT,
                tipo_miembro TEXT,
                UNIQUE(cedula, grupo),
                FOREIGN KEY (cedula) REFERENCES personas (cedula)
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS publicaciones (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                cedula TEXT,
                titulo TEXT,
                revista_libro TEXT,
                doi_url TEXT,
                issn_isbn TEXT,
                año INTEGER,
                tipo TEXT,
                categoria TEXT,
                estado TEXT,
                grupo TEXT,
                fuente TEXT,
                FOREIGN KEY (cedula) REFERENCES personas (cedula)
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS extensiones (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                cedula TEXT,
                actividad TEXT,
                tipo TEXT,
                modalidad TEXT,
                estado TEXT,
                fecha_inicio TEXT,
                fecha_fin TEXT,
                año INTEGER,
                poblacion TEXT,
                grupo TEXT,
                facultad TEXT,
                financiacion_interna TEXT,
                financiacion_externa TEXT,
                fuente TEXT,
                FOREIGN KEY (cedula) REFERENCES personas (cedula)
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS trabajos_grado (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                cedula_director TEXT,
                nombre_director TEXT,
                cedula_estudiante TEXT,
                nombre_estudiante TEXT,
                titulo TEXT,
                programa TEXT,
                año INTEGER,
                estado TEXT,
                fecha_sustentacion TEXT,
                calificacion TEXT,
                facultad TEXT,
                fuente TEXT,
                FOREIGN KEY (cedula_director) REFERENCES personas (cedula)
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS productos_innovacion (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                cedula TEXT,
                tipo_producto TEXT,
                nombre TEXT,
                descripcion TEXT,
                año INTEGER,
                estado TEXT,
                grupo TEXT,
                fuente TEXT,
                FOREIGN KEY (cedula) REFERENCES personas (cedula)
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS proyectos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                cedula TEXT,
                responsable TEXT,
                titulo TEXT,
                objetivo TEXT,
                codigo_cie TEXT,
                tipo TEXT,
                año INTEGER,
                fecha_inicio TEXT,
                fecha_fin TEXT,
                estado TEXT,
                facultad TEXT,
                grupo TEXT,
                valor_aprobado TEXT,
                fuente TEXT,
                FOREIGN KEY (cedula) REFERENCES personas (cedula)
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS propiedad_intelectual (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                responsable TEXT,
                tipo_producto TEXT,
                tipo_patente TEXT,
                nombre_producto TEXT,
                numero_registro TEXT,
                proyecto TEXT,
                fecha_aprobacion TEXT,
                entidad TEXT,
                facultad TEXT,
                fuente TEXT
            )
        ''')
        
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS configuracion (
                clave TEXT PRIMARY KEY,
                valor TEXT
            )
        """)

        cursor.execute("CREATE INDEX IF NOT EXISTS idx_publicaciones_grupo ON publicaciones(grupo)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_extensiones_grupo ON extensiones(grupo)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_productos_innovacion_grupo ON productos_innovacion(grupo)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_proyectos_grupo ON proyectos(grupo)")

        self.conn.commit()
        self._migrar_columnas_adicionales()
        self._reparar_cedulas_duplicadas_huerfanas()

    def _reparar_cedulas_duplicadas_huerfanas(self):
        """
        Repara el daño de un bug histórico en fusionar_personas(): al volver a
        ejecutar la consolidación de duplicados sobre un par ya fusionado, el
        código resolvía cedula_dup a través del mapeo existente y terminaba
        igual a cedula_principal; el DELETE FROM personas final borraba
        entonces a la propia persona "principal" en cada recarga de datos.

        Por cada fila de cedulas_duplicadas cuyo cedula_principal ya no existe
        en personas:
          - Si cedula_duplicada SÍ existe en personas: es la identidad que
            sobrevivió. Se le reasignan los datos (grupos, publicaciones, etc.)
            que quedaron huérfanos bajo cedula_principal, y se borra el mapeo
            (ya no hay un duplicado real, solo una persona).
          - Si ninguna de las dos existe: no hay nada que mostrar todavía (los
            datos quedan huérfanos hasta que se vuelva a cargar el archivo
            fuente con esa cédula); solo se limpia el mapeo inválido.
        """
        cursor = self.conn.cursor()
        filas = cursor.execute(
            "SELECT cedula_duplicada, cedula_principal FROM cedulas_duplicadas"
        ).fetchall()

        tablas_cedula = ["grupos", "publicaciones", "extensiones",
                          "productos_innovacion", "proyectos"]
        reparados = 0
        for cedula_dup, cedula_principal in filas:
            existe_principal = cursor.execute(
                "SELECT 1 FROM personas WHERE cedula = ?", (cedula_principal,)
            ).fetchone()
            if existe_principal:
                continue  # mapeo sano, no tocar

            existe_dup = cursor.execute(
                "SELECT 1 FROM personas WHERE cedula = ?", (cedula_dup,)
            ).fetchone()
            if existe_dup and cedula_dup != cedula_principal:
                for tabla in tablas_cedula:
                    cursor.execute(
                        f"UPDATE OR IGNORE {tabla} SET cedula = ? WHERE cedula = ?",
                        (cedula_dup, cedula_principal),
                    )
                cursor.execute(
                    "UPDATE trabajos_grado SET cedula_director = ? WHERE cedula_director = ?",
                    (cedula_dup, cedula_principal),
                )
                cursor.execute(
                    "UPDATE trabajos_grado SET cedula_estudiante = ? WHERE cedula_estudiante = ?",
                    (cedula_dup, cedula_principal),
                )
                reparados += 1

            cursor.execute(
                "DELETE FROM cedulas_duplicadas WHERE cedula_duplicada = ? AND cedula_principal = ?",
                (cedula_dup, cedula_principal),
            )

        self.conn.commit()
        return reparados

    def _migrar_columnas_adicionales(self):
        """
        Agrega la columna 'datos_adicionales' (JSON con columnas de Excel no
        reconocidas) a las tablas alimentadas por archivo, si todavía no existe.
        CREATE TABLE IF NOT EXISTS no agrega columnas a tablas ya creadas en
        versiones anteriores de la BD, por eso se necesita este ALTER explícito.
        """
        tablas = ["publicaciones", "extensiones", "productos_innovacion", "propiedad_intelectual", "trabajos_grado"]
        cursor = self.conn.cursor()
        for tabla in tablas:
            columnas = {fila[1] for fila in cursor.execute(f"PRAGMA table_info({tabla})")}
            if "datos_adicionales" not in columnas:
                cursor.execute(f"ALTER TABLE {tabla} ADD COLUMN datos_adicionales TEXT")

        # 'origen' distingue si una fila de "grupos" viene directo del
        # archivo de integrantes ('directo') o fue añadida a partir del
        # mapeo semillero->grupo adscrito (ver
        # CargadorDatosIntegrado._grupo_adscrito_semilleros): sin esto, una
        # membresía sintética (la persona en realidad solo está en el
        # semillero) es indistinguible de una membresía real al grupo, y los
        # reportes no podrían avisar "este producto llega por el semillero,
        # no por el grupo". Filas viejas (anteriores a esta columna) quedan
        # NULL y se tratan como 'directo' (todas lo eran, el mapeo semillero
        # no existía todavía).
        columnas_grupos = {fila[1] for fila in cursor.execute("PRAGMA table_info(grupos)")}
        if "origen" not in columnas_grupos:
            cursor.execute("ALTER TABLE grupos ADD COLUMN origen TEXT")

        self.conn.commit()
        self._migrar_indices_unicos()

    # Columnas que identifican una fila como "la misma" entre cargas (para que
    # al volver a cargar un archivo se actualice/acumule en vez de duplicarse).
    _CLAVES_DEDUP = {
        "publicaciones": ("cedula", "titulo", "año"),
        "extensiones": ("cedula", "actividad", "fecha_inicio"),
        "productos_innovacion": ("cedula", "nombre", "año", "grupo"),
        "proyectos": ("cedula", "titulo", "año"),
        "propiedad_intelectual": ("responsable", "nombre_producto", "numero_registro"),
        "trabajos_grado": ("cedula_estudiante", "titulo"),
    }

    def _migrar_indices_unicos(self):
        """
        Crea un índice único por tabla (ver _CLAVES_DEDUP) para poder usar
        INSERT OR REPLACE como mecanismo de "actualizar si existe, agregar si
        es nuevo": al recargar un archivo, los registros ya guardados nunca se
        borran; solo se actualizan si cambiaron y se acumulan los nuevos.
        Antes de crear el índice se eliminan duplicados exactos preexistentes
        (misma combinación de columnas clave), conservando la fila de menor id.
        """
        cursor = self.conn.cursor()
        for tabla, columnas in self._CLAVES_DEDUP.items():
            col_list = ", ".join(columnas)
            cursor.execute(f"""
                DELETE FROM {tabla}
                WHERE id NOT IN (SELECT MIN(id) FROM {tabla} GROUP BY {col_list})
            """)
            cursor.execute(
                f"CREATE UNIQUE INDEX IF NOT EXISTS ux_{tabla}_dedup ON {tabla}({col_list})"
            )
        self.conn.commit()

    # ── Fuentes adicionales (archivos que se acumulan sin reemplazar) ──

    def obtener_fuentes_adicionales(self, clave: str) -> list:
        """Archivos adicionales registrados para una categoría de
        ARCHIVOS_FUENTE_957 (más allá del archivo canónico), cada uno con
        su propia ruta única -- se procesan todos, ninguno reemplaza a otro."""
        row = self.conn.execute(
            "SELECT valor FROM configuracion WHERE clave = ?",
            (f"fuentes_adicionales_{clave}",),
        ).fetchone()
        if not row:
            return []
        try:
            return json.loads(row[0])
        except Exception:
            return []

    def registrar_fuente_adicional(self, clave: str, ruta: str, nombre_original: str):
        from datetime import datetime
        fuentes = self.obtener_fuentes_adicionales(clave)
        fuentes.append({
            "ruta": ruta,
            "nombre_original": nombre_original,
            "fecha_agregado": datetime.now().isoformat(timespec="seconds"),
        })
        self.conn.execute(
            "INSERT OR REPLACE INTO configuracion (clave, valor) VALUES (?, ?)",
            (f"fuentes_adicionales_{clave}", json.dumps(fuentes)),
        )
        self.conn.commit()

    # ── Caché de carga ───────────────────────────────────────────────────────

    def guardar_sello_carga(self, archivos: list):
        """
        Guarda los timestamps de los archivos procesados para detectar cambios.
        archivos: lista de rutas (str o Path).
        """
        import json
        from pathlib import Path as _P
        sello = {}
        for ruta in archivos:
            p = _P(ruta)
            if p.exists():
                sello[str(p)] = p.stat().st_mtime
        self.conn.execute(
            "INSERT OR REPLACE INTO configuracion (clave, valor) VALUES ('sello_carga', ?)",
            (json.dumps(sello),),
        )
        self.conn.commit()

    def cache_valida(self, archivos_a_verificar: list) -> bool:
        """
        Devuelve True si la BD ya tiene datos y ningún archivo fuente
        fue modificado desde la última carga completa.
        """
        import json
        from pathlib import Path as _P
        try:
            row = self.conn.execute(
                "SELECT valor FROM configuracion WHERE clave='sello_carga'"
            ).fetchone()
            if not row:
                return False
            n_personas = self.conn.execute(
                "SELECT COUNT(*) FROM personas"
            ).fetchone()[0]
            if n_personas == 0:
                return False
            sello = json.loads(row[0])
            for ruta in archivos_a_verificar:
                p = _P(ruta)
                if not p.exists():
                    continue
                mtime_actual = p.stat().st_mtime
                mtime_previo = sello.get(str(p))
                if mtime_previo is None or mtime_actual > mtime_previo:
                    return False
            return True
        except Exception:
            return False

    def obtener_cedula_principal(self, cedula):
        if cedula in self.mapa_cedulas:
            return self.mapa_cedulas[cedula]
        
        cursor = self.conn.cursor()
        resultado = cursor.execute(
            'SELECT cedula_principal FROM cedulas_duplicadas WHERE cedula_duplicada = ?',
            (cedula,)
        ).fetchone()
        
        if resultado:
            self.mapa_cedulas[cedula] = resultado[0]
            return resultado[0]
        
        return cedula

    def son_cedulas_duplicadas(self, cedula1, cedula2):
        if cedula1 == cedula2:
            return False
        if cedula1 + "0" == cedula2:
            return True
        if cedula2 + "0" == cedula1:
            return True
        return False
    
    def consolidar_duplicados(self):
        cursor = self.conn.cursor()
        personas = cursor.execute('SELECT cedula, nombre FROM personas ORDER BY nombre').fetchall()

        nombre_por_cedula = {c: n for c, n in personas}

        cedulas_por_nombre = {}
        for cedula, nombre in personas:
            nn = normalizar_nombre(nombre)
            cedulas_por_nombre.setdefault(nn, []).append(cedula)

        pairs = cursor.execute("""
            SELECT DISTINCT a.cedula AS c1, b.cedula AS c2
            FROM personas a
            JOIN personas b ON (b.cedula = a.cedula || '0' OR a.cedula = b.cedula || '0')
            WHERE a.cedula < b.cedula
        """).fetchall()

        adj = {}
        for c1, c2 in pairs:
            adj.setdefault(c1, set()).add(c2)
            adj.setdefault(c2, set()).add(c1)

        duplicados_encontrados = []
        global_processed = set()

        for nombre_norm, grupo_cedulas in cedulas_por_nombre.items():
            if len(grupo_cedulas) < 2:
                continue

            grupo_set = set(grupo_cedulas)

            for cedula in grupo_cedulas:
                if cedula in global_processed:
                    continue

                cluster = set()
                stack = [cedula]
                while stack:
                    cur = stack.pop()
                    if cur in global_processed or cur not in grupo_set:
                        continue
                    global_processed.add(cur)
                    cluster.add(cur)
                    for nb in adj.get(cur, set()):
                        if nb not in global_processed and nb in grupo_set:
                            stack.append(nb)

                if len(cluster) > 1:
                    cedula_principal = min(cluster, key=len)
                    cedulas_secundarias = [c for c in cluster if c != cedula_principal]
                    nombre_principal = nombre_por_cedula[cedula_principal]

                    self.fusionar_personas(cedula_principal, cedulas_secundarias, nombre_principal)

                    duplicados_encontrados.append({
                        'nombre': nombre_principal,
                        'cedula_principal': cedula_principal,
                        'cedulas': list(cluster),
                    })

        return duplicados_encontrados

    def fusionar_personas(self, cedula_principal, cedulas_duplicadas, nombre_principal):
        cursor = self.conn.cursor()
        
        try:
            # Actualizar nombre de la persona principal
            cursor.execute(
                'UPDATE personas SET nombre = ? WHERE cedula = ?',
                (nombre_principal, cedula_principal)
            )
            
            for cedula_dup in cedulas_duplicadas:
                cedula_dup = self.obtener_cedula_principal(cedula_dup)
                if cedula_dup == cedula_principal:
                    # Ya estaba fusionado en una corrida anterior: si se
                    # sigue de largo, las instrucciones de abajo (en
                    # particular el DELETE FROM personas) borrarían a la
                    # propia persona principal por error.
                    continue

                # Registrar duplicado
                cursor.execute('''
                    INSERT OR REPLACE INTO cedulas_duplicadas (cedula_duplicada, cedula_principal)
                    VALUES (?, ?)
                ''', (cedula_dup, cedula_principal))
                
                self.mapa_cedulas[cedula_dup] = cedula_principal
                
                # Mover datos - usar OR IGNORE para evitar conflictos de UNIQUE
                for tabla in ['grupos', 'publicaciones', 'extensiones', 'productos_innovacion', 'proyectos']:
                    try:
                        cursor.execute(f'UPDATE OR IGNORE {tabla} SET cedula = ? WHERE cedula = ?', 
                                     (cedula_principal, cedula_dup))
                    except:
                        pass
                
                # Trabajos de grado
                try:
                    cursor.execute('UPDATE trabajos_grado SET cedula_director = ? WHERE cedula_director = ?', 
                                 (cedula_principal, cedula_dup))
                except:
                    pass
                    
                try:
                    cursor.execute('UPDATE trabajos_grado SET cedula_estudiante = ? WHERE cedula_estudiante = ?', 
                                 (cedula_principal, cedula_dup))
                except:
                    pass
                
                # Copiar email y facultad si están vacíos en la principal
                persona_dup = cursor.execute(
                    'SELECT email, facultad FROM personas WHERE cedula = ?',
                    (cedula_dup,)
                ).fetchone()
                
                if persona_dup:
                    email_dup, facultad_dup = persona_dup
                    cursor.execute('''
                        UPDATE personas 
                        SET email = COALESCE(NULLIF(email, ''), ?),
                            facultad = COALESCE(NULLIF(facultad, ''), ?)
                        WHERE cedula = ?
                    ''', (email_dup, facultad_dup, cedula_principal))
                
                # Eliminar persona duplicada
                cursor.execute('DELETE FROM personas WHERE cedula = ?', (cedula_dup,))
            
            # Hacer commit una sola vez al final
            self.conn.commit()
            
        except sqlite3.OperationalError as e:
            # Si hay error, hacer rollback
            try:
                self.conn.rollback()
            except:
                pass
            # Re-raise el error para que se maneje arriba
            raise e
        except Exception as e:
            # Cualquier otro error
            try:
                self.conn.rollback()
            except:
                pass
            raise e

    def limpiar_datos(self):
        cursor = self.conn.cursor()
        tablas = ['grupos', 'publicaciones', 'extensiones', 'trabajos_grado', 
                  'productos_innovacion', 'proyectos', 'propiedad_intelectual', 
                  'cedulas_duplicadas', 'personas']
        for tabla in tablas:
            cursor.execute(f'DELETE FROM {tabla}')
        self.conn.commit()

    def insertar_persona(self, cedula, nombre, email="", facultad="", tipo=""):
        """Upsert de persona: INSERT si no existe, UPDATE campos vacíos si ya existe."""
        if not cedula:
            return
        cedula = self.obtener_cedula_principal(cedula)
        self.conn.execute(
            """
            INSERT INTO personas (cedula, nombre, email, facultad, tipo_persona)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(cedula) DO UPDATE SET
                email        = COALESCE(NULLIF(excluded.email, ''),        personas.email),
                facultad     = COALESCE(NULLIF(excluded.facultad, ''),     personas.facultad),
                tipo_persona = COALESCE(NULLIF(excluded.tipo_persona, ''), personas.tipo_persona)
            """,
            (cedula, nombre or "", email or "", facultad or "", tipo or ""),
        )
        # No se hace commit aquí; el llamador decide cuándo confirmar.

    def _upsert_personas_batch(self, filas):
        """Batch upsert de personas: [(cedula, nombre, email, facultad, tipo), ...]."""
        if not filas:
            return
        self.conn.executemany(
            """
            INSERT INTO personas (cedula, nombre, email, facultad, tipo_persona)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(cedula) DO UPDATE SET
                email        = COALESCE(NULLIF(excluded.email, ''),        personas.email),
                facultad     = COALESCE(NULLIF(excluded.facultad, ''),     personas.facultad),
                tipo_persona = COALESCE(NULLIF(excluded.tipo_persona, ''), personas.tipo_persona)
            """,
            filas,
        )

    def buscar_personas(self, termino):
        cursor = self.conn.cursor()
        termino_limpio = termino.strip()
        
        if termino_limpio.replace('.', '').replace(',', '').isdigit():
            cedula_limpia = limpiar_cedula(termino_limpio)
            query = '''
                SELECT cedula, nombre, email, facultad, tipo_persona
                FROM personas WHERE cedula LIKE ?
                ORDER BY nombre
            '''
            params = (f'%{cedula_limpia}%',)
        else:
            palabras = termino_limpio.split()
            if len(palabras) == 1:
                query = '''
                    SELECT cedula, nombre, email, facultad, tipo_persona
                    FROM personas WHERE LOWER(nombre) LIKE LOWER(?)
                    ORDER BY nombre
                '''
                params = (f'%{palabras[0]}%',)
            else:
                condiciones = ' AND '.join(['LOWER(nombre) LIKE LOWER(?)' for _ in palabras])
                query = f'''
                    SELECT cedula, nombre, email, facultad, tipo_persona
                    FROM personas WHERE {condiciones}
                    ORDER BY nombre
                '''
                params = tuple([f'%{p}%' for p in palabras])
        
        resultados = cursor.execute(query, params).fetchall()
        
        resultados_filtrados = []
        for resultado in resultados:
            cedula = resultado[0]
            # Solo se excluye si el "principal" al que apunta sigue existiendo
            # de verdad en personas; si no, mostrar esta fila es lo correcto
            # (evita ocultar personas por un mapeo de duplicado roto/huérfano).
            es_duplicada_valida = cursor.execute('''
                SELECT 1 FROM cedulas_duplicadas cd
                JOIN personas pp ON pp.cedula = cd.cedula_principal
                WHERE cd.cedula_duplicada = ?
            ''', (cedula,)).fetchone()

            if not es_duplicada_valida:
                resultados_filtrados.append(resultado)

        return resultados_filtrados

    def obtener_detalle_persona(self, cedula):
        cursor = self.conn.cursor()
        detalle = {}
        
        persona = cursor.execute(
            'SELECT * FROM personas WHERE cedula = ?', (cedula,)
        ).fetchone()
        
        if persona:
            detalle['info'] = {
                'cedula': persona[0],
                'nombre': persona[1],
                'email': persona[2] or 'No disponible',
                'facultad': persona[3] or 'No disponible',
                'tipo': persona[4] or 'No especificado'
            }
        
        detalle['grupos'] = cursor.execute(
            'SELECT * FROM grupos WHERE cedula = ? ORDER BY grupo', (cedula,)
        ).fetchall()
        
        detalle['publicaciones'] = cursor.execute(
            'SELECT * FROM publicaciones WHERE cedula = ? ORDER BY año DESC', (cedula,)
        ).fetchall()
        
        detalle['extensiones'] = cursor.execute(
            'SELECT * FROM extensiones WHERE cedula = ? ORDER BY año DESC', (cedula,)
        ).fetchall()
        
        detalle['trabajos_grado_director'] = cursor.execute(
            'SELECT * FROM trabajos_grado WHERE cedula_director = ? ORDER BY año DESC', (cedula,)
        ).fetchall()
        
        detalle['trabajos_grado_estudiante'] = cursor.execute(
            'SELECT * FROM trabajos_grado WHERE cedula_estudiante = ? ORDER BY año DESC', (cedula,)
        ).fetchall()
        
        detalle['productos_innovacion'] = cursor.execute(
            'SELECT * FROM productos_innovacion WHERE cedula = ? ORDER BY año DESC', (cedula,)
        ).fetchall()
        
        detalle['proyectos'] = cursor.execute(
            'SELECT * FROM proyectos WHERE cedula = ? ORDER BY año DESC', (cedula,)
        ).fetchall()
        
        return detalle

    def obtener_grupos(self):
        cursor = self.conn.cursor()
        return cursor.execute('''
            SELECT DISTINCT grupo FROM grupos 
            WHERE grupo IS NOT NULL AND grupo != ''
            ORDER BY grupo
        ''').fetchall()

    def obtener_integrantes_grupo(self, grupo):
        cursor = self.conn.cursor()
        return cursor.execute('''
            SELECT DISTINCT p.cedula, p.nombre, g.tipo_miembro, p.email, p.facultad
            FROM grupos g
            JOIN personas p ON g.cedula = p.cedula
            WHERE g.grupo = ?
            ORDER BY p.nombre
        ''', (grupo,)).fetchall()

    def _mapeo_semillero_grupo_adscrito(self):
        """Mapeo semillero(normalizado) -> grupo adscrito, guardado por
        CargadorDatosIntegrado._grupo_adscrito_semilleros al procesar
        'Reporte Semilleros con Grupo adscrito.xlsx'. Cacheado en memoria
        porque se consulta una vez por producto en los reportes."""
        if not hasattr(self, '_cache_mapeo_semillero'):
            fila = self.conn.execute(
                "SELECT valor FROM configuracion WHERE clave='mapeo_semillero_grupo_adscrito'"
            ).fetchone()
            try:
                self._cache_mapeo_semillero = json.loads(fila[0]) if fila else {}
            except Exception:
                self._cache_mapeo_semillero = {}
        return self._cache_mapeo_semillero

    def procedencia_grupo(self, cedula, grupo):
        """Para un producto atribuido a 'grupo' por esta 'cedula': ¿la
        persona pertenece a ese grupo de investigación directamente, solo
        por un semillero adscrito a él, o por ambos caminos? GrupLAC no
        rastrea semilleros como entidades propias, así que esta distinción
        le dice al lector del reporte de dónde sale realmente el producto.
        Devuelve 'Grupo', 'Semillero' o 'Grupo y Semillero'."""
        filas = self.conn.execute(
            "SELECT grupo, origen FROM grupos WHERE cedula = ?", (cedula,)
        ).fetchall()

        tiene_directo = any(
            g == grupo and (o is None or o == 'directo') for g, o in filas
        )

        mapa = self._mapeo_semillero_grupo_adscrito()
        tiene_semillero = any(
            "SEMILLERO" in (g or "").upper() and mapa.get(normalize_text(g)) == grupo
            for g, _ in filas
        )

        if tiene_directo and tiene_semillero:
            return "Grupo y Semillero"
        if tiene_semillero:
            return "Semillero"
        return "Grupo"

    def obtener_productos_grupo(self, grupo, filtros=None):
        if filtros is None:
            filtros = ['Publicaciones', 'Extensiones', 'Trabajos de Grado', 'Productos Innovación', 'Proyectos']
        
        cursor = self.conn.cursor()
        cedulas = cursor.execute(
            'SELECT DISTINCT cedula FROM grupos WHERE grupo = ?', (grupo,)
        ).fetchall()
        
        if not cedulas:
            return []
        
        cedulas_list = [c[0] for c in cedulas]
        placeholders = ','.join(['?' for _ in cedulas_list])
        productos = []
        
        if 'Publicaciones' in filtros:
            query = f'''
                SELECT p.cedula, per.nombre, p.titulo, p.año, 'Publicación',
                       p.tipo, p.categoria, p.revista_libro, p.fuente
                FROM publicaciones p
                JOIN personas per ON p.cedula = per.cedula
                WHERE p.cedula IN ({placeholders})
                ORDER BY p.año DESC
            '''
            productos.extend(cursor.execute(query, cedulas_list).fetchall())
        
        if 'Extensiones' in filtros:
            query = f'''
                SELECT e.cedula, per.nombre, e.actividad, e.año, 'Extensión',
                       e.tipo, e.modalidad, e.estado, e.fuente
                FROM extensiones e
                JOIN personas per ON e.cedula = per.cedula
                WHERE e.cedula IN ({placeholders})
                ORDER BY e.año DESC
            '''
            productos.extend(cursor.execute(query, cedulas_list).fetchall())
        
        if 'Trabajos de Grado' in filtros:
            query = f'''
                SELECT t.cedula_director, per.nombre, t.titulo, t.año, 'Trabajo de Grado',
                       t.programa, t.estado, t.nombre_estudiante, t.fuente
                FROM trabajos_grado t
                JOIN personas per ON t.cedula_director = per.cedula
                WHERE t.cedula_director IN ({placeholders})
                ORDER BY t.año DESC
            '''
            productos.extend(cursor.execute(query, cedulas_list).fetchall())
        
        if 'Productos Innovación' in filtros:
            query = f'''
                SELECT pi.cedula, per.nombre, pi.nombre, pi.año, 'Innovación',
                       pi.tipo_producto, pi.estado, '', pi.fuente
                FROM productos_innovacion pi
                JOIN personas per ON pi.cedula = per.cedula
                WHERE pi.cedula IN ({placeholders})
                ORDER BY pi.año DESC
            '''
            productos.extend(cursor.execute(query, cedulas_list).fetchall())
        
        if 'Proyectos' in filtros:
            query = f'''
                SELECT pr.cedula, per.nombre, pr.titulo, pr.año, 'Proyecto',
                       pr.tipo, pr.estado, pr.codigo_cie, pr.fuente
                FROM proyectos pr
                JOIN personas per ON pr.cedula = per.cedula
                WHERE pr.cedula IN ({placeholders})
                ORDER BY pr.año DESC
            '''
            productos.extend(cursor.execute(query, cedulas_list).fetchall())
        
        return productos
    
    def obtener_productos_grupo_detallado(self, grupo, filtros=None, anio_desde=None, anio_hasta=None):
        """Versión detallada que retorna diccionarios con toda la información.
        anio_desde/anio_hasta (opcionales): filtran por período: se incluyen los
        productos con año dentro del rango Y los que no tienen año registrado
        (para no ocultar datos solo porque falte ese dato)."""
        if filtros is None:
            filtros = ['Publicaciones', 'Extensiones', 'Trabajos de Grado', 'Productos Innovación',
                       'Proyectos', 'Propiedad Intelectual']

        cursor = self.conn.cursor()
        cedulas = cursor.execute(
            'SELECT DISTINCT cedula FROM grupos WHERE grupo = ?', (grupo,)
        ).fetchall()

        if not cedulas:
            return []

        cedulas_list = [c[0] for c in cedulas]
        placeholders = ','.join(['?' for _ in cedulas_list])
        productos = []

        filtro_anio_sql = ""
        params_anio = []
        if anio_desde is not None and anio_hasta is not None:
            params_anio = [anio_desde, anio_hasta]

        if 'Publicaciones' in filtros:
            filtro = f" AND (p.año IS NULL OR (p.año >= ? AND p.año <= ?))" if params_anio else ""
            query = f'''
                SELECT p.*, per.nombre as nombre_investigador
                FROM publicaciones p
                JOIN personas per ON p.cedula = per.cedula
                WHERE p.cedula IN ({placeholders}){filtro}
                ORDER BY p.año DESC
            '''
            rows = cursor.execute(query, cedulas_list + params_anio).fetchall()
            for row in rows:
                productos.append({
                    'cedula': row[1],
                    'investigador': row[-1],
                    'titulo': row[2],
                    'revista_libro': row[3],
                    'doi_url': row[4],
                    'issn_isbn': row[5],
                    'año': row[6],
                    'tipo': row[7],
                    'categoria': row[8],
                    'estado': row[9],
                    'grupo': row[10],
                    'fuente': row[11],
                    'datos_adicionales': row[12] if len(row) > 13 else None,
                    'tipo_producto': 'Publicación',
                    'detalle': row[3] or ''
                })

        if 'Extensiones' in filtros:
            filtro = f" AND (e.año IS NULL OR (e.año >= ? AND e.año <= ?))" if params_anio else ""
            query = f'''
                SELECT e.*, per.nombre as nombre_investigador
                FROM extensiones e
                JOIN personas per ON e.cedula = per.cedula
                WHERE e.cedula IN ({placeholders}){filtro}
                ORDER BY e.año DESC
            '''
            rows = cursor.execute(query, cedulas_list + params_anio).fetchall()
            for row in rows:
                productos.append({
                    'cedula': row[1],
                    'investigador': row[-1],
                    'titulo': row[2],
                    'tipo': row[3],
                    'modalidad': row[4],
                    'estado': row[5],
                    'fecha_inicio': row[6],
                    'fecha_fin': row[7],
                    'año': row[8],
                    'poblacion': row[9],
                    'grupo': row[10],
                    'facultad': row[11],
                    'financiacion_interna': row[12],
                    'financiacion_externa': row[13],
                    'fuente': row[14],
                    'datos_adicionales': row[15] if len(row) > 16 else None,
                    'tipo_producto': 'Extensión',
                    'categoria': row[3] or '',
                    'detalle': row[4] or ''
                })

        if 'Trabajos de Grado' in filtros:
            # Solo "conducentes" (trabajo de grado formal que sí se sube a
            # GrupLAC): los "no conducentes" (ej. prácticas no conducentes a
            # título) no cuentan para el reporte del grupo -- solo se
            # muestran al investigador en Búsqueda Persona. calificacion es
            # NULL para el reporte institucional de trabajos de grado
            # posicional (siempre conducente por definición, ese archivo no
            # trae prácticas).
            filtro = f" AND (t.año IS NULL OR (t.año >= ? AND t.año <= ?))" if params_anio else ""
            filtro += " AND (t.calificacion IS NULL OR t.calificacion != 'NO CONDUCENTE')"
            query = f'''
                SELECT t.*, per.nombre as nombre_director
                FROM trabajos_grado t
                JOIN personas per ON t.cedula_director = per.cedula
                WHERE t.cedula_director IN ({placeholders}){filtro}
                ORDER BY t.año DESC
            '''
            rows = cursor.execute(query, cedulas_list + params_anio).fetchall()
            for row in rows:
                productos.append({
                    'cedula': row[1],
                    'investigador': row[-1],
                    'titulo': row[5],
                    'programa': row[6],
                    'año': row[7],
                    'estado': row[8],
                    'fecha_sustentacion': row[9],
                    'calificacion': row[10],
                    'facultad': row[11],
                    'estudiante': row[4],
                    'cedula_estudiante': row[3],
                    'fuente': row[12],
                    'datos_adicionales': row[13] if len(row) > 14 else None,
                    'tipo_producto': 'Trabajo de Grado',
                    'categoria': row[6] or '',
                    'detalle': row[4] or ''
                })

        if 'Productos Innovación' in filtros:
            filtro = f" AND (pi.año IS NULL OR (pi.año >= ? AND pi.año <= ?))" if params_anio else ""
            query = f'''
                SELECT pi.*, per.nombre as nombre_investigador
                FROM productos_innovacion pi
                JOIN personas per ON pi.cedula = per.cedula
                WHERE pi.cedula IN ({placeholders}){filtro}
                ORDER BY pi.año DESC
            '''
            rows = cursor.execute(query, cedulas_list + params_anio).fetchall()
            for row in rows:
                productos.append({
                    'cedula': row[1],
                    'investigador': row[-1],
                    'titulo': row[3],
                    'tipo_producto_detalle': row[2],
                    'descripcion': row[4],
                    'año': row[5],
                    'estado': row[6],
                    'grupo': row[7],
                    'fuente': row[8],
                    'datos_adicionales': row[9] if len(row) > 10 else None,
                    'tipo_producto': 'Innovación',
                    'categoria': row[2] or '',
                    'detalle': row[4][:50] + '...' if row[4] and len(row[4]) > 50 else row[4] or ''
                })

        if 'Proyectos' in filtros:
            filtro = f" AND (pr.año IS NULL OR (pr.año >= ? AND pr.año <= ?))" if params_anio else ""
            query = f'''
                SELECT pr.*, per.nombre as nombre_investigador
                FROM proyectos pr
                JOIN personas per ON pr.cedula = per.cedula
                WHERE pr.cedula IN ({placeholders}){filtro}
                ORDER BY pr.año DESC
            '''
            rows = cursor.execute(query, cedulas_list + params_anio).fetchall()
            for row in rows:
                productos.append({
                    'cedula': row[1],
                    'investigador': row[-1],
                    'responsable': row[2],
                    'titulo': row[3],
                    'objetivo': row[4],
                    'codigo_cie': row[5],
                    'tipo': row[6],
                    'año': row[7],
                    'fecha_inicio': row[8],
                    'fecha_fin': row[9],
                    'estado': row[10],
                    'facultad': row[11],
                    'grupo': row[12],
                    'valor_aprobado': row[13],
                    'fuente': row[14],
                    'tipo_producto': 'Proyecto',
                    'categoria': row[6] or '',
                    'detalle': row[5] or ''
                })

        if 'Propiedad Intelectual' in filtros:
            # No tiene cedula propia (queda registrada por nombre en
            # 'responsable', a veces con varios nombres juntos) -- se asocia
            # al grupo emparejando ese texto contra los nombres de sus
            # integrantes ya normalizados (misma normalización que usa el
            # resto del sistema para nombres).
            nombres_grupo = cursor.execute(
                f'SELECT cedula, nombre FROM personas WHERE cedula IN ({placeholders})', cedulas_list
            ).fetchall()
            mapa_norm = {normalizar_nombre(nom): ced for ced, nom in nombres_grupo if nom}
            rows = cursor.execute('SELECT * FROM propiedad_intelectual').fetchall()
            for row in rows:
                responsable = row[1] or ''
                resp_norm = normalizar_nombre(responsable)
                cedula_match = next((c for n, c in mapa_norm.items() if n and n in resp_norm), None)
                if not cedula_match:
                    continue
                productos.append({
                    'cedula': cedula_match,
                    'investigador': responsable,
                    'titulo': row[4],
                    'tipo_producto_detalle': row[2],
                    'tipo_patente': row[3],
                    'numero_registro': row[5],
                    'proyecto': row[6],
                    'fecha_aprobacion': row[7],
                    'entidad': row[8],
                    'facultad': row[9],
                    'fuente': row[10],
                    'datos_adicionales': row[11] if len(row) > 11 else None,
                    'año': None,
                    'tipo_producto': 'Propiedad Intelectual',
                    'categoria': row[2] or '',
                    'detalle': row[4] or ''
                })

        return productos

    def obtener_estadisticas(self):
        """Obtiene conteos en una sola consulta multi-tabla."""
        row = self.conn.execute("""
            SELECT
                (SELECT COUNT(*)          FROM personas)            AS personas,
                (SELECT COUNT(DISTINCT grupo) FROM grupos)          AS grupos,
                (SELECT COUNT(*)          FROM publicaciones)       AS publicaciones,
                (SELECT COUNT(*)          FROM extensiones)         AS extensiones,
                (SELECT COUNT(*)          FROM trabajos_grado)      AS trabajos,
                (SELECT COUNT(*)          FROM productos_innovacion) AS innovacion,
                (SELECT COUNT(*)          FROM proyectos)           AS proyectos,
                (SELECT COUNT(*)          FROM propiedad_intelectual) AS propiedad
        """).fetchone()
        keys = ["personas", "grupos", "publicaciones", "extensiones",
                "trabajos", "innovacion", "proyectos", "propiedad"]
        return dict(zip(keys, row))
        
    def close(self):
        """Cierra correctamente la conexión SQLite"""
        try:
            self.conn.commit()
            self.conn.close()
        except Exception as e:
            print("Error cerrando BD:", e)       


# ==================== CARGADOR DE DATOS INTEGRADO ====================

class CargadorDatosIntegrado(QThread):
    progreso = pyqtSignal(str)
    finalizado = pyqtSignal(dict)
    duplicados_consolidados = pyqtSignal(list)
    
    def __init__(self, db):
        super().__init__()
        self.db = db
        # self.archivos_directorio = Path(".") Esta estaba anteriormente.
        self.archivos_directorio = obtener_directorio_base()

    # Ver la nota equivalente en loader.py: las categorías de
    # ARCHIVOS_FUENTE_957 (constants.py) no siempre son 1:1 con los
    # extractores de aquí -- este mapeo agrupa las claves de categoría que
    # le corresponden a cada extractor, para sumarles las fuentes
    # adicionales que el usuario acumuló (sin reemplazar) desde "Agregar
    # archivo" en Inicio.
    _CLAVES_POR_EXTRACTOR = {
        'integrantes': ['integrantes'],
        'extension': ['extension'],
        'produccion': ['produccion_2024', 'produccion_2025_ciarp'],
        'trabajos_grado': ['trabajos_grado'],
        'libros': ['libros'],
        'innovacion': ['innovacion'],
        'proyectos': ['proyectos'],
        'propiedad_intelectual': ['cgt0104_2025', 'cgt0104_2024'],
    }

    @staticmethod
    def _serie_o_vacia(df, columna):
        """Como df.get(columna, ''), pero el valor por defecto es una Serie
        vacía ALINEADA al índice de df. pd.Series(dtype=str) suelto tiene su
        propio índice vacío (longitud 0); si una columna del Excel no existe
        y se usa ese valor por defecto dentro de un zip() junto a columnas
        que sí tienen datos, zip() corta TODO el resultado a longitud 0 EN
        SILENCIO -- así se perdían filas completas (confirmado real: 0
        publicaciones cargadas desde la BASE DATOS PRODUCCIÓN CIARP porque
        esa hoja no trae columna 'doi_url'/'doi', y ese único hueco tumbaba
        las 150 filas del zip entero, sin ningún error visible)."""
        if columna in df.columns:
            return df[columna]
        return pd.Series([''] * len(df), index=df.index)

    def _rutas_adicionales(self, extractor):
        rutas = []
        for clave in self._CLAVES_POR_EXTRACTOR.get(extractor, []):
            for fuente in self.db.obtener_fuentes_adicionales(clave):
                p = Path(fuente.get("ruta", ""))
                if p.exists():
                    rutas.append(p)
        return rutas

    def run(self):
        base = self.archivos_directorio
        nomina_nombres = [
            "Listado Integrantes Grupos de Investigación UTP  080825.xlsx",
            "Listado Integrantes Grupos de Investigación UTP - 080825.xlsx",
            "Listado Integrantes Grupos de Investigacion UTP  080825.xlsx",
            "Listado Integrantes Grupos de Investigacion UTP - 080825.xlsx",
            "Consolidado Extensión 2024.xlsx",
            "BASE DATOS PRODUCCIÓN 2024.xlsx",
            "BASE DATOS PRODUCCIÓN  2025  CIARP.xlsx",
            "Trabajos Grado  Trabajo de Grado.xlsx",
            "TrabajosGrado_TrabajoDeGrado 2024.xlsx",
            "Reporte_libros.xlsx",
            "info_productos_innovacion.xlsx",
            "Proyectos de investigación registrados en 2024.xlsx",
            "CGT0104 - No de productos resultados de investigacion 31072025.xlsx",
            "CGT0104 - No de productos resultados de investigacion 31122024.xlsx",
        ]
        archivos_fuente = [str(base / n) for n in nomina_nombres]
        for extractor in self._CLAVES_POR_EXTRACTOR:
            archivos_fuente.extend(str(p) for p in self._rutas_adicionales(extractor))

        if self.db.cache_valida(archivos_fuente):
            self.progreso.emit("Base de datos en caché — omitiendo reproceso")
            self.finalizado.emit(self.db.obtener_estadisticas())
            return

        conn = self.db.conn
        conn.execute("PRAGMA synchronous=OFF")
        try:
            if self.db.obtener_estadisticas()['personas'] == 0:
                # Primera carga sobre una BD vacía: no hay nada que conservar.
                self.db.limpiar_datos()
            conn.commit()
            # En cualquier otro caso NO se borra nada: los extractores leen de
            # nuevo todos los archivos presentes y los índices únicos por
            # categoría (DatabaseManager._CLAVES_DEDUP) hacen que insertar lo
            # mismo otra vez actualice la fila existente en vez de duplicarla,
            # mientras que las filas realmente nuevas simplemente se acumulan.

            # ── Extracción secuencial ────────────────────────────────────────
            # Antes se corría cada extractor en su propio hilo
            # (ThreadPoolExecutor) para ganar velocidad, pero pandas/openpyxl
            # no son seguros para lecturas concurrentes de varios .xlsx desde
            # hilos distintos -- confirmado real: con los archivos actuales
            # (CIARP + Informe Extensión, ambos de varios MB) la extracción en
            # paralelo terminaba el proceso Python entero sin ninguna
            # excepción capturable (crash nativo), mientras que corriendo los
            # mismos 6 extractores uno tras otro no falla y tarda menos de un
            # segundo -- no hay necesidad real de paralelismo aquí.
            self.progreso.emit("Extrayendo datos de archivos fuente…")
            extractores = {
                'integrantes':          self._extraer_integrantes,
                'extensiones':          self._extraer_extensiones,
                'produccion':           self._extraer_produccion,
                'trabajos_grado':       self._extraer_trabajos_grado,
                'libros':               self._extraer_libros,
                'productos_innovacion': self._extraer_productos_innovacion,
            }
            resultados = {}
            for name, fn in extractores.items():
                try:
                    resultados[name] = fn()
                    self.progreso.emit(f"✓ Extraídos datos de {name}")
                except Exception as e:
                    self.progreso.emit(f"Error extrayendo {name}: {e}")
                    resultados[name] = ([], [])

            # ── Inserción secuencial ─────────────────────────────────────────
            self.progreso.emit("Insertando datos en la base…")

            for name, (pers, _) in resultados.items():
                if pers:
                    self.db._upsert_personas_batch(pers)
            conn.commit()

            if resultados.get('integrantes'):
                _, grupos_b = resultados['integrantes']
                if grupos_b:
                    conn.executemany(
                        'INSERT OR IGNORE INTO grupos (cedula,grupo,facultad,tipo_miembro,origen) VALUES(?,?,?,?,?)',
                        grupos_b,
                    )
                    self.progreso.emit(f"Insertados {len(grupos_b)} integrantes de grupos")

            if resultados.get('extensiones'):
                _, ext_b = resultados['extensiones']
                if ext_b:
                    conn.executemany(
                        '''INSERT OR REPLACE INTO extensiones
                           (cedula,actividad,tipo,modalidad,estado,fecha_inicio,fecha_fin,
                            año,poblacion,grupo,facultad,financiacion_interna,financiacion_externa,fuente,
                            datos_adicionales)
                           VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                        ext_b,
                    )
                    self.progreso.emit(f"Insertadas/actualizadas {len(ext_b)} extensiones")

            if resultados.get('produccion'):
                _, pub_b = resultados['produccion']
                if pub_b:
                    conn.executemany(
                        '''INSERT OR REPLACE INTO publicaciones
                           (cedula,titulo,revista_libro,doi_url,issn_isbn,año,
                            tipo,categoria,estado,grupo,fuente,datos_adicionales)
                           VALUES(?,?,?,?,?,?,?,?,?,?,?,?)''',
                        pub_b,
                    )
                    self.progreso.emit(f"Insertadas/actualizadas {len(pub_b)} publicaciones")

            if resultados.get('trabajos_grado'):
                _, tg_b = resultados['trabajos_grado']
                if tg_b:
                    conn.executemany(
                        '''INSERT OR REPLACE INTO trabajos_grado
                           (cedula_director,nombre_director,cedula_estudiante,nombre_estudiante,
                            titulo,programa,año,calificacion,fecha_sustentacion,facultad,fuente,
                            datos_adicionales)
                           VALUES(?,?,?,?,?,?,?,?,?,?,?,?)''',
                        tg_b,
                    )
                    self.progreso.emit(f"Insertados/actualizados {len(tg_b)} trabajos de grado")

            if resultados.get('libros'):
                _, libros_b = resultados['libros']
                if libros_b:
                    conn.executemany(
                        '''INSERT OR REPLACE INTO publicaciones
                           (cedula,titulo,revista_libro,doi_url,issn_isbn,año,
                            tipo,categoria,estado,grupo,fuente)
                           VALUES(?,?,?,?,?,?,?,?,?,?,?)''',
                        libros_b,
                    )
                    self.progreso.emit(f"Insertados/actualizados {len(libros_b)} libros/capítulos")

            if resultados.get('productos_innovacion'):
                _, pi_b = resultados['productos_innovacion']
                if pi_b:
                    conn.executemany(
                        '''INSERT OR REPLACE INTO productos_innovacion
                           (cedula,tipo_producto,nombre,descripcion,año,estado,grupo,fuente,datos_adicionales)
                           VALUES(?,?,?,?,?,?,?,?,?)''',
                        pi_b,
                    )
                    self.progreso.emit(f"Insertados/actualizados {len(pi_b)} productos de innovación")

            conn.commit()

            # ── Cargas secuenciales (dependen del estado de la BD) ──────────
            self.cargar_proyectos()
            self.cargar_propiedad_intelectual()

            self.progreso.emit("Detectando y consolidando duplicados…")
            consolidaciones = self.db.consolidar_duplicados()
            if consolidaciones:
                self.progreso.emit(
                    f"Consolidados {len(consolidaciones)} grupos de duplicados"
                )
                self.duplicados_consolidados.emit(consolidaciones)
            else:
                self.progreso.emit("No se encontraron duplicados")

            stats = self.db.obtener_estadisticas()
            self.db.guardar_sello_carga(archivos_fuente)
            self.finalizado.emit(stats)
        except Exception as e:
            import traceback
            self.progreso.emit(f"Error en carga: {e}")
            self.progreso.emit(traceback.format_exc()[:600])
        finally:
            conn.execute("PRAGMA synchronous=NORMAL")
            conn.commit()

    def _norm_key(k: str) -> str:
        """Normaliza una clave de columna para búsqueda robusta."""
        return normalizar_columna(k)

    @classmethod
    def _col(cls, row: dict, *keys) -> str:
        """
        Primer valor no-vacío buscando la clave tanto en su forma original
        como normalizada (sin tildes, espacios→_, minúsculas).
        Funciona con dicts de to_dict('records') Y con _asdict() de itertuples.
        """
        for k in keys:
            v = row.get(k)
            if v is not None and str(v).strip() not in ('', 'nan', 'None'):
                return v
            k_norm = cls._norm_key(k)
            v = row.get(k_norm)
            if v is not None and str(v).strip() not in ('', 'nan', 'None'):
                return v
        return ''

    @classmethod
    def _normalizar_cols(cls, df):
        """Devuelve el mismo DataFrame con columnas normalizadas en sus nombres."""
        df = df.copy()
        df.columns = [cls._norm_key(c) for c in df.columns]
        return df

    @staticmethod
    def _columnas_extra_json(df, columnas_conocidas: set) -> pd.Series:
        """
        Serializa por fila las columnas del Excel que no están en
        `columnas_conocidas` (columnas nuevas/no mapeadas), para no perder esa
        información aunque todavía no tenga una columna propia en la BD.
        Devuelve '' por fila cuando no hay columnas extra o todas vienen vacías.
        """
        extra_cols = [c for c in df.columns if c not in columnas_conocidas]
        if not extra_cols:
            return pd.Series([''] * len(df), index=df.index)

        def _serializar(fila):
            datos = {c: fila[c] for c in extra_cols if limpiar_texto(fila[c])}
            return json.dumps(datos, ensure_ascii=False) if datos else ''

        return df[extra_cols].apply(_serializar, axis=1)

    @staticmethod
    def _anio(valor):
        """Extrae año entero de un valor; devuelve None si no es posible."""
        if not valor or (isinstance(valor, float) and valor != valor):
            return None
        try:
            return int(str(valor).strip()[:4])
        except (ValueError, TypeError):
            return None

    # ── Integrantes ────────────────────────────────────────────────────────
    def _extraer_integrantes(self):
        archivos = [
            'Listado Integrantes Grupos de Investigación UTP  080825.xlsx',
            'Listado Integrantes Grupos de Investigación UTP - 080825.xlsx',
            'Listado Integrantes Grupos de Investigacion UTP  080825.xlsx',
            'Listado Integrantes Grupos de Investigacion UTP - 080825.xlsx',
            'integrantes.xlsx',
        ]
        rutas = []
        for archivo in archivos:
            ruta = self.archivos_directorio / archivo
            if ruta.exists():
                rutas.append(ruta)
                break
        rutas.extend(self._rutas_adicionales('integrantes'))

        personas_b, grupos_b = [], []
        for ruta in rutas:
            df = self._normalizar_cols(
                pd.read_excel(ruta, engine='openpyxl', dtype=str).fillna('')
            )
            cedulas    = (df['numero_documento'] if 'numero_documento' in df.columns
                          else self._serie_o_vacia(df, 'cedula')).apply(limpiar_cedula)
            nombres    = (df['nombres'] if 'nombres' in df.columns
                          else self._serie_o_vacia(df, 'nombre')).apply(limpiar_texto)
            grupos_s   = (df['nombre_grupo'] if 'nombre_grupo' in df.columns
                          else self._serie_o_vacia(df, 'grupo')).apply(limpiar_texto)
            facultades = self._serie_o_vacia(df, 'facultad').apply(limpiar_texto)
            emails     = self._serie_o_vacia(df, 'email').apply(limpiar_texto)
            tipos      = self._serie_o_vacia(df, 'tipo').apply(limpiar_texto)
            mask = (cedulas.str.len() > 0) & (nombres.str.len() > 0)
            personas_b.extend(zip(cedulas[mask], nombres[mask], emails[mask], facultades[mask], tipos[mask]))
            grupos_b.extend(
                (c, g, f, t, 'directo')
                for c, g, f, t in zip(cedulas[mask], grupos_s[mask], facultades[mask], tipos[mask])
                if g
            )

        grupos_b.extend(self._grupo_adscrito_semilleros(grupos_b))
        return personas_b, grupos_b

    def _grupo_adscrito_semilleros(self, grupos_b):
        """Lee data/Reporte Semilleros con Grupo adscrito.xlsx (columnas
        NOMBRE SEMILLERO / GRUPO ADSCRITO) y, para cada persona ya registrada
        bajo un semillero mapeado ahí, agrega TAMBIÉN su membresía al grupo
        de investigación real al que está adscrito ese semillero. GrupLAC no
        rastrea semilleros como entidades propias (no tienen carpeta
        scrapeada) -- sin esto, la producción de alguien que SOLO aparece
        bajo un semillero (sin membresía directa al grupo real, caso
        confirmado en ~44 personas) queda sin poder atribuirse a ningún
        grupo verificable en Seguimiento Grupos ni en Reportes por Grupo.

        Solo se acepta match normalizado EXACTO entre 'grupo adscrito' y un
        grupo real ya visto en este mismo archivo de integrantes (168/174
        filas del reporte calzan así) -- los 6 restantes apuntan a grupos
        que no tienen match confiable (ninguno pasa de 0.76 de similitud
        contra el grupo interno más parecido) y se descartan en vez de
        arriesgar una atribución equivocada."""
        ruta = self.archivos_directorio / "data" / "Reporte Semilleros con Grupo adscrito.xlsx"
        if not ruta.exists():
            return []

        grupos_reales_norm = {
            normalize_text(g): g for (_, g, _, _, _) in grupos_b
            if "SEMILLERO" not in g.upper()
        }
        if not grupos_reales_norm:
            return []

        try:
            df = self._normalizar_cols(
                pd.read_excel(ruta, engine='openpyxl', dtype=str).fillna('')
            )
        except Exception:
            return []

        nombres_semillero = self._serie_o_vacia(df, 'nombre_semillero').apply(limpiar_texto)
        adscritos = self._serie_o_vacia(df, 'grupo_adscrito').apply(limpiar_texto)

        mapa = {}
        for nombre, adscrito in zip(nombres_semillero, adscritos):
            if not nombre or not adscrito:
                continue
            real = grupos_reales_norm.get(normalize_text(adscrito))
            if real:
                mapa[normalize_text(nombre)] = real

        if not mapa:
            return []

        # Persistido para que los reportes (Reportes por Grupo) puedan
        # más tarde preguntar "¿esta persona también participa en un
        # semillero adscrito a este grupo?" sin tener que releer el Excel --
        # ver DatabaseManager.procedencia_grupo.
        self.db.conn.execute(
            "INSERT OR REPLACE INTO configuracion (clave, valor) VALUES (?, ?)",
            ("mapeo_semillero_grupo_adscrito", json.dumps(mapa, ensure_ascii=False)),
        )
        self.db.conn.commit()

        vistos = {(c, g) for c, g, _, _, _ in grupos_b}
        extra = []
        for cedula, grupo, facultad, tipo, _origen in grupos_b:
            real = mapa.get(normalize_text(grupo))
            if real and (cedula, real) not in vistos:
                vistos.add((cedula, real))
                extra.append((cedula, real, facultad, tipo, 'semillero'))
        return extra

    def cargar_integrantes(self):
        self.progreso.emit('Cargando integrantes de grupos...')
        personas_b, grupos_b = self._extraer_integrantes()
        if personas_b:
            conn = self.db.conn
            self.db._upsert_personas_batch(personas_b)
            conn.executemany(
                'INSERT OR IGNORE INTO grupos (cedula,grupo,facultad,tipo_miembro,origen) VALUES(?,?,?,?,?)',
                grupos_b,
            )
            conn.commit()
            self.progreso.emit(f'Cargados {len(personas_b)} integrantes de grupos')

    # ── Extensiones ──────────────────────────────────────────────
    def _extraer_extensiones(self):
        archivos = [
            'Consolidado Extensión 2024.xlsx',
            'Actividades Extensión enerojulio.xlsx',
            'Actividades Extensión (enero-julio).xlsx',
        ]
        rutas = []
        for archivo in archivos:
            ruta = self.archivos_directorio / archivo
            if ruta.exists():
                rutas.append(ruta)
        rutas.extend(self._rutas_adicionales('extension'))

        personas_b, ext_b = [], []
        for ruta in rutas:
            # "Informe Extensión" (fuente adicional, ej. el que exporta el
            # sistema de Extensión de la universidad) trae una hoja 'Datos'
            # con encabezados propios en vez de 'Consolidado' -- sin este
            # chequeo, pd.read_excel(sheet_name='Consolidado') lanzaba
            # ValueError (hoja inexistente) y la excepción, atrapada en
            # silencio por el ThreadPoolExecutor de run(), tumbaba TODA la
            # categoría "extensiones" (incluida la institucional) en vez de
            # solo saltarse el archivo con formato distinto.
            try:
                xls = pd.ExcelFile(ruta, engine='openpyxl')
            except Exception:
                continue
            if 'Consolidado' not in xls.sheet_names:
                if 'Datos' in xls.sheet_names:
                    p_b, e_b = self._extraer_extension_informe(ruta)
                    personas_b.extend(p_b)
                    ext_b.extend(e_b)
                continue

            df = self._normalizar_cols(
                pd.read_excel(xls, sheet_name='Consolidado', dtype=str).fillna('')
            )

            cedulas = self._serie_o_vacia(df, 'cedula').apply(limpiar_cedula)
            mask = cedulas.str.len() > 0
            if not mask.any():
                continue

            dv = df[mask]
            cv = cedulas[mask]
            nombres   = self._serie_o_vacia(dv, 'nombre_responsable').apply(limpiar_texto).replace('', 'Sin nombre')
            facultades = (dv['facultad_dependencia'] if 'facultad_dependencia' in dv.columns
                          else self._serie_o_vacia(dv, 'facultad')).apply(limpiar_texto)
            fis       = self._serie_o_vacia(dv, 'fecha_inicial').astype(str)
            anios     = fis.str[:4].apply(self._anio)

            grupos_s  = (dv['grupo_semillero_de_investigacion'] if 'grupo_semillero_de_investigacion' in dv.columns
                         else self._serie_o_vacia(dv, 'grupo')).apply(limpiar_texto)
            extra     = self._columnas_extra_json(dv, COLUMNAS_CONOCIDAS_POR_CATEGORIA['extension'])

            personas_b.extend(zip(cv, nombres, pd.Series('', index=dv.index), facultades, pd.Series('Responsable', index=dv.index)))
            ext_b.extend(zip(
                cv,
                self._serie_o_vacia(dv, 'nombre_actividad').apply(limpiar_texto),
                self._serie_o_vacia(dv, 'tipo').apply(limpiar_texto),
                self._serie_o_vacia(dv, 'modalidad').apply(limpiar_texto),
                self._serie_o_vacia(dv, 'estado').apply(limpiar_texto),
                fis,
                self._serie_o_vacia(dv, 'fecha_final').apply(limpiar_texto),
                anios,
                self._serie_o_vacia(dv, 'poblacion_beneficiaria').apply(limpiar_texto),
                grupos_s,
                facultades,
                self._serie_o_vacia(dv, 'financiacion_interna').apply(limpiar_texto),
                self._serie_o_vacia(dv, 'fuente_financiacion_externa').apply(limpiar_texto),
                [ruta.name] * len(dv),
                extra,
            ))
        return personas_b, ext_b

    def _extraer_extension_informe(self, ruta):
        """Lee un 'Informe Extensión' con hoja 'Datos' (formato distinto al
        institucional 'Consolidado Extensión NNNN.xlsx': encabezados propios
        como 'Grupo(s) de investigación', 'Fecha inicial'/'Fecha final',
        'Número documento', 'Nombre responsable'...) y lo mapea a la tabla
        extensiones. Columnas no usadas explícitamente (objetivo, duración,
        resultados obtenidos, beneficiados, correo, id) se guardan en
        'datos_adicionales' para no perderlas."""
        df = self._normalizar_cols(
            pd.read_excel(ruta, sheet_name='Datos', engine='openpyxl', dtype=str).fillna('')
        )

        cedulas = self._serie_o_vacia(df, 'numero_documento').apply(limpiar_cedula)
        mask = cedulas.str.len() > 0
        if not mask.any():
            return [], []

        dv = df[mask]
        cv = cedulas[mask]
        nombres = self._serie_o_vacia(dv, 'nombre_responsable').apply(limpiar_texto).replace('', 'Sin nombre')
        facultades = self._serie_o_vacia(dv, 'facultad_dependencia').apply(limpiar_texto)
        fis = self._serie_o_vacia(dv, 'fecha_inicial').astype(str)
        anios = fis.str[:4].apply(self._anio)
        grupos_s = self._serie_o_vacia(dv, 'grupo_s__de_investigacion').apply(limpiar_texto)

        columnas_conocidas = {
            'numero_documento', 'nombre_responsable', 'facultad_dependencia',
            'fecha_inicial', 'fecha_final', 'grupo_s__de_investigacion',
            'nombre', 'tipo', 'modalidad', 'estado', 'poblacion_beneficiaria',
        }
        columnas_extra = [c for c in dv.columns if c not in columnas_conocidas]
        extra = [
            json.dumps(datos, ensure_ascii=False) if datos else ''
            for datos in (
                {c: limpiar_texto(fila[c]) for c in columnas_extra if limpiar_texto(fila[c])}
                for _, fila in dv.iterrows()
            )
        ]

        personas_b = list(zip(cv, nombres, pd.Series('', index=dv.index), facultades,
                               pd.Series('Responsable', index=dv.index)))
        ext_b = list(zip(
            cv,
            self._serie_o_vacia(dv, 'nombre').apply(limpiar_texto),
            self._serie_o_vacia(dv, 'tipo').apply(limpiar_texto),
            self._serie_o_vacia(dv, 'modalidad').apply(limpiar_texto),
            self._serie_o_vacia(dv, 'estado').apply(limpiar_texto),
            fis,
            self._serie_o_vacia(dv, 'fecha_final').apply(limpiar_texto),
            anios,
            self._serie_o_vacia(dv, 'poblacion_beneficiaria').apply(limpiar_texto),
            grupos_s,
            facultades,
            pd.Series('', index=dv.index),
            pd.Series('', index=dv.index),
            [ruta.name] * len(dv),
            extra,
        ))
        return personas_b, ext_b

    def cargar_extensiones(self):
        self.progreso.emit('Cargando extensiones...')
        personas_b, ext_b = self._extraer_extensiones()
        if ext_b:
            conn = self.db.conn
            self.db._upsert_personas_batch(personas_b)
            conn.executemany(
                '''INSERT OR REPLACE INTO extensiones
                   (cedula,actividad,tipo,modalidad,estado,fecha_inicio,fecha_fin,
                    año,poblacion,grupo,facultad,financiacion_interna,financiacion_externa,fuente,
                    datos_adicionales)
                   VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                ext_b,
            )
            conn.commit()
            self.progreso.emit(f'Cargadas {len(ext_b)} extensiones')

    # ── Producción / Publicaciones ──────────────────────────────
    def _extraer_produccion(self):
        archivos = [
            'BASE DATOS PRODUCCIÓN 2024.xlsx',
            'BASE DATOS PRODUCCIÓN  2025  CIARP.xlsx',
            'BASE DATOS PRODUCCIÓN  2025 - CIARP.xlsx',
        ]
        rutas = []
        for archivo in archivos:
            ruta = self.archivos_directorio / archivo
            if ruta.exists():
                rutas.append(ruta)
        rutas.extend(self._rutas_adicionales('produccion'))

        personas_b, pub_b = [], []
        for ruta in rutas:
            xls = pd.ExcelFile(ruta, engine='openpyxl')
            for sheet in xls.sheet_names:
                df = self._normalizar_cols(
                    pd.read_excel(xls, sheet_name=sheet, dtype=str).fillna('')
                )

                cedulas = self._serie_o_vacia(df, 'cedula').apply(limpiar_cedula)
                mask = cedulas.str.len() > 0
                if not mask.any():
                    continue

                dv = df[mask]
                cv = cedulas[mask]
                nombres   = (dv['autores'] if 'autores' in dv.columns
                             else dv['autor'] if 'autor' in dv.columns
                             else self._serie_o_vacia(dv, 'nombre')).apply(limpiar_texto)
                facultades = (dv['dependencia'] if 'dependencia' in dv.columns
                              else self._serie_o_vacia(dv, 'facultad')).apply(limpiar_texto)
                extra = self._columnas_extra_json(dv, COLUMNAS_CONOCIDAS_POR_CATEGORIA['produccion_2024'])

                fuente = f'{ruta.name} :: {sheet}'
                personas_b.extend(zip(cv, nombres, pd.Series('', index=dv.index), facultades, pd.Series('Autor', index=dv.index)))
                pub_b.extend(zip(
                    cv,
                    (dv['nombre_del_trabajo'] if 'nombre_del_trabajo' in dv.columns
                     else self._serie_o_vacia(dv, 'titulo')).apply(limpiar_texto),
                    (dv['revista_o_libro'] if 'revista_o_libro' in dv.columns
                     else self._serie_o_vacia(dv, 'revista_libro')).apply(limpiar_texto),
                    (dv['doi_url'] if 'doi_url' in dv.columns
                     else self._serie_o_vacia(dv, 'doi')).apply(limpiar_texto),
                    (dv['issn_isbn'] if 'issn_isbn' in dv.columns
                     else self._serie_o_vacia(dv, 'issn')).apply(limpiar_texto),
                    (dv['ano_de_la_publicacion'] if 'ano_de_la_publicacion' in dv.columns
                     else self._serie_o_vacia(dv, 'ano')).apply(self._anio),
                    self._serie_o_vacia(dv, 'tipo').apply(limpiar_texto),
                    self._serie_o_vacia(dv, 'categoria').apply(limpiar_texto),
                    self._serie_o_vacia(dv, 'estado').apply(limpiar_texto),
                    self._serie_o_vacia(dv, 'grupo').apply(limpiar_texto),
                    [fuente] * len(dv),
                    extra,
                ))
        return personas_b, pub_b

    def cargar_produccion(self):
        self.progreso.emit('Cargando publicaciones...')
        personas_b, pub_b = self._extraer_produccion()
        if pub_b:
            conn = self.db.conn
            self.db._upsert_personas_batch(personas_b)
            conn.executemany(
                '''INSERT OR REPLACE INTO publicaciones
                   (cedula,titulo,revista_libro,doi_url,issn_isbn,año,
                    tipo,categoria,estado,grupo,fuente,datos_adicionales)
                   VALUES(?,?,?,?,?,?,?,?,?,?,?,?)''',
                pub_b,
            )
            conn.commit()
            self.progreso.emit(f'Cargadas {len(pub_b)} publicaciones')

    # ── Trabajos de grado ─────────────────────────────────────────
    def _extraer_trabajos_grado(self):
        archivos = [
            'Trabajos Grado  Trabajo de Grado.xlsx',
            'TrabajosGrado_TrabajoDeGrado 2024.xlsx',
            'Trabajos Grado - Trabajo de Grado.xlsx',
        ]
        rutas = []
        for archivo in archivos:
            ruta = self.archivos_directorio / archivo
            if ruta.exists():
                rutas.append(ruta)

        personas_b, tg_b = [], []
        for ruta in rutas:
            df = pd.read_excel(ruta, engine='openpyxl', header=None)
            director_actual = cedula_director = None
            for row in df.itertuples(index=False):
                col_a = limpiar_texto(str(row[0]) if pd.notna(row[0]) else '')
                if col_a and 'Nombre del trabajo' not in col_a:
                    ultima = row[-1] if len(row) > 1 else None
                    if pd.notna(ultima):
                        ced = limpiar_cedula(str(ultima))
                        if ced and len(ced) >= 6:
                            director_actual = col_a
                            cedula_director = ced
                            personas_b.append((ced, col_a, '', '', 'Director'))
                            continue
                if director_actual and cedula_director:
                    titulo = limpiar_texto(str(row[0]) if pd.notna(row[0]) else '')
                    if not titulo or titulo == 'Nombre del trabajo de grado':
                        continue
                    fecha = limpiar_texto(str(row[5]) if len(row) > 5 and pd.notna(row[5]) else '')
                    anio = None
                    if fecha:
                        try:
                            anio = int(fecha.split('/')[-1]) if '/' in fecha else int(fecha.split('-')[0])
                        except (ValueError, IndexError):
                            pass
                    # Columnas del Excel: A=Título, B=Documento Estudiante,
                    # C=Estudiante (nombre), D=Programa, E=Nota, F=Fecha
                    # Sustentación, G=Facultad.
                    tg_b.append((
                        cedula_director, director_actual,
                        limpiar_cedula(str(row[1]) if len(row) > 1 and pd.notna(row[1]) else ''),
                        limpiar_texto(str(row[2]) if len(row) > 2 and pd.notna(row[2]) else ''),
                        titulo,
                        limpiar_texto(str(row[3]) if len(row) > 3 and pd.notna(row[3]) else ''),
                        anio,
                        limpiar_texto(str(row[4]) if len(row) > 4 and pd.notna(row[4]) else ''),
                        fecha,
                        limpiar_texto(str(row[6]) if len(row) > 6 and pd.notna(row[6]) else ''),
                        ruta.name,
                        '',
                    ))

        # Fuentes adicionales (ej. reportes de prácticas): tienen encabezado
        # de columna propio, muy distinto del reporte institucional posicional
        # de arriba, así que se leen por nombre de columna en vez de por
        # posición fija -- ver _extraer_practicas().
        for ruta_extra in self._rutas_adicionales('trabajos_grado'):
            p_b, tg_extra = self._extraer_practicas(ruta_extra)
            personas_b.extend(p_b)
            tg_b.extend(tg_extra)

        return personas_b, tg_b

    def _extraer_practicas(self, ruta):
        """Lee un reporte de prácticas (encabezados tipo 'Código Estudiante',
        'Nombre del Estudiante', 'Escenario de Práctica', 'Nombre de la
        Práctica', 'Cédula/Nombre Docente Guía', etc.) y lo mapea a la tabla
        trabajos_grado: el docente guía cuenta como 'director', la práctica
        como el 'título'. Cualquier columna que no se use explícitamente
        (escenario, NIT, ciudad, modalidad, fecha de inicio...) se guarda
        completa en 'datos_adicionales' para no perderla."""
        df = self._normalizar_cols(
            pd.read_excel(ruta, engine='openpyxl', dtype=str).fillna('')
        )

        def _col(*claves):
            for c in claves:
                if c in df.columns:
                    return df[c]
            return pd.Series([''] * len(df), index=df.index)

        cedulas_docente = _col('cedula_docente_guia', 'cedula_docente', 'documento_docente_guia').apply(limpiar_cedula)
        nombres_docente = _col('nombre_docente_guia', 'docente_guia').apply(limpiar_texto)
        cedulas_estudiante = _col('codigo_estudiante', 'codigo_estudia', 'documento_estudiante', 'cedula_estudiante').apply(limpiar_cedula)
        nombres_estudiante = _col('nombre_del_estudiante', 'nombre_estudiante').apply(limpiar_texto)
        titulos = _col('nombre_de_la_practica', 'nombre_practica').apply(limpiar_texto)
        programas = _col('programa_academico', 'programa').apply(limpiar_texto)
        facultades = _col('facultad').apply(limpiar_texto)
        tipos_practica = _col('tipo_de_practica', 'tipo_practica').apply(limpiar_texto)
        fechas_fin = _col('fecha_de_finalizacion', 'fecha_finalizacion').apply(limpiar_texto)
        fechas_inicio = _col('fecha_de_inicio', 'fecha_inicio').apply(limpiar_texto)

        columnas_conocidas = COLUMNAS_CONOCIDAS_POR_CATEGORIA.get('trabajos_grado', set())
        columnas_extra = [c for c in df.columns if c not in columnas_conocidas]

        personas_b, tg_b = [], []
        for i in range(len(df)):
            titulo = titulos.iloc[i]
            cedula_est = cedulas_estudiante.iloc[i]
            if not titulo or not cedula_est:
                continue
            cedula_doc = cedulas_docente.iloc[i]
            nombre_doc = nombres_docente.iloc[i]

            año = None
            for fecha in (fechas_fin.iloc[i], fechas_inicio.iloc[i]):
                if fecha:
                    try:
                        año = int(fecha.split('-')[0]) if '-' in fecha else int(str(fecha).split('/')[-1])
                        break
                    except (ValueError, IndexError):
                        pass

            if cedula_doc:
                personas_b.append((cedula_doc, nombre_doc, '', facultades.iloc[i], 'Director'))

            fila = df.iloc[i]
            datos_extra = {c: limpiar_texto(fila[c]) for c in columnas_extra if limpiar_texto(fila[c])}
            extra_json = json.dumps(datos_extra, ensure_ascii=False) if datos_extra else ''

            tg_b.append((
                cedula_doc, nombre_doc, cedula_est, nombres_estudiante.iloc[i],
                titulo, programas.iloc[i], año, tipos_practica.iloc[i],
                fechas_fin.iloc[i], facultades.iloc[i], ruta.name, extra_json,
            ))
        return personas_b, tg_b

    def cargar_trabajos_grado(self):
        self.progreso.emit('Cargando trabajos de grado...')
        personas_b, tg_b = self._extraer_trabajos_grado()
        if tg_b:
            conn = self.db.conn
            self.db._upsert_personas_batch(personas_b)
            conn.executemany(
                '''INSERT OR REPLACE INTO trabajos_grado
                   (cedula_director,nombre_director,cedula_estudiante,nombre_estudiante,
                    titulo,programa,año,calificacion,fecha_sustentacion,facultad,fuente)
                   VALUES(?,?,?,?,?,?,?,?,?,?,?)''',
                tg_b,
            )
            conn.commit()
            self.progreso.emit(f'Cargados {len(tg_b)} trabajos de grado')

    # ── Libros ────────────────────────────────────────────────────
    def _extraer_libros(self):
        archivos = [
            'Reporte de libros y capítulos publicados.xlsx',
            'Reporte_libros.xlsx',
        ]
        rutas = []
        for archivo in archivos:
            ruta = self.archivos_directorio / archivo
            if ruta.exists():
                rutas.append(ruta)
        rutas.extend(self._rutas_adicionales('libros'))

        personas_b, pub_b = [], []
        for ruta in rutas:
            df = pd.read_excel(ruta, engine='openpyxl', header=None)
            titulo_actual = tipo_actual = None
            autores_libro = []

            def _flush_libro():
                for aut in autores_libro:
                    pub_b.append((
                        aut['cedula'], titulo_actual, '', '', '', None,
                        'LIBRO', tipo_actual or 'LIBRO', '', '', ruta.name,
                    ))

            for row in df.itertuples(index=False):
                nombre = limpiar_texto(str(row[0]) if pd.notna(row[0]) else '')
                titulo = limpiar_texto(str(row[1]) if len(row) > 1 and pd.notna(row[1]) else '')
                tipo   = limpiar_texto(str(row[3]) if len(row) > 3 and pd.notna(row[3]) else '')
                cedula = limpiar_cedula(str(row[4]) if len(row) > 4 and pd.notna(row[4]) else '')
                if titulo and titulo != titulo_actual:
                    _flush_libro()
                    titulo_actual = titulo
                    tipo_actual   = tipo if tipo else 'LIBRO'
                    autores_libro = []
                if nombre and cedula:
                    personas_b.append((cedula, nombre, '', '', 'Autor'))
                    autores_libro.append({'cedula': cedula})
            _flush_libro()
        return personas_b, pub_b

    def cargar_libros(self):
        self.progreso.emit('Cargando libros y capítulos...')
        personas_b, pub_b = self._extraer_libros()
        if pub_b:
            conn = self.db.conn
            self.db._upsert_personas_batch(personas_b)
            conn.executemany(
                '''INSERT OR REPLACE INTO publicaciones
                   (cedula,titulo,revista_libro,doi_url,issn_isbn,año,
                    tipo,categoria,estado,grupo,fuente)
                   VALUES(?,?,?,?,?,?,?,?,?,?,?)''',
                pub_b,
            )
            conn.commit()
            self.progreso.emit(f'Cargados {len(pub_b)} libros/capítulos')

    # ── Productos de innovación ───────────────────────────────────────
    def _extraer_productos_innovacion(self):
        ruta = self.archivos_directorio / 'info_productos_innovacion.xlsx'
        rutas = [ruta] if ruta.exists() else []
        rutas.extend(self._rutas_adicionales('innovacion'))
        if not rutas:
            return [], []

        batch = []
        for ruta in rutas:
            df = self._normalizar_cols(
                pd.read_excel(ruta, engine='openpyxl', dtype=str).fillna('')
            )
            nombres = (df['nombre'] if 'nombre' in df.columns
                       else df['titulo'] if 'titulo' in df.columns
                       else self._serie_o_vacia(df, 'producto')).apply(limpiar_texto)
            mask = nombres.str.len() > 0
            if not mask.any():
                continue
            dv = df[mask]
            nv = nombres[mask]
            anios = (dv['ano_de_registro'] if 'ano_de_registro' in dv.columns
                     else self._serie_o_vacia(dv, 'ano')).apply(self._anio)
            extra = self._columnas_extra_json(dv, COLUMNAS_CONOCIDAS_POR_CATEGORIA['innovacion'])
            batch.extend(zip(
                ['0000000'] * len(dv),
                (dv['tipo_de_producto'] if 'tipo_de_producto' in dv.columns
                 else self._serie_o_vacia(dv, 'tipo')).apply(limpiar_texto),
                nv,
                self._serie_o_vacia(dv, 'descripcion').apply(limpiar_texto),
                anios,
                self._serie_o_vacia(dv, 'estado').apply(limpiar_texto),
                (dv['grupo_de_investigacion'] if 'grupo_de_investigacion' in dv.columns
                 else self._serie_o_vacia(dv, 'grupo')).apply(limpiar_texto),
                [ruta.name] * len(dv),
                extra,
            ))
        return [], batch

    def cargar_productos_innovacion(self):
        self.progreso.emit('Cargando productos de innovación...')
        _, batch = self._extraer_productos_innovacion()
        if batch:
            conn = self.db.conn
            conn.executemany(
                '''INSERT OR REPLACE INTO productos_innovacion
                   (cedula,tipo_producto,nombre,descripcion,año,estado,grupo,fuente,datos_adicionales)
                   VALUES(?,?,?,?,?,?,?,?,?)''',
                batch,
            )
            conn.commit()
            self.progreso.emit(f'Cargados {len(batch)} productos de innovación')

    def cargar_proyectos(self):
        self.progreso.emit("Cargando proyectos de investigación...")
        
        # ARCHIVOS POSIBLES - buscar con variaciones
        archivos_posibles = [
            "Proyectos de investigación registrados en 2024.xlsx",
            "Proyectos de investigacion registrados en 2024.xlsx",
            "proyectos_investigacion_2024.xlsx",
            "proyectos 2024.xlsx"
        ]
        
        rutas = []
        for archivo in archivos_posibles:
            ruta = self.archivos_directorio / archivo
            if ruta.exists():
                rutas.append(ruta)
                break
        rutas.extend(self._rutas_adicionales('proyectos'))

        if not rutas:
            self.progreso.emit("⚠ No se encontró archivo de proyectos de investigación")
            return

        for ruta in rutas:
          try:
            # USAR DETECCIÓN AUTOMÁTICA DE ENCABEZADOS (como investigacion.py)
            with warnings.catch_warnings():
                warnings.simplefilter("ignore", UserWarning)
                df_raw = pd.read_excel(ruta, sheet_name=0, header=None, dtype=str, engine="openpyxl")
            
            df_raw = df_raw.fillna("")
            
            # Detectar fila de encabezados
            hdr_row = self._find_header_row_proyectos(df_raw)
            
            if hdr_row is None:
                self.progreso.emit(f"⚠ No se detectó encabezado en {ruta.name}. Intentando lectura estándar...")
                # Intento con lectura normal
                df = pd.read_excel(ruta, engine='openpyxl')
                filas_a_procesar = list(df.iterrows())
            else:
                self.progreso.emit(f"✓ Encabezado detectado en fila {hdr_row+1}")

                # Mapear columnas
                colmap = self._pick_columns_proyectos(df_raw.iloc[hdr_row, :])
                self.progreso.emit(f"Columnas detectadas: {sum(1 for v in colmap.values() if v is not None)}/{len(colmap)}")

                # Layout de bloque: algunos reportes traen el proyecto y su
                # primer integrante en la misma fila, y los DEMÁS integrantes
                # en filas siguientes con los campos del proyecto en blanco
                # (viven solo en la fila del bloque) -- confirmado real:
                # "Proyectos ... 2025 a 30062026" trae hasta varios
                # integrantes adicionales por proyecto en filas separadas.
                # Sin "arrastrar" los campos del proyecto hacia abajo, esas
                # filas de integrante quedaban sin título (se descartaban
                # con el "if not titulo" de más abajo) y se perdían todos los
                # coautores salvo el primero.
                campos_proyecto = ['TITULO', 'OBJETIVO', 'CODIGO_CIE', 'ANIO', 'FECHA_INICIO',
                                    'FECHA_FINAL', 'ESTADO', 'TIPO_INV', 'FACULTAD', 'GRUPO',
                                    'VALOR_APROBADO']

                def _valor_col(fila_raw, campo):
                    idx = colmap.get(campo)
                    if idx is None or idx >= len(fila_raw):
                        return ''
                    return limpiar_texto(fila_raw.iloc[idx])

                filas_a_procesar = []
                actual = {campo: '' for campo in campos_proyecto}
                for i in range(hdr_row + 1, len(df_raw)):
                    fila_raw = df_raw.iloc[i]
                    if _valor_col(fila_raw, 'TITULO'):
                        actual = {campo: _valor_col(fila_raw, campo) for campo in campos_proyecto}
                    fila = dict(actual)
                    fila['RESPONSABLE'] = _valor_col(fila_raw, 'RESPONSABLE')
                    fila['CEDULA'] = _valor_col(fila_raw, 'CEDULA')
                    filas_a_procesar.append((i, fila))

            count = 0
            for _, row in filas_a_procesar:
                # Obtener campos (ahora con nombres normalizados si usó detección)
                if hdr_row is not None:
                    responsable = limpiar_texto(row.get('RESPONSABLE', ''))
                    cedula = limpiar_cedula(row.get('CEDULA', ''))
                    titulo = limpiar_texto(row.get('TITULO', ''))
                    objetivo = limpiar_texto(row.get('OBJETIVO', ''))
                    codigo_cie = limpiar_texto(row.get('CODIGO_CIE', ''))
                    tipo = limpiar_texto(row.get('TIPO_INV', ''))
                    año = row.get('ANIO', '')
                    fecha_inicio = limpiar_texto(row.get('FECHA_INICIO', ''))
                    fecha_fin = limpiar_texto(row.get('FECHA_FINAL', ''))
                    estado = limpiar_texto(row.get('ESTADO', ''))
                    facultad = limpiar_texto(row.get('FACULTAD', ''))
                    grupo = limpiar_texto(row.get('GRUPO', ''))
                    valor = limpiar_texto(row.get('VALOR_APROBADO', ''))
                else:
                    # Fallback: buscar por nombres de columnas originales
                    cedula = limpiar_cedula(
                        row.get('Cedula') or row.get('Cédula') or row.get('CEDULA') or 
                        row.get('Documento') or row.get('Número de documento') or ''
                    )
                    responsable = limpiar_texto(
                        row.get('Responsable') or row.get('RESPONSABLE') or 
                        row.get('Investigador principal') or row.get('Nombre') or ''
                    )
                    titulo = limpiar_texto(
                        row.get('Título') or row.get('Titulo') or row.get('TITULO') or
                        row.get('Nombre proyecto') or ''
                    )
                    objetivo = limpiar_texto(row.get('Objetivo') or row.get('OBJETIVO') or '')
                    codigo_cie = limpiar_texto(row.get('Código CIE') or row.get('Codigo CIE') or '')
                    tipo = limpiar_texto(row.get('Tipo') or row.get('TIPO') or '')
                    año = row.get('Año') or row.get('AÑO') or row.get('Ano')
                    fecha_inicio = limpiar_texto(row.get('Fecha inicio') or row.get('FECHA INICIO') or '')
                    fecha_fin = limpiar_texto(row.get('Fecha final') or row.get('FECHA FINAL') or '')
                    estado = limpiar_texto(row.get('Estado') or row.get('ESTADO') or '')
                    facultad = limpiar_texto(row.get('Facultad') or row.get('FACULTAD') or '')
                    grupo = limpiar_texto(row.get('Grupo') or row.get('GRUPO') or '')
                    valor = limpiar_texto(row.get('Valor aprobado') or row.get('VALOR APROBADO') or '')
                
                if not titulo:
                    continue
                
                if not cedula:
                    cedula = '0000000'
                
                if responsable:
                    self.db.insertar_persona(cedula, responsable, '', facultad, 'Investigador')
                
                # Convertir año
                if pd.notna(año) and año:
                    try:
                        año = int(str(año).split('.')[0])
                    except:
                        año = None
                
                cedula_principal = self.db.obtener_cedula_principal(cedula)
                cursor = self.db.conn.cursor()
                cursor.execute('''
                    INSERT OR REPLACE INTO proyectos
                    (cedula, responsable, titulo, objetivo, codigo_cie, tipo, año,
                     fecha_inicio, fecha_fin, estado, facultad, grupo, valor_aprobado, fuente)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    cedula_principal, responsable, titulo, objetivo, codigo_cie, 
                    tipo, año, fecha_inicio, fecha_fin, estado, facultad, 
                    grupo, valor, ruta.name
                ))
                count += 1
            
            self.db.conn.commit()
            self.progreso.emit(f"✓ Cargados {count} proyectos de investigación desde {ruta.name}")
          except Exception as e:
            self.progreso.emit(f"Error en proyectos ({ruta.name}): {str(e)}")

    def _find_header_row_proyectos(self, df: pd.DataFrame, max_scan: int = 50) -> int:
        """Detecta la fila de encabezados en proyectos"""
        keys_any = {"responsable", "responsables", "investigador principal"}
        keys_id = {"cedula", "documento"}
        other_ok = {"codigo cie", "objetivo", "titulo"}
        
        rows = min(max_scan, df.shape[0])
        for r in range(rows):
            vals = [norm_text(str(v)) for v in df.iloc[r, :].tolist()]
            have_resp = any(any(k in v for k in keys_any) for v in vals)
            have_id = any(any(k in v for k in keys_id) for v in vals)
            have_other = any(any(k in v for k in other_ok) for v in vals)
            if have_resp and (have_id or have_other):
                return r
        return None
    
    def _pick_columns_proyectos(self, header: pd.Series) -> dict:
        """Mapea columnas de proyectos"""
        cols = {
            "RESPONSABLE": None, "CEDULA": None, "TITULO": None, "OBJETIVO": None,
            "CODIGO_CIE": None, "ANIO": None, "FECHA_INICIO": None, "FECHA_FINAL": None,
            "ESTADO": None, "TIPO_INV": None, "VALOR_APROBADO": None, 
            "FACULTAD": None, "GRUPO": None
        }
        
        for j, val in header.items():
            key = norm_text(str(val))
            if cols["RESPONSABLE"] is None and ("responsable" in key or "investigador principal" in key):
                cols["RESPONSABLE"] = j
            elif cols["CEDULA"] is None and ("cedula" in key or "documento" in key):
                cols["CEDULA"] = j
            elif cols["TITULO"] is None and ("titulo" in key or "nombre del proyecto" in key or "proyecto" in key):
                cols["TITULO"] = j
            elif cols["OBJETIVO"] is None and "objetivo" in key:
                cols["OBJETIVO"] = j
            elif cols["CODIGO_CIE"] is None and "cie" in key:
                cols["CODIGO_CIE"] = j
            elif cols["ANIO"] is None and ("ano" in key or "year" in key):
                cols["ANIO"] = j
            elif cols["FECHA_INICIO"] is None and "inicio" in key:
                cols["FECHA_INICIO"] = j
            elif cols["FECHA_FINAL"] is None and ("final" in key or "finalizacion" in key):
                cols["FECHA_FINAL"] = j
            elif cols["ESTADO"] is None and "estado" in key:
                cols["ESTADO"] = j
            elif cols["TIPO_INV"] is None and "tipo" in key:
                cols["TIPO_INV"] = j
            elif cols["VALOR_APROBADO"] is None and ("valor" in key or "aprobado" in key):
                cols["VALOR_APROBADO"] = j
            elif cols["FACULTAD"] is None and "facultad" in key:
                cols["FACULTAD"] = j
            elif cols["GRUPO"] is None and "grupo" in key:
                cols["GRUPO"] = j
        
        return cols
    
    def cargar_propiedad_intelectual(self):
        self.progreso.emit("Cargando registros de propiedad intelectual...")
        archivos_posibles = [
            "CGT0104  No de productos resultados de investigacion 31072025.xlsx",
            "CGT0104  No de productos resultados de investigacion 31122024.xlsx"
        ]
        
        rutas = []
        for archivo in archivos_posibles:
            ruta = self.archivos_directorio / archivo
            if ruta.exists():
                rutas.append(ruta)
        rutas.extend(self._rutas_adicionales('propiedad_intelectual'))

        count = 0
        for ruta in rutas:
            try:
                wb = pd.ExcelFile(ruta, engine='openpyxl')
                for sheet_name in wb.sheet_names:
                    if 'Soporte' in sheet_name or 'Reg Propiedad' in sheet_name:
                        df = pd.read_excel(ruta, sheet_name=sheet_name, engine='openpyxl')
                        
                        columnas_conocidas = COLUMNAS_CONOCIDAS_POR_CATEGORIA['cgt0104_2025']
                        col_extra_por_norm = {
                            c: c for c in df.columns
                            if normalizar_columna(c) not in columnas_conocidas
                        }

                        for _, row in df.iterrows():
                            responsable = limpiar_texto(row.get('Responsables') or row.get('Responsable') or '')
                            tipo_producto = limpiar_texto(row.get('Tipo de producto') or '')
                            nombre = limpiar_texto(row.get('Nombre del producto') or row.get('Nombre') or '')

                            if not responsable and not nombre:
                                continue

                            datos_extra = {
                                c: limpiar_texto(row.get(c))
                                for c in col_extra_por_norm
                                if limpiar_texto(row.get(c))
                            }
                            extra_json = json.dumps(datos_extra, ensure_ascii=False) if datos_extra else ''

                            cursor = self.db.conn.cursor()
                            cursor.execute('''
                                INSERT OR REPLACE INTO propiedad_intelectual
                                (responsable, tipo_producto, tipo_patente, nombre_producto,
                                 numero_registro, proyecto, fecha_aprobacion, entidad, facultad, fuente,
                                 datos_adicionales)
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            ''', (
                                responsable,
                                tipo_producto,
                                limpiar_texto(row.get('Tipo de patente') or ''),
                                nombre,
                                limpiar_texto(row.get('No de registro') or ''),
                                limpiar_texto(row.get('Proyecto de investigación') or ''),
                                limpiar_texto(row.get('Fecha de aprobación') or ''),
                                limpiar_texto(row.get('Entidad que lo expide') or ''),
                                limpiar_texto(row.get('Facultad') or ''),
                                f"{ruta.name} :: {sheet_name}",
                                extra_json,
                            ))
                            count += 1

                self.db.conn.commit()
            except Exception as e:
                self.progreso.emit(f"Error en {ruta.name}: {str(e)}")
        
        if count > 0:
            self.progreso.emit(f"Cargados {count} registros de propiedad intelectual")

# ==================== INTERFAZ ====================

class VistaBusqueda(QWidget):
    def __init__(self, db):
        super().__init__()
        self.db = db
        self._analisis_gruplac = AnalisisDuplicados(db.conn)
        self._cache_gruplac = None
        self.setup_ui()

    def _obtener_cache_gruplac(self):
        """Carga (una sola vez por sesión) los integrantes de todos los grupos
        GrupLAC desde reports/excel, para verificar coherencia interno↔GrupLAC."""
        if self._cache_gruplac is None:
            self._cache_gruplac = self._analisis_gruplac._cargar_integrantes_gruplac()
        return self._cache_gruplac
    
    def setup_ui(self):
        layout = QHBoxLayout()
        
        panel_izq = QWidget()
        layout_izq = QVBoxLayout(panel_izq)
        
        self.input_busqueda = QLineEdit()
        self.input_busqueda.setPlaceholderText("Buscar por nombre o cédula...")
        self.input_busqueda.returnPressed.connect(self.buscar)
        
        btn_buscar = QPushButton("Buscar")
        btn_buscar.clicked.connect(self.buscar)
        
        btn_consolidar = QPushButton("Consolidar Duplicados")
        btn_consolidar.clicked.connect(self.consolidar_duplicados_manual)
        btn_consolidar.setStyleSheet("background-color: #e67e22; color: white;")
        
        self.tabla_resultados = QTableWidget()
        self.tabla_resultados.setColumnCount(5)
        self.tabla_resultados.setHorizontalHeaderLabels(['Nombre', 'Cédula', 'Email', 'Facultad', 'Tipo'])
        self.tabla_resultados.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tabla_resultados.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tabla_resultados.horizontalHeader().setStretchLastSection(True)
        self.tabla_resultados.doubleClicked.connect(self.mostrar_detalle)
        
        layout_izq.addWidget(QLabel("Búsqueda de Personas"))
        layout_izq.addWidget(self.input_busqueda)
        layout_izq.addWidget(btn_buscar)
        layout_izq.addWidget(btn_consolidar)
        layout_izq.addWidget(self.tabla_resultados)
        
        panel_der = QWidget()
        layout_der = QVBoxLayout(panel_der)
        
        self.texto_detalle = QTextEdit()
        self.texto_detalle.setReadOnly(True)
        
        layout_der.addWidget(QLabel("Detalle de Persona"))
        layout_der.addWidget(self.texto_detalle)
        
        splitter = QSplitter(Qt.Horizontal)
        splitter.addWidget(panel_izq)
        splitter.addWidget(panel_der)
        splitter.setSizes([400, 600])
        
        layout.addWidget(splitter)
        self.setLayout(layout)
    
    def consolidar_duplicados_manual(self):
        consolidaciones = self.db.consolidar_duplicados()
        
        if consolidaciones:
            mensaje = f"✓ Se consolidaron {len(consolidaciones)} persona(s):\n\n"
            for grupo in consolidaciones:
                mensaje += f"- {grupo['nombre']}\n"
                mensaje += f"  Cédula principal: {grupo['cedula_principal']}\n"
                mensaje += f"  Cédulas consolidadas: {', '.join(grupo['cedulas'])}\n\n"
            
            QMessageBox.information(self, "Consolidación Exitosa", mensaje)
            
            if self.input_busqueda.text():
                self.buscar()
        else:
            QMessageBox.information(self, "Sin Duplicados", "No se encontraron duplicados para consolidar.")
    
    def buscar(self):
        termino = self.input_busqueda.text().strip()
        if not termino:
            return
        
        resultados = self.db.buscar_personas(termino)
        self.tabla_resultados.setRowCount(len(resultados))
        
        for row, persona in enumerate(resultados):
            self.tabla_resultados.setItem(row, 0, QTableWidgetItem(persona[1] or ''))
            self.tabla_resultados.setItem(row, 1, QTableWidgetItem(persona[0] or ''))
            self.tabla_resultados.setItem(row, 2, QTableWidgetItem(persona[2] or ''))
            self.tabla_resultados.setItem(row, 3, QTableWidgetItem(persona[3] or ''))
            self.tabla_resultados.setItem(row, 4, QTableWidgetItem(persona[4] or ''))
    
    def mostrar_detalle(self, index):
        row = index.row()
        cedula = self.tabla_resultados.item(row, 1).text()
        detalle = self.db.obtener_detalle_persona(cedula)
        if not detalle or "info" not in detalle:
            return

        info = detalle["info"]

        def _s(v):
            return str(v) if v else "—"

        def _extra_html(fila):
            """Si la fila trae 'datos_adicionales' (columnas nuevas guardadas
            como JSON, ver views/vista_inicio.py), las muestra como nota."""
            crudo = fila[-1] if fila else None
            if not crudo:
                return ""
            try:
                datos = json.loads(crudo)
            except (TypeError, ValueError):
                return ""
            if not datos:
                return ""
            partes = "; ".join(f"{k}: {v}" for k, v in datos.items())
            return f"<br><span style='color:#888;font-size:9px;'>+ {partes}</span>"

        def _seccion(titulo, color, items_html):
            return (
                f"<div style='margin-top:10px;'>"
                f"<div style='background:{color};color:white;padding:4px 8px;"
                f"border-radius:4px;font-weight:bold;font-size:12px;'>{titulo}</div>"
                f"<div style='padding:4px 8px;'>{items_html}</div></div>"
            )

        html = (
            "<html><body style='font-family:Arial,sans-serif;font-size:11px;"
            "max-width:100%;word-wrap:break-word;'>"
            "<div style='background:#1a365d;color:white;padding:8px;border-radius:6px;"
            "margin-bottom:10px;'>"
            f"<b style='font-size:13px;'>{_s(info['nombre'])}</b><br>"
            f"<span style='font-size:10px;'>Cédula: {_s(info['cedula'])} | "
            f"Facultad: {_s(info['facultad'])}<br>"
            f"Email: {_s(info['email'])} | Tipo: {_s(info['tipo'])}</span></div>"
        )

        grupos_todos = detalle.get("grupos") or []
        grupos_invest = [g for g in grupos_todos if "semillero" not in (g[2] or "").lower()]
        semilleros = [g for g in grupos_todos if "semillero" in (g[2] or "").lower()]

        if grupos_invest:
            items = "".join(
                f"<p style='margin:2px 0;'>&#9679; <b>{_s(g[2])}</b>"
                f" <span style='color:#666;'>({_s(g[4])})</span></p>"
                for g in grupos_invest
            )
            html += _seccion(
                f"Grupos de investigación ({len(grupos_invest)})",
                "#2c3e50", items
            )

        if semilleros:
            items = "".join(
                f"<p style='margin:2px 0;'>&#9679; <b>{_s(g[2])}</b>"
                f" <span style='color:#666;'>({_s(g[4])})</span></p>"
                for g in semilleros
            )
            html += _seccion(
                f"Semilleros de investigación ({len(semilleros)})",
                "#5b6b73", items
            )

        # ── Coherencia interno ↔ GrupLAC ─────────────────────────────────
        nombre_norm = normalizar_nombre(info["nombre"])
        cache_gruplac = self._obtener_cache_gruplac()
        notas_gruplac = []

        grupos_internos_norm = set()
        for g in grupos_invest:
            nombre_grupo = g[2]
            grupo_resuelto = self._analisis_gruplac.resolver_grupo_en_gruplac(nombre_grupo, cache_gruplac)
            grupos_internos_norm.add(self._analisis_gruplac._normalizar_grupo(nombre_grupo))
            if grupo_resuelto is not None:
                grupos_internos_norm.add(grupo_resuelto)

            if grupo_resuelto is None:
                # No hay ningún reporte GrupLAC para este grupo: no se puede
                # confirmar ni desmentir nada, así que no se muestra advertencia.
                continue

            entry = self._analisis_gruplac._nombre_en_gruplac(nombre_norm, nombre_grupo, cache_gruplac)
            if entry is None:
                notas_gruplac.append((
                    "warn",
                    f"No aparece en GrupLAC para el grupo <b>{_s(nombre_grupo)}</b>. Está "
                    "registrado en la base de datos interna de la Vicerrectoría de "
                    "Investigación, Innovación y Extensión, pero no en GrupLAC — favor "
                    "revisar o actualizar."
                ))
            else:
                if entry["activo"]:
                    estado = f"activo desde {entry['desde']}"
                elif entry["hasta"] != "—":
                    estado = f"retirado el {entry['hasta']} (desde {entry['desde']})"
                else:
                    estado = "vinculación registrada en GrupLAC"
                notas_gruplac.append((
                    "ok",
                    f"Sí aparece en GrupLAC para <b>{_s(nombre_grupo)}</b> ({estado})."
                ))

        hallazgos_extra = self._analisis_gruplac.buscar_persona_en_gruplac(
            nombre_norm, cache_gruplac, excluir_grupos_norm=grupos_internos_norm
        )
        for grupo_norm, entry in hallazgos_extra:
            notas_gruplac.append((
                "warn",
                f"Aparece en GrupLAC en el grupo <b>{_s(grupo_norm.title())}</b>, pero no "
                "está registrado en ese grupo en la base de datos interna — favor revisar "
                "o actualizar para que ambos sistemas coincidan."
            ))

        if notas_gruplac:
            items = "".join(
                f"<p style='margin:3px 0;padding:4px 6px;border-radius:4px;"
                f"background:{'#eafaf1' if tipo == 'ok' else '#fdf2e3'};"
                f"color:{'#1e7e34' if tipo == 'ok' else '#9a6700'};'>"
                f"{'✓' if tipo == 'ok' else '⚠'} {texto}</p>"
                for tipo, texto in notas_gruplac
            )
            html += _seccion("Coherencia con GrupLAC", "#6c5ce7", items)

        if detalle.get("publicaciones"):
            items = "".join(
                f"<p style='margin:2px 0;border-bottom:1px solid #eee;padding-bottom:2px;'>"
                f"&#9679; <b>{_s(pub[2])[:120]}</b><br>"
                f"<span style='color:#555;'>Año: {_s(pub[5])} | Tipo: {_s(pub[6])} | "
                f"Revista: {_s(pub[3])[:60]}</span>{_extra_html(pub)}</p>"
                for pub in detalle["publicaciones"]
            )
            html += _seccion(
                f"Publicaciones ({len(detalle['publicaciones'])})",
                "#2e86ab", items
            )

        if detalle.get("extensiones"):
            items = "".join(
                f"<p style='margin:2px 0;'>&#9679; <b>{_s(ext[2])[:100]}</b> "
                f"<span style='color:#555;'>({_s(ext[3])} · {_s(ext[8])})</span>"
                f"{_extra_html(ext)}</p>"
                for ext in detalle["extensiones"]
            )
            html += _seccion(
                f"Extensiones ({len(detalle['extensiones'])})",
                "#f18f01", items
            )

        def _badge_conducente(calificacion):
            cal = (calificacion or "").strip().upper()
            if cal == "NO CONDUCENTE":
                return " <span style='background:#f4d1d1;color:#8a2b2b;border-radius:3px;padding:1px 5px;font-size:10px;'>NO CONDUCENTE</span>"
            if cal == "CONDUCENTE":
                return " <span style='background:#d4ecd8;color:#1e6b3a;border-radius:3px;padding:1px 5px;font-size:10px;'>CONDUCENTE</span>"
            return ""

        if detalle.get("trabajos_grado_director"):
            items = "".join(
                f"<p style='margin:2px 0;'>&#9679; <b>{_s(tg[5])[:100]}</b>{_badge_conducente(tg[10])}<br>"
                f"<span style='color:#555;'>Estudiante: {_s(tg[4])} | Prog: {_s(tg[6])} | "
                f"Año: {_s(tg[7])}</span></p>"
                for tg in detalle["trabajos_grado_director"]
            )
            html += _seccion(
                f"Trabajos de grado dirigidos ({len(detalle['trabajos_grado_director'])})",
                "#3b8c66", items
            )

        if detalle.get("trabajos_grado_estudiante"):
            items = "".join(
                f"<p style='margin:2px 0;'>&#9679; <b>{_s(tg[5])[:100]}</b>{_badge_conducente(tg[10])}<br>"
                f"<span style='color:#555;'>Director: {_s(tg[2])} | Prog: {_s(tg[6])} | "
                f"Año: {_s(tg[7])}</span></p>"
                for tg in detalle["trabajos_grado_estudiante"]
            )
            html += _seccion(
                f"Trabajos de grado (como estudiante) ({len(detalle['trabajos_grado_estudiante'])})",
                "#52796f", items
            )

        if detalle.get("productos_innovacion"):
            items = "".join(
                f"<p style='margin:2px 0;'>&#9679; <b>{_s(p[3])[:100]}</b> "
                f"<span style='color:#555;'>({_s(p[2])} · {_s(p[5])})</span>"
                f"{_extra_html(p)}</p>"
                for p in detalle["productos_innovacion"]
            )
            html += _seccion(
                f"Productos de innovación ({len(detalle['productos_innovacion'])})",
                "#a23b72", items
            )

        if detalle.get("proyectos"):
            items = "".join(
                f"<p style='margin:2px 0;'>&#9679; <b>{_s(pr[3])[:100]}</b> "
                f"<span style='color:#555;'>(Año: {_s(pr[7])} · {_s(pr[10])})</span></p>"
                for pr in detalle["proyectos"]
            )
            html += _seccion(
                f"Proyectos ({len(detalle['proyectos'])})",
                "#c73e1d", items
            )

        html += "</body></html>"
        self.texto_detalle.setHtml(html)


class VistaGrupos(QWidget):
    def __init__(self, db):
        super().__init__()
        self.db = db
        self.productos_completos = []
        self._analisis_gruplac = AnalisisDuplicados(db.conn)
        self._cache_gruplac = None
        self._cache_notas_autor = {}
        self.setup_ui()

    def _obtener_cache_gruplac(self):
        if self._cache_gruplac is None:
            self._cache_gruplac = self._analisis_gruplac._cargar_integrantes_gruplac()
        return self._cache_gruplac

    def setup_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(2)
        layout.setContentsMargins(3, 3, 3, 3)
        
        # ===== BARRA SUPERIOR SIN ESTADÍSTICAS =====
        barra_superior = QWidget()
        barra_superior.setMaximumHeight(35)
        barra_superior.setStyleSheet("background-color: #f8f9fa; border-radius: 3px;")
        layout_barra = QHBoxLayout(barra_superior)
        layout_barra.setContentsMargins(5, 3, 5, 3)
        layout_barra.setSpacing(5)
        
        # Grupo
        layout_barra.addWidget(QLabel("<b>Grupo:</b>"))
        self.combo_grupos = QComboBox()
        self.combo_grupos.setMinimumWidth(200)
        self.combo_grupos.currentTextChanged.connect(self.seleccionar_grupo)
        layout_barra.addWidget(self.combo_grupos)
        
        # Botón refrescar
        btn_refrescar = QPushButton("↻")
        btn_refrescar.setMaximumWidth(30)
        btn_refrescar.setToolTip("Refrescar grupos")
        btn_refrescar.clicked.connect(self.cargar_grupos)
        layout_barra.addWidget(btn_refrescar)
        
        layout_barra.addWidget(QLabel("|"))
        
        # Filtros inline
        self.check_pub = QCheckBox("Pub")
        self.check_pub.setChecked(True)
        self.check_pub.stateChanged.connect(self.seleccionar_grupo)
        layout_barra.addWidget(self.check_pub)
        
        self.check_ext = QCheckBox("Ext")
        self.check_ext.setChecked(True)
        self.check_ext.stateChanged.connect(self.seleccionar_grupo)
        layout_barra.addWidget(self.check_ext)
        
        self.check_tg = QCheckBox("TG")
        self.check_tg.setChecked(True)
        self.check_tg.stateChanged.connect(self.seleccionar_grupo)
        layout_barra.addWidget(self.check_tg)
        
        self.check_innov = QCheckBox("Inn")
        self.check_innov.setChecked(True)
        self.check_innov.stateChanged.connect(self.seleccionar_grupo)
        layout_barra.addWidget(self.check_innov)
        
        self.check_proy = QCheckBox("Proy")
        self.check_proy.setChecked(True)
        self.check_proy.stateChanged.connect(self.seleccionar_grupo)
        layout_barra.addWidget(self.check_proy)

        self.check_pi = QCheckBox("PI")
        self.check_pi.setChecked(True)
        self.check_pi.setToolTip("Propiedad Intelectual")
        self.check_pi.stateChanged.connect(self.seleccionar_grupo)
        layout_barra.addWidget(self.check_pi)

        layout_barra.addWidget(QLabel("|"))

        # Período (año desde - hasta) que se muestra y se exporta
        layout_barra.addWidget(QLabel("Año:"))
        self.spin_anio_desde = QSpinBox()
        self.spin_anio_desde.setRange(1990, 2035)
        self.spin_anio_desde.setValue(2022)
        self.spin_anio_desde.setFixedWidth(60)
        self.spin_anio_desde.setToolTip("Año inicial del período a mostrar/exportar")
        self.spin_anio_desde.valueChanged.connect(self._cambio_rango_anios)
        layout_barra.addWidget(self.spin_anio_desde)

        layout_barra.addWidget(QLabel("-"))
        self.spin_anio_hasta = QSpinBox()
        self.spin_anio_hasta.setRange(1990, 2035)
        self.spin_anio_hasta.setValue(datetime.now().year)
        self.spin_anio_hasta.setFixedWidth(60)
        self.spin_anio_hasta.setToolTip("Año final del período a mostrar/exportar")
        self.spin_anio_hasta.valueChanged.connect(self._cambio_rango_anios)
        layout_barra.addWidget(self.spin_anio_hasta)

        layout_barra.addStretch()
        
        # Botones exportar
        btn_excel = QPushButton("Excel")
        btn_excel.setMaximumWidth(70)
        btn_excel.clicked.connect(self.exportar_excel)
        btn_excel.setStyleSheet("background-color: #27ae60; color: white; font-size: 11px;")
        layout_barra.addWidget(btn_excel)
        
        btn_pdf = QPushButton("PDF")
        btn_pdf.setMaximumWidth(60)
        btn_pdf.clicked.connect(self.exportar_pdf)
        btn_pdf.setStyleSheet("background-color: #e74c3c; color: white; font-size: 11px;")
        layout_barra.addWidget(btn_pdf)
        
        # ===== PANEL PRINCIPAL - 3 COLUMNAS =====
        splitter = QSplitter(Qt.Horizontal)
        
        # COLUMNA 1: Integrantes (20%)
        panel_integrantes = QWidget()
        layout_int = QVBoxLayout(panel_integrantes)
        layout_int.setSpacing(2)
        layout_int.setContentsMargins(2, 2, 2, 2)
        
        lbl_int = QLabel("<b>Integrantes</b>")
        lbl_int.setStyleSheet("color: #1a365d; font-size: 12px;")
        layout_int.addWidget(lbl_int)
        
        self.tabla_integrantes = QTableWidget()
        self.tabla_integrantes.setColumnCount(5)
        self.tabla_integrantes.setHorizontalHeaderLabels(['Nombre', 'Tipo', 'Email', 'Facultad', 'Procedencia'])
        self.tabla_integrantes.horizontalHeader().setStretchLastSection(True)
        self.tabla_integrantes.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tabla_integrantes.verticalHeader().setVisible(False)
        self.tabla_integrantes.setAlternatingRowColors(True)
        self.tabla_integrantes.setStyleSheet("font-size: 10px;")
        layout_int.addWidget(self.tabla_integrantes)
        
        # COLUMNA 2: Productos (55%)
        panel_productos = QWidget()
        layout_prod = QVBoxLayout(panel_productos)
        layout_prod.setSpacing(2)
        layout_prod.setContentsMargins(2, 2, 2, 2)
        
        lbl_prod = QLabel("<b>Productos del Grupo</b>")
        lbl_prod.setStyleSheet("color: #1a365d; font-size: 12px;")
        layout_prod.addWidget(lbl_prod)
        
        self.tabla_productos = QTableWidget()
        self.tabla_productos.setColumnCount(6)
        self.tabla_productos.setHorizontalHeaderLabels([
            'Investigador', 'Título', 'Año', 'Tipo', 'Categoría', 'Estado'
        ])
        self.tabla_productos.horizontalHeader().setStretchLastSection(True)
        self.tabla_productos.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tabla_productos.verticalHeader().setVisible(False)
        self.tabla_productos.setAlternatingRowColors(True)
        self.tabla_productos.setStyleSheet("font-size: 10px;")
        self.tabla_productos.clicked.connect(self.mostrar_detalle_producto)
        layout_prod.addWidget(self.tabla_productos)
        
        # COLUMNA 3: Detalle del Producto (25%)
        panel_detalle = QWidget()
        layout_det = QVBoxLayout(panel_detalle)
        layout_det.setSpacing(2)
        layout_det.setContentsMargins(2, 2, 2, 2)
        
        lbl_det = QLabel("<b>Detalle del Producto</b>")
        lbl_det.setStyleSheet("color: #1a365d; font-size: 12px;")
        layout_det.addWidget(lbl_det)
        
        self.texto_detalle = QTextEdit()
        self.texto_detalle.setReadOnly(True)
        self.texto_detalle.setPlaceholderText("Selecciona un producto para ver detalles...")
        self.texto_detalle.setStyleSheet("font-size: 10px; background-color: #fafafa;")
        layout_det.addWidget(self.texto_detalle)
        
        # Agregar columnas al splitter
        splitter.addWidget(panel_integrantes)
        splitter.addWidget(panel_productos)
        splitter.addWidget(panel_detalle)
        splitter.setSizes([280, 750, 370])
        
        # Agregar todo
        layout.addWidget(barra_superior)
        layout.addWidget(splitter)
        
        self.setLayout(layout)
        self.cargar_grupos()
    
    def cargar_grupos(self):
        self.combo_grupos.clear()
        self.combo_grupos.addItem("-- Seleccionar --")
        grupos = self.db.obtener_grupos()
        for grupo in grupos:
            if grupo[0]:
                self.combo_grupos.addItem(grupo[0])
    
    def _cambio_rango_anios(self):
        if self.spin_anio_desde.value() > self.spin_anio_hasta.value():
            self.spin_anio_hasta.setValue(self.spin_anio_desde.value())
        self.seleccionar_grupo()

    def seleccionar_grupo(self, grupo=None):
        if grupo is None:
            grupo = self.combo_grupos.currentText()
            
        if grupo == "-- Seleccionar --" or not grupo:
            self.tabla_integrantes.setRowCount(0)
            self.tabla_productos.setRowCount(0)
            self.texto_detalle.clear()
            return
        
        # Cargar integrantes
        integrantes = self.db.obtener_integrantes_grupo(grupo)
        self.tabla_integrantes.setRowCount(len(integrantes))
        
        for row, persona in enumerate(integrantes):
            self.tabla_integrantes.setItem(row, 0, QTableWidgetItem(persona[1]))
            self.tabla_integrantes.setItem(row, 1, QTableWidgetItem(persona[2] or ''))
            self.tabla_integrantes.setItem(row, 2, QTableWidgetItem(persona[3] or ''))
            self.tabla_integrantes.setItem(row, 3, QTableWidgetItem(persona[4] or ''))
            self.tabla_integrantes.setItem(row, 4, QTableWidgetItem(
                self.db.procedencia_grupo(persona[0], grupo)))
        
        # Generar reporte automáticamente
        filtros = []
        if self.check_pub.isChecked():
            filtros.append('Publicaciones')
        if self.check_ext.isChecked():
            filtros.append('Extensiones')
        if self.check_tg.isChecked():
            filtros.append('Trabajos de Grado')
        if self.check_innov.isChecked():
            filtros.append('Productos Innovación')
        if self.check_proy.isChecked():
            filtros.append('Proyectos')
        if self.check_pi.isChecked():
            filtros.append('Propiedad Intelectual')

        if not filtros:
            self.tabla_productos.setRowCount(0)
            return

        productos = self.db.obtener_productos_grupo_detallado(
            grupo, filtros,
            anio_desde=self.spin_anio_desde.value(),
            anio_hasta=self.spin_anio_hasta.value(),
        )
        productos = self._agrupar_por_titulo(productos)
        self.productos_completos = productos
        self.tabla_productos.setRowCount(len(productos))
        
        for row, prod in enumerate(productos):
            item_inv = QTableWidgetItem(prod.get('investigador', ''))
            item_inv.setData(Qt.UserRole, prod)
            self.tabla_productos.setItem(row, 0, item_inv)
            
            self.tabla_productos.setItem(row, 1, QTableWidgetItem(prod.get('titulo', '')))
            self.tabla_productos.setItem(row, 2, QTableWidgetItem(str(prod.get('año', '')) if prod.get('año') else ''))
            self.tabla_productos.setItem(row, 3, QTableWidgetItem(prod.get('tipo_producto', '')))
            self.tabla_productos.setItem(row, 4, QTableWidgetItem(prod.get('categoria', '')))
            self.tabla_productos.setItem(row, 5, QTableWidgetItem(prod.get('estado', '')))

    def _agrupar_por_titulo(self, productos):
        """
        Un mismo producto con varios coautores del mismo grupo aparece una vez
        por autor en `productos` (una fila por cédula en la tabla origen).
        Esto los agrupa en una sola entrada por (tipo, título, año), listando
        todos los autores juntos en 'investigador' en vez de repetir el
        producto — así se distingue un producto realmente repetido de uno
        que solo tiene varios autores.
        """
        grupos = {}
        orden = []
        for prod in productos:
            clave = (prod.get('tipo_producto'), normalizar_nombre(prod.get('titulo') or ''), prod.get('año'))
            if clave not in grupos:
                nuevo = dict(prod)
                nuevo['autores'] = []
                grupos[clave] = nuevo
                orden.append(clave)
            nombre = prod.get('investigador')
            autores = grupos[clave]['autores']
            if nombre and not any(a['nombre'] == nombre for a in autores):
                autores.append({
                    'nombre': nombre,
                    'cedula': prod.get('cedula'),
                    'nota_gruplac': self._nota_gruplac_autor(nombre, prod.get('cedula')),
                })

        resultado = []
        for clave in orden:
            g = grupos[clave]
            g['investigador'] = '; '.join(self._formatear_autor(a) for a in g['autores'])
            resultado.append(g)
        return resultado

    @staticmethod
    def _formatear_autor(autor):
        nombre = autor.get('nombre') or ''
        nota = autor.get('nota_gruplac')
        return f"{nombre} {nota}" if nota else nombre

    @staticmethod
    def _texto_datos_adicionales(crudo):
        """Convierte el JSON de 'datos_adicionales' (columnas del Excel que el
        sistema no reconocía al cargar, ver views/vista_inicio.py) a un texto
        legible de una línea, para mostrarlo en el panel de detalle y en los
        exports."""
        if not crudo:
            return ''
        try:
            datos = json.loads(crudo)
        except (TypeError, ValueError):
            return ''
        return '; '.join(f'{k}: {v}' for k, v in datos.items()) if datos else ''

    def _nota_gruplac_autor(self, nombre, cedula):
        """
        Si esta persona también aparece en GrupLAC bajo un grupo donde NO
        está registrada internamente, devuelve una nota corta para mostrar
        junto a su nombre. Reutiliza la misma lógica de "Búsqueda de
        Personas": excluye TODOS los grupos internos de la persona (no solo
        el grupo que se está reportando) — si solo excluyéramos el grupo
        actual, alguien correctamente registrado en GrupLAC bajo OTRO de sus
        propios grupos internos quedaría marcado como inconsistencia falsa
        (esto pasaba con Mauricio Holguín: está en 4 grupos internos, y el
        reporte de uno de ellos lo marcaba mal por no conocer los otros 3).
        """
        if not nombre:
            return ''
        nombre_norm = normalizar_nombre(nombre)
        clave_cache = (nombre_norm, cedula)
        if clave_cache in self._cache_notas_autor:
            return self._cache_notas_autor[clave_cache]

        cache = self._obtener_cache_gruplac()
        grupos_persona = []
        if cedula:
            grupos_persona = [
                r[0] for r in self.db.conn.execute(
                    "SELECT DISTINCT grupo FROM grupos WHERE cedula = ? "
                    "AND grupo IS NOT NULL AND grupo != ''",
                    (cedula,)
                ).fetchall()
            ]

        grupos_internos_norm = set()
        for g in grupos_persona:
            grupos_internos_norm.add(self._analisis_gruplac._normalizar_grupo(g))
            resuelto = self._analisis_gruplac.resolver_grupo_en_gruplac(g, cache)
            if resuelto is not None:
                grupos_internos_norm.add(resuelto)

        hallazgos = self._analisis_gruplac.buscar_persona_en_gruplac(
            nombre_norm, cache, excluir_grupos_norm=grupos_internos_norm
        )
        if not hallazgos:
            nota = ''
        else:
            grupos_extra = ', '.join(sorted({g.title() for g, _ in hallazgos}))
            nota = (
                f"⚠ también aparece en GrupLAC en el grupo {grupos_extra}, "
                "donde no está registrado en la BD interna"
            )

        self._cache_notas_autor[clave_cache] = nota
        return nota

    def mostrar_detalle_producto(self, index):
        row = index.row()
        item = self.tabla_productos.item(row, 0)
        if not item:
            return

        producto = item.data(Qt.UserRole)
        if not producto:
            return

        tipo = producto.get('tipo_producto', '')

        COLORES_TIPO = {
            'Publicación': '#2e86ab',
            'Extensión': '#f18f01',
            'Trabajo de Grado': '#3b8c66',
            'Innovación': '#a23b72',
            'Proyecto': '#c73e1d',
            'Propiedad Intelectual': '#6a4c93',
        }
        color = COLORES_TIPO.get(tipo, '#1a365d')

        def _s(v):
            return str(v).strip() if v and str(v).strip() not in ('', 'None', 'nan') else '—'

        def _fila(label, valor):
            return (
                f"<tr><td style='color:#555;white-space:nowrap;padding:2px 6px 2px 0;"
                f"vertical-align:top;'><b>{label}</b></td>"
                f"<td style='padding:2px 0;word-break:break-word;'>{_s(valor)}</td></tr>"
            )

        autores = producto.get('autores') or []
        label_investigador = "Investigadores" if len(autores) > 1 else "Investigador"
        fila_cedula = "" if len(autores) > 1 else _fila("Cédula", producto.get('cedula'))

        html = (
            "<html><body style='font-family:Arial,sans-serif;font-size:11px;"
            "max-width:100%;word-wrap:break-word;margin:0;padding:4px;'>"
            f"<div style='background:{color};color:white;padding:6px 10px;"
            f"border-radius:5px;margin-bottom:8px;'>"
            f"<b style='font-size:12px;'>{tipo.upper()}</b></div>"
            "<table style='width:100%;border-collapse:collapse;'>"
            + _fila(label_investigador, producto.get('investigador'))
            + fila_cedula
            + _fila("Año", producto.get('año'))
            + "</table>"
            f"<div style='background:{color};color:white;padding:3px 8px;"
            f"border-radius:3px;margin:8px 0 4px 0;font-size:10px;"
            f"font-weight:bold;'>INFORMACIÓN DETALLADA</div>"
            "<table style='width:100%;border-collapse:collapse;'>"
        )

        if tipo == 'Publicación':
            html += (
                f"<tr><td colspan='2' style='padding:2px 0;word-break:break-word;'>"
                f"<b>Título:</b> {_s(producto.get('titulo'))}</td></tr>"
                f"<tr><td colspan='2' style='padding:2px 0;word-break:break-word;"
                f"color:#444;'><b>Revista/Libro:</b> {_s(producto.get('revista_libro'))}</td></tr>"
                + _fila("DOI/URL", producto.get('doi_url'))
                + _fila("ISSN/ISBN", producto.get('issn_isbn'))
                + _fila("Categoría", producto.get('categoria'))
                + _fila("Estado", producto.get('estado'))
            )

        elif tipo == 'Extensión':
            html += (
                f"<tr><td colspan='2' style='padding:2px 0;word-break:break-word;'>"
                f"<b>Actividad:</b> {_s(producto.get('titulo'))}</td></tr>"
                + _fila("Tipo", producto.get('tipo'))
                + _fila("Modalidad", producto.get('modalidad'))
                + _fila("Estado", producto.get('estado'))
                + _fila("Fechas", f"{_s(producto.get('fecha_inicio'))} – {_s(producto.get('fecha_fin'))}")
                + _fila("Población", producto.get('poblacion'))
                + _fila("Financiación interna", producto.get('financiacion_interna'))
                + _fila("Financiación externa", producto.get('financiacion_externa'))
            )

        elif tipo == 'Trabajo de Grado':
            html += (
                f"<tr><td colspan='2' style='padding:2px 0;word-break:break-word;'>"
                f"<b>Título:</b> {_s(producto.get('titulo'))}</td></tr>"
                + _fila("Director", producto.get('investigador'))
                + _fila("Estudiante", producto.get('estudiante'))
                + _fila("Cédula estudiante", producto.get('cedula_estudiante'))
                + _fila("Programa", producto.get('programa'))
                + _fila("Estado", producto.get('estado'))
                + _fila("Fecha sustentación", producto.get('fecha_sustentacion'))
                + _fila("Calificación", producto.get('calificacion'))
            )

        elif tipo == 'Innovación':
            html += (
                f"<tr><td colspan='2' style='padding:2px 0;word-break:break-word;'>"
                f"<b>Nombre:</b> {_s(producto.get('titulo'))}</td></tr>"
                + _fila("Tipo detalle", producto.get('tipo_producto_detalle'))
                + f"<tr><td colspan='2' style='padding:2px 0;word-break:break-word;'>"
                f"<b>Descripción:</b> {_s(producto.get('descripcion'))}</td></tr>"
                + _fila("Estado", producto.get('estado'))
                + _fila("Grupo", producto.get('grupo'))
            )

        elif tipo == 'Proyecto':
            html += (
                f"<tr><td colspan='2' style='padding:2px 0;word-break:break-word;'>"
                f"<b>Título:</b> {_s(producto.get('titulo'))}</td></tr>"
                f"<tr><td colspan='2' style='padding:2px 0;word-break:break-word;"
                f"color:#444;'><b>Objetivo:</b> {_s(producto.get('objetivo'))}</td></tr>"
                + _fila("Código CIE", producto.get('codigo_cie'))
                + _fila("Tipo", producto.get('tipo'))
                + _fila("Fechas", f"{_s(producto.get('fecha_inicio'))} – {_s(producto.get('fecha_fin'))}")
                + _fila("Estado", producto.get('estado'))
                + _fila("Valor aprobado", producto.get('valor_aprobado'))
            )

        elif tipo == 'Propiedad Intelectual':
            html += (
                f"<tr><td colspan='2' style='padding:2px 0;word-break:break-word;'>"
                f"<b>Nombre del producto:</b> {_s(producto.get('titulo'))}</td></tr>"
                + _fila("Tipo de producto", producto.get('tipo_producto_detalle'))
                + _fila("Tipo de patente", producto.get('tipo_patente'))
                + _fila("No. de registro", producto.get('numero_registro'))
                + _fila("Proyecto", producto.get('proyecto'))
                + _fila("Fecha de aprobación", producto.get('fecha_aprobacion'))
                + _fila("Entidad", producto.get('entidad'))
                + _fila("Facultad", producto.get('facultad'))
            )

        datos_adicionales = producto.get('datos_adicionales')
        if datos_adicionales:
            try:
                extra = json.loads(datos_adicionales)
            except (TypeError, ValueError):
                extra = None
            if extra:
                html += (
                    "<tr><td colspan='2' style='padding-top:6px;'>"
                    "<div style='background:#fff8e1;border:1px solid #f0d98c;border-radius:4px;"
                    "padding:4px 6px;'><b style='color:#8a6d00;'>Columnas adicionales del Excel:</b><br>"
                    + "<br>".join(f"<b>{k}:</b> {_s(v)}" for k, v in extra.items())
                    + "</div></td></tr>"
                )

        html += (
            "</table>"
            "<div style='margin-top:8px;padding-top:4px;border-top:1px solid #ddd;"
            f"color:#666;font-size:10px;'>Fuente: {_s(producto.get('fuente'))}</div>"
            "</body></html>"
        )

        self.texto_detalle.setHtml(html)
    
    def exportar_excel(self):
        grupo = self.combo_grupos.currentText()
        if grupo == "-- Seleccionar --" or not grupo:
            QMessageBox.warning(self, "Advertencia", "Selecciona un grupo primero")
            return
        
        if not self.productos_completos:
            QMessageBox.warning(self, "Advertencia", "No hay productos para exportar")
            return
        
        try:
            from datetime import datetime
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            nombre_sugerido = f"reporte_{limpiar_nombre_archivo(grupo)}_{timestamp}.xlsx"
            reports_dir = obtener_directorio_base() / "reports" / "excel"
            reports_dir.mkdir(parents=True, exist_ok=True)

            ruta_str, _ = QFileDialog.getSaveFileName(
                self, "Guardar reporte Excel del grupo",
                str(reports_dir / nombre_sugerido), "Excel (*.xlsx)")
            if not ruta_str:
                return
            ruta = Path(ruta_str)
            if ruta.suffix.lower() != ".xlsx":
                ruta = ruta.with_suffix(".xlsx")

            # Crear workbook
            wb = openpyxl.Workbook()
            
            # ===== HOJA 1: INTEGRANTES =====
            ws_int = wb.active
            ws_int.title = "Integrantes"
            
            # Título
            ws_int['A1'] = f"GRUPO: {grupo} - INTEGRANTES"
            ws_int['A1'].font = Font(bold=True, size=14, color="1a365d")
            ws_int.merge_cells('A1:F1')

            # Encabezados
            encabezados_int = ['Nombre', 'Tipo', 'Cédula', 'Email', 'Facultad', 'Procedencia']
            for col, enc in enumerate(encabezados_int, 1):
                cell = ws_int.cell(row=2, column=col, value=enc)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Datos integrantes
            integrantes = self.db.obtener_integrantes_grupo(grupo)
            for row, persona in enumerate(integrantes, 3):
                ws_int.cell(row=row, column=1, value=persona[1])
                ws_int.cell(row=row, column=2, value=persona[2] or '')
                ws_int.cell(row=row, column=3, value=persona[0])
                ws_int.cell(row=row, column=4, value=persona[3] or '')
                ws_int.cell(row=row, column=5, value=persona[4] or '')
                ws_int.cell(row=row, column=6, value=self.db.procedencia_grupo(persona[0], grupo))

            # Ajustar anchos
            for col in range(1, 7):
                ws_int.column_dimensions[get_column_letter(col)].width = 25
            
            # ===== HOJA 2: PUBLICACIONES =====
            publicaciones = [p for p in self.productos_completos if p.get('tipo_producto') == 'Publicación']
            if publicaciones:
                ws_pub = wb.create_sheet("Publicaciones")

                ws_pub['A1'] = f"GRUPO: {grupo} - PUBLICACIONES"
                ws_pub['A1'].font = Font(bold=True, size=14, color="1a365d")
                ws_pub.merge_cells('A1:I1')

                headers = ['Investigador', 'Título', 'Revista/Libro', 'Año', 'Tipo', 'Categoría',
                           'ISSN/ISBN', 'Estado', 'Datos adicionales']
                for col, h in enumerate(headers, 1):
                    cell = ws_pub.cell(row=2, column=col, value=h)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")

                for row, pub in enumerate(publicaciones, 3):
                    ws_pub.cell(row=row, column=1, value=pub.get('investigador', ''))
                    ws_pub.cell(row=row, column=2, value=pub.get('titulo', ''))
                    ws_pub.cell(row=row, column=3, value=pub.get('revista_libro', ''))
                    ws_pub.cell(row=row, column=4, value=pub.get('año', ''))
                    ws_pub.cell(row=row, column=5, value=pub.get('tipo', ''))
                    ws_pub.cell(row=row, column=6, value=pub.get('categoria', ''))
                    ws_pub.cell(row=row, column=7, value=pub.get('issn_isbn', ''))
                    ws_pub.cell(row=row, column=8, value=pub.get('estado', ''))
                    ws_pub.cell(row=row, column=9, value=self._texto_datos_adicionales(pub.get('datos_adicionales')))

                for col in range(1, 10):
                    ws_pub.column_dimensions[get_column_letter(col)].width = 20
                ws_pub.column_dimensions['B'].width = 50
                ws_pub.column_dimensions['I'].width = 40
            
            # ===== HOJA 3: EXTENSIONES =====
            extensiones = [p for p in self.productos_completos if p.get('tipo_producto') == 'Extensión']
            if extensiones:
                ws_ext = wb.create_sheet("Extensiones")

                ws_ext['A1'] = f"GRUPO: {grupo} - EXTENSIONES"
                ws_ext['A1'].font = Font(bold=True, size=14, color="1a365d")
                ws_ext.merge_cells('A1:H1')

                headers = ['Investigador', 'Actividad', 'Tipo', 'Modalidad', 'Año', 'Población',
                           'Estado', 'Datos adicionales']
                for col, h in enumerate(headers, 1):
                    cell = ws_ext.cell(row=2, column=col, value=h)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="e74c3c", end_color="e74c3c", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")

                for row, ext in enumerate(extensiones, 3):
                    ws_ext.cell(row=row, column=1, value=ext.get('investigador', ''))
                    ws_ext.cell(row=row, column=2, value=ext.get('titulo', ''))
                    ws_ext.cell(row=row, column=3, value=ext.get('tipo', ''))
                    ws_ext.cell(row=row, column=4, value=ext.get('modalidad', ''))
                    ws_ext.cell(row=row, column=5, value=ext.get('año', ''))
                    ws_ext.cell(row=row, column=6, value=ext.get('poblacion', ''))
                    ws_ext.cell(row=row, column=7, value=ext.get('estado', ''))
                    ws_ext.cell(row=row, column=8, value=self._texto_datos_adicionales(ext.get('datos_adicionales')))

                for col in range(1, 9):
                    ws_ext.column_dimensions[get_column_letter(col)].width = 20
                ws_ext.column_dimensions['B'].width = 50
                ws_ext.column_dimensions['H'].width = 40
            
            # ===== HOJA 4: TRABAJOS DE GRADO =====
            trabajos = [p for p in self.productos_completos if p.get('tipo_producto') == 'Trabajo de Grado']
            if trabajos:
                ws_tg = wb.create_sheet("Trabajos de Grado")
                
                ws_tg['A1'] = f"GRUPO: {grupo} - TRABAJOS DE GRADO"
                ws_tg['A1'].font = Font(bold=True, size=14, color="1a365d")
                ws_tg.merge_cells('A1:H1')

                headers = ['Director', 'Título', 'Estudiante', 'Programa', 'Año', 'Estado',
                           'Calificación', 'Datos adicionales']
                for col, h in enumerate(headers, 1):
                    cell = ws_tg.cell(row=2, column=col, value=h)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="9b59b6", end_color="9b59b6", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")

                for row, tg in enumerate(trabajos, 3):
                    ws_tg.cell(row=row, column=1, value=tg.get('investigador', ''))
                    ws_tg.cell(row=row, column=2, value=tg.get('titulo', ''))
                    ws_tg.cell(row=row, column=3, value=tg.get('estudiante', ''))
                    ws_tg.cell(row=row, column=4, value=tg.get('programa', ''))
                    ws_tg.cell(row=row, column=5, value=tg.get('año', ''))
                    ws_tg.cell(row=row, column=6, value=tg.get('estado', ''))
                    ws_tg.cell(row=row, column=7, value=tg.get('calificacion', ''))
                    ws_tg.cell(row=row, column=8, value=self._texto_datos_adicionales(tg.get('datos_adicionales')))

                for col in range(1, 9):
                    ws_tg.column_dimensions[get_column_letter(col)].width = 20
                ws_tg.column_dimensions['B'].width = 50
                ws_tg.column_dimensions['H'].width = 40

            # ===== HOJA 5: PROYECTOS =====
            proyectos = [p for p in self.productos_completos if p.get('tipo_producto') == 'Proyecto']
            if proyectos:
                ws_proy = wb.create_sheet("Proyectos")
                
                ws_proy['A1'] = f"GRUPO: {grupo} - PROYECTOS"
                ws_proy['A1'].font = Font(bold=True, size=14, color="1a365d")
                ws_proy.merge_cells('A1:G1')
                
                headers = ['Investigador', 'Título', 'Código CIE', 'Tipo', 'Año', 'Estado', 'Valor']
                for col, h in enumerate(headers, 1):
                    cell = ws_proy.cell(row=2, column=col, value=h)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="1abc9c", end_color="1abc9c", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")

                for row, proy in enumerate(proyectos, 3):
                    ws_proy.cell(row=row, column=1, value=proy.get('investigador', ''))
                    ws_proy.cell(row=row, column=2, value=proy.get('titulo', ''))
                    ws_proy.cell(row=row, column=3, value=proy.get('codigo_cie', ''))
                    ws_proy.cell(row=row, column=4, value=proy.get('tipo', ''))
                    ws_proy.cell(row=row, column=5, value=proy.get('año', ''))
                    ws_proy.cell(row=row, column=6, value=proy.get('estado', ''))
                    ws_proy.cell(row=row, column=7, value=proy.get('valor_aprobado', ''))

                for col in range(1, 8):
                    ws_proy.column_dimensions[get_column_letter(col)].width = 20
                ws_proy.column_dimensions['B'].width = 50
            
            # ===== HOJA 6: INNOVACIÓN =====
            innovacion = [p for p in self.productos_completos if p.get('tipo_producto') == 'Innovación']
            if innovacion:
                ws_inn = wb.create_sheet("Innovación")

                ws_inn['A1'] = f"GRUPO: {grupo} - PRODUCTOS DE INNOVACIÓN"
                ws_inn['A1'].font = Font(bold=True, size=14, color="1a365d")
                ws_inn.merge_cells('A1:G1')

                headers = ['Investigador', 'Nombre', 'Tipo', 'Año', 'Estado', 'Descripción', 'Datos adicionales']
                for col, h in enumerate(headers, 1):
                    cell = ws_inn.cell(row=2, column=col, value=h)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="f39c12", end_color="f39c12", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")

                for row, inn in enumerate(innovacion, 3):
                    ws_inn.cell(row=row, column=1, value=inn.get('investigador', ''))
                    ws_inn.cell(row=row, column=2, value=inn.get('titulo', ''))
                    ws_inn.cell(row=row, column=3, value=inn.get('tipo_producto_detalle', ''))
                    ws_inn.cell(row=row, column=4, value=inn.get('año', ''))
                    ws_inn.cell(row=row, column=5, value=inn.get('estado', ''))
                    ws_inn.cell(row=row, column=6, value=inn.get('descripcion', ''))
                    ws_inn.cell(row=row, column=7, value=self._texto_datos_adicionales(inn.get('datos_adicionales')))

                for col in range(1, 8):
                    ws_inn.column_dimensions[get_column_letter(col)].width = 20
                ws_inn.column_dimensions['B'].width = 40
                ws_inn.column_dimensions['F'].width = 60
                ws_inn.column_dimensions['G'].width = 40

            # ===== HOJA 7: PROPIEDAD INTELECTUAL =====
            propiedad = [p for p in self.productos_completos if p.get('tipo_producto') == 'Propiedad Intelectual']
            if propiedad:
                ws_pi = wb.create_sheet("Propiedad Intelectual")

                ws_pi['A1'] = f"GRUPO: {grupo} - PROPIEDAD INTELECTUAL"
                ws_pi['A1'].font = Font(bold=True, size=14, color="1a365d")
                ws_pi.merge_cells('A1:H1')

                headers = ['Responsable', 'Nombre del producto', 'Tipo de producto', 'Tipo de patente',
                           'No. de registro', 'Entidad', 'Facultad', 'Datos adicionales']
                for col, h in enumerate(headers, 1):
                    cell = ws_pi.cell(row=2, column=col, value=h)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="6a4c93", end_color="6a4c93", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")

                for row, pi in enumerate(propiedad, 3):
                    ws_pi.cell(row=row, column=1, value=pi.get('investigador', ''))
                    ws_pi.cell(row=row, column=2, value=pi.get('titulo', ''))
                    ws_pi.cell(row=row, column=3, value=pi.get('tipo_producto_detalle', ''))
                    ws_pi.cell(row=row, column=4, value=pi.get('tipo_patente', ''))
                    ws_pi.cell(row=row, column=5, value=pi.get('numero_registro', ''))
                    ws_pi.cell(row=row, column=6, value=pi.get('entidad', ''))
                    ws_pi.cell(row=row, column=7, value=pi.get('facultad', ''))
                    ws_pi.cell(row=row, column=8, value=self._texto_datos_adicionales(pi.get('datos_adicionales')))

                for col in range(1, 9):
                    ws_pi.column_dimensions[get_column_letter(col)].width = 20
                ws_pi.column_dimensions['B'].width = 40
                ws_pi.column_dimensions['H'].width = 40

            # Guardar
            wb.save(str(ruta))

            QMessageBox.information(self, "Éxito",
                f"Excel exportado exitosamente:\n{ruta}\n\n"
                f"Integrantes: {len(integrantes)}\n"
                f"Productos totales: {len(self.productos_completos)}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al exportar Excel:\n{str(e)}")
    
    def exportar_pdf(self):
        """Exporta un reporte PDF del grupo seleccionado"""
        
        # ============= PASO 1: OBTENER NOMBRE DEL GRUPO =============
        grupo = self.combo_grupos.currentText()
        
        # ============= PASO 2: VALIDACIONES =============
        if grupo == "-- Seleccionar --" or not grupo:
            QMessageBox.warning(self, "Advertencia", "Selecciona un grupo primero")
            return
        
        if not self.productos_completos:
            QMessageBox.warning(self, "Advertencia", "No hay productos para exportar")
            return
        
        try:
            # ============= PASO 3: IMPORTACIONES =============
            from datetime import datetime
            from PyQt5.QtPrintSupport import QPrinter
            from PyQt5.QtGui import QTextDocument
            
            # ============= PASO 4: CREAR TIMESTAMP =============
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

            # ============= PASO 5: NOMBRE SUGERIDO Y DIÁLOGO GUARDAR COMO =============
            nombre_sugerido = f"reporte_{limpiar_nombre_archivo(grupo)}_{timestamp}.pdf"
            reports_dir = obtener_directorio_base() / "reports" / "pdf"
            reports_dir.mkdir(parents=True, exist_ok=True)

            ruta_str, _ = QFileDialog.getSaveFileName(
                self, "Guardar reporte PDF del grupo",
                str(reports_dir / nombre_sugerido), "PDF (*.pdf)")
            if not ruta_str:
                return
            ruta = Path(ruta_str)
            if ruta.suffix.lower() != ".pdf":
                ruta = ruta.with_suffix(".pdf")

            # ============= PASO 7: CREAR DOCUMENTO HTML =============
            html = f"""
            <html>
            <head>
                <style>
                    body {{ font-family: Arial, sans-serif; margin: 20px; }}
                    h1 {{ color: #1a365d; border-bottom: 3px solid #1a365d; padding-bottom: 10px; }}
                    h2 {{ color: #2c3e50; margin-top: 30px; border-bottom: 2px solid #ecf0f1; padding-bottom: 5px; }}
                    table {{ width: 100%; border-collapse: collapse; margin-top: 10px; font-size: 9pt; }}
                    th {{ background-color: #1a365d; color: white; padding: 8px; text-align: left; }}
                    td {{ border: 1px solid #ddd; padding: 6px; }}
                    tr:nth-child(even) {{ background-color: #f2f2f2; }}
                    .producto {{ margin-bottom: 15px; padding: 10px; border: 1px solid #ddd; border-radius: 5px; page-break-inside: avoid; }}
                </style>
            </head>
            <body>
                <h1>Reporte del Grupo: {grupo}</h1>
                <p><strong>Fecha:</strong> {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}</p>
            """
            
            # ============= PASO 8: INTEGRANTES =============
            integrantes = self.db.obtener_integrantes_grupo(grupo)
            html += f"""
                <h2>Integrantes ({len(integrantes)})</h2>
                <table>
                    <tr><th>Nombre</th><th>Tipo</th><th>Cédula</th><th>Email</th><th>Procedencia</th></tr>
            """

            for persona in integrantes:
                procedencia = self.db.procedencia_grupo(persona[0], grupo)
                html += f"""
                    <tr>
                        <td>{persona[1]}</td>
                        <td>{persona[2] or ''}</td>
                        <td>{persona[0]}</td>
                        <td>{persona[3] or ''}</td>
                        <td>{procedencia}</td>
                    </tr>
                """

            html += "</table>"

            # ============= PASO 9: PRODUCTOS POR TIPO =============
            tipos_plural = {
                'Publicación': 'Publicaciones',
                'Extensión': 'Extensiones',
                'Trabajo de Grado': 'Trabajos de Grado',
                'Innovación': 'Innovación',
                'Proyecto': 'Proyectos',
                'Propiedad Intelectual': 'Propiedad Intelectual',
            }
            for tipo, plural in tipos_plural.items():
                productos_tipo = [p for p in self.productos_completos if p.get('tipo_producto') == tipo]

                if not productos_tipo:
                    continue

                html += f'<h2>{plural} ({len(productos_tipo)})</h2>'

                for i, prod in enumerate(productos_tipo, 1):
                    html += f'<div class="producto">'
                    html += f'<strong>#{i}</strong> - {prod.get("titulo", "Sin título")}<br>'
                    if tipo == 'Trabajo de Grado':
                        html += f'<em>Director:</em> {prod.get("investigador", "N/A")}'
                        html += f' | <em>Estudiante:</em> {prod.get("estudiante", "N/A")}'
                    else:
                        html += f'<em>Investigador:</em> {prod.get("investigador", "N/A")}'
                    if prod.get('año'):
                        html += f' | <em>Año:</em> {prod["año"]}'
                    extra = self._texto_datos_adicionales(prod.get('datos_adicionales'))
                    if extra:
                        html += f'<br><span style="color:#888;font-size:9pt;">+ {extra}</span>'
                    html += '</div>'

            html += "</body></html>"
            
            # ============= PASO 10: CREAR PDF =============
            documento = QTextDocument()
            documento.setHtml(html)
            
            printer = QPrinter(QPrinter.HighResolution)
            printer.setOutputFormat(QPrinter.PdfFormat)
            printer.setOutputFileName(str(ruta))
            printer.setPageSize(QPrinter.A4)
            printer.setPageMargins(10, 10, 10, 10, QPrinter.Millimeter)

            documento.print_(printer)

            # ============= PASO 11: MENSAJE DE ÉXITO =============
            QMessageBox.information(self, "Éxito", f"PDF exportado:\n{ruta}")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al exportar PDF:\n{str(e)}")


class VentanaPrincipal(QMainWindow):
    def __init__(self):
        super().__init__()
        self.db = DatabaseManager()
        self._tab_registry = {}
        self._loaded_tabs = set()
        self.vista_inicio = None
        self.vista_busqueda = None
        self.vista_grupos = None
        self.setup_ui()
        # Diferido al primer tick del event loop: actualizar widgets antes de
        # que app.exec_() haya arrancado se cuelga de forma reproducible bajo
        # PyQt5 (ver cargar_datos_automaticamente), aunque la llamada sea
        # síncrona y en el hilo principal.
        QTimer.singleShot(0, self.cargar_datos_automaticamente)

    def _crear_tab(self, index):
        if index in self._loaded_tabs:
            return
        self._loaded_tabs.add(index)
        factory, args = self._tab_registry.get(index)
        if factory is None:
            return
        widget = factory(*args)
        self.tabs.blockSignals(True)
        self.tabs.removeTab(index)
        self.tabs.insertTab(index, widget, self._tab_titles[index])
        self.tabs.setCurrentIndex(index)
        self.tabs.blockSignals(False)
        if index == 0:
            self.vista_inicio = widget
            self.vista_inicio.procesar_solicitado.connect(self.cargar_datos_automaticamente)
        elif index == 2:
            self.vista_grupos = widget
        elif index == 1:
            self.vista_busqueda = widget

    _NOMBRE_A_INDICE_PESTANA = {
        "inicio": 0,
        "busqueda_personas": 1,
        "reportes_grupo": 2,
        "seguimiento_grupos": 3,
    }

    def _cambiar_pestana_por_nombre(self, pestana, grupo=None):
        """Callback inyectado en VistaChatbotInvestigacion para que el
        chatbot pueda cambiar de pestaña y, si aplica, preseleccionar un
        grupo en la pestaña destino."""
        idx = self._NOMBRE_A_INDICE_PESTANA.get(pestana)
        if idx is None:
            return
        self.tabs.setCurrentIndex(idx)
        if not grupo:
            return
        widget = self.tabs.widget(idx)
        if hasattr(widget, 'combo_grupos'):
            widget.combo_grupos.setCurrentText(grupo)
        elif hasattr(widget, 'combo_grupo'):
            widget.combo_grupo.setCurrentText(grupo)

    def setup_ui(self):
        self.setWindowTitle("Consolidado de Información")
        self.setGeometry(100, 100, 1400, 800)
        
        widget_central = QWidget()
        self.setCentralWidget(widget_central)
        
        layout = QVBoxLayout(widget_central)
        
        header = QLabel("Consolidado de Información")
        header.setFont(QFont("Arial", 16, QFont.Bold))
        header.setStyleSheet("background-color: #1a365d; color: white; padding: 15px;")
        
        self.tabs = QTabWidget()
        self._tab_titles = [
            "Inicio",
            "Búsqueda de Personas",
            "Reportes por Grupo",
            "Seguimiento Grupos",
            # "Asistente",  # pestaña del chatbot retirada por ahora -- ver
            # chatbot_investigacion.py (código intacto, solo sin punto de
            # entrada). Para reactivarla: descomentar este título y la
            # entrada 5 de _tab_registry más abajo, y el import de
            # VistaChatbotInvestigacion al inicio del archivo.
            # "Estadísticas 957" -- movida fuera de la app principal a
            # petición del usuario: se está reconstruyendo como UI separada
            # en UI_clasificacion/, para integrar más adelante. Ver
            # estadisticas_957.py (código intacto, sin punto de entrada acá).
        ]
        self._tab_registry = {
            0: (lambda db: VistaInicio(db), (self.db,)),
            1: (lambda db: VistaBusqueda(db), (self.db,)),
            2: (lambda db: VistaGrupos(db), (self.db,)),
            3: (lambda db: VistaSeguimientoGrupos(db), (self.db,)),
            # 4: (lambda db, cb: VistaChatbotInvestigacion(db, cambiar_pestana_callback=cb),
            #     (self.db, self._cambiar_pestana_por_nombre)),
            # 4: (lambda db: VistaEstadisticas957(db), (self.db,)),
        }

        for i, title in enumerate(self._tab_titles):
            self.tabs.addTab(QWidget(), title)

        self.tabs.currentChanged.connect(self._crear_tab)
        self._crear_tab(0)

        self.statusBar().showMessage("Listo")
        
        layout.addWidget(header)
        layout.addWidget(self.tabs)
    
    def cargar_datos_automaticamente(self):
        # Corre en el hilo principal (bloqueando la UI unos segundos) en vez
        # de en un QThread: bajo PyQt5, lanzar esta carga con
        # CargadorDatosIntegrado.start() se cuelga de forma reproducible en
        # esta pila (visto en Python 3.11 y 3.12, con y sin cambios propios).
        # Llamarla directamente evita ese cuelgue; el costo es una espera de
        # ~15s con cursor de reloj de arena, aceptable para una operación que
        # el usuario dispara explícitamente al actualizar datos.
        #
        # IMPORTANTE: esta llamada debe ocurrir DESPUÉS de que arrancó el
        # event loop de Qt (app.exec_()) -- ver VentanaPrincipal.__init__,
        # que la difiere con QTimer.singleShot(0, ...). Actualizar widgets
        # (statusBar, QLabel.setText) antes de que el event loop haya
        # corrido su primera iteración también se cuelga de forma
        # reproducible, aunque la llamada sea síncrona y en el hilo
        # principal.
        if getattr(self, "_cargando_datos", False):
            return
        self._cargando_datos = True
        self.statusBar().showMessage("Cargando datos...")
        if self.vista_inicio is not None:
            self.vista_inicio.marcar_procesando("Procesando…")
            self.vista_inicio.btn_procesar.setEnabled(False)
        self.cargador = CargadorDatosIntegrado(self.db)
        self.cargador.progreso.connect(self.actualizar_status)
        self.cargador.finalizado.connect(self.carga_finalizada)
        self.cargador.duplicados_consolidados.connect(self.mostrar_duplicados_consolidados)

        QApplication.setOverrideCursor(Qt.WaitCursor)
        try:
            self.cargador.run()
        finally:
            QApplication.restoreOverrideCursor()
            self._cargando_datos = False

    def actualizar_status(self, mensaje):
        self.statusBar().showMessage(mensaje)
        if self.vista_inicio is not None:
            self.vista_inicio.marcar_procesando(mensaje)

    def mostrar_duplicados_consolidados(self, consolidaciones):
        if not consolidaciones:
            return

        mensaje = f"Se consolidaron {len(consolidaciones)} persona(s) con cédulas duplicadas:\n\n"

        for i, grupo in enumerate(consolidaciones, 1):
            nombre = grupo['nombre']
            cedula_principal = grupo['cedula_principal']
            todas_cedulas = grupo['cedulas']

            mensaje += f"{i}. {nombre}\n"
            mensaje += f"   Cédula principal: {cedula_principal}\n"
            mensaje += f"   Todas las cédulas: {', '.join(todas_cedulas)}\n\n"

        dialogo = QDialog(self)
        dialogo.setWindowTitle("Duplicados Consolidados")
        dialogo.resize(650, 500)
        layout = QVBoxLayout(dialogo)

        titulo = QLabel(f"Consolidación Completada — {len(consolidaciones)} persona(s)")
        titulo.setFont(QFont("Arial", 11, QFont.Bold))
        layout.addWidget(titulo)

        texto = QTextEdit()
        texto.setReadOnly(True)
        texto.setPlainText(mensaje)
        layout.addWidget(texto)

        btn_cerrar = QPushButton("Cerrar")
        btn_cerrar.clicked.connect(dialogo.accept)
        layout.addWidget(btn_cerrar)

        dialogo.exec_()
    
    def carga_finalizada(self, stats):
        self.statusBar().showMessage("Datos actualizados correctamente.")
        if self.vista_inicio is not None:
            self.vista_inicio.procesamiento_finalizado(stats)
            self.vista_inicio.btn_procesar.setEnabled(True)
        self._crear_tab(2)
        if self.vista_grupos is not None:
            self.vista_grupos.cargar_grupos()

    def _generar_reporte_carga(self, stats):
        """
        Genera un Excel de resumen de productos por grupo tras la carga de datos,
        para que cada grupo sepa qué subir a GrupLAC.
        El archivo queda en el directorio base con el timestamp de la carga.
        """
        try:
            import openpyxl
            from datetime import datetime
            from openpyxl.styles import Alignment, Font, PatternFill

            base_dir = obtener_directorio_base()
            timestamp = datetime.now().strftime("%Y%m%d_%H%M")
            nombre = str(base_dir / f"reporte_pendientes_{timestamp}.xlsx")

            cur = self.db.conn.cursor()
            grupos = [
                r[0] for r in cur.execute(
                    "SELECT DISTINCT grupo FROM grupos "
                    "WHERE grupo IS NOT NULL AND grupo != '' "
                    "ORDER BY grupo"
                ).fetchall()
            ]

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Resumen por Grupo"

            fill_h = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
            centrado = Alignment(horizontal="center")
            headers = ["Grupo", "Publicaciones", "Extensiones",
                       "Trabajos de Grado", "Innovación", "Proyectos"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = fill_h
                cell.alignment = centrado

            tablas = ["publicaciones", "extensiones", "trabajos_grado",
                      "productos_innovacion", "proyectos"]
            for fila, grupo in enumerate(grupos, start=2):
                ws.cell(row=fila, column=1, value=grupo)
                for col_idx, tabla in enumerate(tablas, start=2):
                    try:
                        count = cur.execute(
                            f"SELECT COUNT(*) FROM {tabla} WHERE grupo = ?",
                            (grupo,),
                        ).fetchone()[0]
                    except Exception:
                        count = 0
                    cell = ws.cell(row=fila, column=col_idx, value=count)
                    cell.alignment = centrado

            ws.column_dimensions["A"].width = 55
            for letra in "BCDEF":
                ws.column_dimensions[letra].width = 16

            wb.save(nombre)
            self.statusBar().showMessage(f"Reporte de pendientes generado: {nombre}")
        except Exception as e:
            self.statusBar().showMessage(f"No se generó reporte de pendientes: {e}")

    def closeEvent(self, event):
        try:
            self.db.close()
        except Exception:
            pass
        event.accept()


def main():
    app = QApplication(sys.argv)
    ventana = VentanaPrincipal()
    ventana.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
"""
Vista de Seguimiento de Grupos.
Compara datos internos (BD) vs GrupLAC y detecta productos faltantes.
"""

import math
import re
import sys
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date
from pathlib import Path

import pandas as pd
from PyQt5.QtCore import Qt, QAbstractTableModel, QModelIndex, QThread, pyqtSignal
from PyQt5.QtGui import QBrush, QColor, QFont
from PyQt5.QtWidgets import (
    QAbstractItemView, QApplication, QComboBox, QDialog,
    QFileDialog, QHBoxLayout, QHeaderView, QLabel, QLineEdit,
    QMessageBox, QProgressBar, QPushButton, QScrollArea, QSpinBox, QSplitter,
    QTabWidget, QTableView, QTableWidget, QTableWidgetItem,
    QTextEdit, QTreeWidget, QTreeWidgetItem, QVBoxLayout, QWidget,
)
from unidecode import unidecode

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from constants import CATEGORIAS_PRINCIPALES, CLASIF_957, COLORES_CATEGORIA_HEX, VARIANTES_NOMBRES, VENTANAS_957, INDICADORES_957, ORDEN_CATEGORIAS_MINCIENCIAS
from utils import (
    clasificar_producto_957,
    get_orden_categoria,
    get_orden_producto,
    limpiar_nombre_archivo,
    normalizar_nombre_hoja,
    obtener_directorio_base,
    tokenizar_autores,
)

# Alias para compatibilidad con el código interno que usa _tokenizar_autores
_tokenizar_autores = tokenizar_autores



class _DFModel(QAbstractTableModel):
    """Modelo Qt para mostrar DataFrame sin truncar."""
    def __init__(self, df):
        super().__init__()
        self._df = df.reset_index(drop=True)

    def set_df(self, df):
        self.beginResetModel()
        self._df = df.reset_index(drop=True)
        self.endResetModel()

    def rowCount(self, p=QModelIndex()):
        return 0 if p.isValid() else len(self._df)

    def columnCount(self, p=QModelIndex()):
        return 0 if p.isValid() else len(self._df.columns)

    def data(self, idx, role=Qt.DisplayRole):
        if not idx.isValid():
            return None
        val = self._df.iat[idx.row(), idx.column()]
        text = '' if (val is None or (isinstance(val, float) and pd.isna(val))) else str(val)
        if role in (Qt.DisplayRole, Qt.ToolTipRole):
            return text
        return None

    def headerData(self, sec, ori, role=Qt.DisplayRole):
        if role != Qt.DisplayRole:
            return None
        return str(self._df.columns[sec]) if ori == Qt.Horizontal else str(sec + 1)


class ExcelDialog(QDialog):
    """
    Diálogo emergente — muestra todas las hojas de un Excel
    con texto completo (sin truncar) y búsqueda en hoja activa.
    """

    def __init__(self, excel_path, group_name, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"Excel Original  |  {group_name}")
        self.resize(1200, 760)
        self.setModal(True)

        lay = QVBoxLayout(self)
        lay.setContentsMargins(10, 10, 10, 8)
        lay.setSpacing(6)

        hdr = QLabel(f"Datos originales GrupLAC:  {group_name}")
        hdr.setStyleSheet(
            "font-weight:700;font-size:13px;color:white;"
            "padding:8px 14px;background:#003087;border-radius:5px;")
        lay.addWidget(hdr)

        info = QLabel(f"Archivo: {excel_path}")
        info.setStyleSheet("color:#1565C0;font-size:10px;padding:2px;")
        info.setWordWrap(True)
        lay.addWidget(info)

        sb = QHBoxLayout()
        sb.addWidget(QLabel("Buscar en hoja activa:"))
        self._srch = QLineEdit()
        self._srch.setPlaceholderText("Escriba para filtrar filas...")
        self._srch.textChanged.connect(self._on_search)
        sb.addWidget(self._srch, 1)
        self._cnt = QLabel("")
        self._cnt.setStyleSheet("color:#555;font-size:11px;min-width:80px;")
        sb.addWidget(self._cnt)
        lay.addLayout(sb)

        self._tabs = QTabWidget()
        lay.addWidget(self._tabs, 1)

        btn_close = QPushButton("Cerrar")
        btn_close.setFixedWidth(100)
        btn_close.clicked.connect(self.accept)
        bl = QHBoxLayout()
        bl.addStretch()
        bl.addWidget(btn_close)
        lay.addLayout(bl)

        self._dfs: dict = {}
        self._models: dict = {}

        self._load(excel_path)
        self._tabs.currentChanged.connect(self._on_tab)

    def _load(self, path):
        if not HAS_OPENPYXL:
            self._tabs.addTab(QLabel("Instale openpyxl: pip install openpyxl"), "Error")
            return
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as e:
            self._tabs.addTab(QLabel(f"Error al abrir: {e}"), "Error")
            return

        for sname in wb.sheetnames:
            ws = wb[sname]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            hdr_row, hdr_idx = None, 0
            for i, row in enumerate(rows[:10]):
                if any(c is not None and str(c).strip() for c in row):
                    hdr_row = row
                    hdr_idx = i + 1
                    break

            if hdr_row is None:
                hdr_row = [None]
                hdr_idx = 0

            cols, seen = [], {}
            for j, c in enumerate(hdr_row):
                name = str(c).strip() if (c is not None and str(c).strip()) else f"Col_{j+1}"
                base = name
                k = seen.get(base, 0)
                if k:
                    name = f"{base}.{k}"
                seen[base] = k + 1
                cols.append(name)

            data = []
            for row in rows[hdr_idx:]:
                if not any(c is not None for c in row):
                    continue
                r = list(row)
                r = r + [None] * max(0, len(cols) - len(r))
                r = r[:len(cols)]
                data.append(['' if c is None else str(c) for c in r])

            if not data:
                continue

            df = pd.DataFrame(data, columns=cols)
            self._dfs[sname] = df

            w = QWidget()
            wl = QVBoxLayout(w)
            wl.setContentsMargins(4, 4, 4, 4)
            wl.setSpacing(4)

            meta = QLabel(f"{len(df)} filas  |  {len(df.columns)} columnas")
            meta.setStyleSheet("color:#555;font-size:10px;")
            wl.addWidget(meta)

            tv = QTableView()
            tv.setWordWrap(True)
            tv.setTextElideMode(Qt.ElideNone)
            tv.setAlternatingRowColors(True)
            tv.setSelectionBehavior(QAbstractItemView.SelectRows)
            tv.setSortingEnabled(True)
            tv.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
            tv.horizontalHeader().setStretchLastSection(True)
            tv.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)

            model = _DFModel(df)
            tv.setModel(model)
            tv.resizeColumnsToContents()
            wl.addWidget(tv)

            idx = self._tabs.addTab(w, sname)
            self._models[idx] = model

        wb.close()
        if self._tabs.count() == 0:
            self._tabs.addTab(QLabel("No se encontraron hojas con datos."), "-")
        else:
            self._update_cnt()

    def _on_tab(self, _):
        self._srch.clear()
        self._update_cnt()

    def _on_search(self, txt):
        idx = self._tabs.currentIndex()
        model = self._models.get(idx)
        sname = self._tabs.tabText(idx)
        df = self._dfs.get(sname, pd.DataFrame())
        if model is None:
            return
        if txt.strip() and not df.empty:
            mask = df.apply(
                lambda c: c.astype(str).str.contains(txt, case=False, na=False)
            ).any(axis=1)
            model.set_df(df[mask])
            self._cnt.setText(f"{mask.sum()} filas")
        else:
            model.set_df(df)
            self._update_cnt()

    def _update_cnt(self):
        idx = self._tabs.currentIndex()
        sname = self._tabs.tabText(idx)
        df = self._dfs.get(sname, pd.DataFrame())
        self._cnt.setText(f"{len(df)} filas")


# =============================================================================
# CACHE DE VERIFICACION DE FALTANTES (data/cache/verificacion_faltantes.json)
# =============================================================================
_CACHE_FALTANTES_FILENAME = "verificacion_faltantes.json"

_ETIQUETA_CATEGORIA = {
    "publicaciones": "Publicaciones",
    "extensiones": "Extensiones",
    "trabajos_grado": "Trabajos de Grado",
    "proyectos": "Proyectos",
}

# Filtro de estado del panel "Cumplimiento" -> lista de valores de
# estado_verificacion que incluye (None = todos, sin filtrar).
_FILTROS_ESTADO_CUMPLIMIENTO = {
    "pendientes": ["Faltante real", "Registrado en otro grupo"],
    "falta": ["Faltante real"],
    "otro_grupo": ["Registrado en otro grupo"],
    "confirmado": ["Confirmado en BD (mismo grupo)"],
    "segundo_barrido": ["Segundo barrido - mismo grupo", "Segundo barrido - otro grupo"],
    "todos": None,
}

_ETIQUETAS_ESTADO_CUMPLIMIENTO = {
    "Faltante real": "Falta",
    "Registrado en otro grupo": "Falta (en otro grupo)",
    "Confirmado en BD (mismo grupo)": "Confirmado",
    "Segundo barrido - mismo grupo": "Segundo barrido (mismo grupo)",
    "Segundo barrido - otro grupo": "Segundo barrido (otro grupo)",
}


def _cache_faltantes_path():
    return obtener_directorio_base() / "data" / "cache" / _CACHE_FALTANTES_FILENAME


def _aplanar_cache_faltantes(cache):
    """
    Convierte el JSON del caché en una lista plana de registros con
    'grupo_original'. Soporta el formato plano antiguo ({"data": [...]})
    y el formato anidado actual ({"grupos": [{"nombre", "productos": [...]}]}).
    """
    if "data" in cache:
        return cache.get("data") or []
    registros = []
    for grupo in cache.get("grupos", []):
        nombre_grupo = grupo.get("nombre") or grupo.get("nombre_bd", "")
        for producto in grupo.get("productos", []):
            registro = dict(producto)
            registro["grupo_original"] = nombre_grupo
            registros.append(registro)
    return registros


def cargar_df_faltantes():
    """
    Lee data/cache/verificacion_faltantes.json y devuelve (df, mensaje, timestamp).
    df es None si el caché no existe, está vacío o hay un error de lectura;
    en ese caso 'mensaje' explica por qué.
    """
    import json as _json

    cache_file = _cache_faltantes_path()
    if not cache_file.exists():
        return None, (
            "No se encontró data/cache/verificacion_faltantes.json. "
            "Ejecute el script de verificación para generarlo."
        ), ""
    try:
        with open(cache_file, encoding="utf-8") as f:
            cache = _json.load(f)
        registros = _aplanar_cache_faltantes(cache)
        if not registros:
            return None, "El caché está vacío.", ""
        df = pd.DataFrame(registros)
        ts_raw = cache.get("generado") or cache.get("timestamp", "")
        ts = str(ts_raw)[:16].replace("T", " ")
        return df, "", ts
    except Exception as e:
        return None, f"Error leyendo caché: {e}", ""


def _carpeta_gruplac_mas_reciente():
    """Carpeta data/reporte excel_<fecha> más reciente (ver gruplac_scraper.py).
    El nombre trae la fecha en formato YYYYMMDD, así que ordenar por texto
    ya da el orden cronológico correcto."""
    base = obtener_directorio_base() / "data"
    candidatos = sorted(
        (p for p in base.glob("reporte excel_*") if p.is_dir()), reverse=True)
    return candidatos[0] if candidatos else None


# =============================================================================
# HILO: VERIFICACIÓN CONTRA GRUPLAC SCRAPEADO (BD interna vs data/reporte excel_*)
# =============================================================================
class VerificacionGrupLACThread(QThread):
    progreso = pyqtSignal(int, int)
    finalizado = pyqtSignal(dict)
    error = pyqtSignal(str)

    def __init__(self, db, carpeta_gruplac, anio_desde, anio_hasta):
        super().__init__()
        self.db = db
        self.carpeta_gruplac = carpeta_gruplac
        self.anio_desde = anio_desde
        self.anio_hasta = anio_hasta

    def run(self):
        try:
            from comparador_gruplac_scrapeado import ejecutar_y_guardar

            def _cb(i, total):
                self.progreso.emit(i, total)

            resumen = ejecutar_y_guardar(
                self.db, self.carpeta_gruplac, self.anio_desde, self.anio_hasta,
                progreso_callback=_cb)
            self.finalizado.emit(resumen)
        except Exception as e:
            self.error.emit(str(e))


# =============================================================================
# DIALOGO DE DUPLICADOS (Autores y Productos)
# =============================================================================
# DIÁLOGO: DUPLICADOS Y MIEMBROS ACTIVOS/RETIRADOS
# =============================================================================
class DialogoDuplicados(QDialog):
    """Personas en GrupLAC sin registro interno (dirección inversa a Búsqueda
    de Personas). La verificación de productos faltantes/confirmados vive en
    el panel 'Cumplimiento' de Seguimiento Grupos, no en este diálogo."""

    def __init__(self, resultado, db=None, parent=None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("Personas en GrupLAC sin registro interno")
        self.resize(1100, 700)

        layout = QVBoxLayout(self)

        lbl_pers = QLabel(
            "<b>Personas en GrupLAC sin registro interno</b><br>"
            "<small>Dirección inversa a Búsqueda de Personas: integrantes ACTIVOS "
            "(vinculación 'Actual') que GrupLAC sí tiene para el grupo, pero que no "
            "calzan con nadie de la BD interna para ese mismo grupo. Solo reporta — "
            "no agrega nada automáticamente.</small>"
        )
        lbl_pers.setWordWrap(True)
        lbl_pers.setStyleSheet("font-size:11px; color:#1a365d;")
        layout.addWidget(lbl_pers)

        toolbar_pers = QHBoxLayout()
        self.btn_buscar_personas = QPushButton("🔎 Buscar")
        self.btn_buscar_personas.setToolTip(
            "Compara la carpeta 'reporte excel_<fecha>' más reciente contra la BD interna.")
        self.btn_buscar_personas.setStyleSheet(
            "QPushButton{background-color:#2c3e50;color:white;padding:6px 16px;"
            "border-radius:4px;font-weight:bold;}"
            "QPushButton:hover{background-color:#1a252f;}")
        self.btn_buscar_personas.clicked.connect(self._buscar_personas_sin_registro)
        toolbar_pers.addWidget(self.btn_buscar_personas)
        toolbar_pers.addWidget(QLabel("Grupo:"))
        self.combo_pers_grupo = QComboBox()
        self.combo_pers_grupo.addItem("Todos los grupos", None)
        self.combo_pers_grupo.currentIndexChanged.connect(self._filtrar_personas_sin_registro)
        toolbar_pers.addWidget(self.combo_pers_grupo)
        toolbar_pers.addStretch()
        layout.addLayout(toolbar_pers)

        self.tbl_personas = QTableWidget()
        self.tbl_personas.setColumnCount(3)
        self.tbl_personas.setHorizontalHeaderLabels(["Grupo", "Nombre (GrupLAC)", "Vinculación"])
        self.tbl_personas.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.tbl_personas.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.tbl_personas.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl_personas.setSelectionBehavior(QTableWidget.SelectRows)
        self.tbl_personas.setStyleSheet("font-size:10px;")
        self.tbl_personas.setAlternatingRowColors(True)
        layout.addWidget(self.tbl_personas)

        self.lbl_pers_resumen = QLabel("Pulse Buscar para comparar.")
        self.lbl_pers_resumen.setStyleSheet("font-size:10px; color:#555; padding:4px;")
        layout.addWidget(self.lbl_pers_resumen)

        btn = QPushButton("Cerrar")
        btn.clicked.connect(self.accept)
        layout.addWidget(btn)

    def _buscar_personas_sin_registro(self):
        if not self.db:
            QMessageBox.warning(self, "Sin BD", "No hay conexión a la base de datos interna.")
            return
        carpeta = _carpeta_gruplac_mas_reciente()
        if not carpeta:
            QMessageBox.warning(
                self, "No hay datos GrupLAC",
                "No se encontró ninguna carpeta 'data/reporte excel_<fecha>'. "
                "Primero corra 'Actualizar GrupLAC (Web)' en Seguimiento Grupos.")
            return

        from analisis_seguimiento import AnalisisDuplicados

        self.btn_buscar_personas.setEnabled(False)
        self.lbl_pers_resumen.setText("Buscando...")
        QApplication.setOverrideCursor(Qt.WaitCursor)
        try:
            analisis = AnalisisDuplicados(self.db.conn)
            self._df_personas = analisis.personas_gruplac_sin_registro_interno(str(carpeta))
        except Exception as e:
            QApplication.restoreOverrideCursor()
            self.btn_buscar_personas.setEnabled(True)
            QMessageBox.critical(self, "Error", f"Error buscando personas:\n{e}")
            return
        QApplication.restoreOverrideCursor()
        self.btn_buscar_personas.setEnabled(True)

        self.combo_pers_grupo.blockSignals(True)
        self.combo_pers_grupo.clear()
        self.combo_pers_grupo.addItem("Todos los grupos", None)
        for g in sorted(self._df_personas["grupo"].dropna().unique()):
            if g:
                self.combo_pers_grupo.addItem(g, g)
        self.combo_pers_grupo.blockSignals(False)

        self.lbl_pers_resumen.setText(
            f"Carpeta: {carpeta.name} — {len(self._df_personas)} personas activas en GrupLAC "
            f"sin registro interno, en {self._df_personas['grupo'].nunique()} grupos."
        )
        self.lbl_pers_resumen.setStyleSheet(
            "font-size:10px; color:#1a7a4a; padding:4px; font-weight:bold;")
        self._poblar_tabla_personas(self._df_personas)

    def _poblar_tabla_personas(self, df):
        self.tbl_personas.setRowCount(len(df))
        for i, (_, row) in enumerate(df.iterrows()):
            self.tbl_personas.setItem(i, 0, QTableWidgetItem(str(row.get("grupo", ""))))
            self.tbl_personas.setItem(i, 1, QTableWidgetItem(str(row.get("nombre_gruplac", ""))))
            self.tbl_personas.setItem(i, 2, QTableWidgetItem(str(row.get("vinculacion", ""))))

    def _filtrar_personas_sin_registro(self):
        if not hasattr(self, "_df_personas") or self._df_personas is None:
            return
        grupo_sel = self.combo_pers_grupo.currentData()
        df = self._df_personas
        if grupo_sel:
            df = df[df["grupo"] == grupo_sel]
        self._poblar_tabla_personas(df)
# =============================================================================
# THREAD DE PROCESAMIENTO (OPTIMIZADO + RANGO DE AÑOS)
# =============================================================================
class ProcesamientoSeguimientoThread(QThread):
    progreso = pyqtSignal(int, str)
    finalizado = pyqtSignal(dict)
    error = pyqtSignal(str)

    def __init__(self, db, reporte_excel_path, anio_desde=2022, anio_hasta=2025):
        super().__init__()
        self.db = db
        self.reporte_excel_path = Path(reporte_excel_path)
        self.anio_desde = anio_desde
        self.anio_hasta = anio_hasta
        self.anios_filtro = list(range(anio_desde, anio_hasta + 1))
        self._cache_norm = {}
        self._cache_tok = {}

    def _normalizar_nombre(self, nombre):
        key = nombre if isinstance(nombre, str) else str(nombre)
        if key in self._cache_norm:
            return self._cache_norm[key]
        n = ' '.join(str(nombre).split()).strip().lower()
        n = n.replace(':', '').replace('-', ' ').replace(',', '')
        n = unidecode(n)
        res = re.sub(r'\s+', ' ', n).strip()
        self._cache_norm[key] = res
        return res

    def _extraer_anio(self, texto):
        if not texto or pd.isna(texto):
            return None
        anios = re.findall(r'\b(19[89][0-9]|20[0-3][0-9])\b', str(texto))
        validos = [int(a) for a in anios if 1990 <= int(a) <= 2035]
        if not validos:
            return None
        en_filtro = [a for a in validos if a in self.anios_filtro]
        return max(en_filtro) if en_filtro else max(validos)

    def _buscar_carpeta(self, grupo_bd, carpetas):
        gn = self._normalizar_nombre(grupo_bd)
        if not hasattr(self, "_cache_carpeta_norm") or self._cache_carpeta_norm is None:
            self._cache_carpeta_norm = {c: self._normalizar_nombre(c.name) for c in carpetas}
        norm_map = self._cache_carpeta_norm
        for c in carpetas:
            if c.name == grupo_bd:
                return c
        for c in carpetas:
            if norm_map[c] == gn:
                return c
        for longitud in [80, 70, 60, 50, 40]:
            pfx = gn[:longitud].rstrip()
            for c in carpetas:
                if norm_map[c].startswith(pfx):
                    return c
        stop = {'de', 'del', 'la', 'el', 'en', 'y', 'para', 'los', 'las', 'grupo', 'investigacion'}
        palabras = set(gn.split()) - stop
        mejor, max_ok = None, 0
        for c in carpetas:
            pw = set(norm_map[c].split())
            ok = len(palabras & pw)
            if ok > max_ok and ok >= len(palabras) * 0.7:
                max_ok = ok
                mejor = c
        return mejor if mejor and max_ok >= 3 else None

    def _extraer_productos_hoja(self, df, nombre_hoja, nombre_grupo, archivo_nombre):
        productos = []
        if df.empty:
            return productos

        clasificacion = clasificar_producto_957(nombre_hoja)
        if clasificacion is False:
            return productos

        if clasificacion:
            categoria, subcategoria, nombre_oficial = clasificacion
        else:
            categoria, subcategoria, nombre_oficial = 'Sin clasificar', 'Pendiente', nombre_hoja

        columnas = list(df.columns)
        col_titulo = col_anio = col_autores = None
        for idx, col in enumerate(columnas):
            cn = unidecode(str(col).lower())
            if col_titulo is None and any(p in cn for p in ['titulo', 'nombre', 'title']):
                col_titulo = idx
            if col_anio is None and any(p in cn for p in ['ano', 'year', 'fecha']):
                col_anio = idx
            if col_autores is None and any(p in cn for p in ['autor', 'author', 'integrante', 'investigador']):
                col_autores = idx

        # ── Extracción vectorial por columna ─────────────────────────
        n = len(df)
        # Valores por columna de interés
        raw_titulo = df.iloc[:, col_titulo] if col_titulo is not None else pd.Series([None] * n)
        raw_anio   = df.iloc[:, col_anio]   if col_anio   is not None else pd.Series([None] * n)
        raw_autor  = df.iloc[:, col_autores] if col_autores is not None else pd.Series([None] * n)

        # Texto combinado por fila para fallbacks
        textos = [
            df.iloc[:, ci].astype(str).str.strip()
            for ci in range(len(columnas))
        ]

        # Detectar columnas extra (DOI, ISSN, revista, etc.)
        extra_cols = []
        for ci, cn in enumerate(columnas):
            cn_lower = unidecode(cn.lower())
            if any(p in cn_lower for p in ['doi', 'issn', 'isbn', 'revista', 'journal',
                                            'editorial', 'pais', 'country', 'url',
                                            'volumen', 'pagina', 'numero', 'tipo']):
                extra_cols.append(ci)

        for i in range(n):
            try:
                # Título
                titulo = None
                v = raw_titulo.iloc[i]
                if pd.notna(v) and str(v).strip():
                    titulo = str(v).strip()
                if not titulo:
                    for t in (textos[ci].iloc[i] for ci in range(len(columnas))):
                        if t and len(t) > 15 and not any(p in t.lower() for p in ['http', 'doi:', 'issn', 'isbn']):
                            titulo = t
                            break
                if not titulo or len(titulo) < 5:
                    continue

                # Año
                anio = None
                v = raw_anio.iloc[i]
                if pd.notna(v):
                    anio = self._extraer_anio(str(v))
                if anio is None:
                    combined = ' '.join(textos[ci].iloc[i] for ci in range(len(columnas)) if textos[ci].iloc[i])
                    anio = self._extraer_anio(combined)
                if anio is not None and anio not in self.anios_filtro:
                    continue

                # Autores
                autores = None
                v = raw_autor.iloc[i]
                if pd.notna(v):
                    autores = str(v).strip()
                if not autores:
                    for t in (textos[ci].iloc[i] for ci in range(len(columnas))):
                        if t and ',' in t and 10 < len(t) < 500:
                            if not any(p in t.lower() for p in ['http', 'doi', 'issn', 'isbn', '@', '.com']):
                                autores = t
                                break

                # Info extra
                info_extra = {}
                for ci in extra_cols:
                    v = df.iloc[i, ci]
                    if pd.notna(v):
                        info_extra[columnas[ci]] = str(v).strip()

                productos.append({
                    'titulo': titulo[:500],
                    'anio': anio,
                    'autores': autores or 'No especificado',
                    'autores_tokens': _tokenizar_autores(autores or ''),
                    'hoja_original': nombre_hoja,
                    'categoria_957': categoria,
                    'subcategoria_957': subcategoria,
                    'producto_957': nombre_oficial,
                    'grupo': nombre_grupo,
                    'archivo': archivo_nombre,
                    'info_extra': info_extra,
                    'orden_categoria': get_orden_categoria(categoria),
                    'orden_producto': get_orden_producto(nombre_oficial),
                })
            except Exception:
                continue

        return productos

    def _procesar_grupo(self, grupo_bd, carpetas):
        carpeta = self._buscar_carpeta(grupo_bd, carpetas)
        if not carpeta or not carpeta.exists():
            return grupo_bd, None

        archivos = [
            f for f in list(carpeta.glob("*.xlsx")) + list(carpeta.glob("*.xls"))
            if not f.name.startswith("~$")
        ]
        if not archivos:
            return grupo_bd, None

        productos_grupo = []
        resumen = defaultdict(int)

        for archivo in archivos:
            try:
                ef = pd.ExcelFile(archivo, engine='openpyxl')
                for hoja in ef.sheet_names:
                    try:
                        df = pd.read_excel(ef, sheet_name=hoja, dtype=str)
                        cnt = max(0, len(df) - 1) if len(df) > 1 else 0
                        if cnt > 0:
                            resumen[hoja] += cnt
                        prods = self._extraer_productos_hoja(df, hoja, grupo_bd, archivo.name)
                        productos_grupo.extend(prods)
                    except Exception:
                        pass
            except Exception:
                pass

        ruta_excel = str(archivos[0]) if archivos else None

        return grupo_bd, {
            'productos': productos_grupo,
            'resumen': dict(resumen),
            'archivos': len(archivos),
            'carpeta': carpeta.name,
            'ruta_excel': ruta_excel,
            'carpeta_path': str(carpeta),
        }

    def run(self):
        try:
            resultado = {
                'grupos': [],
                'datos_gruplac': {},
                'rutas_excel': {},
            }

            self.progreso.emit(5, "Iniciando análisis...")

            cursor = self.db.conn.cursor()
            grupos_bd = cursor.execute('''
                SELECT DISTINCT grupo FROM grupos
                WHERE grupo IS NOT NULL AND grupo != ''
                AND grupo NOT LIKE '%SEMILLERO%' AND grupo NOT LIKE '%Semillero%' AND grupo NOT LIKE '%semillero%'
                ORDER BY grupo
            ''').fetchall()

            if not grupos_bd:
                self.error.emit("No se encontraron grupos en la base de datos")
                return

            grupos_lista = [g[0] for g in grupos_bd]
            resultado['grupos'] = grupos_lista
            self.progreso.emit(10, f"{len(grupos_lista)} grupos")

            self.progreso.emit(15, "Analizando archivos GrupLAC (paralelo)...")

            if self.reporte_excel_path.exists():
                carpetas = [d for d in self.reporte_excel_path.iterdir() if d.is_dir()]
                self.progreso.emit(18, f"{len(carpetas)} carpetas encontradas")

                total = len(grupos_lista)
                completados = 0

                with ThreadPoolExecutor(max_workers=4) as executor:
                    future_map = {
                        executor.submit(self._procesar_grupo, g, carpetas): g
                        for g in grupos_lista
                    }
                    for future in as_completed(future_map):
                        completados += 1
                        pct = 18 + int((completados / total) * 42)
                        self.progreso.emit(pct, f"Procesando grupos... {completados}/{total}")
                        try:
                            grupo_bd, data = future.result()
                        except Exception:
                            continue

                        if not data or not data['productos']:
                            continue

                        resultado['datos_gruplac'][grupo_bd] = {
                            'archivos': data['archivos'],
                            'productos': data['resumen'],
                            'total': len(data['productos']),
                            'carpeta_encontrada': data['carpeta'],
                            'detalles': [{
                                'tipo': p['producto_957'],
                                'tipo_nombre': p['hoja_original'],
                                'titulo': p['titulo'],
                                'anio': p['anio'],
                                'autores': p['autores'],
                                'categoria_957': p['categoria_957'],
                                'subcategoria_957': p['subcategoria_957'],
                                'info_extra': p.get('info_extra', {}),
                            } for p in data['productos']],
                        }
                        resultado['rutas_excel'][grupo_bd] = {
                            'ruta_excel': data.get('ruta_excel'),
                            'carpeta_path': data.get('carpeta_path'),
                        }

            # Fallback: gruplac_957.db para grupos sin carpeta GrupLAC
            grupos_sin_data = [g for g in grupos_lista if g not in resultado['datos_gruplac']]
            if grupos_sin_data:
                for fb_path in [
                    obtener_directorio_base() / "data" / "db" / "gruplac_957.db",
                    Path("data/db/gruplac_957.db"),
                    obtener_directorio_base() / "data" / "db" / "gruplac_957_ref.db",
                ]:
                    if fb_path.exists():
                        try:
                            import sqlite3 as _sq3
                            fb_conn = _sq3.connect(str(fb_path))
                            fb_cur = fb_conn.cursor()
                            grupos_ref = fb_cur.execute(
                                "SELECT id, nombre FROM grupos ORDER BY nombre"
                            ).fetchall()
                            for g in grupos_sin_data:
                                g_norm = self._normalizar_nombre(g)
                                mejor_id = None
                                mejor_nom = None
                                for rid, rnombre in grupos_ref:
                                    rnorm = self._normalizar_nombre(rnombre)
                                    if rnorm == g_norm:
                                        mejor_id, mejor_nom = rid, rnombre
                                        break
                                if not mejor_id:
                                    for rid, rnombre in grupos_ref:
                                        rnorm = self._normalizar_nombre(rnombre)
                                        if len(g_norm) >= 20 and len(rnorm) >= 20 and g_norm[:60] == rnorm[:60]:
                                            mejor_id, mejor_nom = rid, rnombre
                                            break
                                if mejor_id:
                                    prods = fb_cur.execute(
                                        "SELECT categoria_957, subcategoria_957, tipo_producto_957, "
                                        "titulo, anio, autores, pestana FROM productos_957 WHERE grupo_id=?",
                                        (mejor_id,)
                                    ).fetchall()
                                    if prods:
                                        detalles = []
                                        for p in prods:
                                            anio_val = p[4]
                                            if anio_val is not None:
                                                anio_str = str(anio_val).strip()
                                                anio_val = int(anio_str) if anio_str.isdigit() else None
                                            detalles.append({
                                                'tipo': p[2] or '',
                                                'tipo_nombre': p[6] or '',
                                                'titulo': p[3] or '',
                                                'anio': anio_val,
                                                'autores': p[5] or '',
                                                'categoria_957': p[0] or '',
                                                'subcategoria_957': p[1] or '',
                                                'info_extra': {},
                                            })
                                        resultado['datos_gruplac'][g] = {
                                            'archivos': 0,
                                            'productos': {},
                                            'total': len(detalles),
                                            'carpeta_encontrada': f"gruplac_957.db:{mejor_nom}",
                                            'detalles': detalles,
                                        }
                                        resultado['rutas_excel'][g] = {
                                            'ruta_excel': None,
                                            'carpeta_path': str(fb_path),
                                        }
                            fb_conn.close()
                        except Exception:
                            pass
                        break

            self.progreso.emit(100, "Analisis completado")
            self.finalizado.emit(resultado)

        except Exception as e:
            import traceback
            self.error.emit(f"{str(e)}\n\n{traceback.format_exc()}")


# =============================================================================
# PRODUCTOS INTERNOS POR INDICADOR 957
# =============================================================================

_IND_INFO_957: dict = {
    "TOP":    ("NC_TOP",  "Nuevo Conocimiento TOP",  QColor(26,  82, 118)),
    "TIPO_A": ("NC_A",    "Nuevo Conocimiento A",    QColor(26, 122,  74)),
    "TIPO_B": ("NC_B",    "Nuevo Conocimiento B",    QColor(125,102,   8)),
    "AP":     ("ASC",     "Apropiación Social",      QColor(120,  40,  31)),
    "DPC":    ("DPC",     "Divulgación Pública",     QColor( 60,  60,  60)),
    "FR_A":   ("FRH_A",   "Formación RH Tipo A",     QColor(  0, 100, 148)),
    "FR_B":   ("FRH_B",   "Formación RH Tipo B",     QColor( 80,  80,  80)),
}


class DialogoProductos957(QDialog):
    """
    Dos tabs:
      1. Productos BD interna — árbol agrupado por indicador 957.
      2. Datos paper 957     — grupos reales de medicion_957.xlsx para contrastar.
    """

    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self._df_paper: pd.DataFrame = pd.DataFrame()
        self.setWindowTitle("Indicadores 957 — BD interna vs Paper MinCiencias")
        self.setMinimumSize(1000, 680)
        self.resize(1150, 760)
        self._setup_ui()
        self._cargar_grupos()
        self._cargar_datos_paper()

    # ── UI ───────────────────────────────────────────────────────────────

    def _setup_ui(self):
        lay = QVBoxLayout(self)
        lay.setSpacing(6)

        # — Toolbar global —
        tb = QHBoxLayout()
        tb.addWidget(QLabel("<b>Grupo BD:</b>"))
        self.combo_grupo = QComboBox()
        self.combo_grupo.setMinimumWidth(280)
        tb.addWidget(self.combo_grupo, 3)
        tb.addWidget(QLabel("<b>Año base:</b>"))
        self.spin_año = QSpinBox()
        self.spin_año.setRange(2018, 2035)
        self.spin_año.setValue(2024)
        self.spin_año.setFixedWidth(68)
        tb.addWidget(self.spin_año)
        btn_cargar = QPushButton("Cargar productos")
        btn_cargar.setStyleSheet(
            "QPushButton{background-color:#1a6b3c;color:white;padding:5px 16px;"
            "border-radius:4px;font-weight:bold;font-size:11px;}"
            "QPushButton:hover{background-color:#145a32;}"
        )
        btn_cargar.clicked.connect(self._cargar)
        tb.addWidget(btn_cargar)
        lay.addLayout(tb)

        self.lbl_estado = QLabel("Seleccione un grupo y presione 'Cargar productos'.")
        self.lbl_estado.setStyleSheet("font-size:10px; color:#555; padding:2px;")
        lay.addWidget(self.lbl_estado)

        # — Resumen de λ por indicador —
        self.lbl_resumen = QLabel("")
        self.lbl_resumen.setStyleSheet(
            "font-size:11px; font-family:Consolas; padding:4px 8px; "
            "background:#f0f4f8; border-radius:4px; color:#1a365d;"
        )
        self.lbl_resumen.setWordWrap(True)
        lay.addWidget(self.lbl_resumen)

        # — Tabs principales —
        self.tabs = QTabWidget()
        lay.addWidget(self.tabs, 1)

        # Tab 1: BD interna
        tab_bd = QWidget()
        self._setup_tab_bd(tab_bd)
        self.tabs.addTab(tab_bd, "Productos BD interna")

        # Tab 2: Paper 957
        tab_paper = QWidget()
        self._setup_tab_paper(tab_paper)
        self.tabs.addTab(tab_paper, "Grupos del paper 957")

        # Tab 3: Plan de mejora 957
        tab_plan = QWidget()
        self._setup_tab_plan(tab_plan)
        self.tabs.addTab(tab_plan, "Plan de mejora 957")

        btn_cerrar = QPushButton("Cerrar")
        btn_cerrar.clicked.connect(self.accept)
        blay = QHBoxLayout()
        blay.addStretch()
        blay.addWidget(btn_cerrar)
        lay.addLayout(blay)

    def _setup_tab_bd(self, tab):
        lay = QVBoxLayout(tab)
        lay.setContentsMargins(0, 4, 0, 0)
        self.tree = QTreeWidget()
        self.tree.setHeaderLabels(["Indicador / Producto", "Año", "Tipo", "λ", "Fuente"])
        self.tree.header().setSectionResizeMode(0, QHeaderView.Stretch)
        for c in (1, 2, 3, 4):
            self.tree.header().setSectionResizeMode(c, QHeaderView.ResizeToContents)
        self.tree.setEditTriggers(QTreeWidget.NoEditTriggers)
        self.tree.setAlternatingRowColors(True)
        self.tree.setStyleSheet("font-size:10px;")
        lay.addWidget(self.tree)

    def _setup_tab_plan(self, tab):
        from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
        from matplotlib.figure import Figure

        lay = QVBoxLayout(tab)
        lay.setContentsMargins(0, 4, 0, 0)
        lay.setSpacing(4)

        self.lbl_plan_estado = QLabel(
            "Seleccione un grupo y presione 'Cargar productos' para ver su plan de mejora."
        )
        self.lbl_plan_estado.setWordWrap(True)
        self.lbl_plan_estado.setStyleSheet(
            "font-size:11px; font-weight:bold; color:#1a365d; padding:4px 8px; "
            "background:#e8f0f8; border-radius:4px;"
        )
        lay.addWidget(self.lbl_plan_estado)

        self.tabs_plan = QTabWidget()
        self.tabs_plan.setStyleSheet("QTabWidget{font-size:10px;}")
        lay.addWidget(self.tabs_plan, 1)

        # ════════════════════════════════════════════════════════════════
        # Sub-pestaña 1: Diagnóstico (indicadores vs. umbral + radar)
        # ════════════════════════════════════════════════════════════════
        t1 = QWidget()
        l1 = QVBoxLayout(t1)
        l1.setContentsMargins(2, 2, 2, 2)
        l1.setSpacing(4)

        self.lbl_plan_seccion1 = QLabel(
            "<b>1. ¿Qué indicadores te están deteniendo?</b><br>"
            "<span style='font-size:9px; color:#666;'>"
            "Para cada indicador del modelo 957 se compara TU valor actual contra el "
            "<b>umbral de referencia</b>: el percentil 25 de los grupos de tu misma área de "
            "conocimiento que YA están en la categoría objetivo. "
            "'¿Llegas al umbral?' = Sí → ese indicador ya está al nivel necesario y NO es una "
            "limitante. = No → ese indicador es una brecha que te impide subir de categoría."
            "</span>"
        )
        self.lbl_plan_seccion1.setWordWrap(True)
        l1.addWidget(self.lbl_plan_seccion1)

        split1 = QSplitter(Qt.Horizontal)
        self.tbl_plan_indicadores = QTableWidget()
        _ci = ["Indicador", "Tu valor actual", "Umbral de referencia (P25 del área)", "¿Llegas al umbral?", "Te falta"]
        self.tbl_plan_indicadores.setColumnCount(len(_ci))
        self.tbl_plan_indicadores.setHorizontalHeaderLabels(_ci)
        hh = self.tbl_plan_indicadores.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.Stretch)
        for c in range(1, len(_ci)):
            hh.setSectionResizeMode(c, QHeaderView.ResizeToContents)
        self.tbl_plan_indicadores.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl_plan_indicadores.setAlternatingRowColors(True)
        self.tbl_plan_indicadores.setStyleSheet("font-size:10px;")
        split1.addWidget(self.tbl_plan_indicadores)

        self._fig_radar = Figure(figsize=(4.2, 3.4), dpi=90)
        self._canvas_radar = FigureCanvas(self._fig_radar)
        split1.addWidget(self._canvas_radar)
        split1.setSizes([520, 360])
        l1.addWidget(split1, 1)

        self.tabs_plan.addTab(t1, "1. Diagnóstico")

        # ════════════════════════════════════════════════════════════════
        # Sub-pestaña 2: Recomendaciones y comparación con grupos de referencia
        # ════════════════════════════════════════════════════════════════
        t2 = QWidget()
        l2 = QVBoxLayout(t2)
        l2.setContentsMargins(2, 2, 2, 2)
        l2.setSpacing(4)

        self.lbl_plan_seccion2 = QLabel(
            "<b>2. ¿Con qué productos podrías cerrar cada brecha?</b><br>"
            "<span style='font-size:9px; color:#666;'>"
            "Por cada indicador con brecha se listan TODAS las alternativas de producto "
            "(artículos, libros, capítulos, software, prototipos, patentes, tesis, trabajos de "
            "grado, etc.) que aportan a ese indicador, con la cantidad mínima de cada una para "
            "cerrarla."
            "</span>"
        )
        self.lbl_plan_seccion2.setWordWrap(True)
        l2.addWidget(self.lbl_plan_seccion2)
        self.tbl_plan_recomendaciones = QTableWidget()
        _cr = ["Indicador con brecha", "Te falta", "Producto / opción", "λ por unidad", "Cantidad sugerida"]
        self.tbl_plan_recomendaciones.setColumnCount(len(_cr))
        self.tbl_plan_recomendaciones.setHorizontalHeaderLabels(_cr)
        hh_r = self.tbl_plan_recomendaciones.horizontalHeader()
        hh_r.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        hh_r.setSectionResizeMode(2, QHeaderView.Stretch)
        for c in (1, 3, 4):
            hh_r.setSectionResizeMode(c, QHeaderView.ResizeToContents)
        self.tbl_plan_recomendaciones.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl_plan_recomendaciones.setAlternatingRowColors(True)
        self.tbl_plan_recomendaciones.setWordWrap(True)
        self.tbl_plan_recomendaciones.setStyleSheet("font-size:10px;")
        l2.addWidget(self.tbl_plan_recomendaciones, 1)

        self.lbl_plan_referencia_titulo = QLabel(
            "<b>3. Comparación con grupos reales de tu área que ya están en la categoría objetivo</b>"
        )
        l2.addWidget(self.lbl_plan_referencia_titulo)

        fila_ref = QHBoxLayout()
        fila_ref.addWidget(QLabel("Comparar también contra un grupo elegido por ti (misma área y categoría objetivo):"))
        self.cmb_plan_grupo_ref = QComboBox()
        self.cmb_plan_grupo_ref.addItem("— Automático (grupo más débil) —")
        self.cmb_plan_grupo_ref.setMinimumWidth(280)
        self.cmb_plan_grupo_ref.currentIndexChanged.connect(self._on_cambio_grupo_referencia)
        fila_ref.addWidget(self.cmb_plan_grupo_ref)
        fila_ref.addStretch()
        l2.addLayout(fila_ref)

        self.lbl_plan_grupo_cercano = QLabel("")
        self.lbl_plan_grupo_cercano.setWordWrap(True)
        self.lbl_plan_grupo_cercano.setStyleSheet("font-size:9px; color:#444; padding:2px 4px;")
        l2.addWidget(self.lbl_plan_grupo_cercano)

        self.tbl_plan_comparacion = QTableWidget()
        _cc = ["Indicador", "Tu grupo", "Grupo más débil (mín. puntaje)", "Diferencia",
               "Grupo de perfil similar", "Diferencia", "Grupo seleccionado", "Diferencia"]
        self.tbl_plan_comparacion.setColumnCount(len(_cc))
        self.tbl_plan_comparacion.setHorizontalHeaderLabels(_cc)
        hh_c = self.tbl_plan_comparacion.horizontalHeader()
        hh_c.setSectionResizeMode(0, QHeaderView.Stretch)
        for c in range(1, len(_cc)):
            hh_c.setSectionResizeMode(c, QHeaderView.ResizeToContents)
        self.tbl_plan_comparacion.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl_plan_comparacion.setAlternatingRowColors(True)
        self.tbl_plan_comparacion.setStyleSheet("font-size:10px;")
        l2.addWidget(self.tbl_plan_comparacion, 1)

        self.lbl_plan_referencia = QLabel("")
        self.lbl_plan_referencia.setWordWrap(True)
        self.lbl_plan_referencia.setStyleSheet("font-size:9px; color:#777; padding:2px 4px;")
        l2.addWidget(self.lbl_plan_referencia)

        self.tabs_plan.addTab(t2, "2-3. Recomendaciones")

        # ════════════════════════════════════════════════════════════════
        # Sub-pestaña 3: Mínimos nacionales / cuartiles
        # ════════════════════════════════════════════════════════════════
        t3 = QWidget()
        l3 = QVBoxLayout(t3)
        l3.setContentsMargins(2, 2, 2, 2)
        l3.setSpacing(4)

        self.lbl_plan_seccion4 = QLabel(
            "<b>4. La \"ecuación\" de MinCiencias: requisitos mínimos para la categoría objetivo</b><br>"
            "<span style='font-size:9px; color:#666;'>"
            "Distribución nacional de cada indicador (hoja 'cuartiles' de medicion_957.xlsx) entre "
            "los grupos de tu misma área de conocimiento y categoría objetivo. La columna "
            "<b>'Requisito mínimo'</b> es el valor que, según el documento oficial, un indicador debe "
            "alcanzar para contar a favor de esa categoría. "
            "'¿Cumples?' = Sí → tu valor ya alcanza ese requisito. El gráfico ubica 'Tu valor' "
            "en el percentil nacional aproximado (0=mínimo, 25=Q4, 50=Q3, 75=Q2, 100=máximo) "
            "y la línea punteada marca el percentil exigido para la categoría objetivo."
            "</span>"
        )
        self.lbl_plan_seccion4.setWordWrap(True)
        l3.addWidget(self.lbl_plan_seccion4)

        self.lbl_plan_cuartiles_info = QLabel("")
        self.lbl_plan_cuartiles_info.setWordWrap(True)
        self.lbl_plan_cuartiles_info.setStyleSheet("font-size:9px; color:#777; padding:2px 4px;")
        l3.addWidget(self.lbl_plan_cuartiles_info)

        self._fig_cuartiles = Figure(figsize=(6, 3.2), dpi=90)
        self._canvas_cuartiles = FigureCanvas(self._fig_cuartiles)
        l3.addWidget(self._canvas_cuartiles)

        self.tbl_plan_cuartiles = QTableWidget()
        _ce = ["Indicador", "Tu valor", "Mínimo nacional", "Q4", "Q3", "Q2", "Máximo", "Requisito mínimo", "¿Cumples?"]
        self.tbl_plan_cuartiles.setColumnCount(len(_ce))
        self.tbl_plan_cuartiles.setHorizontalHeaderLabels(_ce)
        hh_e = self.tbl_plan_cuartiles.horizontalHeader()
        hh_e.setSectionResizeMode(0, QHeaderView.Stretch)
        for c in range(1, len(_ce)):
            hh_e.setSectionResizeMode(c, QHeaderView.ResizeToContents)
        self.tbl_plan_cuartiles.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl_plan_cuartiles.setAlternatingRowColors(True)
        self.tbl_plan_cuartiles.setStyleSheet("font-size:10px;")
        l3.addWidget(self.tbl_plan_cuartiles, 1)

        self.tabs_plan.addTab(t3, "4. Mínimos nacionales")

        # ════════════════════════════════════════════════════════════════
        # Sub-pestaña 4: Simulador
        # ════════════════════════════════════════════════════════════════
        t4 = QWidget()
        l4 = QVBoxLayout(t4)
        l4.setContentsMargins(2, 2, 2, 2)
        l4.setSpacing(4)

        self.lbl_plan_seccion5 = QLabel(
            "<b>5. Simulador: ¿qué productos te faltan para llegar a la categoría objetivo?</b><br>"
            "<span style='font-size:9px; color:#666;'>"
            "Para cada indicador que aún no cumple el requisito mínimo de la sección 4 se calcula "
            "un <b>factor de conversión propio de tu grupo</b> (tu valor oficial ÷ tu λ actual, "
            "donde λ = ln(1 + total/ventana), igual que la hoja 'productos' de medicion_957.xlsx). "
            "Indica cuántas unidades adicionales de cada producto piensas generar: el "
            "'Valor simulado' queda en la misma escala que 'Tu valor'/'Requisito mínimo' de la "
            "sección 4 y se recalcula automáticamente, junto con '¿Cumples?'."
            "</span>"
        )
        self.lbl_plan_seccion5.setWordWrap(True)
        l4.addWidget(self.lbl_plan_seccion5)

        self.lbl_plan_simulador_veredicto = QLabel("")
        self.lbl_plan_simulador_veredicto.setWordWrap(True)
        self.lbl_plan_simulador_veredicto.setStyleSheet(
            "font-size:11px; font-weight:bold; padding:4px 8px; "
            "background:#f0f0f0; border-radius:4px;"
        )
        l4.addWidget(self.lbl_plan_simulador_veredicto)

        self._fig_simulador = Figure(figsize=(6, 2.6), dpi=90)
        self._canvas_simulador = FigureCanvas(self._fig_simulador)
        l4.addWidget(self._canvas_simulador)

        self.tbl_plan_simulador = QTableWidget()
        _cs = ["Indicador", "Tu valor", "Requisito mínimo", "Producto candidato",
               "Ya tienes", "Vigencia (años)", "+ Unidades a generar",
               "Valor simulado / ¿Cumples?"]
        self.tbl_plan_simulador.setColumnCount(len(_cs))
        self.tbl_plan_simulador.setHorizontalHeaderLabels(_cs)
        hh_s = self.tbl_plan_simulador.horizontalHeader()
        hh_s.setSectionResizeMode(3, QHeaderView.Stretch)
        for c in range(len(_cs)):
            if c != 3:
                hh_s.setSectionResizeMode(c, QHeaderView.ResizeToContents)
        self.tbl_plan_simulador.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl_plan_simulador.setAlternatingRowColors(True)
        self.tbl_plan_simulador.setWordWrap(True)
        self.tbl_plan_simulador.setStyleSheet("font-size:10px;")
        l4.addWidget(self.tbl_plan_simulador, 1)
        self._simulador_grupos = []

        self.tabs_plan.addTab(t4, "5. Simulador")

    # columnas de indicadores en el mismo orden que el paper
    _CATS_PAPEL   = ["A1", "A", "B", "C", "NR", "SR"]
    _CAT_BG       = {
        "A1": QColor(210, 230, 255),
        "A":  QColor(210, 245, 215),
        "B":  QColor(255, 250, 215),
        "C":  QColor(255, 235, 215),
        "NR": QColor(240, 228, 248),
        "SR": QColor(248, 248, 248),
    }
    _IND_PIVOT = ["NC_TOP", "NC_A", "NC_B", "ASC", "DPC", "FRH_A", "FRH_B"]
    _IND_LABEL_LARGO = {
        "NC_TOP": "Nuevo Conoc. TOP",
        "NC_A":   "Nuevo Conoc. A",
        "NC_B":   "Nuevo Conoc. B",
        "ASC":    "Apropiación Social",
        "DPC":    "Divulgación Pública",
        "FRH_A":  "Formación RH Tipo A",
        "FRH_B":  "Formación RH Tipo B",
    }

    def _setup_tab_paper(self, tab):
        from PyQt5.QtWidgets import QSplitter
        lay = QVBoxLayout(tab)
        lay.setContentsMargins(0, 4, 0, 0)
        lay.setSpacing(4)

        # — Filtros —
        fb = QHBoxLayout()
        fb.addWidget(QLabel("Área:"))
        self.combo_area_paper = QComboBox()
        self.combo_area_paper.setMinimumWidth(190)
        fb.addWidget(self.combo_area_paper, 2)
        fb.addWidget(QLabel("Buscar grupo:"))
        self.le_buscar_paper = QLineEdit()
        self.le_buscar_paper.setPlaceholderText("Nombre del grupo…")
        fb.addWidget(self.le_buscar_paper, 2)
        lay.addLayout(fb)

        self.lbl_paper = QLabel("Cargando datos del paper…")
        self.lbl_paper.setStyleSheet("font-size:10px; color:#555;")
        lay.addWidget(self.lbl_paper)

        # — Splitter vertical: tablas por categoría (arriba) | detalle (abajo) —
        splitter = QSplitter(Qt.Vertical)
        lay.addWidget(splitter, 1)

        # ── Top: tabs por categoría ──────────────────────────────────────
        self.tabs_cat_paper = QTabWidget()
        self._cat_tables: dict = {}
        _cols = (
            ["Nombre del grupo", "Director", "IG"]
            + self._IND_PIVOT
            + ["Área de conocimiento"]
        )
        for cat in self._CATS_PAPEL:
            w = QWidget()
            wl = QVBoxLayout(w)
            wl.setContentsMargins(0, 2, 0, 0)
            tbl = QTableWidget()
            tbl.setColumnCount(len(_cols))
            tbl.setHorizontalHeaderLabels(_cols)
            hh = tbl.horizontalHeader()
            hh.setSectionResizeMode(0, QHeaderView.Stretch)   # Nombre
            hh.setSectionResizeMode(1, QHeaderView.Stretch)   # Director
            hh.setSectionResizeMode(len(_cols)-1, QHeaderView.Stretch)  # Área
            for c in range(2, len(_cols)-1):
                hh.setSectionResizeMode(c, QHeaderView.ResizeToContents)
            tbl.setEditTriggers(QTableWidget.NoEditTriggers)
            tbl.setSelectionBehavior(QTableWidget.SelectRows)
            tbl.setAlternatingRowColors(True)
            tbl.setSortingEnabled(True)
            tbl.setStyleSheet("font-size:10px;")
            tbl.itemSelectionChanged.connect(self._on_grupo_paper_seleccionado)
            wl.addWidget(tbl)
            self._cat_tables[cat] = tbl
            self.tabs_cat_paper.addTab(w, cat)

        splitter.addWidget(self.tabs_cat_paper)

        # ── Bottom: detalle del grupo seleccionado ───────────────────────
        detail = QWidget()
        dl = QVBoxLayout(detail)
        dl.setContentsMargins(0, 0, 0, 0)
        dl.setSpacing(2)

        self.lbl_grupo_paper_sel = QLabel("▶ Seleccione un grupo para ver su detalle completo.")
        self.lbl_grupo_paper_sel.setStyleSheet(
            "font-size:10px; color:#1a365d; font-weight:bold; padding:3px 4px; "
            "background:#e8f0f8; border-radius:3px;"
        )
        dl.addWidget(self.lbl_grupo_paper_sel)

        self.tabs_detalle_paper = QTabWidget()
        self.tabs_detalle_paper.setStyleSheet("QTabWidget{font-size:10px;}")
        dl.addWidget(self.tabs_detalle_paper, 1)

        # Sub-tab: Comparación visual (IG por área)
        tv = QWidget()
        tv_l = QVBoxLayout(tv)
        tv_l.setContentsMargins(0, 2, 0, 0)
        self.lbl_chart_info = QLabel(
            "Seleccione un grupo para ver su posición de IG dentro de su área de conocimiento."
        )
        self.lbl_chart_info.setStyleSheet("font-size:10px; color:#555; padding:2px;")
        self.lbl_chart_info.setWordWrap(True)
        tv_l.addWidget(self.lbl_chart_info)

        from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
        from matplotlib.figure import Figure
        self._fig_ig = Figure(figsize=(6, 3.0), dpi=90)
        self._canvas_ig = FigureCanvas(self._fig_ig)
        tv_l.addWidget(self._canvas_ig, 1)
        self.tabs_detalle_paper.addTab(tv, "Comparación visual (IG)")

        # Sub-tab: Indicadores
        ti = QWidget()
        ti_l = QVBoxLayout(ti)
        ti_l.setContentsMargins(0, 2, 0, 0)
        self.tbl_paper_ind = QTableWidget()
        _ci = ["Indicador", "Nombre completo", "Valor", "Máx. área",
               "Índice (0-1)", "Ponderación", "Aporte IG", "Cuartil"]
        self.tbl_paper_ind.setColumnCount(len(_ci))
        self.tbl_paper_ind.setHorizontalHeaderLabels(_ci)
        hh_i = self.tbl_paper_ind.horizontalHeader()
        hh_i.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        hh_i.setSectionResizeMode(1, QHeaderView.Stretch)
        for c in range(2, len(_ci)):
            hh_i.setSectionResizeMode(c, QHeaderView.ResizeToContents)
        self.tbl_paper_ind.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl_paper_ind.setAlternatingRowColors(True)
        self.tbl_paper_ind.setStyleSheet("font-size:10px;")
        ti_l.addWidget(self.tbl_paper_ind)
        self.tabs_detalle_paper.addTab(ti, "Indicadores")

        # Sub-tab: Productos contabilizados
        tp = QWidget()
        tp_l = QVBoxLayout(tp)
        tp_l.setContentsMargins(0, 2, 0, 0)
        self.tbl_paper_prods = QTableWidget()
        _cp = ["Sección", "Subtipo / tipo de producto", "Total", "En ventana", "λ acumulado"]
        self.tbl_paper_prods.setColumnCount(len(_cp))
        self.tbl_paper_prods.setHorizontalHeaderLabels(_cp)
        hh_p = self.tbl_paper_prods.horizontalHeader()
        hh_p.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        hh_p.setSectionResizeMode(1, QHeaderView.Stretch)
        for c in (2, 3, 4):
            hh_p.setSectionResizeMode(c, QHeaderView.ResizeToContents)
        self.tbl_paper_prods.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl_paper_prods.setAlternatingRowColors(True)
        self.tbl_paper_prods.setStyleSheet("font-size:10px;")
        tp_l.addWidget(self.tbl_paper_prods)
        self.tabs_detalle_paper.addTab(tp, "Productos contabilizados")

        # Sub-tab: Cuartiles
        tc = QWidget()
        tc_l = QVBoxLayout(tc)
        tc_l.setContentsMargins(0, 2, 0, 0)
        self.tbl_paper_cuar = QTableWidget()
        _cc = ["Indicador", "Valor grupo", "Q4 (mín.)", "Q3", "Q2 (med.)", "Máx.", "Cuartil"]
        self.tbl_paper_cuar.setColumnCount(len(_cc))
        self.tbl_paper_cuar.setHorizontalHeaderLabels(_cc)
        hh_c = self.tbl_paper_cuar.horizontalHeader()
        hh_c.setSectionResizeMode(0, QHeaderView.Stretch)
        for c in range(1, len(_cc)):
            hh_c.setSectionResizeMode(c, QHeaderView.ResizeToContents)
        self.tbl_paper_cuar.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl_paper_cuar.setAlternatingRowColors(True)
        self.tbl_paper_cuar.setStyleSheet("font-size:10px;")
        tc_l.addWidget(self.tbl_paper_cuar)
        self.tabs_detalle_paper.addTab(tc, "Cuartiles por indicador")

        splitter.addWidget(detail)
        splitter.setSizes([400, 280])

        # Conectar filtros
        self.combo_area_paper.currentIndexChanged.connect(self._filtrar_paper)
        self.le_buscar_paper.textChanged.connect(self._filtrar_paper)

    # ── Datos BD interna ──────────────────────────────────────────────────

    def _cargar_grupos(self):
        try:
            rows = self.db.conn.execute(
                "SELECT DISTINCT grupo FROM grupos WHERE grupo IS NOT NULL ORDER BY grupo"
            ).fetchall()
            self.combo_grupo.addItems([r[0] for r in rows if r[0]])
        except Exception:
            pass

    def _cargar(self):
        nombre = self.combo_grupo.currentText()
        año    = self.spin_año.value()
        if not nombre:
            return
        try:
            from analisis_seguimiento import SimuladorCategoriaInterna
            sim = SimuladorCategoriaInterna(self.db.conn, año_base=año)
            cedulas = sim._cedulas_del_grupo(nombre)
            if not cedulas:
                self.lbl_estado.setText(f"'{nombre}' no tiene integrantes en la BD.")
                return
            resultado = sim._calcular_indicadores_con_detalle(cedulas)
        except Exception as e:
            self.lbl_estado.setText(f"Error: {e}")
            return

        indicadores       = resultado["indicadores"]
        productos_por_ind = resultado["productos_por_indicador"]
        total_prods       = sum(len(v) for v in productos_por_ind.values())
        self.lbl_estado.setText(
            f"Grupo: {nombre}  |  {len(cedulas)} integrante(s)  |  "
            f"{total_prods} producto(s) en ventana  |  Año base: {año}"
        )
        self._poblar_resumen(indicadores)
        self._poblar_arbol(productos_por_ind)
        self._cargar_plan_mejora(sim, nombre)

    def _limpiar_tablas_plan(self):
        self.tbl_plan_indicadores.setRowCount(0)
        self.tbl_plan_recomendaciones.setRowCount(0)
        self.tbl_plan_comparacion.setRowCount(0)
        self.tbl_plan_cuartiles.setRowCount(0)
        self.tbl_plan_simulador.setRowCount(0)
        self.lbl_plan_grupo_cercano.setText("")
        self.lbl_plan_referencia.setText("")
        self.lbl_plan_cuartiles_info.setText("")
        self.lbl_plan_simulador_veredicto.setText("")
        self._simulador_grupos = []
        for fig, canvas in (
            (self._fig_radar, self._canvas_radar),
            (self._fig_cuartiles, self._canvas_cuartiles),
            (self._fig_simulador, self._canvas_simulador),
        ):
            fig.clear()
            canvas.draw()
        self.cmb_plan_grupo_ref.blockSignals(True)
        self.cmb_plan_grupo_ref.clear()
        self.cmb_plan_grupo_ref.addItem("— Automático (grupo más débil) —")
        self.cmb_plan_grupo_ref.blockSignals(False)
        self._plan_sim = None
        self._plan_nombre = None

    def _cargar_plan_mejora(self, sim, nombre: str):
        try:
            plan = sim.analizar_brechas_area(nombre)
        except Exception as e:
            self.lbl_plan_estado.setText(f"Error calculando el plan de mejora: {e}")
            self._limpiar_tablas_plan()
            return

        if plan is None:
            self.lbl_plan_estado.setText(f"'{nombre}' no tiene integrantes en la BD.")
            self._limpiar_tablas_plan()
            return

        cat_actual    = plan.get("categoria_actual")
        area          = plan.get("area_conocimiento")
        cat_objetivo  = plan.get("categoria_objetivo")

        if cat_objetivo is None:
            self.lbl_plan_estado.setText(
                f"Categoría actual: {cat_actual or 'desconocida'}  |  "
                f"Área: {area or 'desconocida'}  —  {plan.get('mensaje', '')}"
            )
            self._limpiar_tablas_plan()
            return

        self._plan_sim = sim
        self._plan_nombre = nombre

        self.cmb_plan_grupo_ref.blockSignals(True)
        self.cmb_plan_grupo_ref.clear()
        self.cmb_plan_grupo_ref.addItem("— Automático (grupo más débil) —")
        for g in plan.get("grupos_referencia_disponibles", []):
            self.cmb_plan_grupo_ref.addItem(g)
        self.cmb_plan_grupo_ref.setCurrentIndex(0)
        self.cmb_plan_grupo_ref.blockSignals(False)

        fuente = plan.get("fuente_indicadores", "bd_interna")
        if fuente == "oficial":
            fuente_txt = "datos oficiales de MinCiencias (medicion_957.xlsx)"
        else:
            fuente_txt = "BD interna (academia_utp_integrado.db) — el grupo no aparece en medicion_957.xlsx"

        self.lbl_plan_estado.setText(
            f"Categoría actual: {cat_actual}  |  Área: {area}  |  "
            f"Categoría objetivo: {cat_objetivo}  |  "
            f"Grupos de referencia (percentil {sim.percentil}): {plan.get('num_grupos_referencia', 0)}  |  "
            f"Fuente de indicadores: {fuente_txt}"
        )

        indicadores_actuales = plan.get("indicadores_actuales", {})
        umbrales             = plan.get("umbrales_referencia", {})
        brechas              = plan.get("brechas", {})

        if fuente == "oficial":
            self.lbl_plan_seccion1.setText(
                "<b>1. ¿Qué indicadores te están deteniendo?</b><br>"
                "<span style='font-size:9px; color:#666;'>"
                "Valores del <b>índice oficial 957</b> calculado por MinCiencias "
                "(hoja 'indicadores' de medicion_957.xlsx) a partir de los productos del "
                "documento de medición. Se comparan contra el <b>umbral de referencia</b>: "
                "el percentil 25 de los grupos de tu misma área que YA están en la categoría "
                "objetivo (también con datos oficiales). "
                "'¿Llegas al umbral?' = Sí → ese indicador ya está al nivel necesario. "
                "= No → es una brecha que te impide subir de categoría."
                "</span>"
            )
        else:
            self.lbl_plan_seccion1.setText(
                "<b>1. ¿Qué indicadores te están deteniendo?</b><br>"
                "<span style='font-size:9px; color:#666;'>"
                "Este grupo no aparece en el documento oficial de medición "
                "(medicion_957.xlsx), así que sus indicadores se recalculan desde la "
                "<b>BD interna</b> (academia_utp_integrado.db) según las ventanas/λ de "
                "VENTANAS_957. Se comparan contra el <b>umbral de referencia</b>: el "
                "percentil 25 de los grupos de tu misma área que YA están en la categoría "
                "objetivo (también calculados desde BD interna). "
                "'¿Llegas al umbral?' = Sí → ese indicador ya está al nivel necesario. "
                "= No → es una brecha que te impide subir de categoría."
                "</span>"
            )

        # — Tabla 1: indicadores actuales vs. umbral de referencia —
        self.tbl_plan_indicadores.setRowCount(0)
        for ind in INDICADORES_957:
            _, label, _ = _IND_INFO_957.get(ind, (ind, ind, None))
            actual  = indicadores_actuales.get(ind, 0.0)
            umbral  = umbrales.get(ind)
            if umbral is None:
                continue
            cumple  = actual >= umbral
            row = self.tbl_plan_indicadores.rowCount()
            self.tbl_plan_indicadores.insertRow(row)
            self.tbl_plan_indicadores.setItem(row, 0, QTableWidgetItem(label))
            self.tbl_plan_indicadores.setItem(row, 1, QTableWidgetItem(f"{actual:,.3f}"))
            self.tbl_plan_indicadores.setItem(row, 2, QTableWidgetItem(f"{umbral:,.3f}"))
            item_cumple = QTableWidgetItem("Sí" if cumple else "No")
            item_cumple.setForeground(QBrush(QColor("#1a7a4a") if cumple else QColor("#a02020")))
            self.tbl_plan_indicadores.setItem(row, 3, item_cumple)
            diferencia = brechas.get(ind, {}).get("diferencia", 0.0)
            self.tbl_plan_indicadores.setItem(row, 4, QTableWidgetItem(f"{diferencia:,.3f}" if diferencia else "—"))

        self._dibujar_radar_resumen(plan)

        # — Tabla 2: recomendaciones — una fila por cada opción de producto —
        if fuente == "oficial":
            self.lbl_plan_seccion2.setText(
                "<b>2. ¿Con qué productos podrías cerrar cada brecha?</b><br>"
                "<span style='font-size:9px; color:#666;'>"
                "Por cada indicador con brecha se compara, producto por producto (hoja "
                "'productos' de medicion_957.xlsx), lo que YA TIENES contra lo que tiene el "
                "grupo de referencia más débil que ya está en la categoría objetivo. "
                "'Te faltan' = cuántos productos de ese tipo le faltan a tu grupo para "
                "igualar a ese grupo de referencia. Los indicadores de formación de recurso "
                "humano (FR_A/FR_B) no tienen desglose de productos en el documento oficial."
                "</span>"
            )
            self.tbl_plan_recomendaciones.setHorizontalHeaderLabels(
                ["Indicador con brecha", "Brecha del índice oficial", "Producto / opción",
                 "Ya tienes", "Te faltan (vs. grupo de referencia)"]
            )
        else:
            self.lbl_plan_seccion2.setText(
                "<b>2. ¿Con qué productos podrías cerrar cada brecha?</b><br>"
                "<span style='font-size:9px; color:#666;'>"
                "Por cada indicador con brecha se listan TODAS las alternativas de producto "
                "(artículos, libros, capítulos, software, prototipos, patentes, tesis, trabajos de "
                "grado, etc.) que aportan a ese indicador, con la cantidad mínima de cada una para "
                "cerrarla."
                "</span>"
            )
            self.tbl_plan_recomendaciones.setHorizontalHeaderLabels(
                ["Indicador con brecha", "Te falta", "Producto / opción", "λ por unidad", "Cantidad sugerida"]
            )

        self._actualizar_tablas_dependientes_referencia(plan)

        grupos_ref = plan.get("grupos_referencia", [])
        if grupos_ref:
            self.lbl_plan_referencia.setText(
                f"Todos los grupos de referencia usados para el umbral P25 "
                f"({area}, categoría {cat_objetivo}): " + ", ".join(grupos_ref)
            )
        else:
            self.lbl_plan_referencia.setText(plan.get("mensaje") or "")

    def _on_cambio_grupo_referencia(self, index: int):
        """Recalcula las tablas 2-4 al elegir un grupo de referencia manual."""
        sim    = getattr(self, "_plan_sim", None)
        nombre = getattr(self, "_plan_nombre", None)
        if sim is None or not nombre:
            return

        grupo_manual = self.cmb_plan_grupo_ref.currentText() if index > 0 else None
        try:
            plan = sim.analizar_brechas_area(nombre, grupo_referencia_manual=grupo_manual)
        except Exception as e:
            self.lbl_plan_estado.setText(f"Error calculando el plan de mejora: {e}")
            return

        if plan is None or plan.get("categoria_objetivo") is None:
            return

        self._actualizar_tablas_dependientes_referencia(plan)

    def _actualizar_tablas_dependientes_referencia(self, plan: dict):
        """Llena las tablas 2 (recomendaciones), 3 (comparación) y 4 (ecuación
        de cuartiles), cuyo contenido depende del grupo de referencia manual
        elegido en self.cmb_plan_grupo_ref."""
        cat_objetivo = plan.get("categoria_objetivo")
        area         = plan.get("area_conocimiento")
        fuente       = plan.get("fuente_indicadores", "bd_interna")
        indicadores_actuales = plan.get("indicadores_actuales", {})

        # — Tabla 2: recomendaciones — una fila por cada opción de producto —
        recomendaciones = plan.get("recomendaciones", [])
        self.tbl_plan_recomendaciones.setRowCount(0)
        for rec in recomendaciones:
            _, label, _ = _IND_INFO_957.get(rec["indicador"], (rec["indicador"], rec["indicador"], None))
            for i, op in enumerate(rec["opciones"]):
                row = self.tbl_plan_recomendaciones.rowCount()
                self.tbl_plan_recomendaciones.insertRow(row)
                self.tbl_plan_recomendaciones.setItem(row, 0, QTableWidgetItem(label if i == 0 else ""))
                self.tbl_plan_recomendaciones.setItem(row, 1, QTableWidgetItem(f"{rec['diferencia']:,.3f}" if i == 0 else ""))
                self.tbl_plan_recomendaciones.setItem(row, 2, QTableWidgetItem(op["producto"]))
                lam = op.get("lambda_unitario")
                if lam is None:
                    self.tbl_plan_recomendaciones.setItem(row, 3, QTableWidgetItem("—"))
                elif fuente == "oficial":
                    self.tbl_plan_recomendaciones.setItem(row, 3, QTableWidgetItem(f"{int(lam)}"))
                else:
                    self.tbl_plan_recomendaciones.setItem(row, 3, QTableWidgetItem(f"{lam:.2f}"))
                cant = op.get("cantidad_minima")
                item_cant = QTableWidgetItem("—" if cant is None else str(cant))
                if i == 0:
                    f = item_cant.font(); f.setBold(True); item_cant.setFont(f)
                self.tbl_plan_recomendaciones.setItem(row, 4, item_cant)
        self.tbl_plan_recomendaciones.resizeRowsToContents()

        # — Tabla 3: comparación con grupos de referencia (más débil, similar y manual) —
        grupo_minimo  = plan.get("grupo_referencia_minimo")
        grupo_similar = plan.get("grupo_referencia_similar")
        grupo_manual  = plan.get("grupo_referencia_manual")
        self.tbl_plan_comparacion.setRowCount(0)
        if grupo_minimo or grupo_similar or grupo_manual:
            partes = []
            if grupo_minimo:
                partes.append(
                    f"<b>Grupo más débil:</b> {grupo_minimo['nombre']} — el grupo con menor "
                    f"puntaje total entre los que ya están en categoría {cat_objetivo} "
                    f"del área '{area}' (el 'listón más bajo' que ya logró esa categoría)."
                )
            if grupo_similar:
                partes.append(
                    f"<b>Grupo de perfil similar:</b> {grupo_similar['nombre']} — el grupo en "
                    f"categoría {cat_objetivo} cuyo perfil de indicadores es más parecido al tuyo."
                )
            if grupo_manual:
                partes.append(
                    f"<b>Grupo seleccionado:</b> {grupo_manual['nombre']} — grupo de tu misma área "
                    f"y categoría objetivo elegido manualmente como referencia."
                )
            self.lbl_plan_grupo_cercano.setText("<br>".join(partes))

            ind_min  = grupo_minimo["indicadores"] if grupo_minimo else {}
            dif_min  = grupo_minimo["diferencias"] if grupo_minimo else {}
            ind_sim  = grupo_similar["indicadores"] if grupo_similar else {}
            dif_sim  = grupo_similar["diferencias"] if grupo_similar else {}
            ind_man  = grupo_manual["indicadores"] if grupo_manual else {}
            dif_man  = grupo_manual["diferencias"] if grupo_manual else {}

            def _set_dif(row, col, dif):
                item = QTableWidgetItem(f"{dif:+,.3f}")
                if dif > 0:
                    item.setForeground(QBrush(QColor("#a02020")))
                elif dif < 0:
                    item.setForeground(QBrush(QColor("#1a7a4a")))
                self.tbl_plan_comparacion.setItem(row, col, item)

            for ind in INDICADORES_957:
                _, label, _ = _IND_INFO_957.get(ind, (ind, ind, None))
                actual = indicadores_actuales.get(ind, 0.0)
                row = self.tbl_plan_comparacion.rowCount()
                self.tbl_plan_comparacion.insertRow(row)
                self.tbl_plan_comparacion.setItem(row, 0, QTableWidgetItem(label))
                self.tbl_plan_comparacion.setItem(row, 1, QTableWidgetItem(f"{actual:,.3f}"))
                if grupo_minimo:
                    self.tbl_plan_comparacion.setItem(row, 2, QTableWidgetItem(f"{ind_min.get(ind, 0.0):,.3f}"))
                    _set_dif(row, 3, dif_min.get(ind, 0.0))
                else:
                    self.tbl_plan_comparacion.setItem(row, 2, QTableWidgetItem("—"))
                    self.tbl_plan_comparacion.setItem(row, 3, QTableWidgetItem("—"))
                if grupo_similar:
                    self.tbl_plan_comparacion.setItem(row, 4, QTableWidgetItem(f"{ind_sim.get(ind, 0.0):,.3f}"))
                    _set_dif(row, 5, dif_sim.get(ind, 0.0))
                else:
                    self.tbl_plan_comparacion.setItem(row, 4, QTableWidgetItem("—"))
                    self.tbl_plan_comparacion.setItem(row, 5, QTableWidgetItem("—"))
                if grupo_manual:
                    self.tbl_plan_comparacion.setItem(row, 6, QTableWidgetItem(f"{ind_man.get(ind, 0.0):,.3f}"))
                    _set_dif(row, 7, dif_man.get(ind, 0.0))
                else:
                    self.tbl_plan_comparacion.setItem(row, 6, QTableWidgetItem("—"))
                    self.tbl_plan_comparacion.setItem(row, 7, QTableWidgetItem("—"))
        else:
            self.lbl_plan_grupo_cercano.setText(
                "No se encontró ningún grupo interno de referencia para comparar."
            )

        # — Tabla 4: "ecuación" de MinCiencias (requisitos mínimos por cuartiles) —
        self._poblar_tabla_cuartiles(plan)

        # — Tabla 5: simulador de productos a generar —
        self._poblar_tabla_simulador(plan)

    def _poblar_tabla_cuartiles(self, plan: dict):
        self.tbl_plan_cuartiles.setRowCount(0)
        self._dibujar_cuartiles_chart(plan)
        requisitos = plan.get("requisitos_minimos_objetivo")
        cat_objetivo = plan.get("categoria_objetivo")
        grupo_ref_cuartiles = plan.get("grupo_referencia_cuartiles")

        if not requisitos:
            self.lbl_plan_cuartiles_info.setText(
                "No se encontró información de distribución nacional (hoja 'cuartiles' de "
                "medicion_957.xlsx) para un grupo de referencia de tu misma área y categoría objetivo."
            )
            return

        _COL_LABEL = {"min": "Mínimo nacional", "q4": "Q4", "q3": "Q3", "q2": "Q2", "max": "Máximo"}
        columna_objetivo = next(iter(requisitos.values()))["columna_objetivo"]
        self.lbl_plan_cuartiles_info.setText(
            f"Distribución nacional tomada como referencia de '{grupo_ref_cuartiles}' "
            f"(misma área '{plan.get('area_conocimiento')}' y categoría {cat_objetivo}). "
            f"Para alcanzar la categoría {cat_objetivo}, MinCiencias exige que cada indicador "
            f"alcance al menos su valor de '{_COL_LABEL.get(columna_objetivo, columna_objetivo)}'."
        )

        for ind in INDICADORES_957:
            fila = requisitos.get(ind)
            if not fila:
                continue
            _, label, _ = _IND_INFO_957.get(ind, (ind, ind, None))
            row = self.tbl_plan_cuartiles.rowCount()
            self.tbl_plan_cuartiles.insertRow(row)
            self.tbl_plan_cuartiles.setItem(row, 0, QTableWidgetItem(label))
            self.tbl_plan_cuartiles.setItem(row, 1, QTableWidgetItem(f"{fila['tu_valor']:,.3f}"))
            for col_idx, key in ((2, "min"), (3, "q4"), (4, "q3"), (5, "q2"), (6, "max")):
                val = fila.get(key)
                item = QTableWidgetItem("—" if val is None else f"{val:,.3f}")
                if key == columna_objetivo:
                    item.setBackground(QBrush(QColor(255, 245, 200)))
                    f = item.font(); f.setBold(True); item.setFont(f)
                self.tbl_plan_cuartiles.setItem(row, col_idx, item)
            req = fila.get("requisito_minimo")
            item_req = QTableWidgetItem("—" if req is None else f"{req:,.3f}")
            f = item_req.font(); f.setBold(True); item_req.setFont(f)
            self.tbl_plan_cuartiles.setItem(row, 7, item_req)
            cumple = fila.get("cumple")
            if cumple is None:
                item_cumple = QTableWidgetItem("—")
            else:
                item_cumple = QTableWidgetItem("Sí" if cumple else "No")
                item_cumple.setForeground(QBrush(QColor("#1a7a4a") if cumple else QColor("#a02020")))
            self.tbl_plan_cuartiles.setItem(row, 8, item_cumple)

    def _poblar_tabla_simulador(self, plan: dict):
        self.tbl_plan_simulador.setRowCount(0)
        self._simulador_grupos = []
        cat_objetivo = plan.get("categoria_objetivo")
        self._simulador_cat_objetivo = cat_objetivo
        requisitos = plan.get("requisitos_minimos_objetivo")

        if not requisitos:
            self.lbl_plan_simulador_veredicto.setText(
                "No hay información de requisitos mínimos (sección 4) para simular."
            )
            self.lbl_plan_simulador_veredicto.setStyleSheet(
                "font-size:11px; font-weight:bold; color:#666; padding:4px 8px; "
                "background:#f0f0f0; border-radius:4px;"
            )
            self._dibujar_simulador_chart()
            return

        sim = getattr(self, "_plan_sim", None)
        nombre = getattr(self, "_plan_nombre", None)
        indicadores_actuales = plan.get("indicadores_actuales", {})
        opciones_por_indicador = (
            sim.opciones_simulador(nombre, requisitos, indicadores_actuales) if sim else {}
        )

        SPAN_COLS = (0, 1, 2, 7)

        def _fila_resumen(label, tu_valor, requisito, texto_extra, span=5):
            row = self.tbl_plan_simulador.rowCount()
            self.tbl_plan_simulador.insertRow(row)
            item_label = QTableWidgetItem(label)
            f = item_label.font(); f.setBold(True); item_label.setFont(f)
            self.tbl_plan_simulador.setItem(row, 0, item_label)
            self.tbl_plan_simulador.setItem(row, 1, QTableWidgetItem(f"{tu_valor:,.3f}"))
            self.tbl_plan_simulador.setItem(row, 2, QTableWidgetItem(f"{requisito:,.3f}"))
            self.tbl_plan_simulador.setItem(row, 3, QTableWidgetItem(texto_extra))
            self.tbl_plan_simulador.setSpan(row, 3, 1, span)
            return row

        for ind in INDICADORES_957:
            fila = requisitos.get(ind)
            if not fila or fila.get("cumple") or fila.get("requisito_minimo") is None:
                continue

            _, label, _ = _IND_INFO_957.get(ind, (ind, ind, None))
            tu_valor   = fila["tu_valor"]
            requisito  = fila["requisito_minimo"]
            info = opciones_por_indicador.get(ind, {})

            if info.get("sin_desglose"):
                _fila_resumen(
                    label, tu_valor, requisito,
                    "MinCiencias no desglosa este indicador por tipo de producto en el "
                    "documento oficial (formación de recurso humano)."
                )
                continue

            productos = info.get("productos", [])
            ratio = info.get("ratio")
            if not productos or not ratio:
                _fila_resumen(
                    label, tu_valor, requisito,
                    "No se puede calibrar el simulador para este indicador: tu grupo no "
                    "tiene productos registrados en esta sección de medicion_957.xlsx."
                )
                continue

            primera_fila = self.tbl_plan_simulador.rowCount()
            filas_cantidad = []
            for i, p in enumerate(productos):
                row = self.tbl_plan_simulador.rowCount()
                self.tbl_plan_simulador.insertRow(row)
                if i == 0:
                    item_label = QTableWidgetItem(label)
                    f = item_label.font(); f.setBold(True); item_label.setFont(f)
                    self.tbl_plan_simulador.setItem(row, 0, item_label)
                    self.tbl_plan_simulador.setItem(row, 1, QTableWidgetItem(f"{tu_valor:,.3f}"))
                    self.tbl_plan_simulador.setItem(row, 2, QTableWidgetItem(f"{requisito:,.3f}"))
                    item_label.setToolTip(
                        f"Factor de conversión de tu grupo (valor oficial ÷ λ propio): "
                        f"{ratio:,.1f}\nλ actual: {info['lambda_actual_total']:,.3f}  →  "
                        f"λ necesario: {info['lambda_objetivo_total']:,.3f}"
                    )

                producto_txt = p["producto"]
                if i == len(productos) - 1 and info.get("mas_disponibles"):
                    producto_txt += f"  (+{info['mas_disponibles']} opciones más)"
                self.tbl_plan_simulador.setItem(row, 3, QTableWidgetItem(producto_txt))
                self.tbl_plan_simulador.setItem(row, 4, QTableWidgetItem(str(p["total_actual"])))
                self.tbl_plan_simulador.setItem(row, 5, QTableWidgetItem(str(p["ventana"])))

                spin = QSpinBox()
                spin.setRange(0, 999)
                spin.setValue(0)
                spin.valueChanged.connect(self._recalcular_simulador)
                self.tbl_plan_simulador.setCellWidget(row, 6, spin)

                filas_cantidad.append({
                    "row": row,
                    "ventana": p["ventana"],
                    "total_actual": p["total_actual"],
                    "lambda_actual": p["lambda_actual"],
                })

            n = len(productos)
            if n > 1:
                for col in SPAN_COLS:
                    self.tbl_plan_simulador.setSpan(primera_fila, col, n, 1)

            self._simulador_grupos.append({
                "fila": primera_fila,
                "ind": ind,
                "label": label,
                "tu_valor": tu_valor,
                "requisito_minimo": requisito,
                "ratio": ratio,
                "lambda_actual_total": info["lambda_actual_total"],
                "filas_cantidad": filas_cantidad,
            })

        self.tbl_plan_simulador.resizeRowsToContents()
        self._recalcular_simulador()

    def _recalcular_simulador(self, *_args):
        cat_objetivo = getattr(self, "_simulador_cat_objetivo", None)
        grupos = getattr(self, "_simulador_grupos", [])

        if not grupos:
            self.lbl_plan_simulador_veredicto.setText(
                f"Todos los indicadores ya cumplen el requisito mínimo para {cat_objetivo}."
                if cat_objetivo else ""
            )
            self.lbl_plan_simulador_veredicto.setStyleSheet(
                "font-size:11px; font-weight:bold; color:#1a7a4a; padding:4px 8px; "
                "background:#e6f4ea; border-radius:4px;"
            )
            self._dibujar_simulador_chart()
            return

        todos_cumplen = True
        faltan_labels = []
        for g in grupos:
            lambda_sim = g["lambda_actual_total"]
            for fc in g["filas_cantidad"]:
                spin = self.tbl_plan_simulador.cellWidget(fc["row"], 6)
                n = spin.value() if spin else 0
                lam_nuevo = math.log(1 + (fc["total_actual"] + n) / fc["ventana"])
                lambda_sim += lam_nuevo - fc["lambda_actual"]

            valor_sim = g["ratio"] * lambda_sim
            cumple = valor_sim >= g["requisito_minimo"]
            if not cumple:
                todos_cumplen = False
                faltan_labels.append(g["label"])

            item_val = QTableWidgetItem(f"{valor_sim:,.3f}  —  {'Sí' if cumple else 'No'}")
            f = item_val.font(); f.setBold(True); item_val.setFont(f)
            item_val.setForeground(QBrush(QColor("#1a7a4a") if cumple else QColor("#a02020")))
            self.tbl_plan_simulador.setItem(g["fila"], 7, item_val)

        if todos_cumplen:
            self.lbl_plan_simulador_veredicto.setText(
                f"Con esta producción, tu grupo SÍ alcanzaría la categoría {cat_objetivo}."
            )
            self.lbl_plan_simulador_veredicto.setStyleSheet(
                "font-size:11px; font-weight:bold; color:#1a7a4a; padding:4px 8px; "
                "background:#e6f4ea; border-radius:4px;"
            )
        else:
            self.lbl_plan_simulador_veredicto.setText(
                f"Con esta producción, tu grupo AÚN NO alcanzaría {cat_objetivo}. "
                f"Indicadores pendientes: " + ", ".join(faltan_labels)
            )
            self.lbl_plan_simulador_veredicto.setStyleSheet(
                "font-size:11px; font-weight:bold; color:#a02020; padding:4px 8px; "
                "background:#fbeaea; border-radius:4px;"
            )

        self._dibujar_simulador_chart()

    def _dibujar_radar_resumen(self, plan: dict):
        """Gráfico de radar: para cada indicador del 957, % de avance de
        TU grupo respecto al umbral de referencia (P25 del área), con un
        anillo de referencia en 100% = ese umbral."""
        self._fig_radar.clear()
        self._canvas_radar.draw()

        indicadores_actuales = plan.get("indicadores_actuales", {})
        umbrales = plan.get("umbrales_referencia", {})

        inds = [ind for ind in INDICADORES_957 if umbrales.get(ind)]
        if len(inds) < 3:
            return

        labels  = [_IND_INFO_957.get(ind, (ind, ind, None))[0] for ind in inds]
        valores = []
        for ind in inds:
            umbral = umbrales.get(ind) or 0.0
            actual = indicadores_actuales.get(ind, 0.0)
            valores.append((actual / umbral * 100.0) if umbral else 0.0)

        n = len(inds)
        angulos = [i / n * 2 * math.pi for i in range(n)]
        angulos += angulos[:1]
        valores_cerrado = valores + valores[:1]
        r_max = max(120.0, max(valores_cerrado) * 1.05)

        ax = self._fig_radar.add_subplot(111, projection="polar")
        ax.set_theta_offset(math.pi / 2)
        ax.set_theta_direction(-1)

        ax.plot(angulos, valores_cerrado, color="#1a5276", linewidth=1.8,
                marker="o", markersize=4, label="Tu grupo")
        ax.fill(angulos, valores_cerrado, color="#1a5276", alpha=0.15)

        ref = [100.0] * (n + 1)
        ax.plot(angulos, ref, color="#cb4335", linewidth=1.2, linestyle="--",
                label="Umbral (P25 del área) = 100%")

        ax.set_xticks(angulos[:-1])
        ax.set_xticklabels(labels, fontsize=8)
        ax.set_ylim(0, r_max)
        ax.tick_params(axis="y", labelsize=7)
        ax.set_title("Tu grupo vs. umbral de referencia", fontsize=9, fontweight="bold", pad=14)
        ax.legend(loc="upper right", bbox_to_anchor=(1.35, 1.12), fontsize=7)

        self._fig_radar.tight_layout()
        self._canvas_radar.draw()

    _PERCENTIL_COLS = (("min", 0), ("q4", 25), ("q3", 50), ("q2", 75), ("max", 100))

    def _percentil_aprox(self, valor: float, fila: dict) -> float:
        """Convierte un valor de indicador a un percentil nacional aproximado
        (0-100) por interpolación lineal a trozos sobre los puntos
        (0,min) (25,q4) (50,q3) (75,q2) (100,max) de la hoja 'cuartiles'."""
        puntos = [(pct, fila.get(key)) for key, pct in self._PERCENTIL_COLS]
        puntos = [(pct, v) for pct, v in puntos if v is not None]
        if not puntos:
            return 50.0
        if valor <= puntos[0][1]:
            return 0.0
        if valor >= puntos[-1][1]:
            return 100.0
        for (p0, v0), (p1, v1) in zip(puntos, puntos[1:]):
            if v0 <= valor <= v1:
                if v1 == v0:
                    return float(p0)
                return p0 + (valor - v0) / (v1 - v0) * (p1 - p0)
        return 50.0

    def _dibujar_cuartiles_chart(self, plan: dict):
        """Gráfico horizontal: una fila por indicador, posicionando 'tu valor'
        en el percentil nacional aproximado (0=mínimo .. 100=máximo) y una
        línea vertical punteada con la meta para la categoría objetivo."""
        from matplotlib.lines import Line2D

        self._fig_cuartiles.clear()
        self._canvas_cuartiles.draw()

        requisitos = plan.get("requisitos_minimos_objetivo")
        if not requisitos:
            return

        filas = [(ind, requisitos[ind]) for ind in INDICADORES_957 if requisitos.get(ind)]
        if not filas:
            return

        _COL_PCT = dict(self._PERCENTIL_COLS)
        ax = self._fig_cuartiles.add_subplot(111)
        y_labels = []

        for i, (ind, fila) in enumerate(filas):
            bm_key, _, _ = _IND_INFO_957.get(ind, (ind, ind, None))
            y = len(filas) - 1 - i
            y_labels.append(bm_key)

            ax.plot([0, 100], [y, y], color="#dddddd", linewidth=4,
                    solid_capstyle="round", zorder=1)

            pct_actual = self._percentil_aprox(fila["tu_valor"], fila)
            cumple = fila.get("cumple")
            color_punto = "#1a7a4a" if cumple else "#a02020"
            ax.scatter([pct_actual], [y], color=color_punto, s=70, zorder=3,
                       edgecolors="white", linewidths=0.8)

            meta_pct = _COL_PCT.get(fila.get("columna_objetivo"))
            if meta_pct is not None:
                ax.plot([meta_pct, meta_pct], [y - 0.35, y + 0.35],
                        color="#cb4335", linewidth=1.6, linestyle="--", zorder=2)

        ax.set_xlim(-2, 102)
        ax.set_ylim(-0.6, len(filas) - 0.4)
        ax.set_yticks(range(len(filas) - 1, -1, -1))
        ax.set_yticklabels(y_labels, fontsize=8)
        ax.set_xticks([0, 25, 50, 75, 100])
        ax.set_xticklabels(["Mín.", "Q4", "Q3", "Q2", "Máx."], fontsize=8)
        ax.set_xlabel("Percentil nacional aproximado", fontsize=8)
        ax.set_title("Tu valor vs. distribución nacional (línea roja = meta de categoría objetivo)",
                      fontsize=8, fontweight="bold")

        legend_elems = [
            Line2D([0], [0], marker="o", color="w", markerfacecolor="#1a7a4a", markersize=8, label="Cumples"),
            Line2D([0], [0], marker="o", color="w", markerfacecolor="#a02020", markersize=8, label="No cumples"),
            Line2D([0], [0], color="#cb4335", linestyle="--", label="Meta categoría objetivo"),
        ]
        ax.legend(handles=legend_elems, loc="lower right", fontsize=6, framealpha=0.9)

        self._fig_cuartiles.tight_layout()
        self._canvas_cuartiles.draw()

    def _dibujar_simulador_chart(self):
        """Gráfico de barras antes/después: para cada indicador con brecha,
        'tu valor actual' vs. el 'valor simulado' (recalculado en vivo según
        las unidades indicadas en la tabla), con una línea de requisito
        mínimo."""
        from matplotlib.lines import Line2D
        from matplotlib.patches import Patch
        import numpy as np

        self._fig_simulador.clear()
        self._canvas_simulador.draw()

        grupos = getattr(self, "_simulador_grupos", [])
        if not grupos:
            return

        labels, tu_valores, requisitos, valores_sim, colores_sim = [], [], [], [], []
        for g in grupos:
            labels.append(_IND_INFO_957.get(g["ind"], (g["label"], g["label"], None))[0])
            tu_valores.append(g["tu_valor"])
            requisitos.append(g["requisito_minimo"])

            lambda_sim = g["lambda_actual_total"]
            for fc in g["filas_cantidad"]:
                spin = self.tbl_plan_simulador.cellWidget(fc["row"], 6)
                n = spin.value() if spin else 0
                lam_nuevo = math.log(1 + (fc["total_actual"] + n) / fc["ventana"])
                lambda_sim += lam_nuevo - fc["lambda_actual"]
            valor_sim = g["ratio"] * lambda_sim
            valores_sim.append(valor_sim)
            colores_sim.append("#1a7a4a" if valor_sim >= g["requisito_minimo"] else "#a02020")

        x = np.arange(len(labels))
        width = 0.35

        ax = self._fig_simulador.add_subplot(111)
        ax.bar(x - width / 2, tu_valores, width, color="#aab7c4", label="Tu valor actual")
        ax.bar(x + width / 2, valores_sim, width, color=colores_sim, label="Valor simulado")

        for xi, req in zip(x, requisitos):
            ax.plot([xi - width, xi + width], [req, req], color="#cb4335",
                    linewidth=1.8, linestyle="--", zorder=4)

        ax.set_xticks(x)
        ax.set_xticklabels(labels, fontsize=8)
        ax.tick_params(axis="y", labelsize=7)
        ax.set_title("Tu valor actual vs. valor simulado (línea roja = requisito mínimo)",
                      fontsize=8, fontweight="bold")

        legend_elems = [
            Patch(facecolor="#aab7c4", label="Tu valor actual"),
            Patch(facecolor="#1a7a4a", label="Simulado: cumple"),
            Patch(facecolor="#a02020", label="Simulado: no cumple"),
            Line2D([0], [0], color="#cb4335", linestyle="--", label="Requisito mínimo"),
        ]
        ax.legend(handles=legend_elems, loc="upper right", fontsize=6, framealpha=0.9)

        self._fig_simulador.tight_layout()
        self._canvas_simulador.draw()

    def _poblar_resumen(self, indicadores: dict):
        partes = []
        for ind, (bm_key, _, _) in _IND_INFO_957.items():
            val   = indicadores.get(ind, 0.0)
            color = "#1a7a4a" if val > 0 else "#aaa"
            partes.append(f'<span style="color:{color}">{bm_key} <b>{val:.3f}</b></span>')
        self.lbl_resumen.setText("  |  ".join(partes))

    def _poblar_arbol(self, productos_por_ind: dict):
        from constants import INDICADORES_957 as _IND957
        self.tree.clear()
        for ind in _IND957:
            prods        = productos_por_ind.get(ind, [])
            bm_key, label, color_activo = _IND_INFO_957.get(ind, (ind, ind, QColor(80, 80, 80)))
            lambda_total = sum(p.get("lambda", 0) for p in prods)

            item_h = QTreeWidgetItem([
                f"{bm_key}  —  {label}",
                f"λ = {lambda_total:.3f}  |  {len(prods)} prod.",
                "", "", "",
            ])
            f = item_h.font(0); f.setBold(True); item_h.setFont(0, f)
            if prods:
                item_h.setForeground(0, QBrush(color_activo))
                item_h.setForeground(1, QBrush(QColor("#1a7a4a")))
            else:
                item_h.setForeground(0, QBrush(QColor("#bbbbbb")))
                item_h.setForeground(1, QBrush(QColor("#cccccc")))

            for prod in sorted(prods, key=lambda p: -(p.get("lambda", 0))):
                child = QTreeWidgetItem([
                    str(prod.get("titulo", ""))[:110],
                    str(prod.get("año", "")),
                    str(prod.get("tipo", "")),
                    f"{prod.get('lambda', 0):.2f}",
                    str(prod.get("fuente", "")),
                ])
                child.setToolTip(0, str(prod.get("titulo", "")))
                child.setForeground(3, QBrush(QColor("#1a5276")))
                item_h.addChild(child)

            self.tree.addTopLevelItem(item_h)
            if prods:
                item_h.setExpanded(True)

    # ── Datos paper 957 ───────────────────────────────────────────────────

    def _cargar_datos_paper(self):
        from utils import obtener_directorio_base
        excel_path = obtener_directorio_base() / "data" / "output" / "medicion_957.xlsx"

        if not excel_path.exists():
            self.lbl_paper.setText(
                "No se encontró medicion_957.xlsx — "
                "ejecute: python src/parse_pdf_medicion.py"
            )
            return

        try:
            df_g   = pd.read_excel(str(excel_path), sheet_name="grupos")
            df_i   = pd.read_excel(str(excel_path), sheet_name="indicadores")
            df_p   = pd.read_excel(str(excel_path), sheet_name="productos")
            df_c   = pd.read_excel(str(excel_path), sheet_name="cuartiles")

            # Pivotar indicadores (valor_indice) para columnas comparativas
            df_i_main = df_i[df_i["indicador"].isin(self._IND_PIVOT)]
            df_pivot  = (
                df_i_main
                .pivot_table(index="grupo", columns="indicador",
                             values="valor_indice", aggfunc="first")
                .reset_index()
            )
            df_pivot.columns.name = None
            df_wide = df_g.merge(df_pivot, left_on="nombre_grupo", right_on="grupo", how="left")

            self._df_paper      = df_g
            self._df_paper_wide = df_wide
            self._df_paper_ind  = df_i
            self._df_paper_prod = df_p
            self._df_paper_cuar = df_c

            areas = sorted(df_g["area_conocimiento"].dropna().unique().tolist())
            self.combo_area_paper.blockSignals(True)
            self.combo_area_paper.clear()
            self.combo_area_paper.addItem("Todas las áreas")
            self.combo_area_paper.addItems(areas)
            self.combo_area_paper.blockSignals(False)

            self._poblar_tablas_categorias(df_wide)
            total = len(df_wide)
            self.lbl_paper.setText(
                f"{total} grupos del paper MinCiencias Conv. 957  "
                f"— haga clic en un grupo para ver su detalle"
            )
        except Exception as e:
            self.lbl_paper.setText(f"Error cargando medicion_957.xlsx: {e}")

    def _filtrar_paper(self):
        if not hasattr(self, "_df_paper_wide") or self._df_paper_wide.empty:
            return
        df = self._df_paper_wide.copy()

        area = self.combo_area_paper.currentText()
        if area != "Todas las áreas":
            df = df[df["area_conocimiento"].str.contains(area, case=False, na=False)]

        buscar = self.le_buscar_paper.text().strip()
        if buscar:
            df = df[df["nombre_grupo"].str.contains(buscar, case=False, na=False)]

        self._poblar_tablas_categorias(df)
        n_total = len(self._df_paper_wide)
        self.lbl_paper.setText(f"{len(df)} de {n_total} grupos")

    def _poblar_tablas_categorias(self, df: pd.DataFrame):
        for cat, tbl in self._cat_tables.items():
            df_cat = (
                df[df["categoria"] == cat]
                .sort_values("indicador_grupo", ascending=False)
                .reset_index(drop=True)
            )
            bg = self._CAT_BG.get(cat, QColor(248, 248, 248))
            tbl.setSortingEnabled(False)
            tbl.setRowCount(len(df_cat))

            for i, (_, row) in enumerate(df_cat.iterrows()):
                nombre   = str(row.get("nombre_grupo", ""))
                director = str(row.get("director", ""))
                ig_val   = row.get("indicador_grupo")
                area_val = str(row.get("area_conocimiento", ""))
                ig_txt   = f"{ig_val:.4f}" if pd.notna(ig_val) else "—"

                vals = [nombre, director, ig_txt]
                for ind in self._IND_PIVOT:
                    v = row.get(ind)
                    vals.append(f"{v:.4f}" if pd.notna(v) else "—")
                vals.append(area_val)

                for col, txt in enumerate(vals):
                    it = QTableWidgetItem(txt)
                    it.setBackground(bg)
                    if col >= 2:
                        it.setTextAlignment(Qt.AlignCenter)
                    tbl.setItem(i, col, it)

            tbl.setSortingEnabled(True)
            # Actualizar título del tab con el conteo
            idx = self._CATS_PAPEL.index(cat)
            self.tabs_cat_paper.setTabText(idx, f"{cat} ({len(df_cat)})")

    def _on_grupo_paper_seleccionado(self):
        tbl_sender = self.sender()
        sel = tbl_sender.selectedItems()
        if not sel:
            return
        row_idx  = sel[0].row()
        nombre   = tbl_sender.item(row_idx, 0).text()
        ig_txt   = tbl_sender.item(row_idx, 2).text() if tbl_sender.columnCount() > 2 else ""
        cat_idx  = self.tabs_cat_paper.currentIndex()
        cat      = self._CATS_PAPEL[cat_idx] if cat_idx < len(self._CATS_PAPEL) else ""

        self.lbl_grupo_paper_sel.setText(
            f"▶  {nombre}   |   Categoría: {cat}   |   IG: {ig_txt}"
        )
        self._poblar_indicadores_paper(nombre)
        self._poblar_productos_paper(nombre)
        self._poblar_cuartiles_paper(nombre)
        self._dibujar_grafica_ig(nombre)

    def _poblar_indicadores_paper(self, nombre_grupo: str):
        if not hasattr(self, "_df_paper_ind") or self._df_paper_ind.empty:
            return
        # Orden del paper: NC_TOP, NC_A, NC_B, ASC, DPC, FRH_A, FRH_B, cohesion, colaboracion
        df_full  = self._df_paper_ind[self._df_paper_ind["grupo"] == nombre_grupo].copy()
        _ORDEN   = self._IND_PIVOT + ["cohesion", "colaboracion"]
        df_full["_orden"] = df_full["indicador"].apply(
            lambda x: _ORDEN.index(x) if x in _ORDEN else 99
        )
        df = df_full.sort_values("_orden")

        self.tbl_paper_ind.setRowCount(len(df))
        _CUARTIL_COLORS = {
            "Q1": QColor(144, 238, 144),
            "Q2": QColor(255, 255, 153),
            "Q3": QColor(255, 204, 102),
            "Q4": QColor(255, 160, 122),
        }

        for i, (_, row) in enumerate(df.iterrows()):
            ind    = str(row.get("indicador", ""))
            label  = self._IND_LABEL_LARGO.get(ind, ind)
            valor  = row.get("valor_indicador")
            v_max  = row.get("valor_maximo")
            v_idx  = row.get("valor_indice")
            pond   = row.get("ponderacion")
            aporte = row.get("valor_ponderado")

            # Calcular cuartil aproximado desde cuartiles sheet si está disponible
            cuartil_txt = "—"
            if hasattr(self, "_df_paper_cuar") and not self._df_paper_cuar.empty:
                # Mapeo aproximado de nombre indicador a cuartil label
                _IND_A_CUARTIL = {
                    "NC_TOP": "Nuevo Conocimiento TOP",
                    "NC_A":   "Nuevo Conocimiento A",
                    "NC_B":   "Nuevo Conocimiento B",
                    "ASC":    "Apropiación Social",
                    "DPC":    "Divulgación Pública",
                    "FRH_A":  "Formación de Recurso Humano A",
                    "FRH_B":  "Formación de Recurso Humano B",
                }
                c_label = _IND_A_CUARTIL.get(ind)
                if c_label and pd.notna(valor):
                    df_cq = self._df_paper_cuar[
                        (self._df_paper_cuar["grupo"] == nombre_grupo)
                        & (self._df_paper_cuar["cuartil"].str.contains(
                            c_label.split()[0], case=False, na=False))
                    ]
                    if not df_cq.empty:
                        r = df_cq.iloc[0]
                        q4_, q3_, q2_, mx_ = r.get("q4"), r.get("q3"), r.get("q2"), r.get("max")
                        v  = valor
                        if pd.notna(q2_) and v >= q2_:
                            cuartil_txt = "Q1"
                        elif pd.notna(q3_) and v >= q3_:
                            cuartil_txt = "Q2"
                        elif pd.notna(q4_) and v >= q4_:
                            cuartil_txt = "Q3"
                        else:
                            cuartil_txt = "Q4"

            def _f4(v):
                return f"{v:,.4f}" if pd.notna(v) else "—"
            def _f2(v):
                return f"{v:.2f}" if pd.notna(v) else "—"

            items = [
                QTableWidgetItem(ind),
                QTableWidgetItem(label),
                QTableWidgetItem(_f4(valor)),
                QTableWidgetItem(_f4(v_max)),
                QTableWidgetItem(_f4(v_idx)),
                QTableWidgetItem(_f2(pond)),
                QTableWidgetItem(_f4(aporte)),
                QTableWidgetItem(cuartil_txt),
            ]
            for col, it in enumerate(items):
                if col >= 2:
                    it.setTextAlignment(Qt.AlignCenter)
                self.tbl_paper_ind.setItem(i, col, it)

            # Colorear la celda del cuartil
            if cuartil_txt in _CUARTIL_COLORS:
                self.tbl_paper_ind.item(i, 7).setBackground(_CUARTIL_COLORS[cuartil_txt])

    def _poblar_productos_paper(self, nombre_grupo: str):
        if not hasattr(self, "_df_paper_prod") or self._df_paper_prod.empty:
            return
        # Ordenar por sección (mismo orden que paper) luego por λ desc
        df = (
            self._df_paper_prod[self._df_paper_prod["grupo"] == nombre_grupo]
            .copy()
        )
        _ORDEN_SEC = {s: i for i, s in enumerate(self._IND_PIVOT)}
        df["_ord"] = df["seccion"].map(_ORDEN_SEC).fillna(99)
        df = df.sort_values(["_ord", "lambda_val"], ascending=[True, False]).reset_index(drop=True)

        self.tbl_paper_prods.setRowCount(len(df))
        _SEC_BG = {
            "NC_TOP": QColor(235, 245, 255),
            "NC_A":   QColor(235, 255, 240),
            "NC_B":   QColor(255, 255, 230),
            "ASC":    QColor(255, 240, 235),
            "DPC":    QColor(245, 245, 245),
            "FRH_A":  QColor(235, 250, 255),
            "FRH_B":  QColor(248, 248, 248),
        }
        prev_sec = None
        for i, (_, row) in enumerate(df.iterrows()):
            sec    = str(row.get("seccion", ""))
            sub    = str(row.get("subtipo", ""))
            total  = row.get("total")
            vent   = row.get("ventana")
            lam    = row.get("lambda_val")

            def _fi(v):
                if not pd.notna(v): return "—"
                return str(int(v)) if isinstance(v, float) and v == int(v) else str(v)

            bg = _SEC_BG.get(sec, QColor(250, 250, 250))
            # Resaltar primera fila de cada sección con texto en negrita
            bold = (sec != prev_sec)
            prev_sec = sec

            vals = [sec, sub, _fi(total), _fi(vent),
                    f"{lam:.4f}" if pd.notna(lam) else "—"]
            for col, txt in enumerate(vals):
                it = QTableWidgetItem(txt)
                it.setBackground(bg)
                if col >= 2:
                    it.setTextAlignment(Qt.AlignCenter)
                if bold and col == 0:
                    ft = it.font(); ft.setBold(True); it.setFont(ft)
                self.tbl_paper_prods.setItem(i, col, it)

    def _poblar_cuartiles_paper(self, nombre_grupo: str):
        if not hasattr(self, "_df_paper_cuar") or self._df_paper_cuar.empty:
            return
        df = self._df_paper_cuar[self._df_paper_cuar["grupo"] == nombre_grupo].copy()
        self.tbl_paper_cuar.setRowCount(len(df))
        _CUARTIL_COLORS = {
            "Q1": QColor(144, 238, 144),
            "Q2": QColor(255, 255, 153),
            "Q3": QColor(255, 204, 102),
            "Q4": QColor(255, 160, 122),
        }
        for i, (_, row) in enumerate(df.iterrows()):
            cuartil = str(row.get("cuartil", ""))
            vg      = row.get("valor_grupo")
            q4_v    = row.get("q4")
            q3_v    = row.get("q3")
            q2_v    = row.get("q2")
            mx_v    = row.get("max")

            # Determinar cuartil del grupo
            pos = "—"
            if pd.notna(vg) and pd.notna(q2_v):
                if vg >= q2_v:
                    pos = "Q1"
                elif pd.notna(q3_v) and vg >= q3_v:
                    pos = "Q2"
                elif pd.notna(q4_v) and vg >= q4_v:
                    pos = "Q3"
                else:
                    pos = "Q4"

            def _f(v):
                return f"{v:,.2f}" if pd.notna(v) else "—"

            items = [
                QTableWidgetItem(cuartil),
                QTableWidgetItem(_f(vg)),
                QTableWidgetItem(_f(q4_v)),
                QTableWidgetItem(_f(q3_v)),
                QTableWidgetItem(_f(q2_v)),
                QTableWidgetItem(_f(mx_v)),
                QTableWidgetItem(pos),
            ]
            for col, it in enumerate(items):
                if col >= 1:
                    it.setTextAlignment(Qt.AlignCenter)
                self.tbl_paper_cuar.setItem(i, col, it)

            if pos in _CUARTIL_COLORS:
                self.tbl_paper_cuar.item(i, 6).setBackground(_CUARTIL_COLORS[pos])

    def _dibujar_grafica_ig(self, nombre_grupo: str):
        """Dibuja la posición del IG del grupo dentro de su área de
        conocimiento, con bandas Q4/Q3/Q2/Q1 y los demás grupos del área
        coloreados por categoría — para contrastar IG vs categoría asignada."""
        import numpy as np

        self._fig_ig.clear()
        self._canvas_ig.draw()

        if not hasattr(self, "_df_paper_wide") or self._df_paper_wide.empty:
            return
        if not hasattr(self, "_df_paper_cuar") or self._df_paper_cuar.empty:
            return

        df = self._df_paper_wide
        fila = df[df["nombre_grupo"] == nombre_grupo]
        if fila.empty:
            return
        fila    = fila.iloc[0]
        area    = fila.get("area_conocimiento")
        ig_sel  = fila.get("indicador_grupo")
        cat_real = str(fila.get("categoria", ""))

        cuar_area = self._df_paper_cuar[
            (self._df_paper_cuar["grupo"] == nombre_grupo)
            & (self._df_paper_cuar["cuartil"].str.contains(
                "rea de conocimiento", case=False, na=False))
        ]
        if cuar_area.empty or pd.isna(ig_sel):
            self.lbl_chart_info.setText(
                f"No hay umbrales de cuartil de IG disponibles para '{area}'."
            )
            return

        r    = cuar_area.iloc[0]
        q4_v, q3_v, q2_v, max_v = r["q4"], r["q3"], r["q2"], r["max"]
        if pd.isna(q2_v):
            self.lbl_chart_info.setText(
                f"No hay umbrales de cuartil de IG disponibles para '{area}'."
            )
            return

        df_area = df[df["area_conocimiento"] == area]
        xlim_max = float(max(max_v, df_area["indicador_grupo"].max(), ig_sel)) * 1.05

        ax = self._fig_ig.add_subplot(111)

        # Bandas de cuartil (mismo orden/colores que el paper)
        ax.axvspan(0,     q4_v,     color="#f5b7b1", alpha=0.6)
        ax.axvspan(q4_v,  q3_v,     color="#fad7a0", alpha=0.6)
        ax.axvspan(q3_v,  q2_v,     color="#fcf3a1", alpha=0.6)
        ax.axvspan(q2_v,  xlim_max, color="#d5f5e3", alpha=0.6)

        for x, label, color in (
            (q4_v / 2,            "Q4", "#922b21"),
            ((q4_v + q3_v) / 2,   "Q3", "#9c640c"),
            ((q3_v + q2_v) / 2,   "Q2", "#7d6608"),
            ((q2_v + xlim_max)/2, "Q1", "#1e8449"),
        ):
            ax.text(x, 1.04, label, ha="center", va="bottom",
                    color=color, fontweight="bold", fontsize=9,
                    transform=ax.get_xaxis_transform())

        # Puntos de los demás grupos del área, por categoría
        _CAT_COLORS = {
            "A1": "#1a7a4a", "A": "#1a5276", "B": "#b9770e",
            "C": "#cb4335", "NR": "#7f8c8d", "SR": "#000000",
        }
        rng = np.random.default_rng(7)
        for cat, color in _CAT_COLORS.items():
            sub = df_area[(df_area["categoria"] == cat)
                           & (df_area["nombre_grupo"] != nombre_grupo)]
            if sub.empty:
                continue
            y = rng.uniform(0.12, 0.85, size=len(sub))
            ax.scatter(sub["indicador_grupo"], y, marker="x",
                       color=color, label=cat, s=30, alpha=0.85)

        # Resaltar el grupo seleccionado
        ax.scatter([ig_sel], [0.5], marker="o", s=160, facecolors="none",
                   edgecolors="#c0392b", linewidths=2.2, zorder=5,
                   label=f"{cat_real} (seleccionado)")
        ax.annotate(nombre_grupo, (ig_sel, 0.5), xytext=(6, 8),
                     textcoords="offset points", fontsize=8,
                     fontweight="bold", color="#c0392b")

        ax.set_xlim(0, xlim_max)
        ax.set_ylim(0, 1)
        ax.set_yticks([])
        ax.set_xlabel("Indicador de grupo (IG)", fontsize=8)
        ax.set_title(area, fontsize=9, fontweight="bold")
        ax.tick_params(axis="x", labelsize=8)
        ax.legend(loc="lower right", fontsize=6, ncol=4, framealpha=0.9)

        self._fig_ig.tight_layout()
        self._canvas_ig.draw()

        # Zona según el IG vs categoría real asignada
        if ig_sel >= q2_v:
            zona = "A1"
        elif ig_sel >= q3_v:
            zona = "A"
        elif ig_sel >= q4_v:
            zona = "B"
        else:
            zona = "C"

        if zona == cat_real:
            self.lbl_chart_info.setText(
                f"IG = {ig_sel:.4f} en '{area}' → cae en zona {zona}, "
                f"coincide con la categoría asignada ({cat_real})."
            )
        else:
            self.lbl_chart_info.setText(
                f"IG = {ig_sel:.4f} en '{area}' → por su valor caería en zona {zona}, "
                f"pero la categoría asignada fue {cat_real}. "
                f"Compare los indicadores para ver qué pesó en la diferencia."
            )


# =============================================================================
# HILO: ACTUALIZACIÓN GRUPLAC (WEBSCRAPING)
# =============================================================================
class ScrapingGrupLACThread(QThread):
    """Descarga el perfil GrupLAC de cada grupo en BD.xlsx y escribe un .xlsx
    por grupo (mismo layout que data/reporte excel.zip) en dest_root."""
    progreso = pyqtSignal(int, int, str)
    finalizado = pyqtSignal(dict, str)
    error = pyqtSignal(str)

    def __init__(self, bd_path, dest_root):
        super().__init__()
        self.bd_path = bd_path
        self.dest_root = dest_root

    def run(self):
        try:
            from gruplac_scraper import ejecutar_scraping

            def _cb(i, total, nombre):
                self.progreso.emit(i, total, nombre)

            resultados = ejecutar_scraping(
                bd_path=self.bd_path, dest_root=self.dest_root,
                progreso_callback=_cb)
            self.finalizado.emit(resultados, self.dest_root)
        except Exception as e:
            self.error.emit(str(e))


class ResumenIAThread(QThread):
    """Genera el resumen narrativo con un modelo local de Ollama
    (streaming, /api/generate -- no requiere conversación ni herramientas,
    es un solo prompt con los datos ya calculados del tablero)."""
    token = pyqtSignal(str)
    done = pyqtSignal()
    error = pyqtSignal(str)

    def __init__(self, prompt, model="qwen2.5:3b", host="http://localhost:11434"):
        super().__init__()
        self.prompt = prompt
        self.model = model
        self.host = host.rstrip("/")

    def run(self):
        import json
        import requests
        try:
            r = requests.post(
                f"{self.host}/api/generate",
                json={"model": self.model, "prompt": self.prompt, "stream": True,
                      "options": {"temperature": 0.3}},
                stream=True, timeout=180,
            )
            if r.status_code == 404:
                self.error.emit(
                    f"El modelo '{self.model}' no está descargado en Ollama. "
                    f"Corra 'ollama pull {self.model}' primero.")
                return
            r.raise_for_status()
            for line in r.iter_lines(decode_unicode=True):
                if not line:
                    continue
                data = json.loads(line)
                if data.get("response"):
                    self.token.emit(data["response"])
                if data.get("done"):
                    break
            self.done.emit()
        except requests.exceptions.ConnectionError:
            self.error.emit(
                "No se pudo conectar con Ollama en localhost:11434. "
                "¿Está corriendo el servicio? (comando: 'ollama serve')")
        except Exception as e:
            self.error.emit(str(e))


# =============================================================================
# PANORAMA GENERAL (dashboard agregado sobre el caché de verificación)
# =============================================================================
class DialogoPanoramaGeneral(QDialog):
    """Agrega en un solo tablero lo que Cumplimiento ya calcula grupo por
    grupo: % de cumplimiento de cada grupo (ordenados de peor a mejor) y el
    total de faltantes por categoría. Lee el mismo caché que Cumplimiento
    (verificacion_faltantes.json), así que ambas pantallas siempre dicen lo
    mismo -- esto no recalcula nada, solo lo muestra junto."""

    UMBRAL_BUENO = 90
    UMBRAL_ALERTA = 70
    UMBRAL_SERIO = 50

    COLOR_BUENO = "#0ca30c"
    COLOR_ALERTA = "#fab219"
    COLOR_SERIO = "#ec835a"
    COLOR_CRITICO = "#d03b3b"

    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("Panorama General — Cumplimiento GrupLAC")
        self.resize(1000, 800)
        layout = QVBoxLayout(self)

        self._df, mensaje, self._ts = cargar_df_faltantes()
        if self._df is None:
            layout.addWidget(QLabel(mensaje))
            btn = QPushButton("Cerrar")
            btn.clicked.connect(self.accept)
            layout.addWidget(btn)
            return

        self._por_grupo = self._calcular_por_grupo()

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        contenido = QWidget()
        lay = QVBoxLayout(contenido)
        lay.addWidget(self._crear_stats())
        lay.addWidget(self._crear_seccion_resumen_ia())
        lay.addWidget(self._crear_grafico_grupos())
        lay.addWidget(self._crear_grafico_categorias())
        lay.addStretch()
        scroll.setWidget(contenido)
        layout.addWidget(scroll)

        lbl_ts = QLabel(f"Fuente: caché de verificación ({self._ts})")
        lbl_ts.setStyleSheet("font-size:10px; color:#898781; padding:4px;")
        layout.addWidget(lbl_ts)

        btn = QPushButton("Cerrar")
        btn.clicked.connect(self.accept)
        layout.addWidget(btn)

    # ── Datos ────────────────────────────────────────────────────────────
    def _calcular_por_grupo(self):
        """Mismo cálculo de % que usa Cumplimiento (_mostrar_cumplimiento en
        VistaSeguimientoGrupos): confirmados / (confirmados + faltantes),
        sobre el mismo conjunto de grupos que ve el selector principal (sin
        semilleros ni códigos crudos de GrupLAC)."""
        cur = self.db.conn.cursor()
        grupos_validos = [r[0] for r in cur.execute('''
            SELECT DISTINCT grupo FROM grupos
            WHERE grupo IS NOT NULL AND grupo != ''
            AND grupo NOT LIKE '%SEMILLERO%' AND grupo NOT LIKE '%Semillero%' AND grupo NOT LIKE '%semillero%'
        ''').fetchall()]

        resultado = []
        for grupo in grupos_validos:
            df_g = self._df[self._df["grupo_original"] == grupo]
            if df_g.empty:
                continue
            n_conf = int((df_g["estado_verificacion"] == "Confirmado en BD (mismo grupo)").sum())
            n_falt = int(df_g["estado_verificacion"].isin(
                ["Faltante real", "Registrado en otro grupo"]).sum())
            total = n_conf + n_falt
            pct = (n_conf / total * 100) if total else 100.0
            resultado.append({"grupo": grupo, "pct": pct, "n_falt": n_falt, "total": total})
        resultado.sort(key=lambda r: r["pct"])
        return resultado

    def _color_para_pct(self, pct):
        if pct >= self.UMBRAL_BUENO:
            return self.COLOR_BUENO
        if pct >= self.UMBRAL_ALERTA:
            return self.COLOR_ALERTA
        if pct >= self.UMBRAL_SERIO:
            return self.COLOR_SERIO
        return self.COLOR_CRITICO

    # ── Tarjetas resumen ─────────────────────────────────────────────────
    def _crear_stats(self):
        datos = self._por_grupo
        w = QWidget()
        lay = QHBoxLayout(w)
        lay.setContentsMargins(4, 8, 4, 8)

        n_grupos = len(datos)
        pct_prom = sum(d["pct"] for d in datos) / n_grupos if n_grupos else 0.0
        n_criticos = sum(1 for d in datos if d["pct"] < self.UMBRAL_SERIO)
        total_faltantes = sum(d["n_falt"] for d in datos)

        tarjetas = [
            ("Grupos con datos", str(n_grupos), "#0b0b0b"),
            ("Cumplimiento promedio", f"{pct_prom:.0f}%", self._color_para_pct(pct_prom)),
            ("Grupos en estado crítico", str(n_criticos),
             self.COLOR_CRITICO if n_criticos else self.COLOR_BUENO),
            ("Total faltantes", str(total_faltantes), "#0b0b0b"),
        ]
        for etiqueta, valor, color in tarjetas:
            tile = QWidget()
            tl = QVBoxLayout(tile)
            lv = QLabel(valor)
            lv.setStyleSheet(f"font-size:26px; font-weight:600; color:{color};")
            ll = QLabel(etiqueta)
            ll.setStyleSheet("font-size:11px; color:#52514e;")
            tl.addWidget(lv)
            tl.addWidget(ll)
            lay.addWidget(tile)
        return w

    # ── Resumen narrativo (IA local, Ollama) ────────────────────────────
    def _crear_seccion_resumen_ia(self):
        w = QWidget()
        v = QVBoxLayout(w)
        v.setContentsMargins(4, 4, 4, 8)

        header = QHBoxLayout()
        self.btn_resumen = QPushButton("🧠 Generar resumen (IA)")
        self.btn_resumen.setToolTip(
            "Usa un modelo local (Ollama, corre en esta misma máquina -- no "
            "sale nada a internet) para redactar en prosa lo que ya está "
            "calculado arriba: no consulta datos nuevos ni inventa cifras.")
        self.btn_resumen.clicked.connect(self._generar_resumen_ia)
        self.btn_resumen.setStyleSheet(
            "QPushButton{background-color:#4a3aa7;color:white;padding:6px 16px;"
            "border-radius:4px;font-weight:bold;font-size:11px;}"
            "QPushButton:hover{background-color:#3a2d87;}"
            "QPushButton:disabled{background-color:#bdc3c7;}")
        header.addWidget(self.btn_resumen)
        self.lbl_resumen_estado = QLabel("")
        self.lbl_resumen_estado.setStyleSheet("font-size:10px; color:#898781;")
        header.addWidget(self.lbl_resumen_estado)
        header.addStretch()
        v.addLayout(header)

        self.txt_resumen = QTextEdit()
        self.txt_resumen.setReadOnly(True)
        self.txt_resumen.setPlaceholderText(
            "Pulsa 'Generar resumen (IA)' para redactar, con un modelo "
            "local, un resumen en prosa de los mismos datos de este tablero.")
        self.txt_resumen.setStyleSheet("font-size:11px;")
        self.txt_resumen.setMaximumHeight(160)
        v.addWidget(self.txt_resumen)
        return w

    def _construir_contexto_ia(self):
        """Arma el texto de datos que se le pasa al modelo -- exactamente lo
        mismo que ya se calculó para las tarjetas y el gráfico, sin
        consultar nada adicional."""
        datos = self._por_grupo
        n_grupos = len(datos)
        pct_prom = sum(d["pct"] for d in datos) / n_grupos if n_grupos else 0.0
        peores = datos[:15]

        df_falt = self._df[self._df["estado_verificacion"].isin(
            ["Faltante real", "Registrado en otro grupo"])]
        por_cat = df_falt["categoria"].value_counts()
        cat_txt = "\n".join(
            f"- {_ETIQUETA_CATEGORIA.get(c, c)}: {n} faltantes"
            for c, n in por_cat.items()) or "- Sin faltantes registrados"

        peores_txt = "\n".join(
            f"- {d['grupo']}: {d['pct']:.0f}% cumplimiento ({d['n_falt']} faltantes)"
            for d in peores) or "- Sin datos"

        return (
            f"Total de grupos con datos: {n_grupos}\n"
            f"Cumplimiento promedio: {pct_prom:.0f}%\n\n"
            f"Faltantes por categoría de producto (todos los grupos):\n{cat_txt}\n\n"
            f"Los {len(peores)} grupos con menor cumplimiento:\n{peores_txt}\n"
        )

    def _generar_resumen_ia(self):
        prompt = (
            "Eres un asistente que ayuda a interpretar datos de cumplimiento "
            "de grupos de investigación universitarios frente a GrupLAC "
            "(plataforma de Minciencias, Colombia). Con los datos a "
            "continuación, escribe un resumen breve en español (máximo 120 "
            "palabras), en prosa corrida, sin listas ni markdown, que "
            "identifique: 1) el panorama general de cumplimiento, 2) qué "
            "categoría de producto concentra más faltantes, y 3) qué 2 o 3 "
            "grupos requieren atención prioritaria. No inventes datos que "
            "no estén en el contexto.\n\n"
            f"DATOS:\n{self._construir_contexto_ia()}\n\nRESUMEN:"
        )

        self.txt_resumen.clear()
        self.btn_resumen.setEnabled(False)
        self.lbl_resumen_estado.setText("Generando con IA local (puede tardar unos segundos)...")

        self._thread_resumen = ResumenIAThread(prompt)
        self._thread_resumen.token.connect(self._on_token_resumen)
        self._thread_resumen.done.connect(self._on_fin_resumen)
        self._thread_resumen.error.connect(self._on_error_resumen)
        self._thread_resumen.start()

    def _on_token_resumen(self, token):
        cursor = self.txt_resumen.textCursor()
        cursor.movePosition(cursor.End)
        self.txt_resumen.setTextCursor(cursor)
        self.txt_resumen.insertPlainText(token)

    def _on_fin_resumen(self):
        self.btn_resumen.setEnabled(True)
        self.lbl_resumen_estado.setText("")

    def _on_error_resumen(self, mensaje):
        self.btn_resumen.setEnabled(True)
        self.lbl_resumen_estado.setText("")
        QMessageBox.critical(self, "Error generando resumen", mensaje)

    # ── Gráfico 1: cumplimiento por grupo ───────────────────────────────
    def _crear_grafico_grupos(self):
        from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
        from matplotlib.figure import Figure
        from matplotlib.patches import Patch

        datos = self._por_grupo
        n = len(datos)
        fig = Figure(figsize=(9, max(4, n * 0.22)), facecolor="#fcfcfb")
        ax = fig.add_subplot(111)
        ax.set_facecolor("#fcfcfb")

        nombres = [d["grupo"] for d in datos]
        valores = [d["pct"] for d in datos]
        colores = [self._color_para_pct(v) for v in valores]

        y = list(range(n))
        ax.barh(y, valores, color=colores, height=0.65, zorder=3)
        ax.set_yticks(y)
        ax.set_yticklabels(nombres, fontsize=7, color="#0b0b0b")
        ax.invert_yaxis()  # peor arriba
        ax.set_xlim(0, 100)
        ax.set_xlabel("% Cumplimiento", fontsize=9, color="#52514e")
        # suptitle (ligado a la FIGURA completa) en vez de ax.set_title
        # (ligado al eje): con nombres de grupo largos el eje queda corrido
        # muy a la derecha y un título anclado a él se sale del lienzo.
        fig.suptitle("Cumplimiento por grupo (peor a mejor)", x=0.01, ha="left",
                      fontsize=11, color="#0b0b0b")
        ax.tick_params(axis="x", colors="#898781", labelsize=8)
        ax.tick_params(axis="y", length=0)
        ax.grid(axis="x", color="#e1e0d9", linewidth=1, zorder=0)
        ax.set_axisbelow(True)
        for spine in ("top", "right", "left"):
            ax.spines[spine].set_visible(False)
        ax.spines["bottom"].set_color("#c3c2b7")

        # Etiquetar solo los peores 10 -- son los que hay que priorizar; el
        # resto ya se lee por el eje y el color (etiquetar los 127 sería
        # ruido, no información).
        for i, d in enumerate(datos[:10]):
            ax.text(d["pct"] + 1.5, i, f'{d["pct"]:.0f}%', va="center",
                     fontsize=7, color="#52514e")

        leyenda = [
            Patch(facecolor=self.COLOR_CRITICO, label=f"Crítico (<{self.UMBRAL_SERIO}%)"),
            Patch(facecolor=self.COLOR_SERIO,
                  label=f"Bajo ({self.UMBRAL_SERIO}-{self.UMBRAL_ALERTA}%)"),
            Patch(facecolor=self.COLOR_ALERTA,
                  label=f"Medio ({self.UMBRAL_ALERTA}-{self.UMBRAL_BUENO}%)"),
            Patch(facecolor=self.COLOR_BUENO, label=f"Bueno (≥{self.UMBRAL_BUENO}%)"),
        ]
        ax.legend(handles=leyenda, loc="lower right", fontsize=7, frameon=False)

        fig.tight_layout(rect=(0, 0, 1, 0.99))
        canvas = FigureCanvas(fig)
        canvas.setMinimumHeight(int(max(400, n * 22)))
        return canvas

    # ── Gráfico 2: faltantes por categoría ──────────────────────────────
    def _crear_grafico_categorias(self):
        from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
        from matplotlib.figure import Figure

        df_falt = self._df[self._df["estado_verificacion"].isin(
            ["Faltante real", "Registrado en otro grupo"])]
        por_cat = df_falt["categoria"].value_counts()
        categorias = [_ETIQUETA_CATEGORIA.get(c, c) for c in por_cat.index]
        valores = list(por_cat.values)

        fig = Figure(figsize=(9, 3), facecolor="#fcfcfb")
        ax = fig.add_subplot(111)
        ax.set_facecolor("#fcfcfb")
        y = list(range(len(categorias)))
        ax.barh(y, valores, color="#2a78d6", height=0.5, zorder=3)
        ax.set_yticks(y)
        ax.set_yticklabels(categorias, fontsize=9, color="#0b0b0b")
        ax.invert_yaxis()
        ax.set_title("Faltantes por categoría (todos los grupos)", fontsize=11,
                      color="#0b0b0b", loc="left")
        ax.tick_params(axis="x", colors="#898781", labelsize=8)
        ax.tick_params(axis="y", length=0)
        ax.grid(axis="x", color="#e1e0d9", linewidth=1, zorder=0)
        ax.set_axisbelow(True)
        for spine in ("top", "right", "left"):
            ax.spines[spine].set_visible(False)
        ax.spines["bottom"].set_color("#c3c2b7")
        maximo = max(valores) if valores else 0
        for i, v in enumerate(valores):
            ax.text(v + maximo * 0.01, i, str(v), va="center", fontsize=8, color="#52514e")

        fig.tight_layout()
        canvas = FigureCanvas(fig)
        canvas.setMinimumHeight(260)
        return canvas


# =============================================================================
# VISTA PRINCIPAL
# =============================================================================
class VistaSeguimientoGrupos(QWidget):

    def __init__(self, db):
        super().__init__()
        self.db = db
        # "reports/excel" es un scrape viejo (dic-2025) que quedó de una
        # corrida anterior -- el botón "Ver Excel" y la comparación de esta
        # pestaña deben abrir/usar siempre el scrape MÁS RECIENTE de
        # data/reporte excel_<fecha> (mismo que usa Cumplimiento), no ese
        # directorio fijo desactualizado.
        carpeta_reciente = _carpeta_gruplac_mas_reciente()
        self.reporte_excel_path = str(carpeta_reciente) if carpeta_reciente else "reports/excel"
        self.datos_comparacion = None
        self.thread_procesamiento = None
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(5)
        layout.setContentsMargins(5, 5, 5, 5)

        # === HEADER ===
        header_layout = QHBoxLayout()

        titulo = QLabel("Seguimiento de Grupos — Internos vs GrupLAC")
        tf = QFont()
        tf.setPointSize(12)
        tf.setBold(True)
        titulo.setFont(tf)
        titulo.setStyleSheet("color: #1a365d;")
        header_layout.addWidget(titulo)
        header_layout.addStretch()

        # Rango de años
        header_layout.addWidget(QLabel("Año desde:"))
        self.spin_anio_desde = QSpinBox()
        self.spin_anio_desde.setRange(1990, 2035)
        self.spin_anio_desde.setValue(2022)
        self.spin_anio_desde.setFixedWidth(70)
        self.spin_anio_desde.setToolTip("Año inicial del rango")
        header_layout.addWidget(self.spin_anio_desde)

        header_layout.addWidget(QLabel("hasta:"))
        self.spin_anio_hasta = QSpinBox()
        self.spin_anio_hasta.setRange(1990, 2035)
        self.spin_anio_hasta.setValue(date.today().year)
        self.spin_anio_hasta.setFixedWidth(70)
        self.spin_anio_hasta.setToolTip("Año final del rango")
        header_layout.addWidget(self.spin_anio_hasta)

        self.spin_anio_desde.valueChanged.connect(lambda v: self.spin_anio_hasta.setMinimum(v))
        self.spin_anio_hasta.valueChanged.connect(lambda v: self.spin_anio_desde.setMaximum(v))

        self.btn_comparar = QPushButton("Comparar")
        self.btn_comparar.clicked.connect(self.iniciar_comparacion)
        self.btn_comparar.setStyleSheet(
            "QPushButton{background-color:#3498db;color:white;padding:8px 20px;"
            "border-radius:4px;font-weight:bold;font-size:12px;}"
            "QPushButton:hover{background-color:#2980b9;}")
        header_layout.addWidget(self.btn_comparar)

        self.btn_duplicados = QPushButton("Personas sin registro (GrupLAC)")
        self.btn_duplicados.setToolTip(
            "Integrantes activos que GrupLAC tiene para un grupo pero que no "
            "calzan con nadie de la BD interna. La verificación de productos "
            "(faltantes/confirmados) está en el panel 'Cumplimiento' de la derecha.")
        self.btn_duplicados.clicked.connect(self.mostrar_duplicados)
        self.btn_duplicados.setStyleSheet(
            "QPushButton{background-color:#8e44ad;color:white;padding:8px 20px;"
            "border-radius:4px;font-weight:bold;font-size:12px;}"
            "QPushButton:hover{background-color:#6c3483;}"
            "QPushButton:disabled{background-color:#bdc3c7;}")
        header_layout.addWidget(self.btn_duplicados)

        # Botón "Panorama General" retirado de la UI a propósito: el resumen
        # narrativo actual (ResumenIAThread) es un solo prompt de una vía
        # contra Ollama local, no un chat interactivo que pueda navegar
        # pestañas o responder preguntas de seguimiento -- se va a rehacer
        # como módulo aparte (chatbot con tool-calling) antes de volver a
        # exponerlo. DialogoPanoramaGeneral y mostrar_panorama_general
        # quedan intactos en el código, solo sin punto de entrada.

        self.btn_scraping_gruplac = QPushButton("Actualizar GrupLAC (Web)")
        self.btn_scraping_gruplac.clicked.connect(self.iniciar_scraping_gruplac)
        self.btn_scraping_gruplac.setToolTip(
            "Descarga de nuevo el perfil GrupLAC de cada grupo (BD.xlsx) desde "
            "scienti.minciencias.gov.co y guarda los .xlsx en una carpeta nueva "
            "con la fecha de hoy, sin tocar los archivos existentes.")
        self.btn_scraping_gruplac.setStyleSheet(
            "QPushButton{background-color:#d35400;color:white;padding:8px 20px;"
            "border-radius:4px;font-weight:bold;font-size:12px;}"
            "QPushButton:hover{background-color:#a04000;}"
            "QPushButton:disabled{background-color:#bdc3c7;}")
        header_layout.addWidget(self.btn_scraping_gruplac)

        layout.addLayout(header_layout)

        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setStyleSheet(
            "QProgressBar{border:2px solid #bdc3c7;border-radius:5px;"
            "text-align:center;height:25px;}"
            "QProgressBar::chunk{background-color:#3498db;}")
        layout.addWidget(self.progress_bar)

        self.lbl_status = QLabel("")
        self.lbl_status.setStyleSheet("color:#7f8c8d;font-size:10px;")
        layout.addWidget(self.lbl_status)

        # === PANEL PRINCIPAL ===
        splitter = QSplitter(Qt.Horizontal)

        # Panel Izquierdo
        panel_izq = QWidget()
        layout_izq = QVBoxLayout(panel_izq)
        layout_izq.setSpacing(3)
        layout_izq.setContentsMargins(3, 3, 3, 3)

        sel_layout = QHBoxLayout()
        sel_layout.addWidget(QLabel("<b>Grupo:</b>"))
        self.combo_grupo = QComboBox()
        self.combo_grupo.setMinimumWidth(250)
        self.combo_grupo.currentTextChanged.connect(self.mostrar_detalle_grupo)
        sel_layout.addWidget(self.combo_grupo, 1)

        # Boton Ver Excel
        self.btn_ver_excel = QPushButton("Ver Excel")
        self.btn_ver_excel.clicked.connect(self._abrir_excel_grupo)
        self.btn_ver_excel.setEnabled(False)
        self.btn_ver_excel.setToolTip(
            "Abre el Excel de GrupLAC del grupo seleccionado\n"
            "mostrando todas las hojas completas (sin truncar)")
        self.btn_ver_excel.setStyleSheet(
            "QPushButton{background-color:#4527A0;color:white;padding:6px 14px;"
            "border-radius:4px;font-weight:bold;font-size:11px;}"
            "QPushButton:hover{background-color:#311B92;}"
            "QPushButton:disabled{background-color:#95a5a6;}")
        sel_layout.addWidget(self.btn_ver_excel)
        layout_izq.addLayout(sel_layout)

        lbl_tabla = QLabel(
            "<b>Comparacion por Grupo</b> — "
            "Internos (BD) vs GrupLAC")
        lbl_tabla.setStyleSheet("color:#1a365d;font-size:10px;")
        layout_izq.addWidget(lbl_tabla)

        self.tabla_comparacion = QTableWidget()
        self.tabla_comparacion.setColumnCount(1)
        self.tabla_comparacion.setHorizontalHeaderLabels(['Grupo'])
        self.tabla_comparacion.horizontalHeader().setStretchLastSection(False)
        self.tabla_comparacion.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tabla_comparacion.setSelectionBehavior(QTableWidget.SelectRows)
        self.tabla_comparacion.verticalHeader().setVisible(False)
        self.tabla_comparacion.setAlternatingRowColors(True)
        self.tabla_comparacion.setStyleSheet("font-size:10px;")
        self.tabla_comparacion.itemSelectionChanged.connect(self.seleccion_tabla_cambiada)
        layout_izq.addWidget(self.tabla_comparacion)

        # Panel Derecho
        panel_der = QWidget()
        layout_der = QVBoxLayout(panel_der)
        layout_der.setSpacing(3)
        layout_der.setContentsMargins(3, 3, 3, 3)

        lbl_det = QLabel("<b>Cumplimiento del Grupo Seleccionado</b>")
        lbl_det.setStyleSheet("color:#1a365d;font-size:11px;")
        layout_der.addWidget(lbl_det)

        self.lbl_cumplimiento = QLabel("Seleccione un grupo.")
        self.lbl_cumplimiento.setStyleSheet("font-size:11px; color:#555; padding:4px; font-weight:bold;")
        layout_der.addWidget(self.lbl_cumplimiento)

        toolbar_cumpl = QHBoxLayout()
        self.btn_cumpl_recargar = QPushButton("🔄 Recargar caché")
        self.btn_cumpl_recargar.setToolTip("Releer data/cache/verificacion_faltantes.json")
        self.btn_cumpl_recargar.setStyleSheet(
            "QPushButton{background-color:#2c3e50;color:white;padding:4px 10px;"
            "border-radius:4px;font-weight:bold;font-size:10px;}"
            "QPushButton:hover{background-color:#1a252f;}")
        self.btn_cumpl_recargar.clicked.connect(self._recargar_cumplimiento)
        toolbar_cumpl.addWidget(self.btn_cumpl_recargar)

        self.btn_cumpl_verificar = QPushButton("Verificar contra GrupLAC (nuevo)")
        self.btn_cumpl_verificar.setToolTip(
            "Compara la BD interna contra los .xlsx scrapeados en la carpeta "
            "'reporte excel_<fecha>' más reciente y regenera el caché de "
            "verificación (data/cache/verificacion_faltantes.json).")
        self.btn_cumpl_verificar.setStyleSheet(
            "QPushButton{background-color:#d35400;color:white;padding:4px 10px;"
            "border-radius:4px;font-weight:bold;font-size:10px;}"
            "QPushButton:hover{background-color:#a04000;}"
            "QPushButton:disabled{background-color:#bdc3c7;}")
        self.btn_cumpl_verificar.clicked.connect(self._iniciar_verificacion_cumplimiento)
        toolbar_cumpl.addWidget(self.btn_cumpl_verificar)

        self.btn_cumpl_exportar = QPushButton("Exportar Excel")
        self.btn_cumpl_exportar.setStyleSheet(
            "QPushButton{background-color:#27ae60;color:white;padding:4px 10px;"
            "border-radius:4px;font-weight:bold;font-size:10px;}"
            "QPushButton:hover{background-color:#1e8449;}"
            "QPushButton:disabled{background-color:#bdc3c7;}")
        self.btn_cumpl_exportar.clicked.connect(self._exportar_excel_cumplimiento)
        self.btn_cumpl_exportar.setEnabled(False)
        toolbar_cumpl.addWidget(self.btn_cumpl_exportar)
        layout_der.addLayout(toolbar_cumpl)

        toolbar_cumpl_estado = QHBoxLayout()
        toolbar_cumpl_estado.addWidget(QLabel("Estado:"))
        self.combo_cumpl_estado = QComboBox()
        self.combo_cumpl_estado.addItem("Pendientes (falta subir)", "pendientes")
        self.combo_cumpl_estado.addItem("Falta", "falta")
        self.combo_cumpl_estado.addItem("Falta (en otro grupo)", "otro_grupo")
        self.combo_cumpl_estado.addItem("Confirmado", "confirmado")
        self.combo_cumpl_estado.addItem("Segundo barrido (revisar)", "segundo_barrido")
        self.combo_cumpl_estado.addItem("Todos", "todos")
        self.combo_cumpl_estado.currentIndexChanged.connect(self._refiltrar_tabla_cumplimiento)
        toolbar_cumpl_estado.addWidget(self.combo_cumpl_estado)
        toolbar_cumpl_estado.addStretch()
        layout_der.addLayout(toolbar_cumpl_estado)

        self.progress_cumplimiento = QProgressBar()
        self.progress_cumplimiento.setVisible(False)
        layout_der.addWidget(self.progress_cumplimiento)

        self.tabla_cumplimiento = QTableWidget()
        cols_cumpl = ["Producto", "Tipo", "Responsable", "Estado", "Grupo encontrado"]
        self.tabla_cumplimiento.setColumnCount(len(cols_cumpl))
        self.tabla_cumplimiento.setHorizontalHeaderLabels(cols_cumpl)
        self.tabla_cumplimiento.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.tabla_cumplimiento.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.tabla_cumplimiento.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.tabla_cumplimiento.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.tabla_cumplimiento.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeToContents)
        self.tabla_cumplimiento.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tabla_cumplimiento.setSelectionBehavior(QTableWidget.SelectRows)
        self.tabla_cumplimiento.verticalHeader().setVisible(False)
        self.tabla_cumplimiento.setAlternatingRowColors(True)
        self.tabla_cumplimiento.setStyleSheet("font-size:10px;")
        self.tabla_cumplimiento.setWordWrap(True)
        layout_der.addWidget(self.tabla_cumplimiento)

        splitter.addWidget(panel_izq)
        splitter.addWidget(panel_der)
        splitter.setSizes([550, 550])
        layout.addWidget(splitter, 1)

        # Estadisticas
        stats = QHBoxLayout()
        stats.setSpacing(20)
        self.lbl_total_grupos = QLabel("Total grupos: 0")
        self.lbl_total_grupos.setStyleSheet("font-weight:bold;color:#2c3e50;")
        stats.addWidget(self.lbl_total_grupos)
        self.lbl_con_gruplac = QLabel("Con GrupLAC: 0")
        self.lbl_con_gruplac.setStyleSheet("font-weight:bold;color:#27ae60;")
        stats.addWidget(self.lbl_con_gruplac)
        self.lbl_rango = QLabel("")
        self.lbl_rango.setStyleSheet("color:#7f8c8d;font-size:10px;")
        stats.addWidget(self.lbl_rango)
        stats.addStretch()
        layout.addLayout(stats)
        self.setLayout(layout)
        self.cargar_grupos_inicial()

    # ── Abrir Excel del grupo seleccionado ────────────────────────────────────
    def _abrir_excel_grupo(self):
        grupo = self.combo_grupo.currentText().strip()
        if not grupo or grupo == "-- Seleccionar grupo --":
            QMessageBox.warning(self, "Aviso", "Seleccione un grupo primero.")
            return

        ruta = None
        if self.datos_comparacion:
            info = self.datos_comparacion.get('rutas_excel', {}).get(grupo, {})
            ruta = info.get('ruta_excel')
            carpeta_path = info.get('carpeta_path')

            if (not ruta or not Path(ruta).exists()) and carpeta_path:
                for f in Path(carpeta_path).glob("*.xlsx"):
                    if not f.name.startswith("~$"):
                        ruta = str(f)
                        break

        # Busqueda por nombre en disco si no se encontro
        if not ruta or not Path(ruta).exists():
            base = Path(self.reporte_excel_path)
            if base.exists():
                gn = unidecode(grupo.lower().strip())
                for d in base.iterdir():
                    if d.is_dir():
                        dn = unidecode(d.name.lower().strip())
                        if dn[:50] == gn[:50] or gn[:50] in dn or dn[:50] in gn[:50]:
                            xlsxs = [f for f in d.glob("*.xlsx") if not f.name.startswith("~$")]
                            if xlsxs:
                                ruta = str(xlsxs[0])
                                break

        if not ruta or not Path(ruta).exists():
            ans = QMessageBox.question(
                self, "Excel no encontrado",
                f"No se encontro el Excel para:\n{grupo}\n\n"
                "Desea buscarlo manualmente?",
                QMessageBox.Yes | QMessageBox.No)
            if ans == QMessageBox.Yes:
                ruta, _ = QFileDialog.getOpenFileName(
                    self, "Seleccionar Excel", self.reporte_excel_path,
                    "Excel (*.xlsx *.xls)")
                if not ruta:
                    return
            else:
                return

        ExcelDialog(ruta, grupo, self).exec_()

    # ── Datos ─────────────────────────────────────────────────────────────────
    def cargar_grupos_inicial(self):
        try:
            cursor = self.db.conn.cursor()
            grupos_bd = cursor.execute('''
                SELECT DISTINCT grupo FROM grupos
                WHERE grupo IS NOT NULL AND grupo != ''
                AND grupo NOT LIKE '%SEMILLERO%' AND grupo NOT LIKE '%Semillero%' AND grupo NOT LIKE '%semillero%'
                ORDER BY grupo
            ''').fetchall()
            if grupos_bd:
                self.combo_grupo.clear()
                self.combo_grupo.addItem("-- Seleccionar grupo --")
                for g in grupos_bd:
                    self.combo_grupo.addItem(g[0])
                self.lbl_status.setText(f"{len(grupos_bd)} grupos cargados")
                self.lbl_total_grupos.setText(f"Total grupos: {len(grupos_bd)}")
            else:
                self.lbl_status.setText("No se encontraron grupos en la base de datos")
        except Exception as e:
            self.lbl_status.setText(f"Error: {str(e)}")

    def iniciar_comparacion(self):
        ad = self.spin_anio_desde.value()
        ah = self.spin_anio_hasta.value()
        if ad > ah:
            QMessageBox.warning(self, "Rango invalido",
                                "El año inicial no puede ser mayor que el año final.")
            return

        self.btn_comparar.setEnabled(False)
        self.btn_ver_excel.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.lbl_status.setText(f"Iniciando comparacion {ad}-{ah}...")
        self.lbl_rango.setText(f"Rango: {ad}-{ah}")

        self.thread_procesamiento = ProcesamientoSeguimientoThread(
            self.db, self.reporte_excel_path, ad, ah)
        self.thread_procesamiento.progreso.connect(self.actualizar_progreso)
        self.thread_procesamiento.finalizado.connect(self.mostrar_comparacion)
        self.thread_procesamiento.error.connect(self.mostrar_error)
        self.thread_procesamiento.start()

    def actualizar_progreso(self, valor, mensaje):
        self.progress_bar.setValue(valor)
        self.lbl_status.setText(mensaje)

    def mostrar_comparacion(self, resultado):
        self.datos_comparacion = resultado
        grupos = resultado['grupos']
        datos_gruplac = resultado['datos_gruplac']

        self.combo_grupo.clear()
        self.combo_grupo.addItem("-- Seleccionar grupo --")
        self.combo_grupo.addItems(grupos)
        self.tabla_comparacion.setRowCount(len(grupos))

        total_con_gruplac = 0

        for row, grupo in enumerate(grupos):
            glac = datos_gruplac.get(grupo, {}).get('total', 0)

            if glac > 0:
                total_con_gruplac += 1

            self.tabla_comparacion.setItem(row, 0, QTableWidgetItem(grupo))

        self.tabla_comparacion.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)

        self.lbl_total_grupos.setText(f"Total grupos: {len(grupos)}")
        self.lbl_con_gruplac.setText(f"Con GrupLAC: {total_con_gruplac}")

        self.btn_comparar.setEnabled(True)
        self.btn_ver_excel.setEnabled(True)
        self.progress_bar.setVisible(False)
        ad = self.spin_anio_desde.value()
        ah = self.spin_anio_hasta.value()
        self.lbl_status.setText(
            f"Completado [{ad}-{ah}] — {total_con_gruplac} grupos con GrupLAC")

    def mostrar_error(self, mensaje):
        QMessageBox.critical(self, "Error", f"Error al comparar:\n{mensaje}")
        self.btn_comparar.setEnabled(True)
        self.progress_bar.setVisible(False)
        self.lbl_status.setText("Error en la comparacion")

    # ── Actualización GrupLAC (webscraping) ─────────────────────────────────
    def iniciar_scraping_gruplac(self):
        bd_path = obtener_directorio_base() / "data" / "BD.xlsx"
        if not bd_path.exists():
            QMessageBox.warning(self, "BD.xlsx no encontrado",
                                 f"No se encontró {bd_path}.")
            return

        fecha = date.today().strftime("%Y%m%d")
        dest_root = obtener_directorio_base() / "data" / f"reporte excel_{fecha}"
        ans = QMessageBox.question(
            self, "Actualizar GrupLAC",
            "Esto descarga de nuevo el perfil GrupLAC de cada grupo listado en "
            "BD.xlsx directamente desde scienti.minciencias.gov.co (~128 "
            "grupos, unos 4-5 minutos, con espera entre cada uno para no "
            "saturar el servidor).\n\n"
            f"Los archivos nuevos quedarán en una carpeta aparte:\n{dest_root}\n\n"
            "No se borra ni modifica nada de lo que ya existe. ¿Continuar?",
            QMessageBox.Yes | QMessageBox.No)
        if ans != QMessageBox.Yes:
            return

        self.btn_scraping_gruplac.setEnabled(False)
        self.btn_comparar.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.lbl_status.setText("Iniciando actualización GrupLAC...")

        self.thread_scraping = ScrapingGrupLACThread(str(bd_path), str(dest_root))
        self.thread_scraping.progreso.connect(self._progreso_scraping_gruplac)
        self.thread_scraping.finalizado.connect(self._fin_scraping_gruplac)
        self.thread_scraping.error.connect(self._error_scraping_gruplac)
        self.thread_scraping.start()

    def _progreso_scraping_gruplac(self, i, total, nombre):
        pct = int(i / total * 100) if total else 0
        self.progress_bar.setValue(pct)
        self.lbl_status.setText(f"[{i}/{total}] Descargando GrupLAC: {nombre}")

    def _fin_scraping_gruplac(self, resultados, dest_root):
        self.btn_scraping_gruplac.setEnabled(True)
        self.btn_comparar.setEnabled(True)
        self.progress_bar.setVisible(False)
        ok = len(resultados.get("ok", []))
        errores = resultados.get("error", [])
        self.lbl_status.setText(
            f"Actualización GrupLAC completa — {ok} ok, {len(errores)} con error")
        msg = (f"Actualización GrupLAC completa.\n\nOK: {ok}\n"
               f"Con error: {len(errores)}\n\nGuardado en:\n{dest_root}")
        if errores:
            detalle = "\n".join(f"- {n}: {m}" for n, m in errores[:15])
            if len(errores) > 15:
                detalle += f"\n... y {len(errores) - 15} más"
            msg += f"\n\nGrupos con error:\n{detalle}"
        QMessageBox.information(self, "Actualización GrupLAC", msg)

    def _error_scraping_gruplac(self, mensaje):
        self.btn_scraping_gruplac.setEnabled(True)
        self.btn_comparar.setEnabled(True)
        self.progress_bar.setVisible(False)
        QMessageBox.critical(self, "Error",
                              f"Error en la actualización GrupLAC:\n{mensaje}")

    def seleccion_tabla_cambiada(self):
        sel = self.tabla_comparacion.selectedItems()
        if sel:
            fila = sel[0].row()
            grupo = self.tabla_comparacion.item(fila, 0).text()
            self.combo_grupo.setCurrentText(grupo)

    # ── Detalle ───────────────────────────────────────────────────────────────
    def mostrar_detalle_grupo(self, grupo):
        if not grupo or grupo == "-- Seleccionar grupo --" or not self.datos_comparacion:
            self.tabla_cumplimiento.setRowCount(0)
            self.lbl_cumplimiento.setText("Seleccione un grupo.")
            self.lbl_cumplimiento.setStyleSheet("font-size:11px; color:#555; padding:4px; font-weight:bold;")
            return
        self._mostrar_cumplimiento(grupo)

    def _mostrar_cumplimiento(self, grupo):
        """Muestra el % de cumplimiento del grupo y, en la tabla de abajo,
        el detalle de verificación contra GrupLAC para ese grupo -- misma
        fuente (data/cache/verificacion_faltantes.json) que antes usaba el
        popup 'Duplicados' (ahora retirado), así que ambas pantallas dicen
        siempre lo mismo. El grupo viene de combo_grupo, que ya excluye
        semilleros y códigos crudos de GrupLAC -- por eso acá tampoco
        aparecen."""
        df, mensaje, ts = cargar_df_faltantes()
        if df is None:
            self._df_cumplimiento_grupo = None
            self._ts_cumplimiento = ""
            self.tabla_cumplimiento.setRowCount(0)
            self.lbl_cumplimiento.setText(mensaje)
            color = "#555" if "vacío" in mensaje else "#a00"
            self.lbl_cumplimiento.setStyleSheet(f"font-size:11px; color:{color}; padding:4px;")
            self.btn_cumpl_exportar.setEnabled(False)
            return

        df_grupo = df[df["grupo_original"] == grupo]
        if df_grupo.empty:
            self._df_cumplimiento_grupo = None
            self._ts_cumplimiento = ts
            self.tabla_cumplimiento.setRowCount(0)
            self.lbl_cumplimiento.setText(
                f"Sin datos de verificación para '{grupo}' en el caché ({ts}).")
            self.lbl_cumplimiento.setStyleSheet("font-size:11px; color:#555; padding:4px;")
            self.btn_cumpl_exportar.setEnabled(False)
            return

        self._df_cumplimiento_grupo = df_grupo
        self._ts_cumplimiento = ts
        self.btn_cumpl_exportar.setEnabled(True)

        es_confirmado = df_grupo["estado_verificacion"] == "Confirmado en BD (mismo grupo)"
        faltantes = df_grupo[df_grupo["estado_verificacion"].isin(
            ["Faltante real", "Registrado en otro grupo"])]
        n_faltantes = len(faltantes)
        n_confirmados = int(es_confirmado.sum())
        total_relevante = n_confirmados + n_faltantes
        pct = (n_confirmados / total_relevante * 100) if total_relevante else 100.0

        self.lbl_cumplimiento.setText(
            f"Cumplimiento: {pct:.0f}% ({n_confirmados}/{total_relevante} confirmados) — "
            f"Faltantes: {n_faltantes}  [Caché: {ts}]"
        )
        color = "#1a7a4a" if n_faltantes == 0 else "#a00"
        self.lbl_cumplimiento.setStyleSheet(
            f"font-size:11px; color:{color}; padding:4px; font-weight:bold;")

        self._poblar_tabla_cumplimiento(df_grupo)

    def _poblar_tabla_cumplimiento(self, df_grupo):
        filtro_key = self.combo_cumpl_estado.currentData()
        estados = _FILTROS_ESTADO_CUMPLIMIENTO.get(filtro_key)
        df = df_grupo if estados is None else df_grupo[
            df_grupo["estado_verificacion"].isin(estados)]

        COLOR_FALTANTE = QColor(255, 200, 200)
        COLOR_OTRO = QColor(200, 220, 255)
        COLOR_CONFIRMADO = QColor(210, 245, 215)
        COLOR_REVISION = QColor(255, 245, 200)
        colores_estado = {
            "Faltante real": COLOR_FALTANTE,
            "Registrado en otro grupo": COLOR_OTRO,
            "Confirmado en BD (mismo grupo)": COLOR_CONFIRMADO,
            "Segundo barrido - mismo grupo": COLOR_REVISION,
            "Segundo barrido - otro grupo": COLOR_REVISION,
        }

        self.tabla_cumplimiento.setRowCount(len(df))
        for i, (_, row) in enumerate(df.iterrows()):
            estado_raw = str(row.get("estado_verificacion", ""))
            tipo = _ETIQUETA_CATEGORIA.get(str(row.get("categoria", "")), row.get("categoria", ""))
            estado = _ETIQUETAS_ESTADO_CUMPLIMIENTO.get(estado_raw, estado_raw)
            color = colores_estado.get(estado_raw, QColor(255, 255, 255))
            items = [
                QTableWidgetItem(str(row.get("producto", ""))),
                QTableWidgetItem(str(tipo)),
                QTableWidgetItem(str(row.get("responsable", ""))),
                QTableWidgetItem(estado),
                QTableWidgetItem(str(row.get("grupo_encontrado", ""))),
            ]
            for col, item in enumerate(items):
                item.setBackground(color)
                self.tabla_cumplimiento.setItem(i, col, item)

    def _refiltrar_tabla_cumplimiento(self):
        if getattr(self, "_df_cumplimiento_grupo", None) is None:
            return
        self._poblar_tabla_cumplimiento(self._df_cumplimiento_grupo)

    def _recargar_cumplimiento(self):
        grupo = self.combo_grupo.currentText()
        if not grupo or grupo == "-- Seleccionar grupo --":
            QMessageBox.information(self, "Recargar caché", "Seleccione un grupo primero.")
            return
        self._mostrar_cumplimiento(grupo)

    def _iniciar_verificacion_cumplimiento(self):
        if not self.db:
            QMessageBox.warning(self, "Sin BD", "No hay conexión a la base de datos interna.")
            return
        carpeta = _carpeta_gruplac_mas_reciente()
        if not carpeta:
            QMessageBox.warning(
                self, "No hay datos GrupLAC",
                "No se encontró ninguna carpeta 'data/reporte excel_<fecha>'. "
                "Primero corra 'Actualizar GrupLAC (Web)'.")
            return

        anio_desde = self.spin_anio_desde.value()
        anio_hasta = self.spin_anio_hasta.value()

        ans = QMessageBox.question(
            self, "Verificar contra GrupLAC",
            f"Esto compara la BD interna (año {anio_desde}-{anio_hasta}) contra "
            f"los perfiles scrapeados en:\n{carpeta}\n\n"
            "y reemplaza el caché de verificación actual (data/cache/"
            "verificacion_faltantes.json). ¿Continuar?",
            QMessageBox.Yes | QMessageBox.No)
        if ans != QMessageBox.Yes:
            return

        self.btn_cumpl_verificar.setEnabled(False)
        self.btn_cumpl_recargar.setEnabled(False)
        self.progress_cumplimiento.setVisible(True)
        self.progress_cumplimiento.setValue(0)
        self.lbl_cumplimiento.setText("Verificando contra GrupLAC...")

        self.thread_verificacion_cumplimiento = VerificacionGrupLACThread(
            self.db, str(carpeta), anio_desde, anio_hasta)
        self.thread_verificacion_cumplimiento.progreso.connect(
            self._progreso_verificacion_cumplimiento)
        self.thread_verificacion_cumplimiento.finalizado.connect(
            self._fin_verificacion_cumplimiento)
        self.thread_verificacion_cumplimiento.error.connect(
            self._error_verificacion_cumplimiento)
        self.thread_verificacion_cumplimiento.start()

    def _progreso_verificacion_cumplimiento(self, i, total):
        pct = int(i / total * 100) if total else 0
        self.progress_cumplimiento.setValue(pct)
        self.lbl_cumplimiento.setText(f"Verificando... [{i}/{total}]")

    def _fin_verificacion_cumplimiento(self, resumen):
        self.btn_cumpl_verificar.setEnabled(True)
        self.btn_cumpl_recargar.setEnabled(True)
        self.progress_cumplimiento.setVisible(False)
        grupo = self.combo_grupo.currentText()
        if grupo and grupo != "-- Seleccionar grupo --":
            self._mostrar_cumplimiento(grupo)
        QMessageBox.information(
            self, "Verificación completa",
            f"Total comparado: {resumen['total']}\n"
            f"Confirmados: {resumen['confirmados']}\n"
            f"Faltantes: {resumen['faltantes']}\n"
            f"Segundo barrido (revisión manual): {resumen['revision']}\n\n"
            f"Guardado en:\n{resumen['ruta_cache']}")

    def _error_verificacion_cumplimiento(self, mensaje):
        self.btn_cumpl_verificar.setEnabled(True)
        self.btn_cumpl_recargar.setEnabled(True)
        self.progress_cumplimiento.setVisible(False)
        QMessageBox.critical(self, "Error", f"Error en la verificación:\n{mensaje}")

    def _exportar_excel_cumplimiento(self):
        """Exporta a Excel el detalle de verificación del grupo seleccionado
        (mismo formato que antes generaba el popup 'Duplicados'), respetando
        el filtro de Estado activo."""
        if getattr(self, "_df_cumplimiento_grupo", None) is None:
            QMessageBox.information(self, "Exportar", "No hay datos para exportar.")
            return

        from comparador_faltantes import ComparadorFaltantes

        filtro_key = self.combo_cumpl_estado.currentData()
        estados = _FILTROS_ESTADO_CUMPLIMIENTO.get(filtro_key)
        df = self._df_cumplimiento_grupo
        if estados is not None:
            df = df[df["estado_verificacion"].isin(estados)]

        if df.empty:
            QMessageBox.information(self, "Exportar", "No hay productos para exportar con el filtro actual.")
            return

        base_dir = obtener_directorio_base()
        reports_dir = base_dir / "reports" / "excel"
        reports_dir.mkdir(parents=True, exist_ok=True)

        grupo = self.combo_grupo.currentText()
        timestamp = date.today().strftime("%Y%m%d")
        sufijo = f"_{limpiar_nombre_archivo(grupo)[:20]}" if grupo else ""
        nombre_sugerido = f"REPORTE_FALTANTES_{timestamp}{sufijo}.xlsx"

        ruta_str, _ = QFileDialog.getSaveFileName(
            self, "Guardar reporte de faltantes",
            str(reports_dir / nombre_sugerido), "Excel (*.xlsx)")
        if not ruta_str:
            return
        ruta = Path(ruta_str)
        if ruta.suffix.lower() != ".xlsx":
            ruta = ruta.with_suffix(".xlsx")

        try:
            ComparadorFaltantes.generar_reporte_excel(df, ruta)
            QMessageBox.information(
                self, "Exportado",
                f"Reporte guardado en:\n{ruta}\n\n"
                f"Total productos: {len(df)}\n"
                f"Faltantes: {len(df[df['es_faltante']])}\n"
                f"Segundo barrido: {len(df[df['necesita_revision']])}"
            )
        except Exception as e:
            QMessageBox.critical(self, "Error", f"No se pudo generar el reporte:\n{e}")

    # ── Exportar Excel ────────────────────────────────────────────────────────
    # ── Exportar PDF ──────────────────────────────────────────────────────────
    # ── Diagnostico ───────────────────────────────────────────────────────────
    def mostrar_duplicados(self):
        dlg = DialogoDuplicados(self.datos_comparacion or {}, self.db, self)
        dlg.exec_()

    def mostrar_panorama_general(self):
        if not self.db:
            QMessageBox.warning(self, "Sin datos", "No hay base de datos disponible.")
            return
        dlg = DialogoPanoramaGeneral(self.db, self)
        dlg.exec_()

    def _cargar_cache_faltantes(self):
        """Carga el caché de verificación de faltantes (panel Cumplimiento)."""
        import json as _json
        cache_file = obtener_directorio_base() / "data" / "cache" / "verificacion_faltantes.json"
        if not cache_file.exists():
            return None
        try:
            with open(cache_file, encoding="utf-8") as f:
                cache = _json.load(f)
            return cache.get("data", [])
        except Exception:
            return None

    def export_context_for_llm(self, modo="general"):
        if not getattr(self, "datos_comparacion", None):
            return "[Seguimiento Grupos]\nAun no hay datos. Ejecuta Comparar primero.\n"

        grupos = self.datos_comparacion.get("grupos", [])
        datos_gruplac = self.datos_comparacion.get("datos_gruplac", {})
        ad = self.spin_anio_desde.value()
        ah = self.spin_anio_hasta.value()

        registros = self._cargar_cache_faltantes()
        if registros is None:
            return (
                f"[Seguimiento Grupos]\nRango: {ad}-{ah}\nTotal grupos: {len(grupos)}\n\n"
                "Sin datos de verificación de faltantes. Seleccione un grupo y pulse "
                "'Verificar contra GrupLAC (nuevo)' en el panel Cumplimiento.\n"
            )

        por_grupo = defaultdict(lambda: {'reales': [], 'en_otros': []})
        for r in registros:
            estado = r.get("estado_verificacion", "")
            grupo = r.get("grupo_original", "")
            if estado == "Faltante real":
                por_grupo[grupo]['reales'].append(r)
            elif estado == "Registrado en otro grupo":
                por_grupo[grupo]['en_otros'].append(r)

        if modo == "grupo":
            grupo_sel = self.combo_grupo.currentText().strip()
            if not grupo_sel or grupo_sel not in grupos:
                return "[Seguimiento Grupos]\nNo hay grupo valido seleccionado.\n"
            clasif = por_grupo.get(grupo_sel, {'reales': [], 'en_otros': []})
            reales = clasif['reales']
            en_otros = clasif['en_otros']
            falt_txt = [
                f"{i}. [{r.get('categoria','')}] {r.get('producto','')}"
                for i, r in enumerate(reales[:60], 1)
            ]
            if len(reales) > 60:
                falt_txt.append(f"... ({len(reales)-60} mas)")
            return (
                f"[Seguimiento Grupos - Detalle]\nGrupo: {grupo_sel}\nRango: {ad}-{ah}\n"
                f"GrupLAC total: {datos_gruplac.get(grupo_sel, {}).get('total', 0)}\n"
                f"Faltantes reales: {len(reales)}\nEn otro grupo: {len(en_otros)}\n\n"
                "Faltantes reales:\n" +
                ("\n".join(falt_txt) if falt_txt else "Sin faltantes") + "\n"
            )

        ranking = sorted(
            ((g, len(por_grupo.get(g, {}).get('reales', []))) for g in grupos),
            key=lambda x: x[1], reverse=True
        )
        top = "\n".join(f"- {g}: {n}" for g, n in ranking[:15]) or "Sin datos."
        tr = sum(len(v['reales']) for v in por_grupo.values())
        to = sum(len(v['en_otros']) for v in por_grupo.values())
        return (
            f"[Seguimiento Grupos - Resumen]\nRango: {ad}-{ah}\n"
            f"Total grupos: {len(grupos)}\nFaltantes reales: {tr}\nEn otro grupo: {to}\n\n"
            f"Top grupos faltantes:\n{top}\n"
        )

"""
Vista Visor GrupLAC – Conv. 957  |  v5 – PyQt5 embebible
─────────────────────────────────────────────────────────
• Lee de gruplac_957.db generado por build_gruplac_dbs.py
• Muestra Excel original completo (sin truncar, todas las celdas)
• Checklist 100% MANUAL: la persona chequea y pulsa "Guardar Estado"
• PDF se genera SOLO desde el estado guardado
• Si una categoría no está chuleada → aparece como Pendiente en PDF
• Versión PyQt5 para integrar como pestaña en main_10.py

Dependencias:
    pip install PyQt5 pandas openpyxl reportlab
"""

import os
import re
import shutil
import sqlite3
import unicodedata
from datetime import datetime

import pandas as pd

from PyQt5.QtCore import Qt, QAbstractTableModel, QModelIndex
from PyQt5.QtGui import QColor, QBrush
from PyQt5.QtWidgets import (
    QWidget, QDialog,
    QHBoxLayout, QVBoxLayout, QSplitter,
    QLabel, QLineEdit, QListWidget, QListWidgetItem,
    QComboBox, QPushButton, QTableView, QMessageBox,
    QGroupBox, QFrame, QTabWidget, QSpinBox,
    QFileDialog, QCheckBox, QHeaderView, QAbstractItemView,
)

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font as XLFont, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer,
        Table, TableStyle, HRFlowable,
    )
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_LEFT, TA_RIGHT
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

# ─────────────────────────────────────────────────────────────────────────────
# Configuración
# ─────────────────────────────────────────────────────────────────────────────
DEFAULT_DB_PATH  = "data/db/gruplac_957.db"
DEFAULT_DATA_DIR = "webscrapinzip"

CATEGORIES_957 = [
    "1. Generación de nuevo conocimiento",
    "2. Desarrollo tecnológico e innovación",
    "3. Apropiación social del conocimiento",
    "4. Divulgación pública de la ciencia",
    "5. Formación de recurso humano en CTeI",
]

MONTHS_ES = {
    1:"enero", 2:"febrero", 3:"marzo",    4:"abril",
    5:"mayo",  6:"junio",   7:"julio",    8:"agosto",
    9:"septiembre", 10:"octubre", 11:"noviembre", 12:"diciembre",
}

# ── Paleta ────────────────────────────────────────────────────────────────────
C_BLUE_D = "#003087"
C_BLUE_M = "#1565C0"
C_BLUE_L = "#E8F0FE"
C_BLUE_P = "#F0F4FF"
C_ORG    = "#E65100"
C_ORG_S  = "#FF8F00"
C_ORG_P  = "#FFF3E0"
C_WHITE  = "#FFFFFF"
C_GREY   = "#555555"
C_GRN    = "#1B5E20"
C_GRN_B  = "#E8F5E9"
C_RED    = "#B71C1C"

CAT_BG = {
    CATEGORIES_957[0]: QColor("#E8F0FE"),
    CATEGORIES_957[1]: QColor("#E3F2FD"),
    CATEGORIES_957[2]: QColor("#FFF3E0"),
    CATEGORIES_957[3]: QColor("#FFE0B2"),
    CATEGORIES_957[4]: QColor("#EDE7F6"),
}
CAT_FG = {
    CATEGORIES_957[0]: QColor(C_BLUE_D),
    CATEGORIES_957[1]: QColor(C_BLUE_M),
    CATEGORIES_957[2]: QColor(C_ORG),
    CATEGORIES_957[3]: QColor(C_ORG_S),
    CATEGORIES_957[4]: QColor("#4527A0"),
}

TABLE_COLS = [
    ("Categoría 957",    "categoria_957"),
    ("Subcategoría",     "subcategoria_957"),
    ("Tipo de Producto", "tipo_producto_957"),
    ("Título",           "titulo"),
    ("Año",              "anio"),
    ("Autores",          "autores"),
    ("Pestaña",          "pestana"),
]

# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────
def _decode(s):
    return re.sub(r'#U([0-9a-fA-F]{4})',
                  lambda m: chr(int(m.group(1), 16)), str(s))

def _norm(s):
    s = _decode(s).upper().strip()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    return re.sub(r'\s+', ' ', re.sub(r'[^A-Z0-9 ]', ' ', s)).strip()

def find_excel(gname, data_dir):
    if not os.path.isdir(data_dir):
        return None
    tgt = _norm(gname)
    idx = {}
    for e in os.listdir(data_dir):
        fp = os.path.join(data_dir, e)
        if os.path.isdir(fp):
            idx[_norm(e)] = fp
    if tgt in idx:
        return _first_xlsx(idx[tgt])
    for nk, fp in idx.items():
        if tgt in nk or nk in tgt:
            x = _first_xlsx(fp)
            if x: return x
    tw = set(tgt.split())
    best = max(idx.items(),
               key=lambda kv: len(tw & set(kv[0].split())),
               default=(None, None))
    if best[0] and len(tw & set(best[0].split())) >= 2:
        return _first_xlsx(best[1])
    return None

def _first_xlsx(folder):
    try:
        for f in os.listdir(folder):
            if f.lower().endswith('.xlsx'):
                return os.path.join(folder, f)
    except Exception:
        pass
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Modelo Qt — texto completo sin truncar  (PyQt5)
# ─────────────────────────────────────────────────────────────────────────────
class DFModel(QAbstractTableModel):
    def __init__(self, df, cat_col=None):
        super().__init__()
        self._df  = df.reset_index(drop=True)
        self._cat = cat_col

    def set_df(self, df):
        self.beginResetModel()
        self._df = df.reset_index(drop=True)
        self.endResetModel()

    def rowCount(self, p=QModelIndex()):
        return 0 if p.isValid() else len(self._df)

    def columnCount(self, p=QModelIndex()):
        return 0 if p.isValid() else len(self._df.columns)

    def data(self, idx, role=Qt.DisplayRole):
        if not idx.isValid():
            return None
        val  = self._df.iat[idx.row(), idx.column()]
        text = "" if (val is None or (isinstance(val, float) and pd.isna(val))) \
               else str(val)
        if role == Qt.DisplayRole:
            return text
        if role == Qt.ToolTipRole:
            return text
        if role == Qt.BackgroundRole and self._cat:
            if self._cat in self._df.columns:
                cat = str(self._df.iat[
                    idx.row(), self._df.columns.get_loc(self._cat)])
                if cat in CAT_BG:
                    return QBrush(CAT_BG[cat])
        return None

    def headerData(self, sec, ori, role=Qt.DisplayRole):
        if role != Qt.DisplayRole:
            return None
        if ori == Qt.Horizontal:
            return str(self._df.columns[sec])
        return str(sec + 1)


# ─────────────────────────────────────────────────────────────────────────────
# Repositorio
# ─────────────────────────────────────────────────────────────────────────────
class Repo:
    def __init__(self, db_path):
        if not os.path.exists(db_path):
            raise FileNotFoundError(f"No se encontró la BD: {db_path}")
        tmp = os.path.join(os.path.dirname(os.path.abspath(db_path)),
                           "_tmp_viewer957.db")
        shutil.copy(db_path, tmp)
        self.conn = sqlite3.connect(tmp)
        self._tmp = tmp

    def groups(self):
        return pd.read_sql_query(
            "SELECT id, nombre FROM grupos ORDER BY nombre;", self.conn)

    def cat_counts(self):
        return pd.read_sql_query(
            """SELECT grupo_id, categoria_957, COUNT(*) n
               FROM productos_957
               WHERE categoria_957 IS NOT NULL
                 AND TRIM(categoria_957)<>''
                 AND categoria_957<>'Sin clasificar'
               GROUP BY grupo_id, categoria_957;""",
            self.conn)

    def products(self, gid, yf=None, yt=None):
        q = ("SELECT categoria_957,subcategoria_957,tipo_producto_957,"
             "titulo,anio,autores,pestana FROM productos_957 WHERE grupo_id=?")
        p = [gid]
        if yf: q += " AND CAST(NULLIF(TRIM(anio),'') AS INTEGER)>=?"; p.append(yf)
        if yt: q += " AND CAST(NULLIF(TRIM(anio),'') AS INTEGER)<=?"; p.append(yt)
        q += " ORDER BY categoria_957, anio DESC;"
        return pd.read_sql_query(q, self.conn, params=p)

    def excel_rel(self, gid):
        r = self.conn.execute(
            "SELECT ruta_excel FROM grupos WHERE id=?", (gid,)).fetchone()
        return r[0] if r else None

    def close(self):
        try: self.conn.close()
        except Exception: pass


# ─────────────────────────────────────────────────────────────────────────────
# Diálogo: Excel original — texto COMPLETO, sin elide, todas las hojas
# ─────────────────────────────────────────────────────────────────────────────
class ExcelDialog(QDialog):
    def __init__(self, excel_path, group_name, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"Excel Original  ·  {group_name}")
        self.resize(1180, 740)
        self.setModal(True)

        lay = QVBoxLayout(self)
        lay.setContentsMargins(10, 10, 10, 8)
        lay.setSpacing(6)

        hdr = QLabel(f"📊  Datos originales GrupLAC:  {group_name}")
        hdr.setStyleSheet(
            f"font-weight:700;font-size:13px;color:{C_WHITE};"
            f"padding:8px 14px;background:{C_BLUE_D};border-radius:5px;")
        lay.addWidget(hdr)

        info = QLabel(f"📂  {os.path.basename(excel_path)}")
        info.setStyleSheet(f"color:{C_BLUE_M};font-size:11px;padding:2px;")
        lay.addWidget(info)

        # Buscador
        sb = QHBoxLayout()
        sb.addWidget(QLabel("🔍 Buscar en hoja activa:"))
        self._srch = QLineEdit()
        self._srch.setPlaceholderText("Escriba para filtrar filas…")
        self._srch.textChanged.connect(self._on_search)
        sb.addWidget(self._srch, 1)
        self._cnt = QLabel("")
        self._cnt.setStyleSheet(f"color:{C_GREY};font-size:11px;min-width:80px;")
        sb.addWidget(self._cnt)
        lay.addLayout(sb)

        self._tabs   = QTabWidget()
        lay.addWidget(self._tabs, 1)

        bc = QPushButton("✕  Cerrar")
        bc.setFixedWidth(110)
        bc.clicked.connect(self.accept)
        bl = QHBoxLayout()
        bl.addStretch(); bl.addWidget(bc)
        lay.addLayout(bl)

        self._dfs:    dict = {}
        self._models: dict = {}

        self._load(excel_path)
        self._tabs.currentChanged.connect(self._on_tab)

    def _load(self, path):
        if not HAS_OPENPYXL:
            self._tabs.addTab(
                QLabel("⚠️  Instale openpyxl:  pip install openpyxl"), "Error")
            return
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as e:
            self._tabs.addTab(QLabel(f"⚠️  Error al abrir:\n{e}"), "Error")
            return

        for sname in wb.sheetnames:
            ws   = wb[sname]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            hdr_row = None
            hdr_idx = 0
            for i, row in enumerate(rows[:10]):
                if any(c is not None and str(c).strip() for c in row):
                    hdr_row = row
                    hdr_idx = i + 1
                    break

            if hdr_row is None:
                hdr_row = [None]
                hdr_idx = 0

            cols = []
            seen: dict = {}
            for j, c in enumerate(hdr_row):
                name = str(c).strip() if (c is not None and str(c).strip()) \
                       else f"Col_{j+1}"
                base = name
                k = seen.get(base, 0)
                if k:
                    name = f"{base}.{k}"
                seen[base] = k + 1
                cols.append(name)

            data = []
            for row in rows[hdr_idx:]:
                if not any(c is not None for c in row):
                    continue
                r = list(row)
                r = r + [None] * max(0, len(cols) - len(r))
                r = r[:len(cols)]
                r = ["" if c is None else str(c) for c in r]
                data.append(r)

            if not data:
                continue

            df = pd.DataFrame(data, columns=cols)
            self._dfs[sname] = df

            w  = QWidget()
            wl = QVBoxLayout(w)
            wl.setContentsMargins(4, 4, 4, 4)
            wl.setSpacing(4)

            meta = QLabel(f"{len(df)} filas  ·  {len(df.columns)} columnas")
            meta.setStyleSheet(f"color:{C_GREY};font-size:10px;")
            wl.addWidget(meta)

            tv = QTableView()
            tv.setWordWrap(True)
            tv.setTextElideMode(Qt.ElideNone)
            tv.setAlternatingRowColors(True)
            tv.setSelectionBehavior(QAbstractItemView.SelectRows)
            tv.setSortingEnabled(True)
            tv.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
            tv.horizontalHeader().setStretchLastSection(True)
            tv.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)

            model = DFModel(df)
            tv.setModel(model)
            tv.resizeColumnsToContents()

            wl.addWidget(tv)

            idx = self._tabs.addTab(w, sname)
            self._models[idx] = model

        wb.close()
        if self._tabs.count() == 0:
            self._tabs.addTab(QLabel("No se encontraron hojas con datos."), "—")
        else:
            self._update_cnt()

    def _on_tab(self, _):
        self._srch.clear()
        self._update_cnt()

    def _on_search(self, txt):
        idx   = self._tabs.currentIndex()
        model = self._models.get(idx)
        sname = self._tabs.tabText(idx)
        df    = self._dfs.get(sname, pd.DataFrame())
        if model is None:
            return
        if txt.strip() and not df.empty:
            mask = df.apply(
                lambda c: c.astype(str).str.contains(txt, case=False, na=False)
            ).any(axis=1)
            model.set_df(df[mask])
            self._cnt.setText(f"{mask.sum()} filas")
        else:
            model.set_df(df)
            self._update_cnt()

    def _update_cnt(self):
        idx   = self._tabs.currentIndex()
        sname = self._tabs.tabText(idx)
        df    = self._dfs.get(sname, pd.DataFrame())
        self._cnt.setText(f"{len(df)} filas")


# ─────────────────────────────────────────────────────────────────────────────
# Fila de categoría — checkbox manual + info BD como referencia
# ─────────────────────────────────────────────────────────────────────────────
class CatRow(QWidget):
    def __init__(self, cat, on_change=None):
        super().__init__()
        self.cat = cat
        self._on_change = on_change

        lay = QHBoxLayout(self)
        lay.setContentsMargins(6, 3, 6, 3)
        lay.setSpacing(8)

        blt = QFrame()
        blt.setFixedSize(12, 12)
        blt.setStyleSheet(
            f"background:{CAT_FG[cat].name()};border-radius:3px;")
        lay.addWidget(blt)

        lbl = QLabel(cat)
        lbl.setStyleSheet(f"font-size:12px;color:{C_BLUE_D};")
        lay.addWidget(lbl, 1)

        self._info = QLabel("—")
        self._info.setStyleSheet(
            f"font-size:10px;color:{C_GREY};min-width:160px;")
        lay.addWidget(self._info)

        self.chk = QCheckBox("✔  Cumple")
        self.chk.setStyleSheet(
            "QCheckBox{font-size:12px;font-weight:bold;}"
            f"QCheckBox:checked{{color:{C_GRN};}}"
            f"QCheckBox:!checked{{color:{C_RED};}}")
        self.chk.stateChanged.connect(self._toggled)
        lay.addWidget(self.chk)

        self._badge = QLabel()
        self._badge.setFixedWidth(120)
        self._badge.setAlignment(Qt.AlignCenter | Qt.AlignVCenter)
        self._set_badge(False)
        lay.addWidget(self._badge)

    def _toggled(self, state):
        ok = (state == Qt.Checked)
        self._set_badge(ok)
        if self._on_change:
            self._on_change()

    def _set_badge(self, ok):
        if ok:
            self._badge.setText("✔  Cumple")
            self._badge.setStyleSheet(
                f"font-weight:700;font-size:12px;color:{C_GRN};"
                f"background:{C_GRN_B};border-radius:4px;padding:2px 8px;")
        else:
            self._badge.setText("✗  Pendiente")
            self._badge.setStyleSheet(
                f"font-weight:700;font-size:12px;color:{C_RED};"
                "background:#FFEBEE;border-radius:4px;padding:2px 8px;")

    def set_info(self, n):
        self._info.setText(
            f"{'▶' if n > 0 else '·'}  {n} producto(s) en BD")
        self._info.setStyleSheet(
            f"font-size:10px;color:{'#1B5E20' if n > 0 else C_GREY};"
            "min-width:160px;")

    def set_checked(self, v):
        self.chk.blockSignals(True)
        self.chk.setChecked(v)
        self.chk.blockSignals(False)
        self._set_badge(v)

    def is_checked(self):
        return self.chk.isChecked()


# ─────────────────────────────────────────────────────────────────────────────
# Export Excel
# ─────────────────────────────────────────────────────────────────────────────
def export_excel(group_name, df, filepath):
    if not HAS_OPENPYXL:
        return False, "openpyxl no instalado"
    try:
        wb   = openpyxl.Workbook()
        thin = Side(style='thin', color="BBBBBB")
        brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
        XF = {
            CATEGORIES_957[0]: PatternFill("solid", fgColor="E8F0FE"),
            CATEGORIES_957[1]: PatternFill("solid", fgColor="E3F2FD"),
            CATEGORIES_957[2]: PatternFill("solid", fgColor="FFF3E0"),
            CATEGORIES_957[3]: PatternFill("solid", fgColor="FFE0B2"),
            CATEGORIES_957[4]: PatternFill("solid", fgColor="EDE7F6"),
        }
        XC = {
            CATEGORIES_957[0]: "003087", CATEGORIES_957[1]: "1565C0",
            CATEGORIES_957[2]: "E65100", CATEGORIES_957[3]: "FF8F00",
            CATEGORIES_957[4]: "4527A0",
        }
        HF = PatternFill("solid", fgColor="003087")
        ws = wb.active; ws.title = "Productos"
        ws.merge_cells("A1:G1")
        c = ws["A1"]
        c.value = f"Productos Clasificados – Conv. 957  |  {group_name}"
        c.font = XLFont(name="Calibri", size=14, bold=True, color="FFFFFF")
        c.fill = HF; c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        ws.merge_cells("A2:G2")
        c = ws["A2"]
        c.value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  ·  Corte: 30/11/2025"
        c.font = XLFont(name="Calibri", size=10, italic=True, color="555555")
        c.alignment = Alignment(horizontal="center")
        hdrs = [c[0] for c in TABLE_COLS]
        wids = [40, 30, 30, 65, 8, 28, 22]
        for ci, (h, w) in enumerate(zip(hdrs, wids), 1):
            cell = ws.cell(row=3, column=ci, value=h)
            cell.font = XLFont(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.fill = HF; cell.border = brd
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            ws.column_dimensions[cell.column_letter].width = w
        for ri, (_, row) in enumerate(df.iterrows(), 4):
            cat = str(row.get(hdrs[0], ""))
            for ci, h in enumerate(hdrs, 1):
                v = row.get(h, "")
                cell = ws.cell(row=ri, column=ci,
                               value="" if pd.isna(v) else str(v))
                cell.font = XLFont(name="Calibri", size=10)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = brd
                if cat in XF: cell.fill = XF[cat]
                if ci == 1 and cat in XC:
                    cell.font = XLFont(name="Calibri", size=10,
                                       bold=True, color=XC[cat])
            ws.row_dimensions[ri].height = 20
        ws2 = wb.create_sheet("Resumen")
        ws2.merge_cells("A1:C1")
        c2 = ws2["A1"]
        c2.value = f"Resumen – {group_name}"
        c2.font = XLFont(name="Calibri", size=13, bold=True, color="FFFFFF")
        c2.fill = HF; c2.alignment = Alignment(horizontal="center")
        for ci, h in enumerate(["Categoría","Cantidad","Estado"], 1):
            cc = ws2.cell(row=2, column=ci, value=h)
            cc.font = XLFont(bold=True, color="FFFFFF"); cc.fill = HF
            cc.alignment = Alignment(horizontal="center")
        ws2.column_dimensions["A"].width = 52
        ws2.column_dimensions["B"].width = 12
        ws2.column_dimensions["C"].width = 10
        cat_col = hdrs[0]
        cc_map  = df[cat_col].value_counts() if cat_col in df.columns else {}
        for ri2, cat in enumerate(CATEGORIES_957, 3):
            cnt = int(cc_map.get(cat, 0)) if hasattr(cc_map,'get') else 0
            c1 = ws2.cell(row=ri2, column=1, value=cat)
            c1.font = XLFont(size=10)
            c2_ = ws2.cell(row=ri2, column=2, value=cnt)
            c2_.alignment = Alignment(horizontal="center")
            c3_ = ws2.cell(row=ri2, column=3, value="✓" if cnt else "✗")
            c3_.font = XLFont(size=12, bold=True,
                               color="1B5E20" if cnt else "B71C1C")
            c3_.alignment = Alignment(horizontal="center")
            if cat in XF:
                for co in (c1, c2_, c3_): co.fill = XF[cat]
        tr = 3 + len(CATEGORIES_957)
        ws2.cell(row=tr, column=1, value="TOTAL").font = XLFont(bold=True)
        tc = ws2.cell(row=tr, column=2, value=len(df))
        tc.font = XLFont(bold=True); tc.alignment = Alignment(horizontal="center")
        wb.save(filepath)
        return True, ""
    except Exception as e:
        return False, str(e)


# ─────────────────────────────────────────────────────────────────────────────
# PDF Certificado
# ─────────────────────────────────────────────────────────────────────────────
def gen_pdf(group_name, filepath, saved_state, product_counts=None):
    if not HAS_REPORTLAB:
        return False, "reportlab no instalado.  pip install reportlab"

    product_counts = product_counts or {}
    cats_met = [c for c in CATEGORIES_957 if saved_state.get(c, False)]
    all_met  = len(cats_met) == len(CATEGORIES_957)
    n_met    = len(cats_met)
    total_p  = sum(product_counts.get(c, 0) for c in CATEGORIES_957)
    pct      = int(n_met / len(CATEGORIES_957) * 100)
    now      = datetime.now()
    dstr     = f"{now.day} de {MONTHS_ES[now.month]} de {now.year}"
    folio    = f"CERT-957-{now.strftime('%Y%m%d%H%M%S')}"

    try:
        NAVY = rl_colors.HexColor("#002060")
        B1   = rl_colors.HexColor("#003087")
        B2   = rl_colors.HexColor("#1565C0")
        BL   = rl_colors.HexColor("#D6E4FF")
        ORG  = rl_colors.HexColor("#E65100")
        GLD  = rl_colors.HexColor("#FFC107")
        WHT  = rl_colors.white
        GY1  = rl_colors.HexColor("#37474F")
        GY2  = rl_colors.HexColor("#78909C")
        GY3  = rl_colors.HexColor("#ECEFF1")
        GRN  = rl_colors.HexColor("#1B5E20")
        GRN2 = rl_colors.HexColor("#43A047")
        RED  = rl_colors.HexColor("#B71C1C")

        RBG = [rl_colors.HexColor(x) for x in
               ["#EEF2FF","#E3F2FD","#FFF3E0","#FFE0B2","#EDE7F6"]]
        RFG = [rl_colors.HexColor(x) for x in
               ["#003087","#1565C0","#E65100","#FF8F00","#4527A0"]]

        def ps(n, **kw):
            return ParagraphStyle(n, **kw)

        s_cert = ps("c",  fontName="Helvetica-Bold",  fontSize=22,
                    textColor=WHT, alignment=TA_CENTER, spaceAfter=0)
        s_res  = ps("r",  fontName="Helvetica",        fontSize=10,
                    textColor=rl_colors.HexColor("#B3D0FF"),
                    alignment=TA_CENTER, spaceAfter=0, spaceBefore=2)
        s_fol  = ps("f",  fontName="Helvetica-Oblique",fontSize=8,
                    textColor=GY2, alignment=TA_RIGHT,  spaceAfter=0)
        s_body = ps("b",  fontName="Helvetica",        fontSize=10,
                    textColor=GY1, alignment=TA_JUSTIFY,
                    spaceAfter=8, leading=16)
        s_bc   = ps("bc", fontName="Helvetica",        fontSize=10,
                    textColor=GY1, alignment=TA_CENTER, spaceAfter=8)
        s_grp  = ps("g",  fontName="Helvetica-Bold",   fontSize=15,
                    textColor=B1,  alignment=TA_CENTER,
                    spaceAfter=6, spaceBefore=6)
        s_sec  = ps("s",  fontName="Helvetica-Bold",   fontSize=11,
                    textColor=NAVY, spaceAfter=4, spaceBefore=6)
        s_ok   = ps("ok", fontName="Helvetica-Bold",   fontSize=13,
                    textColor=GRN, alignment=TA_CENTER,
                    spaceAfter=8, spaceBefore=4)
        s_pt   = ps("pt", fontName="Helvetica-Bold",   fontSize=13,
                    textColor=ORG, alignment=TA_CENTER,
                    spaceAfter=8, spaceBefore=4)
        s_no   = ps("no", fontName="Helvetica-Bold",   fontSize=13,
                    textColor=RED, alignment=TA_CENTER,
                    spaceAfter=8, spaceBefore=4)
        s_ftr  = ps("ft", fontName="Helvetica-Oblique",fontSize=8,
                    textColor=GY2, alignment=TA_CENTER, spaceAfter=0)

        doc = SimpleDocTemplate(
            filepath, pagesize=A4,
            rightMargin=2.2*cm, leftMargin=2.2*cm,
            topMargin=1.8*cm,   bottomMargin=1.5*cm,
            title=f"Certificado Conv. 957 – {group_name}",
        )
        W = doc.width
        story = []

        def band(color, h):
            t = Table([[""]], colWidths=[W])
            t.setStyle(TableStyle([
                ("BACKGROUND",(0,0),(-1,-1), color),
                ("ROWHEIGHTS",(0,0),(-1,-1), h),
            ]))
            return t

        header_data = [
            [Paragraph(
                "Universidad Tecnológica de Pereira  ·  "
                "Vicerrectoría de Investigaciones",
                ps("ib", fontName="Helvetica", fontSize=8,
                   textColor=rl_colors.HexColor("#BBDEFB"),
                   alignment=TA_CENTER))],
            [""],
            [Paragraph("CERTIFICADO DE CUMPLIMIENTO", s_cert)],
            [""],
        ]

        header_tbl = Table(header_data, colWidths=[W])
        header_tbl.setStyle(TableStyle([
            ("BACKGROUND",    (0,0),(0,0), NAVY),
            ("TOPPADDING",    (0,0),(0,0), 7),
            ("BOTTOMPADDING", (0,0),(0,0), 7),
            ("BACKGROUND",    (0,1),(0,1), ORG),
            ("ROWHEIGHT",     (0,1),(0,1), 5),
            ("TOPPADDING",    (0,1),(0,1), 0),
            ("BOTTOMPADDING", (0,1),(0,1), 0),
            ("BACKGROUND",    (0,2),(0,2), B1),
            ("TOPPADDING",    (0,2),(0,2), 14),
            ("BOTTOMPADDING", (0,2),(0,2), 4),
            ("BACKGROUND",    (0,3),(0,3), B1),
            ("TOPPADDING",    (0,3),(0,3), 2),
            ("BOTTOMPADDING", (0,3),(0,3), 14),
            ("BACKGROUND",    (0,4),(0,4), GLD) if len(header_data) > 4 else ("BACKGROUND",(0,3),(0,3),B1),
            ("LEFTPADDING",   (0,0),(-1,-1), 8),
            ("RIGHTPADDING",  (0,0),(-1,-1), 8),
            ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
        ]))
        story.append(header_tbl)
        story.append(Spacer(1, 0.35*cm))

        story.append(Paragraph(
            f"N.° de verificación: <b>{folio}</b>  &nbsp;|&nbsp;  "
            f"Fecha de emisión: <b>{dstr}</b>", s_fol))
        story.append(Spacer(1, 0.3*cm))

        story.append(Paragraph(
            "Por medio del presente documento se <b>certifica</b> que el "
            "grupo de investigación:", s_bc))

        grp_t = Table([[Paragraph(f'"{group_name}"', s_grp)]], colWidths=[W])
        grp_t.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,-1), BL),
            ("BOX",(0,0),(-1,-1), 1.5, B2),
            ("TOPPADDING",(0,0),(-1,-1), 12),
            ("BOTTOMPADDING",(0,0),(-1,-1), 12),
            ("LEFTPADDING",(0,0),(-1,-1), 16),
            ("RIGHTPADDING",(0,0),(-1,-1), 16),
        ]))
        story.append(grp_t)
        story.append(Spacer(1, 0.25*cm))

        story.append(Paragraph(
            "ha cumplido con la <b>Resolución 7729 del 31 de diciembre de 2020</b> "
            "al haber actualizado la información del <b>GrupLAC</b> con corte al "
            "<b>30 de noviembre de 2025</b>, en el marco de la "
            "<b>Convocatoria 957</b> de Minciencias para la clasificación de "
            "grupos de investigación, desarrollo tecnológico e innovación.", s_body))

        cw4 = (W - 3*0.2*cm) / 4
        pct_c = GRN if pct == 100 else (ORG if pct >= 60 else RED)

        def met(val, lbl, col=B1):
            return Table([
                [Paragraph(str(val),
                           ps(f"mv{lbl[:3]}", fontName="Helvetica-Bold",
                              fontSize=24, textColor=col, alignment=TA_CENTER))],
                [Spacer(1, 0.18*cm)],
                [Paragraph(lbl,
                           ps(f"ml{lbl[:3]}", fontName="Helvetica", fontSize=8,
                              textColor=GY2, alignment=TA_CENTER))],
            ], colWidths=[cw4], rowHeights=[None, 0.18*cm, None])

        mx = Table([[met(n_met,    "Categorías\ncumplidas"),
                     met(5,        "Total\ncategorías"),
                     met(total_p,  "Productos\nregistrados"),
                     met(f"{pct}%","Porcentaje de\ncumplimiento", pct_c)]],
                   colWidths=[cw4]*4)
        mx.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,-1), GY3),
            ("LINEBEFORE",(1,0),(-1,-1), 1, rl_colors.HexColor("#CFD8DC")),
            ("TOPPADDING",(0,0),(-1,-1), 8),
            ("BOTTOMPADDING",(0,0),(-1,-1), 8),
            ("ALIGN",(0,0),(-1,-1),"CENTER"),
            ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ]))
        story.append(Spacer(1, 0.2*cm))
        story.append(mx)
        story.append(Spacer(1, 0.4*cm))

        story.append(Paragraph(
            "▪  Detalle por categoría — Convocatoria 957", s_sec))

        th = lambda txt: Paragraph(
            txt, ps(f"th{txt[:2]}", fontName="Helvetica-Bold",
                    fontSize=9, textColor=WHT, alignment=TA_CENTER))

        cw = [0.7*cm, W-0.7*cm-2.2*cm-2.6*cm, 2.2*cm, 2.6*cm]
        td = [[th("#"), th("Categoría"), th("Productos"), th("Estado")]]

        for i, cat in enumerate(CATEGORIES_957):
            met_  = saved_state.get(cat, False)
            cnt   = product_counts.get(cat, 0)
            stxt  = "✔  Cumple" if met_ else "✘  Pendiente"
            stcol = GRN if met_ else RED
            td.append([
                Paragraph(str(i+1),
                          ps(f"tn{i}", fontName="Helvetica-Bold", fontSize=10,
                             textColor=RFG[i], alignment=TA_CENTER)),
                Paragraph(cat,
                          ps(f"tc{i}", fontName="Helvetica", fontSize=9,
                             textColor=RFG[i])),
                Paragraph(str(cnt),
                          ps(f"tp{i}", fontName="Helvetica-Bold", fontSize=11,
                             textColor=RFG[i], alignment=TA_CENTER)),
                Paragraph(stxt,
                          ps(f"ts{i}", fontName="Helvetica-Bold", fontSize=9,
                             textColor=stcol, alignment=TA_CENTER)),
            ])

        ctbl = Table(td, colWidths=cw)
        cts  = TableStyle([
            ("BACKGROUND",(0,0),(-1,0), B1),
            ("TOPPADDING",(0,0),(-1,-1), 7),
            ("BOTTOMPADDING",(0,0),(-1,-1), 7),
            ("LEFTPADDING",(1,1),(1,-1), 10),
            ("GRID",(0,0),(-1,-1), 0.4, rl_colors.HexColor("#CFD8DC")),
            ("BOX",(0,0),(-1,-1), 1.2, B1),
            ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ])
        for i in range(len(CATEGORIES_957)):
            cts.add("BACKGROUND",(0,i+1),(-1,i+1), RBG[i])
        ctbl.setStyle(cts)
        story.append(ctbl)
        story.append(Spacer(1, 0.35*cm))

        if all_met:
            res = Table([[Paragraph(
                "✔  El grupo CUMPLE PLENAMENTE con todas las categorías "
                "requeridas por la Convocatoria 957.", s_ok)]],
                colWidths=[W])
            res.setStyle(TableStyle([
                ("BACKGROUND",(0,0),(-1,-1), rl_colors.HexColor("#E8F5E9")),
                ("BOX",(0,0),(-1,-1), 1.5, GRN2),
                ("TOPPADDING",(0,0),(-1,-1), 10),
                ("BOTTOMPADDING",(0,0),(-1,-1), 10),
            ]))
        elif n_met > 0:
            falt = " · ".join(c.split(". ",1)[-1]
                               for c in CATEGORIES_957
                               if not saved_state.get(c, False))
            res  = Table([[Paragraph(
                f"El grupo cumple <b>{n_met} de 5</b> categorías. "
                f"Pendientes: {falt}.", s_pt)]],
                colWidths=[W])
            res.setStyle(TableStyle([
                ("BACKGROUND",(0,0),(-1,-1), rl_colors.HexColor("#FFF3E0")),
                ("BOX",(0,0),(-1,-1), 1.5, ORG),
                ("TOPPADDING",(0,0),(-1,-1), 10),
                ("BOTTOMPADDING",(0,0),(-1,-1), 10),
            ]))
        else:
            res = Table([[Paragraph(
                "El grupo NO ha completado ninguna categoría de la "
                "Convocatoria 957.", s_no)]],
                colWidths=[W])
            res.setStyle(TableStyle([
                ("BACKGROUND",(0,0),(-1,-1), rl_colors.HexColor("#FFEBEE")),
                ("BOX",(0,0),(-1,-1), 1.5, RED),
                ("TOPPADDING",(0,0),(-1,-1), 10),
                ("BOTTOMPADDING",(0,0),(-1,-1), 10),
            ]))
        story.append(res)
        story.append(Spacer(1, 0.4*cm))

        half = (W - 2*cm) / 2
        def sig(rol, inst):
            return Table([
                [Paragraph("______________________________",
                           ps(f"sl{rol[:3]}", fontName="Helvetica", fontSize=9,
                              textColor=GY1, alignment=TA_CENTER))],
                [Paragraph(rol,
                           ps(f"st{rol[:3]}", fontName="Helvetica-Bold",
                              fontSize=9, textColor=NAVY, alignment=TA_CENTER))],
                [Paragraph(inst,
                           ps(f"sb{rol[:3]}", fontName="Helvetica", fontSize=8,
                              textColor=GY2, alignment=TA_CENTER))],
            ], colWidths=[half])

        sigs = Table(
            [[sig("Director de Investigaciones",
                  "Universidad Tecnológica de Pereira")]],
            colWidths=[half+cm, half+cm])
        sigs.setStyle(TableStyle([
            ("TOPPADDING",(0,0),(-1,-1), 8),
            ("BOTTOMPADDING",(0,0),(-1,-1), 8),
            ("ALIGN",(0,0),(-1,-1),"CENTER"),
        ]))
        story.append(sigs)

        story.append(HRFlowable(
            width="100%", thickness=0.5,
            color=rl_colors.HexColor("#CFD8DC"), spaceAfter=4))
        story.append(Paragraph(
            f"Certificado generado automáticamente el {dstr} · "
            "Sistema de Verificación GrupLAC · UTP.  "
            "Resolución 7729/2020.  "
            f"Código: {folio}", s_ftr))
        story.append(Spacer(1, 0.2*cm))
        story.append(band(ORG, 8))

        doc.build(story)
        return True, ""

    except Exception as e:
        import traceback
        return False, f"{e}\n{traceback.format_exc()}"


# ─────────────────────────────────────────────────────────────────────────────
# Widget principal — embebible como pestaña en main_10.py
# ─────────────────────────────────────────────────────────────────────────────
class VisorGrupLAC957(QWidget):
    """
    Visor GrupLAC Conv. 957 como QWidget embebible.
    Uso:
        self.vista_visor = VisorGrupLAC957()
        self.tabs.addTab(self.vista_visor, "Visor GrupLAC 957")
    """

    def __init__(self, db_path=DEFAULT_DB_PATH, data_dir=DEFAULT_DATA_DIR, parent=None):
        super().__init__(parent)
        self.data_dir    = data_dir
        self.current_gid = None
        self.current_nm  = ""
        self._cur_df     = pd.DataFrame()
        self._saved:  dict = {}   # {gid: {cat: bool}}
        self._dirty:  dict = {}   # {gid: bool}

        # Indicador de error de BD
        self._repo_error = None

        try:
            self.repo      = Repo(db_path)
            self.grps_df   = self.repo.groups()
            self.counts_df = self.repo.cat_counts()
            self.compliance = self._build_compliance()
        except FileNotFoundError as e:
            self._repo_error = str(e)
            self.repo = None
            self.grps_df    = pd.DataFrame(columns=["id","nombre"])
            self.counts_df  = pd.DataFrame()
            self.compliance = {}

        self._build_ui()

        if self._repo_error:
            self._show_db_error()
        elif self.group_list.count():
            self.group_list.setCurrentRow(0)

    # ── Error de BD ──────────────────────────────────────────────────────────
    def _show_db_error(self):
        self.grp_hdr.setText(
            f"⚠️  BD no encontrada: {self._repo_error}")
        self.grp_hdr.setStyleSheet(
            f"font-weight:700;font-size:13px;color:{C_WHITE};"
            f"padding:9px 14px;background:#B71C1C;border-radius:5px;")

    def _build_compliance(self):
        cmp = {int(g): {c: 0 for c in CATEGORIES_957}
               for g in self.grps_df["id"]}
        for _, row in self.counts_df.iterrows():
            gid = int(row["grupo_id"])
            cat = str(row["categoria_957"]).strip()
            if gid in cmp and cat in cmp[gid]:
                cmp[gid][cat] = int(row["n"])
        return cmp

    # ── UI ───────────────────────────────────────────────────────────────────
    def _build_ui(self):
        rl  = QHBoxLayout(self)
        rl.setContentsMargins(8, 8, 8, 8)
        spl = QSplitter(Qt.Horizontal)

        # ── Izquierda ─────────────────────────────────────────────────────────
        left = QWidget(); left.setFixedWidth(280)
        ll = QVBoxLayout(left)
        ll.setContentsMargins(0, 0, 4, 0); ll.setSpacing(6)

        hdr = QLabel("🔬  Grupos de Investigación")
        hdr.setStyleSheet(
            f"font-weight:700;font-size:13px;color:{C_BLUE_D};"
            f"padding:6px;background:{C_BLUE_L};border-radius:4px;")
        ll.addWidget(hdr)

        hint = QLabel("✅ = todas guardadas  ·  ☐ = incompleto")
        hint.setStyleSheet(f"font-size:10px;color:{C_GREY};padding:2px;")
        ll.addWidget(hint)

        self.srch = QLineEdit()
        self.srch.setPlaceholderText("🔍  Buscar grupo…")
        self.srch.textChanged.connect(self._filter_list)
        ll.addWidget(self.srch)

        self.group_list = QListWidget()
        self.group_list.currentItemChanged.connect(self._on_select)
        ll.addWidget(self.group_list, 1)

        self.list_ctr = QLabel("")
        self.list_ctr.setStyleSheet(f"font-size:10px;color:{C_GREY};padding:2px;")
        ll.addWidget(self.list_ctr)
        spl.addWidget(left)

        # ── Derecha ───────────────────────────────────────────────────────────
        right = QWidget()
        rr = QVBoxLayout(right)
        rr.setContentsMargins(4, 0, 0, 0); rr.setSpacing(6)

        self.grp_hdr = QLabel("Seleccione un grupo de investigación")
        self.grp_hdr.setStyleSheet(
            f"font-weight:700;font-size:14px;color:{C_WHITE};"
            f"padding:9px 14px;background:{C_BLUE_D};border-radius:5px;")
        rr.addWidget(self.grp_hdr)

        # Checklist manual
        chk_box = QGroupBox(
            "✔  Verificación manual — marque las categorías verificadas y guarde")
        chk_box.setStyleSheet(
            f"QGroupBox{{border:2px solid {C_BLUE_M};}}")
        cb_lay = QVBoxLayout(chk_box)
        cb_lay.setContentsMargins(6, 4, 6, 4); cb_lay.setSpacing(2)

        self.cat_rows: dict = {}
        for c in CATEGORIES_957:
            row = CatRow(c, on_change=self._mark_dirty)
            self.cat_rows[c] = row
            cb_lay.addWidget(row)

        sbar = QHBoxLayout(); sbar.setSpacing(8)
        self.save_lbl = QLabel("Sin guardar")
        self.save_lbl.setStyleSheet(
            f"font-size:11px;color:{C_GREY};font-style:italic;")
        sbar.addWidget(self.save_lbl, 1)
        self.btn_save = QPushButton("💾  Guardar Estado")
        self.btn_save.setObjectName("btnSave")
        self.btn_save.setStyleSheet(
            f"background:{C_GRN};color:{C_WHITE};border:none;"
            "border-radius:5px;padding:6px 14px;font-size:12px;font-weight:bold;")
        self.btn_save.setToolTip(
            "Guarda el estado actual de los checkboxes.\n"
            "El PDF se generará usando SOLO el estado guardado.")
        self.btn_save.clicked.connect(self._save_state)
        sbar.addWidget(self.btn_save)
        cb_lay.addLayout(sbar)

        self.global_ind = QLabel("—")
        self.global_ind.setAlignment(Qt.AlignCenter)
        self.global_ind.setStyleSheet(
            f"font-size:12px;font-weight:bold;padding:5px;"
            f"border-radius:4px;background:{C_BLUE_P};color:{C_BLUE_D};")
        cb_lay.addWidget(self.global_ind)
        rr.addWidget(chk_box)

        # Filtros
        fbar = QHBoxLayout(); fbar.setSpacing(6)
        fbar.addWidget(QLabel("Categoría:"))
        self.cat_cb = QComboBox()
        self.cat_cb.addItem("-- Todas --")
        for c in CATEGORIES_957:
            self.cat_cb.addItem(c)
        self.cat_cb.currentIndexChanged.connect(self._refresh_table)
        fbar.addWidget(self.cat_cb, 1)
        fbar.addWidget(QLabel("Año desde:"))
        self.yf = QSpinBox(); self.yf.setRange(1990,2030); self.yf.setValue(2000)
        self.yf.valueChanged.connect(self._refresh_table)
        fbar.addWidget(self.yf)
        fbar.addWidget(QLabel("hasta:"))
        self.yt = QSpinBox(); self.yt.setRange(1990,2030); self.yt.setValue(2025)
        self.yt.valueChanged.connect(self._refresh_table)
        fbar.addWidget(self.yt)
        rr.addLayout(fbar)

        self.tbl_lbl = QLabel("Productos clasificados")
        self.tbl_lbl.setStyleSheet(
            f"font-weight:600;color:{C_BLUE_M};font-size:12px;")
        rr.addWidget(self.tbl_lbl)

        self.table = QTableView()
        self.table.setAlternatingRowColors(True)
        self.table.setSortingEnabled(True)
        self.table.setWordWrap(True)
        self.table.setTextElideMode(Qt.ElideNone)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.verticalHeader().setSectionResizeMode(
            QHeaderView.ResizeToContents)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tbl_model = DFModel(
            pd.DataFrame(columns=[c[0] for c in TABLE_COLS]),
            cat_col="Categoría 957")
        self.table.setModel(self.tbl_model)
        rr.addWidget(self.table, 1)

        self.tbl_foot = QLabel("Total: 0 productos")
        self.tbl_foot.setStyleSheet(f"color:{C_GREY};font-size:11px;")
        rr.addWidget(self.tbl_foot)

        # Botones
        btn_bar = QHBoxLayout(); btn_bar.setSpacing(8)

        self.btn_view = QPushButton("📂  Ver Excel Original")
        self.btn_view.setStyleSheet(
            "background:#4527A0;color:white;border:none;"
            "border-radius:5px;padding:6px 14px;font-size:12px;font-weight:bold;")
        self.btn_view.setToolTip(
            "Abre el Excel original mostrando todas las celdas sin truncar")
        self.btn_view.clicked.connect(self._open_excel)
        btn_bar.addWidget(self.btn_view)
        btn_bar.addStretch()

        self.btn_exp = QPushButton("📊  Exportar Excel")
        self.btn_exp.setStyleSheet(
            "background:#1B5E20;color:white;border:none;"
            "border-radius:5px;padding:6px 14px;font-size:12px;font-weight:bold;")
        self.btn_exp.clicked.connect(self._export_excel)
        btn_bar.addWidget(self.btn_exp)

        self.btn_pdf = QPushButton("📄  Generar Certificado PDF")
        self.btn_pdf.setStyleSheet(
            f"background:{C_ORG};color:white;border:none;"
            "border-radius:5px;padding:6px 14px;font-size:12px;font-weight:bold;")
        self.btn_pdf.setToolTip(
            "Genera el certificado usando el estado GUARDADO.\n"
            "Primero guarda el estado con '💾 Guardar Estado'.")
        self.btn_pdf.clicked.connect(self._export_pdf)
        btn_bar.addWidget(self.btn_pdf)
        rr.addLayout(btn_bar)

        spl.addWidget(right)
        spl.setStretchFactor(0, 0); spl.setStretchFactor(1, 1)
        rl.addWidget(spl)

        self._populate_list()

    # ── Lista ─────────────────────────────────────────────────────────────────
    def _populate_list(self):
        self.group_list.clear()
        for _, g in self.grps_df.iterrows():
            gid  = int(g["id"])
            name = str(g["nombre"])
            sv   = self._saved.get(gid, {})
            ok   = bool(sv) and all(sv.get(c, False) for c in CATEGORIES_957)
            item = QListWidgetItem(("✅  " if ok else "☐   ") + name)
            item.setData(Qt.UserRole, gid)
            per = self.compliance.get(gid, {})
            tip = "\n".join(
                f"BD: {per.get(c,0)} prod. | Guardado: {'✔' if sv.get(c) else '✗'}"
                for c in CATEGORIES_957)
            item.setToolTip(tip)
            if ok:
                item.setForeground(QBrush(QColor(C_BLUE_D)))
            self.group_list.addItem(item)
        total = self.grps_df.shape[0]
        self.list_ctr.setText(f"{total} grupos")

    def _filter_list(self, txt):
        t = txt.strip().lower()
        vis = 0
        for i in range(self.group_list.count()):
            item = self.group_list.item(i)
            hide = bool(t) and t not in item.text().lower()
            item.setHidden(hide)
            if not hide: vis += 1
        self.list_ctr.setText(f"{vis}/{self.grps_df.shape[0]} grupos")

    def _refresh_list_item(self, gid):
        sv   = self._saved.get(gid, {})
        ok   = bool(sv) and all(sv.get(c, False) for c in CATEGORIES_957)
        ns   = self.grps_df.loc[self.grps_df["id"]==gid,"nombre"]
        name = str(ns.iloc[0]) if not ns.empty else ""
        for i in range(self.group_list.count()):
            item = self.group_list.item(i)
            if item.data(Qt.UserRole) == gid:
                item.setText(("✅  " if ok else "☐   ") + name)
                col = QColor(C_BLUE_D) if ok else QColor("#1A1A2E")
                item.setForeground(QBrush(col))
                break

    # ── Selección ─────────────────────────────────────────────────────────────
    def _on_select(self, cur, _prev):
        if not cur: return
        gid = int(cur.data(Qt.UserRole))
        self.current_gid = gid
        ns = self.grps_df.loc[self.grps_df["id"]==gid,"nombre"]
        self.current_nm = str(ns.iloc[0]) if not ns.empty else ""
        self.grp_hdr.setText(f"🔬  {self.current_nm}")
        self.grp_hdr.setStyleSheet(
            f"font-weight:700;font-size:14px;color:{C_WHITE};"
            f"padding:9px 14px;background:{C_BLUE_D};border-radius:5px;")
        self._load_ui_from_saved()
        self._refresh_table()

    # ── Estado del checklist ──────────────────────────────────────────────────
    def _mark_dirty(self):
        if self.current_gid is not None:
            self._dirty[self.current_gid] = True
            self._update_save_lbl()
            self._update_global()

    def _update_save_lbl(self):
        gid = self.current_gid
        if gid is None:
            self.save_lbl.setText("Sin cargar"); return
        dirty = self._dirty.get(gid, False)
        sv    = self._saved.get(gid)
        n     = sum(1 for c in CATEGORIES_957 if sv and sv.get(c)) if sv else 0
        if dirty:
            self.save_lbl.setText("⚠️  Cambios SIN GUARDAR")
            self.save_lbl.setStyleSheet(
                f"font-size:11px;color:{C_ORG};font-weight:bold;")
        elif sv is not None:
            self.save_lbl.setText(f"✔  Guardado: {n}/{len(CATEGORIES_957)} marcadas")
            self.save_lbl.setStyleSheet(
                f"font-size:11px;color:{C_GRN};font-style:normal;")
        else:
            self.save_lbl.setText("Sin guardar aún para este grupo")
            self.save_lbl.setStyleSheet(
                f"font-size:11px;color:{C_GREY};font-style:italic;")

    def _load_ui_from_saved(self):
        gid  = self.current_gid
        per  = self.compliance.get(gid, {c:0 for c in CATEGORIES_957})
        sv   = self._saved.get(gid, {})
        for c, row in self.cat_rows.items():
            row.set_info(per.get(c, 0))
            row.set_checked(sv.get(c, False))
        self._dirty[gid] = False
        self._update_save_lbl()
        self._update_global()

    def _save_state(self):
        gid = self.current_gid
        if gid is None:
            QMessageBox.warning(self, "Aviso", "Seleccione un grupo primero.")
            return
        state = {c: self.cat_rows[c].is_checked() for c in CATEGORIES_957}
        self._saved[gid] = state
        self._dirty[gid] = False
        self._update_save_lbl()
        self._update_global()
        self._refresh_list_item(gid)
        n = sum(1 for v in state.values() if v)
        QMessageBox.information(
            self, "✔  Estado guardado",
            f"Estado guardado para:\n«{self.current_nm}»\n\n"
            f"Categorías marcadas como cumplidas: {n}/{len(CATEGORIES_957)}\n\n"
            "Ahora puede generar el certificado PDF.")

    def _update_global(self):
        gid   = self.current_gid
        sv    = self._saved.get(gid, {})
        dirty = self._dirty.get(gid, False)
        n     = sum(1 for c in CATEGORIES_957 if sv.get(c, False))
        if dirty:
            self.global_ind.setText("⚠️  Cambios pendientes de guardar")
            self.global_ind.setStyleSheet(
                f"font-size:12px;font-weight:bold;padding:5px;"
                f"border-radius:4px;background:{C_ORG_P};color:{C_ORG};")
        elif n == len(CATEGORIES_957):
            self.global_ind.setText(
                f"✅  GUARDADO — Cumple {n}/{len(CATEGORIES_957)} categorías")
            self.global_ind.setStyleSheet(
                f"font-size:12px;font-weight:bold;padding:5px;"
                f"border-radius:4px;background:{C_GRN_B};color:{C_GRN};")
        elif n > 0:
            self.global_ind.setText(
                f"⚠️  GUARDADO — {n}/{len(CATEGORIES_957)} categorías marcadas")
            self.global_ind.setStyleSheet(
                f"font-size:12px;font-weight:bold;padding:5px;"
                f"border-radius:4px;background:{C_ORG_P};color:{C_ORG};")
        elif gid in self._saved:
            self.global_ind.setText("☐  Sin categorías marcadas (estado guardado)")
            self.global_ind.setStyleSheet(
                f"font-size:12px;font-weight:bold;padding:5px;"
                f"border-radius:4px;background:#FFEBEE;color:{C_RED};")
        else:
            self.global_ind.setText("☐  Sin estado guardado")
            self.global_ind.setStyleSheet(
                f"font-size:12px;font-weight:bold;padding:5px;"
                f"border-radius:4px;background:{C_BLUE_P};color:{C_BLUE_D};")

    # ── Tabla de productos ────────────────────────────────────────────────────
    def _refresh_table(self):
        gid = self.current_gid
        if gid is None or self.repo is None: return
        yf  = self.yf.value() if self.yf.value() > 1990 else None
        yt  = self.yt.value() if self.yt.value() < 2030 else None
        df  = self.repo.products(gid, yf, yt)
        sel = self.cat_cb.currentText()
        if sel != "-- Todas --" and "categoria_957" in df.columns:
            df = df[df["categoria_957"] == sel].copy()
        out = pd.DataFrame({
            disp: (df[col].values if col in df.columns else [""]*len(df))
            for disp, col in TABLE_COLS
        })
        self.tbl_lbl.setText(
            f"Productos clasificados — {self.current_nm}"
            + (f"  [filtro: {sel}]" if sel != "-- Todas --" else ""))
        self.tbl_model.set_df(out)
        self.table.resizeColumnsToContents()
        self.tbl_foot.setText(
            f"Total: {len(out)} productos  "
            f"({'filtro: '+sel if sel != '-- Todas --' else 'sin filtro'})")
        self._cur_df = out

    # ── Ver Excel original ────────────────────────────────────────────────────
    def _open_excel(self):
        if self.current_gid is None:
            QMessageBox.warning(self, "Aviso", "Seleccione un grupo primero.")
            return
        path = find_excel(self.current_nm, self.data_dir)
        if not path and self.repo:
            rel = self.repo.excel_rel(self.current_gid)
            if rel:
                for base in (self.data_dir, os.getcwd(), "."):
                    cand = os.path.join(base, rel)
                    if os.path.exists(cand):
                        path = cand; break
        if not path or not os.path.exists(path):
            ans = QMessageBox.question(
                self, "Archivo no encontrado",
                f"No se encontró el Excel para:\n«{self.current_nm}»\n\n"
                "¿Desea buscarlo manualmente?",
                QMessageBox.Yes | QMessageBox.No)
            if ans == QMessageBox.Yes:
                path, _ = QFileDialog.getOpenFileName(
                    self, "Seleccionar Excel", self.data_dir,
                    "Excel (*.xlsx *.xls)")
                if not path: return
            else:
                return
        ExcelDialog(path, self.current_nm, self).exec_()

    # ── Exportar Excel clasificado ────────────────────────────────────────────
    def _export_excel(self):
        if self.current_gid is None:
            QMessageBox.warning(self, "Aviso", "Seleccione un grupo primero.")
            return
        df = self._cur_df
        if df.empty and self.repo:
            raw = self.repo.products(self.current_gid)
            df  = pd.DataFrame({
                d:(raw[c] if c in raw.columns else "")
                for d,c in TABLE_COLS})
        path, _ = QFileDialog.getSaveFileName(
            self, "Guardar Excel",
            f"{self.current_nm[:40]}_conv957.xlsx",
            "Excel (*.xlsx)")
        if not path: return
        ok, msg = export_excel(self.current_nm, df, path)
        if ok:
            QMessageBox.information(self, "Listo", f"Excel guardado en:\n{path}")
        else:
            QMessageBox.critical(self, "Error", f"No se pudo guardar:\n{msg}")

    # ── Generar PDF ───────────────────────────────────────────────────────────
    def _export_pdf(self):
        gid = self.current_gid
        if gid is None:
            QMessageBox.warning(self, "Aviso", "Seleccione un grupo primero.")
            return
        if self._dirty.get(gid, False):
            QMessageBox.warning(
                self, "⚠️  Cambios sin guardar",
                "Tiene cambios en el checklist que aún NO se han guardado.\n\n"
                "Pulse «💾 Guardar Estado» primero.\n\n"
                "El PDF se genera con el estado GUARDADO, no con "
                "los cambios actuales.")
            return
        if gid not in self._saved:
            QMessageBox.warning(
                self, "Sin estado guardado",
                "Aún no ha guardado el estado de verificación de este grupo.\n\n"
                "1. Marque las categorías verificadas.\n"
                "2. Pulse «💾 Guardar Estado».\n"
                "3. Luego genere el certificado PDF.")
            return

        sv  = self._saved[gid]
        per = self.compliance.get(gid, {c:0 for c in CATEGORIES_957})

        path, _ = QFileDialog.getSaveFileName(
            self, "Guardar Certificado PDF",
            f"Certificado_957_{self.current_nm[:35]}.pdf",
            "PDF (*.pdf)")
        if not path: return

        ok, msg = gen_pdf(self.current_nm, path, sv, product_counts=per)
        if ok:
            n = sum(1 for v in sv.values() if v)
            QMessageBox.information(
                self, "✔  PDF generado",
                f"Certificado guardado en:\n{path}\n\n"
                f"Categorías reflejadas: {n}/{len(CATEGORIES_957)}")
        else:
            QMessageBox.critical(
                self, "Error al generar PDF",
                f"No se pudo generar:\n\n{msg}\n\n"
                "pip install reportlab")

    # ── Limpieza al cerrar ────────────────────────────────────────────────────
    def closeEvent(self, e):
        if self.repo:
            self.repo.close()
        super().closeEvent(e)
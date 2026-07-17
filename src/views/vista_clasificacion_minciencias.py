"""
Vista de Clasificación de Productos GrupLAC según Convocatoria 957 MinCiencias.
"""

import re
import traceback
from collections import defaultdict
from datetime import datetime

import pandas as pd
from pathlib import Path
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtGui import QBrush, QColor, QFont
from PyQt5.QtWidgets import (
    QCheckBox, QComboBox, QGroupBox, QHBoxLayout, QHeaderView,
    QLabel, QMessageBox, QProgressBar, QPushButton, QSpinBox,
    QSplitter, QTableWidget, QTableWidgetItem, QTextEdit,
    QTreeWidget, QTreeWidgetItem, QVBoxLayout, QWidget,
)
from unidecode import unidecode

from constants import CLASIF_957, COLORES_CATEGORIA
from utils import (
    clasificar_producto_957,
    get_orden_categoria,
    get_orden_producto,
    normalizar_nombre_hoja,
    obtener_directorio_base,
)


def _carpeta_gruplac_mas_reciente():
    """Carpeta data/reporte excel_<fecha> más reciente. Mismo criterio que
    vista_seguimiento_grupos._carpeta_gruplac_mas_reciente -- el nombre trae
    la fecha en formato YYYYMMDD, así que ordenar por texto ya da el orden
    cronológico correcto."""
    base = obtener_directorio_base() / "data"
    candidatos = sorted(
        (p for p in base.glob("reporte excel_*") if p.is_dir()), reverse=True)
    return candidatos[0] if candidatos else None


class ProcesamientoClasificacionThread(QThread):
    """Thread para procesar y clasificar productos de GrupLAC según Conv. 957."""

    progreso = pyqtSignal(int, str)
    finalizado = pyqtSignal(dict)
    error = pyqtSignal(str)

    def __init__(self, db, reporte_excel_path, años_filtro=None, incluir_sin_año=True):
        super().__init__()
        self.db = db
        self.reporte_excel_path = Path(reporte_excel_path)
        self.años_filtro = años_filtro or [2022, 2023, 2024, 2025]
        self.incluir_sin_año = incluir_sin_año

    def _buscar_carpeta_grupo(self, grupo_bd, carpetas_disponibles):
        """Busca la carpeta que corresponde a un grupo usando coincidencia normalizada."""
        grupo_bd_norm = normalizar_nombre_hoja(grupo_bd)

        for carpeta in carpetas_disponibles:
            if carpeta.name == grupo_bd:
                return carpeta

        for carpeta in carpetas_disponibles:
            if normalizar_nombre_hoja(carpeta.name) == grupo_bd_norm:
                return carpeta

        for longitud in [80, 70, 60, 50, 40, 30]:
            grupo_corto = grupo_bd_norm[:longitud].rstrip()
            for carpeta in carpetas_disponibles:
                carpeta_norm = normalizar_nombre_hoja(carpeta.name)
                if carpeta_norm.startswith(grupo_corto) or grupo_corto.startswith(
                    carpeta_norm[:longitud]
                ):
                    return carpeta

        return None

    def _extraer_año(self, texto):
        """Extrae el año de un texto; busca valores entre 1990 y 2035."""
        if not texto or pd.isna(texto):
            return None
        años = re.findall(r"\b(19[89]\d|20[0-3]\d)\b", str(texto).strip())
        validos = [int(a) for a in años if 1990 <= int(a) <= 2035]
        if not validos:
            return None
        en_filtro = [a for a in validos if a in self.años_filtro]
        return max(en_filtro) if en_filtro else max(validos)

    def _extraer_productos_de_hoja(self, df, nombre_hoja, nombre_grupo, archivo_nombre):
        """Extrae todos los productos de una hoja de Excel."""
        if df.empty:
            return []

        clasificacion = clasificar_producto_957(nombre_hoja)
        if clasificacion:
            categoria, subcategoria, nombre_oficial = clasificacion
        else:
            categoria = "Sin clasificar"
            subcategoria = "Pendiente de revisión"
            nombre_oficial = nombre_hoja

        columnas = list(df.columns)
        col_titulo = col_año = col_autores = None

        for idx, col in enumerate(columnas):
            col_norm = unidecode(str(col).lower())
            if col_titulo is None and any(
                p in col_norm for p in ["titulo", "nombre", "title"]
            ):
                col_titulo = idx
            if col_año is None and any(
                p in col_norm for p in ["ano", "year", "fecha"]
            ):
                col_año = idx
            if col_autores is None and any(
                p in col_norm for p in ["autor", "author", "integrante"]
            ):
                col_autores = idx

        productos = []
        for _, row in df.iterrows():
            try:
                textos_fila = []
                datos_completos = {}
                for i, col_nombre in enumerate(columnas):
                    valor = row.iloc[i] if i < len(row) else None
                    if pd.notna(valor):
                        val_str = str(valor).strip()
                        if val_str and val_str.lower() not in ("nan", "none", ""):
                            datos_completos[str(col_nombre)] = val_str
                            textos_fila.append(val_str)

                if not textos_fila:
                    continue

                # Título
                titulo = None
                if col_titulo is not None and col_titulo < len(row):
                    val = row.iloc[col_titulo]
                    if pd.notna(val) and str(val).strip():
                        titulo = str(val).strip()
                if not titulo:
                    for txt in textos_fila:
                        if len(txt) > 15 and not any(
                            p in txt.lower() for p in ["http", "doi:", "issn", "isbn"]
                        ):
                            titulo = txt
                            break
                if not titulo or len(titulo) < 5:
                    continue

                # Año
                año = None
                if col_año is not None and col_año < len(row):
                    año = self._extraer_año(row.iloc[col_año])
                if año is None:
                    año = self._extraer_año(" ".join(textos_fila))

                if año is not None and año not in self.años_filtro:
                    continue
                if año is None and not self.incluir_sin_año:
                    continue

                # Autores
                autores = None
                if col_autores is not None and col_autores < len(row):
                    val = row.iloc[col_autores]
                    if pd.notna(val):
                        autores = str(val).strip()
                if not autores:
                    for txt in textos_fila:
                        if (
                            "," in txt
                            and 10 < len(txt) < 500
                            and not any(
                                p in txt.lower()
                                for p in ["http", "doi", "issn", "isbn", "@", ".com"]
                            )
                        ):
                            autores = txt
                            break

                productos.append({
                    "titulo": titulo[:500],
                    "año": año,
                    "autores": autores or "No especificado",
                    "hoja_original": nombre_hoja,
                    "categoria_957": categoria,
                    "subcategoria_957": subcategoria,
                    "producto_957": nombre_oficial,
                    "grupo": nombre_grupo,
                    "archivo": archivo_nombre,
                    "datos_completos": datos_completos,
                    "orden_categoria": get_orden_categoria(categoria),
                    "orden_producto": get_orden_producto(nombre_oficial),
                })
            except Exception:
                continue

        return productos

    def _procesar_archivo_excel(self, archivo_path, nombre_grupo):
        """Procesa un archivo Excel completo y devuelve la lista de productos."""
        productos = []
        try:
            excel_file = pd.ExcelFile(archivo_path, engine="openpyxl")
            for hoja in excel_file.sheet_names:
                try:
                    df = pd.read_excel(
                        archivo_path, sheet_name=hoja, engine="openpyxl", dtype=str
                    )
                    productos.extend(
                        self._extraer_productos_de_hoja(
                            df, hoja, nombre_grupo, archivo_path.name
                        )
                    )
                except Exception:
                    pass
        except Exception:
            pass
        return productos

    def run(self):
        try:
            self.progreso.emit(5, "Iniciando clasificación según Conv. 957...")

            cursor = self.db.conn.cursor()
            grupos_bd = cursor.execute(
                """
                SELECT DISTINCT grupo FROM grupos
                WHERE grupo IS NOT NULL AND grupo != ''
                  AND grupo NOT LIKE '%SEMILLERO%'
                  AND grupo NOT LIKE '%Semillero%'
                  AND grupo NOT LIKE '%semillero%'
                ORDER BY grupo
                """
            ).fetchall()

            if not grupos_bd:
                self.error.emit("No se encontraron grupos en la base de datos")
                return

            grupos_lista = [g[0] for g in grupos_bd]
            self.progreso.emit(10, f"{len(grupos_lista)} grupos encontrados")

            if not self.reporte_excel_path.exists():
                self.error.emit(f"No se encontró la carpeta: {self.reporte_excel_path}")
                return

            carpetas_disponibles = [
                d for d in self.reporte_excel_path.iterdir() if d.is_dir()
            ]
            self.progreso.emit(15, f"{len(carpetas_disponibles)} carpetas de grupos")

            todos_productos = []
            productos_por_grupo = defaultdict(list)
            total = len(grupos_lista)

            for i, grupo_bd in enumerate(grupos_lista):
                pct = 15 + int((i / total) * 70)
                self.progreso.emit(pct, f"[{i + 1}/{total}] {grupo_bd[:50]}...")

                carpeta = self._buscar_carpeta_grupo(grupo_bd, carpetas_disponibles)
                if not carpeta or not carpeta.exists():
                    continue

                archivos = [
                    f
                    for f in list(carpeta.glob("*.xlsx")) + list(carpeta.glob("*.xls"))
                    if not f.name.startswith("~$")
                ]
                if not archivos:
                    continue

                for archivo in archivos:
                    prods = self._procesar_archivo_excel(archivo, grupo_bd)
                    productos_por_grupo[grupo_bd].extend(prods)
                    todos_productos.extend(prods)

            self.progreso.emit(90, "Calculando estadísticas...")

            stats = {
                "total_productos": len(todos_productos),
                "grupos_procesados": len([g for g in productos_por_grupo if productos_por_grupo[g]]),
                "por_categoria": defaultdict(int),
                "por_subcategoria": defaultdict(int),
                "por_producto": defaultdict(int),
                "por_año": defaultdict(int),
                "por_grupo": {g: len(p) for g, p in productos_por_grupo.items()},
            }
            for prod in todos_productos:
                stats["por_categoria"][prod["categoria_957"]] += 1
                stats["por_subcategoria"][prod["subcategoria_957"]] += 1
                stats["por_producto"][prod["producto_957"]] += 1
                stats["por_año"][prod["año"] if prod["año"] else "Sin año"] += 1

            for key in ("por_categoria", "por_subcategoria", "por_producto", "por_año"):
                stats[key] = dict(stats[key])

            self.progreso.emit(100, f"{len(todos_productos):,} productos clasificados")
            self.finalizado.emit({
                "productos": todos_productos,
                "por_grupo": dict(productos_por_grupo),
                "estadisticas": stats,
            })

        except Exception as e:
            self.error.emit(f"{e}\n\n{traceback.format_exc()}")


class VistaClasificacionMinCiencias(QWidget):
    """Vista principal para clasificación según Convocatoria 957 MinCiencias."""

    def __init__(self, db):
        super().__init__()
        self.db = db
        # "reports/excel" quedó de un scrape viejo (dic-2025) -- usar siempre
        # el más reciente de data/reporte excel_<fecha> (ver misma nota en
        # VistaSeguimientoGrupos.__init__).
        carpeta_reciente = _carpeta_gruplac_mas_reciente()
        self.reporte_excel_path = str(carpeta_reciente) if carpeta_reciente else "reports/excel"
        self.datos_analisis = None
        self.productos_mostrados = []
        self.thread_procesamiento = None
        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(8)
        layout.setContentsMargins(8, 8, 8, 8)

        # Cabecera
        header = QHBoxLayout()
        titulo = QLabel("Clasificación Conv. 957 MinCiencias")
        titulo.setFont(QFont("Arial", 14, QFont.Bold))
        titulo.setStyleSheet("color: #1a365d;")
        header.addWidget(titulo)
        header.addStretch()

        header.addWidget(QLabel("Años:"))
        self.spin_año_inicio = QSpinBox()
        self.spin_año_inicio.setRange(2000, 2030)
        self.spin_año_inicio.setValue(2022)
        header.addWidget(self.spin_año_inicio)
        header.addWidget(QLabel("a"))
        self.spin_año_fin = QSpinBox()
        self.spin_año_fin.setRange(2000, 2030)
        self.spin_año_fin.setValue(2025)
        header.addWidget(self.spin_año_fin)

        self.chk_sin_año = QCheckBox("Incluir sin año")
        self.chk_sin_año.setChecked(True)
        header.addWidget(self.chk_sin_año)

        btn_style_primary = (
            "QPushButton { background-color: #2E86AB; color: white; "
            "padding: 8px 20px; border-radius: 4px; font-weight: bold; }"
            "QPushButton:hover { background-color: #1a5276; }"
        )
        btn_style_success = (
            "QPushButton { background-color: #27ae60; color: white; "
            "padding: 8px 20px; border-radius: 4px; font-weight: bold; }"
            "QPushButton:hover { background-color: #1e8449; }"
            "QPushButton:disabled { background-color: #95a5a6; }"
        )
        self.btn_clasificar = QPushButton("Clasificar")
        self.btn_clasificar.setStyleSheet(btn_style_primary)
        self.btn_clasificar.clicked.connect(self.iniciar_clasificacion)
        header.addWidget(self.btn_clasificar)

        self.btn_exportar = QPushButton("Exportar Excel")
        self.btn_exportar.setStyleSheet(btn_style_success)
        self.btn_exportar.setEnabled(False)
        self.btn_exportar.clicked.connect(self.exportar_excel)
        header.addWidget(self.btn_exportar)

        layout.addLayout(header)

        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setStyleSheet(
            "QProgressBar { border: 2px solid #bdc3c7; border-radius: 5px; "
            "text-align: center; height: 25px; }"
            "QProgressBar::chunk { background-color: #2E86AB; }"
        )
        layout.addWidget(self.progress_bar)

        self.lbl_status = QLabel("Presione 'Clasificar' para iniciar el análisis")
        self.lbl_status.setStyleSheet("color: #7f8c8d; font-size: 11px;")
        layout.addWidget(self.lbl_status)

        # Panel principal
        splitter = QSplitter(Qt.Horizontal)

        # Árbol izquierdo
        panel_izq = QWidget()
        lay_izq = QVBoxLayout(panel_izq)
        lay_izq.setContentsMargins(0, 0, 0, 0)
        lay_izq.addWidget(QLabel("<b>Estructura Conv. 957</b>"))
        self.tree_clasificacion = QTreeWidget()
        self.tree_clasificacion.setHeaderLabels(["Categoría / Subcategoría / Producto", "Cant."])
        self.tree_clasificacion.setColumnWidth(0, 320)
        self.tree_clasificacion.setColumnWidth(1, 50)
        self.tree_clasificacion.setStyleSheet("font-size: 10px;")
        self.tree_clasificacion.itemClicked.connect(self.filtrar_por_arbol)
        lay_izq.addWidget(self.tree_clasificacion)

        resumen_box = QGroupBox("Resumen")
        resumen_box.setStyleSheet("QGroupBox { font-weight: bold; color: #1a365d; }")
        lay_res = QVBoxLayout(resumen_box)
        self.lbl_total = QLabel("Total: 0")
        self.lbl_grupos = QLabel("Grupos: 0")
        self.lbl_sin_clasificar = QLabel("Sin clasificar: 0")
        for lbl in (self.lbl_total, self.lbl_grupos, self.lbl_sin_clasificar):
            lbl.setStyleSheet("font-size: 11px;")
            lay_res.addWidget(lbl)
        lay_izq.addWidget(resumen_box)

        # Tabla derecha
        panel_der = QWidget()
        lay_der = QVBoxLayout(panel_der)
        lay_der.setContentsMargins(0, 0, 0, 0)

        filtros = QHBoxLayout()
        filtros.addWidget(QLabel("Categoría:"))
        self.combo_categoria = QComboBox()
        self.combo_categoria.addItem("-- Todas --")
        self.combo_categoria.currentTextChanged.connect(self.aplicar_filtros)
        filtros.addWidget(self.combo_categoria, 1)

        filtros.addWidget(QLabel("Grupo:"))
        self.combo_grupo = QComboBox()
        self.combo_grupo.addItem("-- Todos --")
        self.combo_grupo.currentTextChanged.connect(self.aplicar_filtros)
        filtros.addWidget(self.combo_grupo, 1)

        filtros.addWidget(QLabel("Año:"))
        self.combo_año = QComboBox()
        self.combo_año.addItem("-- Todos --")
        self.combo_año.currentTextChanged.connect(self.aplicar_filtros)
        filtros.addWidget(self.combo_año)

        btn_limpiar = QPushButton("Limpiar filtros")
        btn_limpiar.clicked.connect(self.limpiar_filtros)
        filtros.addWidget(btn_limpiar)
        lay_der.addLayout(filtros)

        self.tabla = QTableWidget()
        self.tabla.setColumnCount(7)
        self.tabla.setHorizontalHeaderLabels([
            "Categoría 957", "Subcategoría", "Producto",
            "Título", "Año", "Grupo", "Hoja Original",
        ])
        self.tabla.horizontalHeader().setStretchLastSection(True)
        self.tabla.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tabla.setSelectionBehavior(QTableWidget.SelectRows)
        self.tabla.setAlternatingRowColors(True)
        self.tabla.setStyleSheet("font-size: 10px;")
        self.tabla.itemSelectionChanged.connect(self.mostrar_detalle)
        lay_der.addWidget(self.tabla)

        self.txt_detalle = QTextEdit()
        self.txt_detalle.setReadOnly(True)
        self.txt_detalle.setMaximumHeight(120)
        self.txt_detalle.setStyleSheet("background-color: #f8f9fa; font-size: 10px;")
        lay_der.addWidget(self.txt_detalle)

        splitter.addWidget(panel_izq)
        splitter.addWidget(panel_der)
        splitter.setSizes([380, 720])
        layout.addWidget(splitter, 1)
        self.setLayout(layout)

    def iniciar_clasificacion(self):
        año_inicio = self.spin_año_inicio.value()
        año_fin = self.spin_año_fin.value()
        if año_inicio > año_fin:
            QMessageBox.warning(self, "Error", "El año de inicio debe ser ≤ al año final")
            return

        self.btn_clasificar.setEnabled(False)
        self.btn_exportar.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)

        self.thread_procesamiento = ProcesamientoClasificacionThread(
            self.db,
            self.reporte_excel_path,
            list(range(año_inicio, año_fin + 1)),
            self.chk_sin_año.isChecked(),
        )
        self.thread_procesamiento.progreso.connect(self._on_progreso)
        self.thread_procesamiento.finalizado.connect(self.mostrar_resultados)
        self.thread_procesamiento.error.connect(self._on_error)
        self.thread_procesamiento.start()

    def _on_progreso(self, valor, mensaje):
        self.progress_bar.setValue(valor)
        self.lbl_status.setText(mensaje)

    def _on_error(self, mensaje):
        QMessageBox.critical(self, "Error", mensaje)
        self.btn_clasificar.setEnabled(True)
        self.progress_bar.setVisible(False)

    def mostrar_resultados(self, resultado):
        self.datos_analisis = resultado
        productos = resultado["productos"]
        stats = resultado["estadisticas"]

        self.lbl_total.setText(f"Total: {stats['total_productos']:,}")
        self.lbl_grupos.setText(f"Grupos: {stats['grupos_procesados']:,}")
        self.lbl_sin_clasificar.setText(
            f"Sin clasificar: {stats['por_categoria'].get('Sin clasificar', 0):,}"
        )

        self._construir_arbol(productos)

        self.combo_categoria.blockSignals(True)
        self.combo_categoria.clear()
        self.combo_categoria.addItem("-- Todas --")
        for cat in sorted(stats["por_categoria"].keys(), key=get_orden_categoria):
            self.combo_categoria.addItem(cat)
        self.combo_categoria.blockSignals(False)

        self.combo_grupo.blockSignals(True)
        self.combo_grupo.clear()
        self.combo_grupo.addItem("-- Todos --")
        for grupo in sorted({p["grupo"] for p in productos}):
            self.combo_grupo.addItem(grupo[:70])
        self.combo_grupo.blockSignals(False)

        self.combo_año.blockSignals(True)
        self.combo_año.clear()
        self.combo_año.addItem("-- Todos --")
        self.combo_año.addItem("Sin año")
        for año in sorted(
            (a for a in stats["por_año"] if isinstance(a, int)), reverse=True
        ):
            self.combo_año.addItem(str(año))
        self.combo_año.blockSignals(False)

        self.productos_mostrados = productos
        self._actualizar_tabla(productos)
        self.btn_clasificar.setEnabled(True)
        self.btn_exportar.setEnabled(True)
        self.progress_bar.setVisible(False)

    def _construir_arbol(self, productos):
        self.tree_clasificacion.clear()
        conteo: dict = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
        for prod in productos:
            conteo[prod["categoria_957"]][prod["subcategoria_957"]][prod["producto_957"]] += 1

        cats_vistas: set = set()
        for nombre_prod, (cat, subcat) in CLASIF_957.items():
            if cat not in cats_vistas:
                total_cat = sum(
                    sum(conteo[cat][s].values()) for s in conteo[cat]
                )
                item_cat = QTreeWidgetItem([cat, str(total_cat)])
                color = COLORES_CATEGORIA.get(cat, "#7F8C8D")
                item_cat.setForeground(0, QBrush(QColor(color)))
                item_cat.setFont(0, QFont("Arial", 10, QFont.Bold))
                item_cat.setData(0, Qt.UserRole, {"tipo": "categoria", "valor": cat})
                self.tree_clasificacion.addTopLevelItem(item_cat)
                cats_vistas.add(cat)

                subcats_vistas: set = set()
                for np2, (c2, s2) in CLASIF_957.items():
                    if c2 != cat or s2 in subcats_vistas:
                        continue
                    total_subcat = sum(conteo[cat][s2].values())
                    item_sub = QTreeWidgetItem([f"  {s2}", str(total_subcat)])
                    item_sub.setData(
                        0, Qt.UserRole,
                        {"tipo": "subcategoria", "categoria": cat, "valor": s2},
                    )
                    item_cat.addChild(item_sub)
                    subcats_vistas.add(s2)

                    for np3, (c3, s3) in CLASIF_957.items():
                        if c3 == cat and s3 == s2:
                            cant = conteo[cat][s2][np3]
                            if cant > 0:
                                item_p = QTreeWidgetItem([f"    • {np3}", str(cant)])
                                item_p.setData(
                                    0, Qt.UserRole,
                                    {"tipo": "producto", "categoria": cat,
                                     "subcategoria": s2, "valor": np3},
                                )
                                item_sub.addChild(item_p)

                item_cat.setExpanded(True)

        if "Sin clasificar" in conteo:
            total_sin = sum(
                sum(conteo["Sin clasificar"][s].values())
                for s in conteo["Sin clasificar"]
            )
            item_sin = QTreeWidgetItem(["Sin clasificar", str(total_sin)])
            item_sin.setForeground(0, QBrush(QColor("#7F8C8D")))
            item_sin.setData(
                0, Qt.UserRole, {"tipo": "categoria", "valor": "Sin clasificar"}
            )
            self.tree_clasificacion.addTopLevelItem(item_sin)

    def filtrar_por_arbol(self, item, _column):
        data = item.data(0, Qt.UserRole)
        if not data or not self.datos_analisis:
            return
        todos = self.datos_analisis["productos"]
        tipo = data["tipo"]
        if tipo == "categoria":
            filtrado = [p for p in todos if p["categoria_957"] == data["valor"]]
        elif tipo == "subcategoria":
            filtrado = [
                p for p in todos
                if p["categoria_957"] == data["categoria"]
                and p["subcategoria_957"] == data["valor"]
            ]
        else:
            filtrado = [p for p in todos if p["producto_957"] == data["valor"]]
        self.productos_mostrados = filtrado
        self._actualizar_tabla(filtrado)
        self.lbl_status.setText(f"Mostrando {len(filtrado):,} productos")

    def aplicar_filtros(self):
        if not self.datos_analisis:
            return
        productos = self.datos_analisis["productos"]
        cat = self.combo_categoria.currentText()
        grupo = self.combo_grupo.currentText()
        año = self.combo_año.currentText()

        if cat != "-- Todas --":
            productos = [p for p in productos if p["categoria_957"] == cat]
        if grupo != "-- Todos --":
            productos = [p for p in productos if p["grupo"].startswith(grupo[:70])]
        if año == "Sin año":
            productos = [p for p in productos if p["año"] is None]
        elif año != "-- Todos --":
            try:
                año_num = int(año)
                productos = [p for p in productos if p["año"] == año_num]
            except ValueError:
                pass

        self.productos_mostrados = productos
        self._actualizar_tabla(productos)
        self.lbl_status.setText(f"Mostrando {len(productos):,} productos")

    def limpiar_filtros(self):
        self.combo_categoria.setCurrentIndex(0)
        self.combo_grupo.setCurrentIndex(0)
        self.combo_año.setCurrentIndex(0)
        if self.datos_analisis:
            self.productos_mostrados = self.datos_analisis["productos"]
            self._actualizar_tabla(self.productos_mostrados)

    def _actualizar_tabla(self, productos):
        ordenados = sorted(
            productos,
            key=lambda p: (p["orden_categoria"], p["orden_producto"], p["titulo"]),
        )
        self.tabla.setRowCount(len(ordenados))
        for row, prod in enumerate(ordenados):
            color = COLORES_CATEGORIA.get(prod["categoria_957"], "#7F8C8D")
            item_cat = QTableWidgetItem(prod["categoria_957"])
            item_cat.setForeground(QBrush(QColor(color)))
            item_cat.setFont(QFont("Arial", 9, QFont.Bold))
            self.tabla.setItem(row, 0, item_cat)
            self.tabla.setItem(row, 1, QTableWidgetItem(prod["subcategoria_957"]))
            self.tabla.setItem(row, 2, QTableWidgetItem(prod["producto_957"]))
            self.tabla.setItem(row, 3, QTableWidgetItem(prod["titulo"][:100]))
            item_año = QTableWidgetItem(str(prod["año"]) if prod["año"] else "N/A")
            item_año.setTextAlignment(Qt.AlignCenter)
            self.tabla.setItem(row, 4, item_año)
            self.tabla.setItem(row, 5, QTableWidgetItem(prod["grupo"][:50]))
            self.tabla.setItem(row, 6, QTableWidgetItem(prod["hoja_original"]))

        self.tabla.setColumnWidth(0, 200)
        self.tabla.setColumnWidth(1, 150)
        self.tabla.setColumnWidth(2, 200)
        self.tabla.horizontalHeader().setSectionResizeMode(3, QHeaderView.Stretch)
        self.tabla.setColumnWidth(4, 50)
        self.tabla.setColumnWidth(5, 150)
        self.tabla.setColumnWidth(6, 120)

    def mostrar_detalle(self):
        seleccion = self.tabla.selectedItems()
        if not seleccion or not self.productos_mostrados:
            return
        fila = seleccion[0].row()
        ordenados = sorted(
            self.productos_mostrados,
            key=lambda p: (p["orden_categoria"], p["orden_producto"], p["titulo"]),
        )
        if fila >= len(ordenados):
            return
        prod = ordenados[fila]
        self.txt_detalle.setHtml(
            f"<b style='color:#1a365d;'>TÍTULO:</b> {prod['titulo']}<br><br>"
            f"<b style='color:#2E86AB;'>CLASIFICACIÓN 957:</b><br>"
            f"&nbsp;&nbsp;• Categoría: {prod['categoria_957']}<br>"
            f"&nbsp;&nbsp;• Subcategoría: {prod['subcategoria_957']}<br>"
            f"&nbsp;&nbsp;• Producto: {prod['producto_957']}<br><br>"
            f"<b style='color:#27ae60;'>DATOS:</b><br>"
            f"&nbsp;&nbsp;• Año: {prod['año'] or 'N/A'}<br>"
            f"&nbsp;&nbsp;• Grupo: {prod['grupo']}<br>"
            f"&nbsp;&nbsp;• Hoja original: {prod['hoja_original']}<br>"
            f"&nbsp;&nbsp;• Autores: {prod['autores'][:100]}"
        )

    def exportar_excel(self):
        if not self.datos_analisis:
            QMessageBox.warning(self, "Advertencia", "No hay datos para exportar")
            return
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill

            nombre_archivo = (
                f"clasificacion_957_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            wb = openpyxl.Workbook()

            # Hoja 1: todos los productos
            ws = wb.active
            ws.title = "Productos Conv. 957"
            ws["A1"] = "CLASIFICACIÓN DE PRODUCTOS - CONVOCATORIA 957 MINCIENCIAS"
            ws["A1"].font = Font(bold=True, size=14, color="1a365d")
            ws.merge_cells("A1:H1")
            ws["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            ws["A2"].font = Font(italic=True, size=10)

            for col, h in enumerate(
                ["Categoría 957", "Subcategoría", "Producto Reconocido",
                 "Título", "Año", "Grupo", "Hoja Original", "Autores"],
                start=1,
            ):
                cell = ws.cell(row=4, column=col, value=h)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    start_color="2E86AB", end_color="2E86AB", fill_type="solid"
                )
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

            productos = self.datos_analisis["productos"]
            ordenados = sorted(
                productos,
                key=lambda p: (p["orden_categoria"], p["orden_producto"], p["titulo"]),
            )
            for row, prod in enumerate(ordenados, start=5):
                ws.cell(row=row, column=1, value=prod["categoria_957"])
                ws.cell(row=row, column=2, value=prod["subcategoria_957"])
                ws.cell(row=row, column=3, value=prod["producto_957"])
                ws.cell(row=row, column=4, value=prod["titulo"])
                ws.cell(row=row, column=5, value=prod["año"] or "")
                ws.cell(row=row, column=6, value=prod["grupo"])
                ws.cell(row=row, column=7, value=prod["hoja_original"])
                ws.cell(row=row, column=8, value=prod["autores"])

            for col_letter, ancho in zip("ABCDEFGH", [35, 25, 35, 60, 8, 40, 25, 40]):
                ws.column_dimensions[col_letter].width = ancho
            ws.auto_filter.ref = f"A4:H{4 + len(productos)}"

            # Hoja 2: resumen por categoría
            ws2 = wb.create_sheet("Resumen por Categoría")
            ws2["A1"] = "RESUMEN POR CATEGORÍA Y SUBCATEGORÍA"
            ws2["A1"].font = Font(bold=True, size=12)
            for col, h in enumerate(["Categoría", "Subcategoría", "Producto", "Cantidad"], 1):
                cell = ws2.cell(row=3, column=col, value=h)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    start_color="3B8C66", end_color="3B8C66", fill_type="solid"
                )

            conteo: dict = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
            for prod in productos:
                conteo[prod["categoria_957"]][prod["subcategoria_957"]][prod["producto_957"]] += 1

            fila = 4
            for nombre_prod, (cat, subcat) in CLASIF_957.items():
                cant = conteo[cat][subcat][nombre_prod]
                if cant > 0:
                    ws2.cell(row=fila, column=1, value=cat)
                    ws2.cell(row=fila, column=2, value=subcat)
                    ws2.cell(row=fila, column=3, value=nombre_prod)
                    ws2.cell(row=fila, column=4, value=cant)
                    fila += 1
            for subcat in conteo.get("Sin clasificar", {}):
                for pnombre, cant in conteo["Sin clasificar"][subcat].items():
                    ws2.cell(row=fila, column=1, value="Sin clasificar")
                    ws2.cell(row=fila, column=2, value=subcat)
                    ws2.cell(row=fila, column=3, value=pnombre)
                    ws2.cell(row=fila, column=4, value=cant)
                    fila += 1
            for col_letter, ancho in zip("ABCD", [40, 30, 40, 12]):
                ws2.column_dimensions[col_letter].width = ancho
            if fila > 4:
                ws2.auto_filter.ref = f"A3:D{fila - 1}"

            # Hoja 3: por grupo
            ws3 = wb.create_sheet("Por Grupo")
            ws3["A1"] = "PRODUCTOS POR GRUPO DE INVESTIGACIÓN"
            ws3["A1"].font = Font(bold=True, size=12)
            for col, h in enumerate(["Grupo", "Categoría", "Cantidad"], 1):
                cell = ws3.cell(row=3, column=col, value=h)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    start_color="A23B72", end_color="A23B72", fill_type="solid"
                )

            por_grupo: dict = defaultdict(lambda: defaultdict(int))
            for prod in productos:
                por_grupo[prod["grupo"]][prod["categoria_957"]] += 1

            fila = 4
            for grupo in sorted(por_grupo):
                for cat in sorted(por_grupo[grupo], key=get_orden_categoria):
                    ws3.cell(row=fila, column=1, value=grupo)
                    ws3.cell(row=fila, column=2, value=cat)
                    ws3.cell(row=fila, column=3, value=por_grupo[grupo][cat])
                    fila += 1
            for col_letter, ancho in zip("ABC", [50, 40, 12]):
                ws3.column_dimensions[col_letter].width = ancho
            if fila > 4:
                ws3.auto_filter.ref = f"A3:C{fila - 1}"

            wb.save(nombre_archivo)
            QMessageBox.information(
                self,
                "Éxito",
                f"Excel exportado:\n{nombre_archivo}\n\n"
                "Hojas incluidas:\n"
                "• Productos Conv. 957 (ordenados)\n"
                "• Resumen por Categoría\n"
                "• Por Grupo",
            )

        except Exception as e:
            QMessageBox.critical(
                self, "Error", f"Error al exportar:\n{e}\n\n{traceback.format_exc()}"
            )

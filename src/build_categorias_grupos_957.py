"""
build_categorias_grupos_957.py

Genera data/cache/categorias_grupos_957.json a partir de la hoja
"Datos Básicos" de cada Excel en la carpeta 'data/reporte excel_<fecha>'
MÁS RECIENTE (scrape actual de GrupLAC).

Para cada grupo extrae:
  - categoria_asignada: valor de "Clasificación" (primera línea, p.ej. "A1", "B")
  - area_conocimiento:  primer segmento de "Área de conocimiento"
                        (p.ej. "Humanidades -- Arte -- ..." → "Humanidades")

Sirve como respaldo de medicion_957.xlsx para grupos que no aparecen en el
documento oficial de medición (que solo cubre 75 de los 125 grupos) y como
fuente única para "Estadísticas 957" (src/estadisticas_957.py), que solo
necesita la categoría VIGENTE reportada por GrupLAC, no el cuartil oficial.

Uso:
    python src/build_categorias_grupos_957.py
"""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl

from utils import obtener_directorio_base
from views.vista_seguimiento_grupos import _carpeta_gruplac_mas_reciente


def _valor_datos_basicos(ws, claves) -> str | None:
    """Busca en la hoja 'Datos Básicos' una fila cuya primera celda
    contenga alguna de las claves dadas (case-insensitive) y devuelve
    el valor de la segunda celda."""
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        etiqueta = str(row[0])
        if any(clave.lower() in etiqueta.lower() for clave in claves):
            return row[1] if len(row) > 1 else None
    return None


def _hoja_datos_basicos(wb):
    for nombre in wb.sheetnames:
        if "datos" in nombre.lower():
            return wb[nombre]
    return None


def construir_cache(ruta_base: Path | None = None) -> dict:
    if ruta_base is None:
        ruta_base = _carpeta_gruplac_mas_reciente()
    if ruta_base is None:
        return {}

    grupos = {}
    for carpeta in sorted(ruta_base.iterdir()):
        if not carpeta.is_dir():
            continue
        archivos = list(carpeta.glob("*.xlsx"))
        if not archivos:
            continue

        try:
            wb = openpyxl.load_workbook(archivos[0], read_only=True, data_only=True)
            ws = _hoja_datos_basicos(wb)
            if ws is None:
                continue

            clasif = _valor_datos_basicos(ws, ["Clasificaci"])
            area   = _valor_datos_basicos(ws, ["rea de conocimiento"])
            wb.close()
        except Exception:
            continue

        categoria = None
        if clasif:
            categoria = str(clasif).split("\n")[0].strip()

        area_principal = None
        if area:
            area_principal = str(area).split("--")[0].strip()

        if categoria or area_principal:
            grupos[carpeta.name] = {
                "categoria_asignada": categoria,
                "area_conocimiento": area_principal,
            }

    return grupos


def main():
    grupos = construir_cache()
    cache_dir = obtener_directorio_base() / "data" / "cache"
    cache_dir.mkdir(parents=True, exist_ok=True)
    out_path = cache_dir / "categorias_grupos_957.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump({"grupos": grupos}, f, ensure_ascii=False, indent=2)
    print(f"{len(grupos)} grupos escritos en {out_path}")


if __name__ == "__main__":
    main()

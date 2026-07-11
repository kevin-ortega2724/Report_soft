"""
Comparador de productos contra gruplac_957.db.
Toma los productos de Supervision_plano.xlsx y verifica si están
registrados en gruplac_957.db, clasificándolos en:
  - Confirmado en BD (mismo grupo)
  - Registrado en otro grupo
  - Segundo barrido (revisión manual)
  - Faltante real
"""

import io
import re
import sqlite3
import unicodedata
import zipfile
from difflib import SequenceMatcher
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

THRESHOLD_CONFIRMED = 0.82
THRESHOLD_REVIEW = 0.62
MIN_COMMON_TOKENS = 2

POSSIBLE_PRODUCT_COLS = [
    "producto", "nombre_producto", "titulo",
    "title", "nombre", "descripcion", "producto_nombre",
]

STOPWORDS = {
    "de", "la", "el", "y", "en", "con", "para", "una", "un",
    "del", "los", "las", "por", "sobre", "al", "se", "que",
    "o", "a", "e", "su", "sus", "es", "son", "lo", "le",
    "nos", "si", "mas", "pero", "como", "cuando", "donde",
}

_NOTA_SUF = re.compile(
    r"\s*\((?:registrado|distinci[oó]n|libro|cap[ií]tulo|[eé]nfasis)[^)]*\)\s*$",
    re.IGNORECASE,
)
_ELLIPSIS = re.compile(r"\.{2,}\s*$")


def normalize_text(text: str) -> str:
    if text is None:
        return ""
    text = str(text).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^\w\s]", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def tokenize(text: str) -> set:
    return {w for w in normalize_text(text).split()
            if len(w) > 2 and w not in STOPWORDS}


def clean_supervision_title(title: str) -> str:
    t = str(title).strip()
    t = _NOTA_SUF.sub("", t)
    t = _ELLIPSIS.sub("", t)
    return t.strip().rstrip(".")


def combined_score(a_norm, b_norm, a_tok, b_tok):
    """Devuelve (score_combinado, seq, jaccard, tokens_comunes)."""
    seq = SequenceMatcher(None, a_norm, b_norm).ratio()
    jac = (len(a_tok & b_tok) / len(a_tok | b_tok) if (a_tok and b_tok) else 0.0)
    return 0.65 * seq + 0.35 * jac, seq, jac, len(a_tok & b_tok)


def _find_col(columns, candidates):
    norm_map = {c["name"].strip().lower(): c["name"] for c in columns}
    for cand in candidates:
        if cand.strip().lower() in norm_map:
            return norm_map[cand.strip().lower()]
    return None


class ComparadorFaltantes:
    """Compara productos de Supervision_plano.xlsx o faltantes.zip contra gruplac_957.db."""

    def __init__(self, supervision_path: str | None, gruplac_db_path: str):
        self.supervision_path = Path(supervision_path) if supervision_path else None
        self.db_path = Path(gruplac_db_path)
        self.db_index = []
        self.groups_index = {}

    # ── Indexar gruplac_957.db ────────────────────────────────────────

    def build_db_index(self):
        """Indexa todos los productos de gruplac_957.db (productos_957 + otras tablas)."""
        conn = sqlite3.connect(str(self.db_path))
        conn.row_factory = sqlite3.Row
        index = []

        # Cargar grupos
        grupos_rows = conn.execute("SELECT id, nombre FROM grupos").fetchall()
        self.groups_index = {r["id"]: r["nombre"] for r in grupos_rows}

        # Tabla principal con grupo_id
        try:
            rows = conn.execute(
                "SELECT grupo_id, grupo, titulo FROM productos_957"
            ).fetchall()
            for row in rows:
                pt = str(row["titulo"] or "").strip()
                if len(pt) < 5:
                    continue
                index.append({
                    "table": "productos_957",
                    "product_value": pt,
                    "product_norm": normalize_text(pt),
                    "product_tokens": tokenize(pt),
                    "group_value": str(row["grupo"] or "").strip(),
                    "group_norm": normalize_text(str(row["grupo"] or "")),
                    "grupo_id": row["grupo_id"],
                })
        except Exception:
            pass

        # Otras tablas (fallback)
        for rt in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' "
            "AND name NOT LIKE 'sqlite_%' AND name != 'productos_957' "
            "ORDER BY name"
        ).fetchall():
            tname = rt["name"]
            cols = [dict(r) for r in
                    conn.execute(f"PRAGMA table_info('{tname}')").fetchall()]
            pc = _find_col(cols, POSSIBLE_PRODUCT_COLS)
            if not pc:
                continue
            try:
                product_rows = conn.execute(
                    f'SELECT "{pc}" AS pv FROM "{tname}"'
                ).fetchall()
            except Exception:
                continue
            for row in product_rows:
                pt = str(row["pv"] or "").strip()
                if len(pt) < 5:
                    continue
                index.append({
                    "table": tname,
                    "product_value": pt,
                    "product_norm": normalize_text(pt),
                    "product_tokens": tokenize(pt),
                    "group_value": "",
                    "group_norm": "",
                    "grupo_id": None,
                })

        conn.close()
        self.db_index = index
        # Índice agrupado por grupo_id para búsqueda eficiente
        self.db_index_by_group = {}
        for rec in index:
            gid = rec.get("grupo_id")
            self.db_index_by_group.setdefault(gid, []).append(rec)
        return index

    # ── Cargar Supervision_plano.xlsx ─────────────────────────────────

    def load_supervision_data(self) -> pd.DataFrame:
        """Carga Supervision_plano.xlsx y normaliza grupos."""
        df = pd.read_excel(str(self.supervision_path))
        df["_grupo_norm"] = df["grupo"].apply(
            lambda g: normalize_text(str(g)) if pd.notna(g) else ""
        )
        df["_producto_limpio"] = df["producto"].apply(
            lambda p: clean_supervision_title(str(p)) if pd.notna(p) else ""
        )
        return df

    # ── Cargar desde faltantes.zip ────────────────────────────────────

    @staticmethod
    def load_from_zip(zip_path: str) -> pd.DataFrame:
        """
        Carga productos de todos los grupos desde faltantes.zip.
        Lee data/archivogrupos/[NN. GRUPO]/reporte*.xlsx y extrae títulos
        de las hojas Publicaciones, Extensiones, Trabajos de Grado y Proyectos.
        """
        _PREFIX = re.compile(r"^\d+\.\s*")
        # hoja → (palabra clave columna de título, etiqueta categoría)
        SHEETS = {
            "Publicaciones":     ("titulo",    "Publicaciones"),
            "Extensiones":       ("actividad", "Extensiones"),
            "Trabajos de Grado": ("titulo",    "Trabajos de Grado"),
            "Proyectos":         ("titulo",    "Proyectos"),
        }

        rows = []
        with zipfile.ZipFile(str(zip_path)) as z:
            # Solo data/archivogrupos/ (excluir data1/)
            xlsx_files = [
                f for f in z.namelist()
                if "/data/archivogrupos/" in f
                and "/data1/" not in f
                and f.endswith(".xlsx")
            ]

            for zip_entry in xlsx_files:
                parts = zip_entry.split("/")
                if len(parts) < 4:
                    continue
                # Estructura: faltantes/data/archivogrupos/NN. GRUPO/archivo.xlsx
                folder_name = parts[-2]
                grupo = _PREFIX.sub("", folder_name).strip()
                if not grupo:
                    continue

                try:
                    with z.open(zip_entry) as fp:
                        xl = pd.ExcelFile(io.BytesIO(fp.read()))
                        for sheet_name, (col_kw, categoria) in SHEETS.items():
                            if sheet_name not in xl.sheet_names:
                                continue
                            try:
                                df_sh = xl.parse(sheet_name, header=1)
                            except Exception:
                                continue
                            df_sh.columns = [str(c).strip() for c in df_sh.columns]

                            # Buscar columna de título por palabra clave normalizada
                            title_col = None
                            for c in df_sh.columns:
                                if col_kw in normalize_text(c):
                                    title_col = c
                                    break
                            if title_col is None:
                                continue

                            for val in df_sh[title_col].dropna():
                                producto = clean_supervision_title(str(val).strip())
                                if len(producto) >= 5:
                                    rows.append({
                                        "grupo": grupo,
                                        "producto": producto,
                                        "categoria": categoria,
                                        "hoja": sheet_name,
                                        "_grupo_norm": normalize_text(grupo),
                                        "_producto_limpio": producto,
                                    })
                except Exception:
                    continue

        if not rows:
            return pd.DataFrame(
                columns=["grupo", "producto", "categoria", "hoja",
                         "_grupo_norm", "_producto_limpio"]
            )
        return pd.DataFrame(rows).drop_duplicates(subset=["_grupo_norm", "_producto_limpio"])

    # ── Búsqueda de un producto en el índice ─────────────────────────

    def _search_in_records(self, pn: str, pt: set, records: list,
                            grupo_id) -> dict | None:
        """Busca en una lista específica de registros. Retorna resultado o None."""
        exact_other, exact_same = [], []
        best: tuple | None = None

        for rec in records:
            dn = rec["product_norm"]
            dt = rec["product_tokens"]
            if not dn:
                continue
            same = (rec.get("grupo_id") is not None and rec["grupo_id"] == grupo_id)

            # Coincidencia exacta
            if pn == dn:
                (exact_same if same else exact_other).append(rec)
                continue

            # Contención
            lp, ld = len(pn), len(dn)
            if min(lp, ld) >= 10 and (pn in dn or dn in pn):
                if min(lp, ld) / max(lp, ld) >= 0.70:
                    if same:
                        exact_same.append(rec)
                    else:
                        return self._make_result(
                            "Registrado en otro grupo", rec, 0.95,
                            "Coincidencia por contención"
                        )
                    continue

            # Filtro barato de tokens en común ANTES del score costoso: la
            # intersección de conjuntos es O(tokens), mientras que
            # combined_score hace un SequenceMatcher caracter a caracter --
            # descartar acá el 90%+ de pares que no comparten ni una palabra
            # es lo que hace viable comparar contra miles de títulos.
            min_tok = MIN_COMMON_TOKENS if len(pt) >= 4 else 1
            if len(pt & dt) < min_tok:
                continue

            # Score combinado (SequenceMatcher + Jaccard)
            comb, seq, jac, common = combined_score(pn, dn, pt, dt)
            if best is None or comb > best[0]:
                best = (comb, common, rec, same)

        if exact_other:
            groups = sorted({r["group_value"] for r in exact_other if r["group_value"]})
            return self._make_result(
                "Registrado en otro grupo", exact_other[0], 1.0,
                "Coincidencia exacta en otro grupo", grupos_extra=groups
            )
        if exact_same:
            return self._make_result(
                "Confirmado en BD (mismo grupo)", exact_same[0], 1.0,
                "Coincidencia exacta en el mismo grupo"
            )
        if best:
            comb, common, br, same = best
            det = f"comb={comb:.4f} | tokens={common}"
            if comb >= THRESHOLD_CONFIRMED:
                est = ("Confirmado en BD (mismo grupo)" if same
                       else "Registrado en otro grupo")
                return self._make_result(est, br, comb, det)
            if comb >= THRESHOLD_REVIEW:
                est = ("Segundo barrido - mismo grupo" if same
                       else "Segundo barrido - otro grupo")
                return self._make_result(est, br, comb, det)

        return None

    def search_product(self, product_name: str, grupo_id: int) -> dict:
        """
        Busca un producto en el índice.
        Primero busca dentro del mismo grupo (rápido).
        Si no se encuentra, busca en todos los grupos (lento).
        """
        pn = normalize_text(product_name)
        pt = tokenize(product_name)

        # Fase 1: buscar solo en el mismo grupo (≈ 500–1000 registros)
        if grupo_id is not None:
            group_records = self.db_index_by_group.get(grupo_id)
            if group_records:
                result = self._search_in_records(pn, pt, group_records, grupo_id)
                if result:
                    return result

        # Fase 2: buscar en otros grupos (solo si no se encontró en fase 1)
        other_records = []
        for gid, recs in self.db_index_by_group.items():
            if gid != grupo_id:
                other_records.extend(recs)

        # También incluir registros sin grupo_id
        if self.db_index_by_group.get(None):
            other_records.extend(self.db_index_by_group[None])

        if other_records:
            result = self._search_in_records(pn, pt, other_records, grupo_id)
            if result:
                return result

        # No encontrado
        return {
            "estado_verificacion": "Faltante real",
            "grupo_encontrado": "",
            "tabla_origen": "",
            "texto_coincidente": "",
            "similitud": 0.0,
            "detalle_verificacion": "No se encontró en la BD",
            "necesita_revision": False,
            "es_faltante": True,
        }

    @staticmethod
    def _make_result(estado, rec, score, detalle, grupos_extra=None):
        grupos = (" | ".join(grupos_extra) if grupos_extra
                  else (rec.get("group_value", "") if rec else ""))
        return {
            "estado_verificacion": estado,
            "grupo_encontrado": grupos,
            "tabla_origen": rec.get("table", "") if rec else "",
            "texto_coincidente": rec.get("product_value", "") if rec else "",
            "similitud": round(score, 4),
            "detalle_verificacion": detalle,
            "necesita_revision": "Segundo barrido" in estado,
            "es_faltante": estado == "Faltante real",
        }

    # ── Comparar todos los grupos ─────────────────────────────────────

    def compare_all_groups(self, progress_callback=None,
                           df_supervision: pd.DataFrame | None = None) -> pd.DataFrame:
        """
        Compara productos contra gruplac_957.db.
        Si df_supervision es None, carga desde Supervision_plano.xlsx.
        Si se pasa df_supervision (cargado desde ZIP), lo usa directamente.
        """
        self.build_db_index()
        if df_supervision is not None:
            df_sup = df_supervision
        else:
            df_sup = self.load_supervision_data()

        # Construir mapping grupo_norm -> grupo_id desde gruplac_957.db
        norm_to_id = {}
        conn = sqlite3.connect(str(self.db_path))
        conn.row_factory = sqlite3.Row
        for row in conn.execute("SELECT id, nombre FROM grupos").fetchall():
            key = normalize_text(row["nombre"])
            norm_to_id[key] = row["id"]
        conn.close()

        results = []
        total = len(df_sup)
        for idx, (_, row) in enumerate(df_sup.iterrows()):
            grupo_norm = row["_grupo_norm"]
            producto = row["_producto_limpio"]
            if not producto:
                continue

            grupo_id = norm_to_id.get(grupo_norm)
            if grupo_id is None:
                # Buscar fuzzy en norm_to_id
                best_id, best_score = None, 0
                for key, gid in norm_to_id.items():
                    score = SequenceMatcher(None, grupo_norm, key).ratio()
                    if score > best_score and score > 0.6:
                        best_score = score
                        best_id = gid
                grupo_id = best_id

            result = self.search_product(producto, grupo_id)

            results.append({
                "grupo_original": row.get("grupo", ""),
                "producto": producto,
                "categoria": row.get("categoria", ""),
                "hoja": row.get("hoja", ""),
                "grupo_id_match": grupo_id,
                **result,
            })

            if progress_callback and (idx + 1) % 10 == 0:
                progress_callback(idx + 1, total)

        return pd.DataFrame(results)

    # ── Generar reporte Excel (formato original verificar_faltantes.py) ──

    @staticmethod
    def generar_reporte_excel(df_resultados, ruta_salida):
        """
        Genera un Excel con una sola hoja 'Faltantes Detalle': todos los
        productos faltantes o en revisión (segundo barrido).
        """
        wb = Workbook()
        AZUL_MED = "2E75B6"
        BLANCO = "FFFFFF"

        def borde_fino():
            s = Side(style="thin", color="AAAAAA")
            return Border(left=s, right=s, top=s, bottom=s)

        def estilo_encabezado(celda, bg=AZUL_MED, fg=BLANCO, negrita=True, tamanio=11):
            celda.fill = PatternFill("solid", fgColor=bg)
            celda.font = Font(name="Calibri", bold=negrita, color=fg, size=tamanio)
            celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            celda.border = borde_fino()

        ws_det = wb.active
        ws_det.title = "Faltantes Detalle"
        ws_det.sheet_view.showGridLines = False
        anchos = [55, 28, 20, 20, 14, 12]
        cols_det = ["Producto", "Grupo", "Categoría", "Hoja", "Estado", "Similitud"]
        for i, (col, ancho) in enumerate(zip(cols_det, anchos), 1):
            ws_det.column_dimensions[get_column_letter(i)].width = ancho
            c = ws_det.cell(row=1, column=i, value=col)
            estilo_encabezado(c)
        ws_det.row_dimensions[1].height = 24

        ws_det.merge_cells("A2:F2")
        c = ws_det["A2"]
        c.value = "⚠  Estos productos deben ser registrados en GrupLAC"
        c.fill = PatternFill("solid", fgColor="FFF2CC")
        c.font = Font(name="Calibri", bold=True, color="7F6000", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws_det.row_dimensions[2].height = 20

        faltantes_df = df_resultados[
            df_resultados["es_faltante"].fillna(False) | df_resultados["necesita_revision"].fillna(False)
        ].copy()
        fila_det = 3
        for _, row in faltantes_df.sort_values(["grupo_original", "producto"]).iterrows():
            score = row.get("similitud", 0) or 0
            bg = "FFE5E5" if row.get("es_faltante", False) else "FFF2CC"
            vals = [
                str(row.get("producto", "")),
                str(row.get("grupo_original", "")),
                str(row.get("categoria", "")),
                str(row.get("hoja", "")),
                str(row.get("estado_verificacion", "")),
                f"{score:.0%}" if isinstance(score, (int, float)) and score > 0 else "",
            ]
            for col, val in enumerate(vals, 1):
                c = ws_det.cell(row=fila_det, column=col, value=val)
                c.fill = PatternFill("solid", fgColor=bg)
                c.font = Font(name="Calibri", size=9)
                c.alignment = Alignment(vertical="center", wrap_text=True,
                                        horizontal="left" if col in (1, 2, 3, 4, 5) else "center")
                c.border = borde_fino()
            ws_det.row_dimensions[fila_det].height = 30
            fila_det += 1

        wb.save(str(ruta_salida))
        return ruta_salida
